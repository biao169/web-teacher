from __future__ import annotations

import csv
import hashlib
import io
import json
import time
import zipfile
from pathlib import Path
from typing import Any

from .models import SCHEMA_VERSION, TABLES
from .repository import Query, Repository


EXPORT_ROW_LIMIT = 100000
EXPORT_EXCLUDED_TABLES: set[str] = set()


def exportable_tables():
    return tuple(table for table in TABLES if table.name not in EXPORT_EXCLUDED_TABLES)


def export_bundle(repo: Repository, output: str | Path, source_platform: str = "ubuntu", site_url: str = "") -> Path:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    files: dict[str, bytes] = {}
    table_counts: dict[str, int] = {}

    for table in exportable_tables():
        rows = repo.list(table.name, Query(limit=EXPORT_ROW_LIMIT))
        table_counts[table.name] = len(rows)
        files[f"content/{table.name}.json"] = json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")
        files[f"tabular/{table.name}.csv"] = csv_bytes(rows, table.field_names)
        if table.name in {"site_settings", "navigation_items"}:
            files[f"yaml/{table.name}.yaml"] = yaml_bytes(rows)

    files["media/media_manifest.json"] = files.get("content/media_assets.json", b"[]")
    media_rows = json.loads(files["media/media_manifest.json"].decode("utf-8"))
    files["media/media_manifest.csv"] = csv_bytes(media_rows, ["object_key", "title", "category", "mime_type", "size", "checksum"])

    excel_note = ""
    try:
        files["tabular/all_tables.xlsx"] = excel_bytes(repo)
    except Exception as exc:  # pragma: no cover - depends on optional openpyxl
        excel_note = f"Excel export skipped: {exc}"

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "exported_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "source_platform": source_platform,
        "site_url": site_url,
        "table_counts": table_counts,
        "media_count": len(media_rows),
        "media_total_bytes": sum(int(row.get("size") or 0) for row in media_rows),
        "excel_note": excel_note,
    }
    checksums = {name: hashlib.sha256(content).hexdigest() for name, content in files.items()}
    manifest["checksum"] = checksums
    files["manifest.json"] = json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8")

    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in sorted(files.items()):
            archive.writestr(name, content)
    return output_path


def export_json(repo: Repository) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "tables": {table.name: repo.list(table.name, Query(limit=EXPORT_ROW_LIMIT)) for table in exportable_tables()},
    }


def csv_bytes(rows: list[dict[str, Any]], fields: list[str]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue().encode("utf-8-sig")


def yaml_bytes(rows: list[dict[str, Any]]) -> bytes:
    try:
        import yaml

        return yaml.safe_dump(rows, allow_unicode=True, sort_keys=False).encode("utf-8")
    except Exception:
        lines: list[str] = []
        for row in rows:
            lines.append("-")
            for key, value in row.items():
                scalar = "" if value is None else str(value).replace("\n", "\\n")
                lines.append(f"  {key}: {json.dumps(scalar, ensure_ascii=False)}")
        return ("\n".join(lines) + "\n").encode("utf-8")


def excel_bytes(repo: Repository) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook(write_only=True)
    for table in exportable_tables():
        sheet = workbook.create_sheet(title=table.name[:31])
        fields = table.field_names
        sheet.append(fields)
        for row in repo.list(table.name, Query(limit=EXPORT_ROW_LIMIT)):
            sheet.append([row.get(field, "") for field in fields])
    default = workbook["Sheet"] if "Sheet" in workbook.sheetnames else None
    if default:
        workbook.remove(default)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()

