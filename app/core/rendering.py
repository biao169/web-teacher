from __future__ import annotations

import csv
import io
import json
import base64
import hashlib
import mimetypes
import os
import re
import shutil
import time
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.error import URLError
from urllib.parse import parse_qs, quote, urlencode, urlparse
from urllib.request import Request, urlopen

from .exporting import csv_bytes, excel_bytes, export_json
from .media import image_tag, media_storage_kind, media_url, r2_preferred_key
from .models import TABLES, TABLE_MAP, Table, int_value
from .repository import Query, Repository
from .security import (
    esc,
    hash_password,
    parse_cookie_header,
    read_signed_session,
    render_limited_html,
    render_plain_or_limited_html,
    safe_href,
    safe_slug,
    same_origin_post_allowed,
    signed_session_token,
    stable_uid,
    text_only,
    verify_password,
)


ResponseTuple = tuple[int, list[tuple[str, str]], bytes]
MEDIA_STATS_CACHE_TTL_SECONDS = 300
MEDIA_STATS_CACHE_PATH = Path(".cache") / "media_stats.json"
MEDIA_SCAN_EXTENSIONS = {
    ".apng", ".avif", ".bmp", ".gif", ".ico", ".jpeg", ".jpg", ".png", ".svg", ".tif", ".tiff", ".webp",
    ".avi", ".m4v", ".mov", ".mp4", ".mpeg", ".mpg", ".ogv", ".webm",
    ".aac", ".flac", ".m4a", ".mp3", ".ogg", ".wav",
    ".csv", ".doc", ".docx", ".json", ".md", ".pdf", ".ppt", ".pptx", ".txt", ".xls", ".xlsx", ".yaml", ".yml",
}
ASSET_VERSION = "20260822-media-confirm"
I18N_DICTIONARY_FILENAME = "i18n_dictionary.json"
I18N_DICTIONARY_R2_KEY = "i18n/i18n_dictionary.json"
I18N_DICTIONARY_CACHE: dict[str, Any] = {"path": "", "mtime": -1.0, "data": None}
LOGIN_RATE_LIMIT_CACHE_PATH = Path(".cache") / "login_rate_limits.json"
LOGIN_RATE_LIMIT_WINDOW_SECONDS = 15 * 60
LOGIN_RATE_LIMIT_LOCK_SECONDS = 15 * 60
LOGIN_RATE_LIMIT_MAX_FAILURES = 5
AUTH_SECRET_MIN_LENGTH = 32
PUBLICATION_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
PROFILE_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
PROJECT_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
PATENT_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
STUDENT_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
NEWS_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
COURSE_SUGGESTION_CACHE: dict[str, Any] = {"ts": 0.0, "ttl": 30, "payload": None}
TRANSLATION_FALLBACK_PROVIDERS = "auto,mymemory,argos_local"
MICROSOFT_TRANSLATOR_DEFAULT_ENDPOINT = "https://api.cognitive.microsofttranslator.com"
TRANSLATION_BUNDLE_MAX_ITEMS = 6
TRANSLATION_BUNDLE_MAX_CHARS = 6000
TRANSLATION_BUNDLE_MARKER = "TS_FIELD"
AUTH_COOKIE_NAME = "teacher_site_auth"
AUTH_SESSION_SECONDS = 60 * 60 * 8
AUTH_TABLES = {"auth_roles", "auth_users", "auth_permissions"}
AUTH_MODULE_LABELS = {
    "admin": "后台入口",
    "export": "导入与导出",
    "media_tools": "媒体工具",
}

AUTH_ROLE_DEFAULTS = (
    {
        "uid": "role-super-admin",
        "name": "高级管理员",
        "level": 100,
        "description": "拥有站点、权限、内容、媒体、导出等全部后台权限。",
        "visibility_scopes": "public,authenticated,staff,owner",
        "is_system": 1,
        "is_active": 1,
        "sort_order": 10,
    },
    {
        "uid": "role-admin",
        "name": "普通管理员",
        "level": 80,
        "description": "默认可维护主要内容与媒体，不默认管理账号权限。",
        "visibility_scopes": "public,authenticated,staff",
        "is_system": 1,
        "is_active": 1,
        "sort_order": 20,
    },
    {
        "uid": "role-staff",
        "name": "员工",
        "level": 40,
        "description": "默认可查看后台并维护动态、留言、学生等日常内容。",
        "visibility_scopes": "public,authenticated,staff",
        "is_system": 1,
        "is_active": 1,
        "sort_order": 30,
    },
    {
        "uid": "role-visitor",
        "name": "访客用户",
        "level": 10,
        "description": "用于登录后查看受限前台内容，默认不能进入后台。",
        "visibility_scopes": "public,authenticated",
        "is_system": 1,
        "is_active": 1,
        "sort_order": 40,
    },
)

CONTENT_ADMIN_TABLES = {
    "navigation_items",
    "profiles",
    "research_interests",
    "publications",
    "projects",
    "patents",
    "students",
    "student_category_displays",
    "news",
    "courses",
    "messages",
    "media_assets",
    "translation_cache",
    "autofetch_logs",
}
EXPORT_MAIN_TABLES = (
    "site_settings",
    "global_settings",
    "navigation_items",
    "profiles",
    "research_interests",
    "publications",
    "projects",
    "patents",
    "students",
    "student_category_displays",
    "news",
    "courses",
    "messages",
    "media_assets",
    "translation_cache",
)
EXPORT_TABLE_GROUPS = (
    ("site", "站点与导航", ("site_settings", "global_settings", "navigation_items")),
    ("people", "教师与学生", ("profiles", "students", "student_category_displays")),
    ("research", "科研成果", ("research_interests", "publications", "projects", "patents")),
    ("content", "动态课程留言", ("news", "courses", "messages")),
    ("media", "媒体与翻译", ("media_assets", "translation_cache")),
)

ENGLISH_FIELD_OVERRIDES = {
    ("site_settings", "site_name"): "site_name_en",
    ("navigation_items", "title"): "title_en",
    ("profiles", "name"): "name_en",
    ("profiles", "bio"): "bio_en",
    ("research_interests", "name"): "name_en",
    ("students", "name"): "name_en",
    ("student_category_displays", "label"): "label_en",
}

FRONTEND_TRANSLATION_FIELDS = {
    "site_settings": ("site_name", "hero_title", "hero_subtitle", "seo_title", "seo_description", "footer_text"),
    "navigation_items": ("title",),
    "profiles": ("name", "role", "title", "organization", "lab", "office", "bio", "education", "experience", "recruiting"),
    "research_interests": ("name", "description"),
    "publications": ("title", "source_citation", "authors", "venue", "publication_type", "author_role", "index_type", "display_tags"),
    "projects": ("name", "source", "fund_name", "principal", "members", "status", "summary"),
    "patents": ("name", "country", "patent_type", "inventors", "owner", "legal_status", "summary"),
    "students": ("name", "degree", "category", "grade", "direction", "status", "destination", "awards", "bio"),
    "student_category_displays": ("label",),
    "news": ("title", "category"),
    "courses": ("name", "semester", "audience", "summary"),
}

FRONTEND_TABLE_URLS = {
    "site_settings": "/",
    "navigation_items": "/",
    "profiles": "/team/{uid}",
    "research_interests": "/",
    "publications": "/publications",
    "projects": "/projects",
    "patents": "/patents",
    "students": "/students",
    "student_category_displays": "/students",
    "news": "/news/{slug}",
    "courses": "/courses",
    "media_assets": "/",
    "messages": "/contact",
}


def admin_modules() -> list[str]:
    return ["admin", "export", "media_tools", *[table.name for table in TABLES]]


def auth_secret(repo: Repository, env: dict[str, str]) -> str:
    configured = str(env.get("TEACHER_SITE_AUTH_SECRET") or "").strip()
    if configured:
        return configured
    site = active_site(repo)
    seed = str(site.get("uid") or site.get("site_name") or "teacher-site-dev-secret")
    return hashlib.sha256(f"teacher-site:{seed}:local-dev".encode("utf-8")).hexdigest()


def production_auth_secret_required(env: dict[str, str]) -> bool:
    if truthy(env.get("TEACHER_SITE_REQUIRE_AUTH_SECRET"), default=False):
        return True
    platform = str(env.get("PLATFORM") or "").strip().lower()
    if platform == "cloudflare":
        return True
    host = str(env.get("_HOST") or "").split(",", 1)[0].strip().lower()
    site_url = str(env.get("SITE_URL") or "").strip().lower()
    local_hosts = ("127.0.0.1", "localhost", "[::1]", "::1")
    is_local_host = any(host.startswith(item) for item in local_hosts)
    is_local_url = any(site_url.startswith(f"http://{item}") for item in ("127.0.0.1", "localhost"))
    is_https = str(env.get("_SCHEME") or "").lower() == "https" or site_url.startswith("https://")
    return is_https and not is_local_host and not is_local_url


def auth_secret_config_issue(env: dict[str, str]) -> str:
    if not production_auth_secret_required(env):
        return ""
    configured = str(env.get("TEACHER_SITE_AUTH_SECRET") or "").strip()
    if not configured:
        return "生产环境必须配置 TEACHER_SITE_AUTH_SECRET，否则登录会话无法安全签名。"
    if len(configured) < AUTH_SECRET_MIN_LENGTH:
        return f"TEACHER_SITE_AUTH_SECRET 至少需要 {AUTH_SECRET_MIN_LENGTH} 个字符，请使用随机生成的强密钥。"
    weak_values = {"teacher-site-dev-secret", "changeme", "change-me", "password", "secret"}
    if configured.lower() in weak_values:
        return "TEACHER_SITE_AUTH_SECRET 不能使用默认或弱口令。"
    return ""


def auth_config_response(env: dict[str, str], api: bool = False) -> ResponseTuple | None:
    issue = auth_secret_config_issue(env)
    if not issue:
        return None
    if api:
        return json_response({"ok": False, "message": issue}, 503)
    body = f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>生产密钥未配置</title><link rel="stylesheet" href="/assets/site.css?v={ASSET_VERSION}"></head>
<body><main class="compact-page"><section class="notice"><h1>生产密钥未配置</h1><p>{esc(issue)}</p><p>请在运行环境中配置强随机 <code>TEACHER_SITE_AUTH_SECRET</code> 后重启网站。</p></section></main></body></html>"""
    return html_response(body, 503)


def ensure_auth_defaults(repo: Repository) -> None:
    for role in AUTH_ROLE_DEFAULTS:
        existing = repo.get("auth_roles", role["uid"])
        if not existing:
            repo.save("auth_roles", role)
    roles = {str(row.get("uid") or ""): row for row in repo.list("auth_roles", Query(limit=1000))}
    modules = admin_modules()
    for role_uid, role in roles.items():
        if not role_uid:
            continue
        level = int_value(role.get("level"), 0)
        for index, module in enumerate(modules, 1):
            uid = stable_uid("perm", f"{role_uid}:{module}")
            if repo.get("auth_permissions", uid):
                continue
            if level >= 100:
                flags = {"can_view": 1, "can_create": 1, "can_edit": 1, "can_delete": 1, "can_export": 1}
            elif level >= 80:
                can_manage = module in CONTENT_ADMIN_TABLES or module in {"admin", "export", "media_tools"}
                flags = {
                    "can_view": 1 if can_manage else 0,
                    "can_create": 1 if module in CONTENT_ADMIN_TABLES else 0,
                    "can_edit": 1 if module in CONTENT_ADMIN_TABLES else 0,
                    "can_delete": 1 if module in CONTENT_ADMIN_TABLES - {"translation_cache", "autofetch_logs"} else 0,
                    "can_export": 1 if module == "export" else 0,
                }
            elif level >= 40:
                staff_tables = {"news", "messages", "students", "student_category_displays", "media_assets"}
                flags = {
                    "can_view": 1 if module in staff_tables or module in {"admin", "media_tools"} else 0,
                    "can_create": 1 if module in {"news", "media_assets"} else 0,
                    "can_edit": 1 if module in staff_tables else 0,
                    "can_delete": 0,
                    "can_export": 0,
                }
            else:
                flags = {"can_view": 0, "can_create": 0, "can_edit": 0, "can_delete": 0, "can_export": 0}
            repo.save("auth_permissions", {"uid": uid, "role_uid": role_uid, "module": module, "sort_order": index, **flags})


def auth_users_exist(repo: Repository) -> bool:
    return any(str(row.get("status") or "active") == "active" for row in repo.list("auth_users", Query(limit=1000)))


def current_auth(repo: Repository, env: dict[str, str]) -> dict[str, Any]:
    cached = env.get("_AUTH_USER")
    if isinstance(cached, dict):
        return cached
    cookies = parse_cookie_header(env.get("_COOKIE", ""))
    payload = read_signed_session(cookies.get(AUTH_COOKIE_NAME, ""), auth_secret(repo, env))
    uid = text_only(payload.get("uid"), 120).strip()
    user = repo.get("auth_users", uid) if uid else {}
    if not user or str(user.get("status") or "active") != "active":
        return {}
    role = repo.get("auth_roles", user.get("role_uid") or "") or {}
    if not role or not truthy(role.get("is_active"), default=True):
        return {}
    return {"user": user, "role": role, "payload": payload}


def current_user(repo: Repository, env: dict[str, str]) -> dict[str, Any]:
    return current_auth(repo, env).get("user") or {}


def auth_role_level(repo: Repository, env: dict[str, str]) -> int:
    return int_value((current_auth(repo, env).get("role") or {}).get("level"), 0)


def auth_visibility_scopes(repo: Repository, env: dict[str, str]) -> set[str]:
    auth = current_auth(repo, env)
    role = auth.get("role") or {}
    raw = text_only(role.get("visibility_scopes"), 300)
    scopes = {item.strip() for item in raw.replace(";", ",").split(",") if item.strip()}
    if scopes:
        return scopes
    level = int_value(role.get("level"), 0)
    if level >= 100:
        return {"public", "authenticated", "staff", "owner"}
    if level >= 40:
        return {"public", "authenticated", "staff"}
    if level > 0:
        return {"public", "authenticated"}
    return {"public"}


def role_permission(repo: Repository, role_uid: str, module: str) -> dict[str, Any]:
    target_uid = stable_uid("perm", f"{role_uid}:{module}")
    row = repo.get("auth_permissions", target_uid)
    if row:
        return row
    for item in repo.list("auth_permissions", Query(limit=1000)):
        if str(item.get("role_uid") or "") == role_uid and str(item.get("module") or "") == module:
            return item
    return {}


def has_permission(repo: Repository, env: dict[str, str], module: str, action: str = "can_view") -> bool:
    auth = current_auth(repo, env)
    role = auth.get("role") or {}
    role_uid = str(role.get("uid") or "")
    if not role_uid:
        return False
    if int_value(role.get("level"), 0) >= 100:
        return True
    perm = role_permission(repo, role_uid, module)
    return truthy(perm.get(action), default=False)


def auth_cookie(repo: Repository, env: dict[str, str], user: dict[str, Any]) -> str:
    now = int(time.time())
    payload = {
        "uid": str(user.get("uid") or ""),
        "iat": now,
        "exp": now + AUTH_SESSION_SECONDS,
        "csrf": stable_uid("csrf", f"{user.get('uid')}-{now}-{time.time_ns()}"),
    }
    secure = "; Secure" if str(env.get("_SCHEME") or "").lower() == "https" else ""
    return f"{AUTH_COOKIE_NAME}={signed_session_token(payload, auth_secret(repo, env))}; Path=/; Max-Age={AUTH_SESSION_SECONDS}; HttpOnly; SameSite=Lax{secure}"


def clear_auth_cookie() -> str:
    return f"{AUTH_COOKIE_NAME}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax"


def login_rate_key(username: str, env: dict[str, str]) -> str:
    host = str(env.get("_HOST") or "").split(",", 1)[0].strip().lower()
    remote = str(env.get("_REMOTE_ADDR") or "").split(",", 1)[0].strip()
    seed = f"{host}|{remote}|{username.strip().lower()}"
    return hashlib.sha256(seed.encode("utf-8", "ignore")).hexdigest()


def login_rate_state(env: dict[str, str]) -> dict[str, Any]:
    cached = env.get("_LOGIN_RATE_STATE")
    if isinstance(cached, dict):
        return cached
    if env.get("PLATFORM") == "cloudflare":
        state: dict[str, Any] = {}
    else:
        try:
            state = json.loads(LOGIN_RATE_LIMIT_CACHE_PATH.read_text(encoding="utf-8"))
            if not isinstance(state, dict):
                state = {}
        except (OSError, json.JSONDecodeError):
            state = {}
    env["_LOGIN_RATE_STATE"] = state
    return state


def save_login_rate_state(env: dict[str, str], state: dict[str, Any]) -> None:
    if env.get("PLATFORM") == "cloudflare":
        return
    now = time.time()
    cleaned: dict[str, Any] = {}
    for key, value in state.items():
        if not isinstance(value, dict):
            continue
        try:
            marker = float(value.get("last_at") or value.get("locked_until") or 0)
        except (TypeError, ValueError):
            continue
        if now - marker < 24 * 3600:
            cleaned[key] = value
    try:
        LOGIN_RATE_LIMIT_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        LOGIN_RATE_LIMIT_CACHE_PATH.write_text(json.dumps(cleaned, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    except OSError:
        return


def login_rate_wait_seconds(username: str, env: dict[str, str]) -> int:
    if not username:
        return 0
    state = login_rate_state(env)
    entry = state.get(login_rate_key(username, env))
    if not isinstance(entry, dict):
        return 0
    locked_until = float(entry.get("locked_until") or 0)
    remaining = int(locked_until - time.time())
    return max(0, remaining)


def record_login_failure(username: str, env: dict[str, str]) -> None:
    if not username:
        return
    state = login_rate_state(env)
    key = login_rate_key(username, env)
    now = time.time()
    entry = state.get(key) if isinstance(state.get(key), dict) else {}
    first_at = float(entry.get("first_at") or now)
    if now - first_at > LOGIN_RATE_LIMIT_WINDOW_SECONDS:
        first_at = now
        failures = 0
    else:
        failures = int(entry.get("failures") or 0)
    failures += 1
    locked_until = float(entry.get("locked_until") or 0)
    if failures >= LOGIN_RATE_LIMIT_MAX_FAILURES:
        locked_until = max(locked_until, now + LOGIN_RATE_LIMIT_LOCK_SECONDS)
    state[key] = {"first_at": first_at, "last_at": now, "failures": failures, "locked_until": locked_until}
    save_login_rate_state(env, state)


def clear_login_failures(username: str, env: dict[str, str]) -> None:
    if not username:
        return
    state = login_rate_state(env)
    key = login_rate_key(username, env)
    if key in state:
        state.pop(key, None)
        save_login_rate_state(env, state)


def redirect_with_cookie(location: str, cookie: str) -> ResponseTuple:
    return 302, security_headers() + [("location", location), ("set-cookie", cookie)], b""


def admin_auth_response(repo: Repository, path: str, env: dict[str, str]) -> ResponseTuple | None:
    if path in {"/admin/login", "/admin/setup"}:
        return None
    if path == "/admin/logout":
        return None
    if not auth_users_exist(repo):
        return redirect("/admin/setup")
    if not current_auth(repo, env):
        return redirect(f"/admin/login?next={quote(path or '/admin')}")
    if not has_permission(repo, env, "admin", "can_view"):
        return html_response(admin_denied_html(repo, env, "后台入口", "当前账号没有进入后台的权限。"), 403)
    return None


def api_auth_response(repo: Repository, env: dict[str, str], module: str, action: str = "can_view") -> ResponseTuple | None:
    if not same_origin_post_allowed("POST" if str(env.get("_METHOD") or "").upper() == "POST" else "GET", env):
        return json_response({"ok": False, "message": "跨站请求已被拦截。"}, 403)
    if not auth_users_exist(repo) or not current_auth(repo, env):
        return json_response({"ok": False, "message": "请先登录后台。"}, 401)
    if not has_permission(repo, env, module, action):
        return json_response({"ok": False, "message": "当前账号没有执行该操作的权限。"}, 403)
    return None


def row_visible_to_current_user(repo: Repository, env: dict[str, str], row: dict[str, Any]) -> bool:
    visibility = str(row.get("visibility") or "public")
    if visibility == "public":
        return True
    if visibility == "hidden":
        return False
    return bool(current_auth(repo, env)) and visibility in auth_visibility_scopes(repo, env)


def visible_list(repo: Repository, env: dict[str, str], table: str, query: Query) -> list[dict[str, Any]]:
    unrestricted = Query(
        q=query.q,
        filters=query.filters,
        public_only=False,
        limit=query.limit,
        order_by=query.order_by,
        descending=query.descending,
    )
    return [row for row in repo.list(table, unrestricted) if row_visible_to_current_user(repo, env, row)]


def visible_get(repo: Repository, env: dict[str, str], table: str, key: str) -> dict[str, Any]:
    row = repo.get(table, key) or {}
    return row if row and row_visible_to_current_user(repo, env, row) else {}


def route_request(repo: Repository, method: str, path: str, query_string: str = "", body: bytes = b"", env: dict[str, str] | None = None) -> ResponseTuple:
    env = env or {}
    query = _query(query_string)
    if path != "/" and path.endswith("/"):
        path = path.rstrip("/")
    env = {**env, "_LANG": "en" if query.get("lang") == "en" else "zh", "_PATH": path, "_METHOD": method.upper()}
    if path.startswith(("/admin", "/api/admin", "/api/export")) or path in {"/login", "/register", "/logout"}:
        config_guard = auth_config_response(env, api=path.startswith("/api"))
        if config_guard:
            return config_guard
    if path.startswith("/admin") or path.startswith("/api/admin") or path.startswith("/api/export/"):
        ensure_auth_defaults(repo)
    if method.upper() == "POST" and (path.startswith("/admin") or path.startswith("/api/admin")) and not same_origin_post_allowed(method, env):
        if path.startswith("/api/"):
            return json_response({"ok": False, "message": "跨站请求已被拦截。"}, 403)
        return html_response(admin_denied_html(repo, env, "请求被拦截", "检测到跨站来源，后台写入请求已被拒绝。"), 403)

    if path == "/sitemap.xml":
        return xml_response(sitemap_xml(repo, env.get("SITE_URL", "")))
    if path == "/sitemap.txt":
        return text_response(sitemap_txt(repo, env.get("SITE_URL", "")))
    if path == "/sitemap-index.xml":
        return xml_response(sitemap_index_xml(env.get("SITE_URL", "")))
    if path == "/robots.txt":
        return text_response(robots_txt(env.get("SITE_URL", "")))
    if path in {"/.well-known/security.txt", "/security.txt"}:
        return text_response(security_txt(env.get("SITE_URL", "")))
    if path == "/llms.txt":
        return text_response(llms_txt(repo, env.get("SITE_URL", "")))
    if path == "/login":
        ensure_auth_defaults(repo)
        return front_login_route(repo, method, query, body, env)
    if path == "/register":
        ensure_auth_defaults(repo)
        return front_register_route(repo, method, query, body, env)
    if path == "/logout":
        return redirect_with_cookie(lang_url("/", env), clear_auth_cookie())
    if path == "/api/admin/media/summary":
        denied = api_auth_response(repo, env, "media_assets", "can_view")
        if denied:
            return denied
        refresh = query.get("refresh") == "1"
        return media_json_response(media_summary_payload(repo, env, refresh=refresh), refresh=refresh)
    if path == "/api/admin/media/file-size":
        denied = api_auth_response(repo, env, "media_assets", "can_view")
        if denied:
            return denied
        refresh = query.get("refresh") == "1"
        return media_json_response(media_file_size_payload(query.get("key", ""), env, refresh=refresh), refresh=refresh)
    if path == "/api/admin/media/options":
        denied = api_auth_response(repo, env, "media_assets", "can_view")
        if denied:
            return denied
        return json_response(media_options_payload(repo, env, query))
    if path == "/api/admin/media/upload" and method == "POST":
        denied = api_auth_response(repo, env, "media_assets", "can_create")
        if denied:
            return denied
        return json_response(media_upload_payload(repo, body, env))
    if path == "/api/admin/media/crop" and method == "POST":
        denied = api_auth_response(repo, env, "media_assets", "can_edit")
        if denied:
            return denied
        return json_response(media_crop_payload(repo, body, env))
    if path == "/api/admin/publications/parse" and method == "POST":
        denied = api_auth_response(repo, env, "publications", "can_edit")
        if denied:
            return denied
        return json_response(publication_parse_payload(body))
    if path == "/api/admin/publications/duplicates":
        denied = api_auth_response(repo, env, "publications", "can_view")
        if denied:
            return denied
        return json_response(publication_duplicates_payload(repo, query))
    if path == "/api/admin/publications/lookup" and method == "POST":
        denied = api_auth_response(repo, env, "publications", "can_edit")
        if denied:
            return denied
        return json_response(publication_lookup_payload(repo, body))
    if path == "/api/admin/publications/citations" and method == "POST":
        denied = api_auth_response(repo, env, "publications", "can_edit")
        if denied:
            return denied
        return json_response(publication_citations_payload(body))
    if path == "/api/admin/publications/suggestions":
        denied = api_auth_response(repo, env, "publications", "can_view")
        if denied:
            return denied
        return json_response(publication_suggestions_payload(repo))
    if path == "/api/admin/profiles/suggestions":
        denied = api_auth_response(repo, env, "profiles", "can_view")
        if denied:
            return denied
        return json_response(profile_suggestions_payload(repo))
    if path == "/api/admin/projects/duplicates":
        denied = api_auth_response(repo, env, "projects", "can_view")
        if denied:
            return denied
        return json_response(project_duplicates_payload(repo, query))
    if path == "/api/admin/projects/suggestions":
        denied = api_auth_response(repo, env, "projects", "can_view")
        if denied:
            return denied
        return json_response(project_suggestions_payload(repo))
    if path == "/api/admin/patents/duplicates":
        denied = api_auth_response(repo, env, "patents", "can_view")
        if denied:
            return denied
        return json_response(patent_duplicates_payload(repo, query))
    if path == "/api/admin/patents/lookup" and method == "POST":
        denied = api_auth_response(repo, env, "patents", "can_edit")
        if denied:
            return denied
        return json_response(patent_lookup_payload(repo, body))
    if path == "/api/admin/patents/suggestions":
        denied = api_auth_response(repo, env, "patents", "can_view")
        if denied:
            return denied
        return json_response(patent_suggestions_payload(repo))
    if path == "/api/admin/students/suggestions":
        denied = api_auth_response(repo, env, "students", "can_view")
        if denied:
            return denied
        return json_response(student_suggestions_payload(repo))
    if path == "/api/admin/news/suggestions":
        denied = api_auth_response(repo, env, "news", "can_view")
        if denied:
            return denied
        return json_response(news_suggestions_payload(repo))
    if path == "/api/admin/courses/suggestions":
        denied = api_auth_response(repo, env, "courses", "can_view")
        if denied:
            return denied
        return json_response(course_suggestions_payload(repo))
    if path == "/api/admin/translation/status":
        denied = api_auth_response(repo, env, "translation_cache", "can_view")
        if denied:
            return denied
        return json_response(translation_job_status_payload(repo, env))
    if path == "/api/admin/translation/start" and method == "POST":
        denied = api_auth_response(repo, env, "translation_cache", "can_edit")
        if denied:
            return denied
        return json_response(translation_job_start(repo, body, env))
    if path == "/api/admin/translation/stop" and method == "POST":
        denied = api_auth_response(repo, env, "translation_cache", "can_edit")
        if denied:
            return denied
        return json_response(translation_job_stop(repo, env))
    if path == "/api/admin/translation/auto-step" and method == "POST":
        denied = api_auth_response(repo, env, "translation_cache", "can_edit")
        if denied:
            return denied
        return json_response(translation_auto_translate_step(repo, body, env))
    if path == "/api/admin/translation/inline" and method == "POST":
        denied = api_auth_response(repo, env, "translation_cache", "can_edit")
        if denied:
            return denied
        return json_response(translation_inline_payload(translation_inline_update(repo, body)))
    if path.startswith("/api/export/"):
        denied = api_auth_response(repo, env, "export", "can_export")
        if denied:
            return denied
        return export_api_route(repo, path, query, env)

    if path.startswith("/admin"):
        return admin_route(repo, method, path, query, body, env)

    if method == "POST" and path == "/contact":
        data = _form(body)
        if data.get("website"):
            return redirect("/contact")
        repo.save(
            "messages",
            {
                "uid": stable_uid("msg", f"{data.get('email','')}-{time.time_ns()}"),
                "name": text_only(data.get("name"), 80),
                "email": text_only(data.get("email"), 160),
                "message_type": text_only(data.get("message_type") or "other", 40),
                "subject": text_only(data.get("subject") or "网站留言", 160),
                "content": text_only(data.get("content"), 2000),
                "status": "new",
                "visibility": "staff",
            },
        )
        return html_response(layout(repo, t(env, "message_submitted"), f'<section class="notice"><h1>{esc(t(env, "message_submitted"))}</h1><p>{esc(t(env, "message_thanks"))}</p><p><a class="button" href="{esc(lang_url("/", env))}">{esc(t(env, "back_home"))}</a></p></section>', env))

    routes = {
        "/": home_page,
        "/team": team_page,
        "/publications": publications_page,
        "/featured-publications": lambda r, q, e: publications_page(r, {**q, "featured": "1"}, e),
        "/projects": projects_page,
        "/patents": patents_page,
        "/students": students_page,
        "/news": news_page,
        "/courses": lambda r, q, e: list_page(r, q, e, "courses", t(e, "courses"), ["name", "semester", "audience", "summary"], t(e, "course_search"), [("semester", t(e, "semester")), ("audience", t(e, "audience"))]),
        "/contact": contact_page,
    }
    if path in routes:
        return html_response(routes[path](repo, query, env))
    if path.startswith("/team/"):
        return html_response(team_detail_page(repo, path.removeprefix("/team/"), env))
    if path.startswith("/news/"):
        return html_response(news_detail_page(repo, path.removeprefix("/news/"), env))
    return html_response(layout(repo, t(env, "not_found"), f'<section class="notice"><h1>404</h1><p>{esc(t(env, "page_missing"))}</p></section>', env), 404)


def home_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    lang = current_lang(env)
    site = active_site(repo)
    site_display = front_row(repo, env, "site_settings", site)
    profile_uid = str(site.get("homepage_profile_uid") or "")
    profile = visible_get(repo, env, "profiles", profile_uid) if profile_uid else {}
    profile = profile or (visible_list(repo, env, "profiles", Query(limit=1)) or [{}])[0]
    profile_display = front_row(repo, env, "profiles", profile)
    interests = visible_list(repo, env, "research_interests", Query(limit=12))
    publication_display_style = str(active_global(repo).get("publication_display_style") or "gbt")
    publication_limit = max(1, int_value(site.get("homepage_publication_limit"), 5))
    pubs = latest_publications(visible_list(repo, env, "publications", Query(filters={"is_featured": 1}, limit=200, order_by="year", descending=True)), publication_limit)
    if not pubs:
        pubs = latest_publications(visible_list(repo, env, "publications", Query(limit=200, order_by="year", descending=True)), publication_limit)
    projects = visible_list(repo, env, "projects", Query(filters={"is_featured": 1}, limit=4))
    patents = visible_list(repo, env, "patents", Query(filters={"is_featured": 1}, limit=4))
    students = visible_list(repo, env, "students", Query(filters={"is_featured": 1}, limit=6))
    news = visible_list(repo, env, "news", Query(limit=int_value(site.get("homepage_news_limit"), 4), order_by="published_at", descending=True))
    actions = nav(repo, "home_hero", env)
    profile_name = front_value(repo, env, "profiles", profile, "name", 500) or localized_name(profile, lang)
    site_name = front_value(repo, env, "site_settings", site, "site_name", 200) or localized_site_name(site, lang)
    content = f"""
    <section class="hero">
      <div class="hero-main">
        <div class="identity">
          <figure class="home-teacher-photo-cell">{image_tag(profile.get("avatar_key"), profile_name, "teacher-photo", env.get("PUBLIC_MEDIA_BASE_URL", ""), lang)}</figure>
          <div>
            <p class="eyebrow">{esc(profile_display.get("organization") or site_name)}</p>
            <h1>{esc(profile_name or site_display.get("hero_title") or site_name)}</h1>
            <p class="subtitle">{esc(" / ".join([x for x in [profile_display.get("title"), profile_display.get("lab")] if x]))}</p>
          </div>
        </div>
        <div class="profile-text">{front_paragraphs(repo, env, "profiles", profile, "bio") or front_paragraphs(repo, env, "site_settings", site, "hero_subtitle")}</div>
        <div class="actions">{''.join(button(item, env, repo) for item in actions)}</div>
      </div>
      <aside class="facts">
        <h2>{esc(t(env, "academic_profile"))}</h2>
        <dl>{fact(t(env, "email"), profile.get("email"))}{fact(t(env, "office"), profile_display.get("office"))}{fact("ORCID", profile.get("orcid"))}</dl>
        {external_links(profile)}
        {f'<div class="recruiting"><strong>{esc(t(env, "recruiting"))}</strong>{front_paragraphs(repo, env, "profiles", profile, "recruiting")}</div>' if profile.get("recruiting") else ""}
      </aside>
    </section>
    {section(t(env, "research_interests"), '<div class="tags">' + ''.join(f'<span title="{esc(front_value(repo, env, "research_interests", item, "description", 500))}">{esc(front_value(repo, env, "research_interests", item, "name", 200))}</span>' for item in interests) + '</div>') if interests else ""}
    <section class="two-col">
      <div>{section_head(t(env, "featured_publications"), lang_url("/publications", env), env)}{home_publication_list(pubs, publication_display_style, repo, env)}</div>
      <div>{section_head(t(env, "latest_news"), lang_url("/news", env), env)}{news_list(news, repo=repo, env=env)}</div>
    </section>
    <section class="three-col">
      <div>{section_head(t(env, "projects"), lang_url("/projects", env), env)}{home_project_items(projects, repo, env)}</div>
      <div>{section_head(t(env, "patents"), lang_url("/patents", env), env)}{home_patent_items(patents, repo, env)}</div>
      <div>{section_head(t(env, "students"), lang_url("/students", env), env)}{home_student_items(students, repo, env)}</div>
    </section>
    """
    return layout(repo, str(site_display.get("seo_title") or site_name), content, env, site)


def team_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    lang = current_lang(env)
    all_rows = visible_list(repo, env, "profiles", Query(limit=200))
    filter_specs = [("role", t(env, "role")), ("title", t(env, "title")), ("organization", t(env, "organization")), ("lab", t(env, "team"))]
    filters = query_filters(query, [name for name, _label in filter_specs])
    rows = visible_list(repo, env, "profiles", Query(q=query.get("q", ""), filters=filters, limit=200))
    cards = []
    for row in rows:
        display_row = front_row(repo, env, "profiles", row)
        identity = team_identity(display_row)
        display = front_value(repo, env, "profiles", row, "name", 500) or localized_name(row, lang)
        detail_href = lang_url(f'/team/{row.get("uid") or row.get("id") or ""}', env)
        summary = front_value(repo, env, "profiles", row, "bio", 520).strip()
        summary_html = f'<p class="team-row team-row-summary team-summary">{esc(summary)}</p>' if summary else '<p class="team-row team-row-summary team-summary is-empty"></p>'
        cards.append(
            f"""<article class="person-card team-card">
            {image_tag(row.get("avatar_key"), display, "person-avatar", env.get("PUBLIC_MEDIA_BASE_URL", ""), lang)}
            <div class="person-body">
              <div class="team-row team-row-title"><h2>{esc(display)}</h2>{identity}</div>
              {summary_html}
              <div class="team-row team-row-links"><div class="person-links team-links">{profile_links(row)}</div><a class="profile-open-button" href="{detail_href}" target="_blank" rel="noreferrer" title="打开完整信息">详情</a></div>
            </div>
          </article>"""
        )
    toolbar = compact_filter_form(query, t(env, "team_search"), filter_options(all_rows, filter_specs, repo, env, "profiles"), env)
    return layout(repo, t(env, "team_members"), '<div class="compact-page">' + toolbar + '<section class="people-list team-list">' + ("".join(cards) or empty(env)) + "</section></div>", env)


def team_detail_page(repo: Repository, uid: str, env: dict[str, str]) -> str:
    lang = current_lang(env)
    row = visible_get(repo, env, "profiles", uid)
    if not row:
        return layout(repo, t(env, "not_found"), f'<section class="notice"><h1>{esc(t(env, "member_missing"))}</h1></section>', env)
    display_row = front_row(repo, env, "profiles", row)
    identity = team_identity(display_row)
    details_html = profile_detail_sections(display_row)
    bio = front_value(repo, env, "profiles", row, "bio", 5000).strip()
    bio_html = f'<section class="team-detail-bio"><h2>{esc(t(env, "bio"))}</h2><div>{paragraphs(bio)}</div></section>' if bio else ""
    display = front_value(repo, env, "profiles", row, "name", 500) or localized_name(row, lang)
    body = f"""<div class="compact-page team-profile-page">
    <section class="team-detail-hero">
      {image_tag(row.get("avatar_key"), display, "team-detail-photo", env.get("PUBLIC_MEDIA_BASE_URL", ""), lang)}
      <div class="team-detail-main">
        <div class="team-row team-row-title"><h1>{esc(display)}</h1>{identity}</div>
        <div class="person-links team-links">{profile_links(row)}</div>
      </div>
    </section>
    {bio_html}
    {details_html}
    <p class="team-detail-back"><a class="button ghost" href="{esc(lang_url("/team", env))}">{esc(t(env, "back_team"))}</a></p>
    </div>"""
    return layout(repo, str(display or t(env, "team_members")), body, env)


def publications_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    q = query.get("q", "")
    publication_display_style = str(active_global(repo).get("publication_display_style") or "gbt")
    all_rows = visible_list(repo, env, "publications", Query(limit=500, order_by="year", descending=True))
    filters = {}
    if query.get("year"):
        filters["year"] = query["year"]
    if query.get("author_role"):
        filters["author_role"] = query["author_role"]
    if query.get("publication_type"):
        filters["publication_type"] = query["publication_type"]
    if query.get("venue"):
        filters["venue"] = query["venue"]
    if query.get("index_type"):
        filters["index_type"] = query["index_type"]
    if query.get("featured"):
        filters["is_featured"] = 1
    rows = visible_list(repo, env, "publications", Query(q=q, filters=filters, limit=500, order_by="year", descending=True))
    filter_specs = [("year", t(env, "year")), ("venue", t(env, "venue")), ("publication_type", t(env, "publication_type")), ("author_role", t(env, "author_role")), ("index_type", t(env, "index_type"))]
    reset_href = "?lang=en" if current_lang(env) == "en" else "?"
    toolbar = f"""
    <form class="filters filters-wide publication-filters compact-filterbar" method="get">
      {lang_hidden(env)}
      <input class="filter-search" name="q" value="{esc(q)}" placeholder="{esc(t(env, "publication_search"))}">
      {filter_selects(query, filter_options(all_rows, filter_specs, repo, env, "publications"))}
      <button>{esc(t(env, "search"))}</button>
      <a class="button ghost filter-reset" href="{esc(reset_href)}">{esc(t(env, "reset"))}</a>
    </form>
    <div class="copy-toolbar citation-copy-toolbar"><label class="copy-select-all"><input type="checkbox" id="select-all-citations">{esc(t(env, "select_all"))}</label><select id="citation-style">{citation_style_options(publication_display_style, env)}</select><button type="button" id="copy-selected">{esc(t(env, "copy"))}</button><span id="copy-status"></span></div>
    """
    return layout(repo, t(env, "publications"), '<div class="compact-page">' + toolbar + publication_list(rows, selectable=True, compact=True, display_style=publication_display_style, repo=repo, env=env) + "</div>", env)


def list_page(repo: Repository, query: dict[str, str], env: dict[str, str], table: str, title: str, fields: list[str], search_placeholder: str = "关键词", filter_specs: list[tuple[str, str]] | None = None) -> str:
    filter_specs = filter_specs or []
    all_rows = visible_list(repo, env, table, Query(limit=500))
    filters = query_filters(query, [name for name, _label in filter_specs])
    rows = visible_list(repo, env, table, Query(q=query.get("q", ""), filters=filters, limit=500))
    items = []
    for index, row in enumerate(rows, 1):
        display_row = front_row(repo, env, table, row)
        meta = " / ".join(str(display_row.get(field) or "") for field in fields[1:4] if display_row.get(field))
        summary = front_paragraphs(repo, env, table, row, fields[-1]) if row.get(fields[-1]) else ""
        items.append(f'<article class="compact-item"><div class="item-index"><span class="item-number">{index}</span></div><div class="compact-body"><h2>{esc(display_row.get(fields[0]))}</h2><div class="compact-meta">{esc(meta)}</div>{summary}</div></article>')
    toolbar = compact_filter_form(query, search_placeholder, filter_options(all_rows, filter_specs, repo, env, table), env)
    body = f'<div class="compact-page">{toolbar}<section class="compact-list">{"".join(items) or empty(env)}</section></div>'
    return layout(repo, title, body, env)


def projects_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    all_rows = visible_list(repo, env, "projects", Query(limit=500, order_by="sort_order", descending=True))
    filter_specs = [("source", t(env, "source")), ("fund_name", t(env, "fund_name")), ("status", t(env, "status"))]
    filters = query_filters(query, [name for name, _label in filter_specs])
    rows = visible_list(repo, env, "projects", Query(q=query.get("q", ""), filters=filters, limit=500, order_by="sort_order", descending=True))
    total = len(rows)
    cards = [project_card(row, total - index + 1, env, repo) for index, row in enumerate(rows, 1)]
    toolbar = compact_filter_form(query, t(env, "project_search"), filter_options(all_rows, filter_specs, repo, env, "projects"), env)
    body = f'<div class="compact-page">{toolbar}<section class="compact-list project-list">{"".join(cards) or empty(env)}</section></div>'
    return layout(repo, t(env, "projects"), body, env)


def project_card(row: dict[str, Any], index: int, env: dict[str, str], repo: Repository | None = None) -> str:
    display_row = front_row(repo, env, "projects", row) if repo else row
    source = text_only(display_row.get("source"), 160).strip()
    fund_name = text_only(display_row.get("fund_name"), 180).strip()
    number = text_only(row.get("project_number"), 120).strip()
    funding_parts = []
    if source:
        funding_parts.append(f'<span class="project-source">{esc(source)}</span>')
    if source and fund_name:
        funding_parts.append('<span class="project-funding-sep">-</span>')
    if fund_name:
        funding_parts.append(f'<span class="project-fund">{esc(fund_name)}</span>')
    period = front_project_period(row.get("start_date"), row.get("end_date"), env)
    amount = front_project_amount(row.get("amount"), env)
    status = text_only(display_row.get("status"), 80).strip()
    facts = "".join(
        project_fact(label, value, class_name)
        for label, value, class_name in (
            (t(env, "project_period"), period, "project-period"),
            (t(env, "project_number"), number, "project-number-fact"),
            (t(env, "project_amount"), amount, "project-amount"),
            (t(env, "status"), status, "project-status"),
        )
        if value
    )
    return f"""<article class="compact-item project-card">
      <div class="item-index"><span class="item-number">{index}</span></div>
      <div class="compact-body project-body">
        <div class="project-topline"><div class="project-funding">{"".join(funding_parts)}</div></div>
        <div class="project-mainline"><h2>{esc(display_row.get("name"))}</h2><div class="project-facts">{facts}</div></div>
      </div>
    </article>"""


def project_fact(label: str, value: str, class_name: str) -> str:
    return f'<span class="project-fact {esc(class_name)}"><span class="project-fact-label">{esc(label)}</span><span class="project-fact-value">{esc(value)}</span></span>'


def front_project_period(start: Any, end: Any, env: dict[str, str]) -> str:
    start_text = text_only(start, 40).strip()
    end_text = text_only(end, 40).strip()
    if start_text and end_text:
        return f"{start_text} - {end_text}"
    return start_text or end_text


def front_project_amount(value: Any, env: dict[str, str]) -> str:
    text = text_only(value, 80).strip()
    if not text:
        return ""
    parsed = parse_project_amount(text)
    if parsed and current_lang(env) == "en":
        amount_cny, _unit = parsed
        return f"{format_k_cny_amount(amount_cny)}k CNY"
    if parsed:
        amount_cny, unit = parsed
        if unit in {"yuan", "cny"}:
            return f"{format_money_amount(amount_cny)} 元"
        return f"{format_decimal_amount(amount_cny / Decimal('10000'))} {t(env, 'project_amount_unit')}".strip()
    return text


def parse_project_amount(value: str) -> tuple[Decimal, str] | None:
    text = text_only(value, 80).strip().replace(",", "")
    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(万元|万|元|人民币|cny|rmb|10k\s*cny)?", text, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        amount = Decimal(match.group(1))
    except InvalidOperation:
        return None
    raw_unit = (match.group(2) or "").strip().casefold()
    if raw_unit in {"元", "人民币", "cny", "rmb"}:
        return amount, "yuan"
    return amount * Decimal("10000"), "wan"


def format_money_amount(value: Decimal) -> str:
    if value == value.to_integral_value():
        return f"{int(value):,}"
    normalized = value.quantize(Decimal("0.01")).normalize()
    return f"{normalized:,f}"


def format_k_cny_amount(value: Decimal) -> str:
    amount = value / Decimal("1000")
    if amount == amount.to_integral_value():
        return str(int(amount))
    normalized = amount.quantize(Decimal("0.01")).normalize()
    return f"{normalized:f}"


def format_decimal_amount(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    return f"{value.normalize():f}"


def patents_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    all_rows = visible_list(repo, env, "patents", Query(limit=500))
    filter_specs = [("patent_type", t(env, "patent_type")), ("legal_status", t(env, "legal_status")), ("country", t(env, "country"))]
    filters = query_filters(query, [name for name, _label in filter_specs])
    rows = visible_list(repo, env, "patents", Query(q=query.get("q", ""), filters=filters, limit=500, order_by="sort_order", descending=True))
    cards = []
    total = len(rows)
    for index, row in enumerate(rows, 1):
        cards.append(patent_card(row, total - index + 1, env, repo))
    toolbar = compact_filter_form(query, t(env, "patent_search"), filter_options(all_rows, filter_specs, repo, env, "patents"), env)
    body = f'<div class="compact-page">{toolbar}<section class="compact-list patent-list">{"".join(cards) or empty(env)}</section></div>'
    return layout(repo, t(env, "patents"), body, env)


def patent_card(row: dict[str, Any], index: int, env: dict[str, str], repo: Repository | None = None) -> str:
    display_row = front_row(repo, env, "patents", row) if repo else row
    tags = patent_tags(display_row)
    details = [
        patent_detail(t(env, "inventors"), display_row.get("inventors")),
        patent_detail(t(env, "owner"), display_row.get("owner")),
        patent_number_date(t(env, "application_info"), row.get("application_number"), row.get("application_date")),
        patent_number_date(t(env, "grant_info"), row.get("grant_number"), row.get("grant_date")),
    ]
    detail_html = "".join(item for item in details if item)
    summary = paragraphs(display_row.get("summary")) if display_row.get("summary") else ""
    return f"""<article class="compact-item patent-card">
      <div class="item-index"><span class="item-number">{index}</span></div>
      <div class="compact-body patent-body">
        <div class="patent-title-row"><h2>{esc(display_row.get("name"))}</h2>{tags}</div>
        <dl class="patent-detail-grid">{detail_html}</dl>
        {summary}
      </div>
    </article>"""


def patent_detail(label: str, value: Any) -> str:
    text = text_only(value, 300).strip()
    if not text:
        return ""
    return f'<div class="patent-detail-item"><dt class="patent-detail-label">{esc(label)}</dt><dd class="patent-detail-value">{esc(text)}</dd></div>'


def patent_number_date(label: str, number: Any, date: Any) -> str:
    number_text = text_only(number, 220).strip()
    date_text = text_only(date, 80).strip()
    if not number_text and not date_text:
        return ""
    date_html = f'<span class="patent-date">{esc(date_text)}</span>' if date_text else ""
    separator = '<span class="patent-date-sep">·</span>' if number_text and date_text else ""
    return f'<div class="patent-detail-item patent-number-date"><dt class="patent-detail-label">{esc(label)}</dt><dd class="patent-detail-value">{esc(number_text)}{separator}{date_html}</dd></div>'


def patent_tags(row: dict[str, Any]) -> str:
    tags = []
    for key in ["patent_type", "legal_status", "country"]:
        value = text_only(row.get(key), 80).strip()
        if value:
            tags.append(f'<span class="patent-tag">{esc(value)}</span>')
    return '<div class="patent-tags">' + "".join(tags) + "</div>" if tags else ""


def students_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    lang = current_lang(env)
    all_rows = visible_list(repo, env, "students", Query(limit=300, order_by="sort_order", descending=True))
    filter_specs = [("degree", t(env, "degree")), ("category", t(env, "category")), ("grade", t(env, "grade")), ("status", t(env, "status"))]
    filters = query_filters(query, [name for name, _label in filter_specs])
    rows = visible_list(repo, env, "students", Query(q=query.get("q", ""), filters=filters, limit=300, order_by="sort_order", descending=True))
    rows = sorted(rows, key=lambda row: (int_value(row.get("sort_order"), int_value(row.get("id"), 0)), int_value(row.get("id"), 0)), reverse=True)
    rows_by_group: dict[str, dict[str, Any]] = {}
    group_mode = student_group_mode(query)
    category_meta = student_category_display_map(repo, env) if group_mode == "category" else {}
    for row in rows:
        display_row = front_row(repo, env, "students", row)
        group = student_group_info(row, display_row, group_mode, category_meta, env)
        rows_by_group.setdefault(group["key"], {"title": group["title"], "sort": group["sort"], "rows": []})
        rows_by_group[group["key"]]["rows"].append(row)
    sections = []
    for group in sorted(rows_by_group.values(), key=lambda item: item["sort"]):
        group_rows = group["rows"]
        group_total = len(group_rows)
        cards = [student_card(row, group_total - index + 1, repo, env, lang) for index, row in enumerate(group_rows, 1)]
        sections.append(f"""<section class="student-group-section">
          <div class="student-group-head"><h2>{esc(group["title"])}</h2><span>{len(cards)} {esc(t(env, "people_count_unit"))}</span></div>
          <div class="people-list student-group-list">{"".join(cards)}</div>
        </section>""")
    toolbar = student_filter_form(query, t(env, "student_search"), filter_options(all_rows, filter_specs, repo, env, "students"), env)
    return layout(repo, t(env, "students"), f'<div class="compact-page student-page">{toolbar}<div class="student-groups">{"".join(sections) or empty(env)}</div></div>', env)


def student_card(row: dict[str, Any], number: int, repo: Repository, env: dict[str, str], lang: str) -> str:
    display_row = front_row(repo, env, "students", row)
    display = front_value(repo, env, "students", row, "name", 500) or localized_name(row, lang)
    meta = " / ".join([x for x in [display_row.get("degree"), display_row.get("category"), display_row.get("grade"), display_row.get("status")] if x])
    summary = text_only(display_row.get("bio") or display_row.get("direction"), 360)
    return f"""<article class="person-card student-card">
        <div class="student-visual">{image_tag(row.get("avatar_key"), display, "person-avatar", env.get("PUBLIC_MEDIA_BASE_URL", ""), lang)}<span class="student-card-number">{number}</span></div>
        <div class="person-body"><div class="person-head"><h2>{esc(display)}</h2><span class="meta">{esc(meta)}</span></div><p class="person-summary">{esc(summary)}</p><div class="person-links">{profile_links(row)}</div></div>
        </article>"""


def student_group_mode(query: dict[str, str]) -> str:
    value = text_only(query.get("group_by"), 40).strip()
    return value if value in {"category", "degree"} else "category"


def student_filter_form(query: dict[str, str], placeholder: str, filter_groups: list[tuple[str, str, list[Any]]], env: dict[str, str]) -> str:
    reset_href = "?lang=en" if current_lang(env) == "en" else "?"
    group_options = [
        ("category", t(env, "group_by_category")),
        ("degree", t(env, "group_by_degree")),
    ]
    group_style = filter_select_style(t(env, "student_group_mode"), group_options, student_group_mode(query))
    return f"""<form class="filters compact-filterbar student-filterbar" method="get">
      {lang_hidden(env)}
      <input class="filter-search" name="q" value="{esc(query.get("q", ""))}" placeholder="{esc(placeholder)}">
      <select name="group_by" class="filter-select student-group-select" style="{group_style}"><option value="">{esc(t(env, "student_group_mode"))}</option>{select_options(group_options, student_group_mode(query))}</select>
      {filter_selects(query, filter_groups)}
      <button>{esc(t(env, "search"))}</button>
      <a class="button ghost filter-reset" href="{esc(reset_href)}">{esc(t(env, "reset"))}</a>
    </form>"""


def student_category_display_map(repo: Repository, env: dict[str, str]) -> dict[str, tuple[int, str]]:
    rows = repo.list("student_category_displays", Query(filters={"enabled": 1}, limit=200, order_by="display_order"))
    result: dict[str, tuple[int, str]] = {}
    for row in rows:
        order = int_value(row.get("display_order"), int_value(row.get("id"), 9999))
        label = front_value(repo, env, "student_category_displays", row, "label", 240) or text_only(row.get("label") or row.get("key"), 240).strip()
        tokens = [row.get("key"), row.get("label"), row.get("label_en")]
        tokens.extend(split_publication_tag_text(text_only(row.get("keywords"), 500)))
        for token in tokens:
            normalized = text_only(token, 120).strip().casefold()
            if normalized:
                result.setdefault(normalized, (order, label))
    return result


def student_group_info(row: dict[str, Any], display_row: dict[str, Any], mode: str, category_meta: dict[str, tuple[int, str]], env: dict[str, str]) -> dict[str, Any]:
    field = "degree" if mode == "degree" else "category"
    raw = text_only(row.get(field), 160).strip()
    display = text_only(display_row.get(field), 160).strip()
    if mode == "category":
        matched = student_category_match(raw, category_meta)
        if matched:
            order, label = matched
            return {"key": f"category:{order}:{label}", "title": label, "sort": (0, order, label.casefold())}
    title = display or raw or t(env, "student_group_uncategorized")
    sort_key = filter_option_sort_key(field, title) if title != t(env, "student_group_uncategorized") else (99, 0, title.casefold())
    if len(sort_key) == 2:
        sort_key = (1, sort_key[0], sort_key[1])
    return {"key": f"{field}:{title.casefold()}", "title": title, "sort": sort_key}


def student_category_match(value: str, category_meta: dict[str, tuple[int, str]]) -> tuple[int, str] | None:
    normalized = text_only(value, 160).strip().casefold()
    if not normalized:
        return None
    if normalized in category_meta:
        return category_meta[normalized]
    for token, meta in category_meta.items():
        if token and (token in normalized or normalized in token):
            return meta
    return None


def news_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    all_rows = visible_list(repo, env, "news", Query(limit=100, order_by="published_at", descending=True))
    filter_specs = [("category", t(env, "category"))]
    filters = query_filters(query, [name for name, _label in filter_specs])
    rows = visible_list(repo, env, "news", Query(q=query.get("q", ""), filters=filters, limit=100, order_by="published_at", descending=True))
    toolbar = compact_filter_form(query, t(env, "news_search"), filter_options(all_rows, filter_specs, repo, env, "news"), env)
    return layout(repo, t(env, "news"), f'<div class="compact-page">{toolbar}{news_list(rows, detail=True, repo=repo, env=env)}</div>', env)


def news_detail_page(repo: Repository, slug: str, env: dict[str, str]) -> str:
    row = visible_get(repo, env, "news", safe_slug(slug))
    if not row:
        return layout(repo, t(env, "not_found"), f'<section class="notice"><h1>{esc(t(env, "news_missing"))}</h1></section>', env)
    display_row = front_row(repo, env, "news", row)
    image = image_tag(row.get("cover_key"), display_row.get("title", ""), "news-cover", env.get("PUBLIC_MEDIA_BASE_URL", "")) if row.get("cover_key") else ""
    content_format = text_only(row.get("content_format"), 40).strip() or "plain"
    content_source = row.get("content") if content_format == "html" else display_row.get("content")
    body = f'<div class="compact-page"><article class="article news-detail-article"><p class="meta">{esc(display_row.get("category"))} / {esc(row.get("published_at"))}</p><h1>{esc(display_row.get("title"))}</h1>{image}<div class="article-content">{render_plain_or_limited_html(content_source, content_format)}</div></article></div>'
    return layout(repo, str(display_row.get("title") or t(env, "news")), body, env)


def contact_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    message_types = [
        ("recruiting", t(env, "message_type_recruiting")),
        ("cooperation", t(env, "message_type_cooperation")),
        ("paper", t(env, "message_type_paper")),
        ("course", t(env, "message_type_course")),
        ("other", t(env, "message_type_other")),
    ]
    return layout(
        repo,
        t(env, "contact"),
        f"""<div class="compact-page">
        <form class="edit-form contact-form" method="post" action="{esc(lang_url("/contact", env))}">
          <input class="honeypot" name="website" tabindex="-1" autocomplete="off">
          <label>{esc(t(env, "name"))}<input name="name" autocomplete="name"></label>
          <label>{esc(t(env, "email"))}<input name="email" type="email" autocomplete="email"></label>
          <label>{esc(t(env, "message_type"))}<select name="message_type">{navigation_pair_options(message_types, "other")}</select></label>
          <label>{esc(t(env, "subject"))}<input name="subject" required></label>
          <label>{esc(t(env, "content"))}<textarea name="content" rows="7" required></textarea></label>
          <button>{esc(t(env, "submit"))}</button>
        </form></div>""",
        env,
    )


def news_rich_editor_tool_page() -> str:
    return f"""<!doctype html><html lang="zh-CN"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>动态富文本编辑器</title>
<link rel="stylesheet" href="/assets/site.css?v={ASSET_VERSION}">
<link rel="stylesheet" href="/assets/news-rich-editor.css?v={ASSET_VERSION}">
</head><body class="news-rich-tool-page">
<main class="news-rich-window-shell">
  <header class="news-rich-window-header">
    <div><strong>动态富文本编辑器</strong><span>独立窗口编辑，应用后回写到原动态编辑页。</span></div>
    <div class="news-rich-window-actions">
      <button class="button light" type="button" data-window-refresh>重新读取原文</button>
      <button class="button secondary" type="button" data-window-apply>应用到正文</button>
      <button class="button ghost" type="button" data-window-close>关闭窗口</button>
    </div>
  </header>
  <section class="news-rich-toolbar" role="toolbar" aria-label="富文本工具栏">
    <select data-rich-block title="段落样式"><option value="p">正文</option><option value="h2">二级标题</option><option value="h3">三级标题</option><option value="blockquote">引用</option><option value="pre">代码块</option></select>
    <select data-rich-font title="字体"><option value="">默认字体</option><option value="Microsoft YaHei">微软雅黑</option><option value="SimSun">宋体</option><option value="Arial">Arial</option><option value="Times New Roman">Times</option></select>
    <select data-rich-size title="字号"><option value="">字号</option><option value="2">小</option><option value="3">正文</option><option value="4">中</option><option value="5">大</option><option value="6">特大</option></select>
    <button type="button" data-rich-cmd="bold" title="加粗"><b>B</b></button>
    <button type="button" data-rich-cmd="italic" title="斜体"><i>I</i></button>
    <button type="button" data-rich-cmd="underline" title="下划线"><u>U</u></button>
    <button type="button" data-rich-cmd="strikeThrough" title="删除线"><s>S</s></button>
    <button type="button" data-rich-cmd="insertUnorderedList" title="无序列表">列表</button>
    <button type="button" data-rich-cmd="insertOrderedList" title="有序列表">编号</button>
    <button type="button" data-rich-cmd="justifyLeft" title="靠左">左</button>
    <button type="button" data-rich-cmd="justifyCenter" title="居中">中</button>
    <button type="button" data-rich-cmd="justifyRight" title="靠右">右</button>
    <button type="button" data-rich-float="left" title="图片左悬浮">图左</button>
    <button type="button" data-rich-float="right" title="图片右悬浮">图右</button>
    <button type="button" data-rich-float="center" title="图片居中">居中图</button>
    <button type="button" data-rich-float="none" title="取消图片悬浮">取消浮动</button>
    <label title="文字颜色"><span>字色</span><input type="color" data-rich-color value="#173b32"></label>
    <label title="背景色"><span>背景</span><input type="color" data-rich-bg value="#f4faf6"></label>
    <button type="button" data-rich-link>链接</button>
    <button type="button" data-rich-media>插入媒体</button>
    <button type="button" data-rich-clear>清除格式</button>
  </section>
  <section class="news-rich-window-workspace">
    <article class="news-rich-window-editor" contenteditable="true" data-rich-editor spellcheck="true"></article>
    <aside class="news-rich-side">
      <section class="news-rich-media-tools" data-media-tools hidden>
        <header><strong>媒体属性</strong><button class="button ghost" type="button" data-media-unselect>取消选择</button></header>
        <div class="media-tools-grid">
          <label>宽度<input type="text" inputmode="decimal" data-media-width placeholder="自动 / 60% / 320px"></label>
          <label>高度<input type="text" inputmode="decimal" data-media-height placeholder="自动 / 180px"></label>
          <label>最小宽<input type="text" inputmode="decimal" data-media-min-width value="48px"></label>
          <label>最小高<input type="text" inputmode="decimal" data-media-min-height value="32px"></label>
          <label class="media-tools-check"><input type="checkbox" data-media-ratio checked> 锁定比例</label>
          <button type="button" data-media-size-apply>应用尺寸</button>
        </div>
        <div class="media-tools-actions">
          <button type="button" data-media-align="left">靠左</button>
          <button type="button" data-media-align="center">居中</button>
          <button type="button" data-media-align="right">靠右</button>
          <button type="button" data-media-align="none">普通</button>
          <button type="button" data-media-stick-mode="sticky">区域吸顶</button>
          <button type="button" data-media-stick-mode="page">全页吸顶</button>
          <button type="button" data-media-stick-mode="none">取消吸顶</button>
        </div>
        <small>先设置靠左、居中、靠右或普通，再设置吸顶；全页吸顶会把媒体提升为编辑区一级元素，取消吸顶不改变原有对齐。</small>
      </section>
      <section class="news-rich-window-preview"><strong>预览</strong><div data-rich-preview></div></section>
    </aside>
  </section>
  <footer class="news-rich-window-footer">
    <span data-rich-status>正在连接原动态编辑页。</span>
    <span>可粘贴图片/视频文件，上传后进入媒体库并插入正文。</span>
  </footer>
</main>
<section class="rich-media-dialog" data-media-dialog hidden>
  <div class="rich-media-panel">
    <header><strong>插入媒体</strong><button class="button ghost" type="button" data-media-close>关闭</button></header>
    <form class="rich-media-search" data-media-search><input name="q" placeholder="搜索媒体标题、key、分类"><button type="submit">搜索</button></form>
    <div class="rich-media-list" data-media-list></div>
    <div class="rich-media-upload">
      <label>本地上传<input type="file" data-media-file accept="image/*,video/*,.svg,.pdf,.doc,.docx"></label>
      <button class="button secondary" type="button" data-media-upload>上传并插入</button>
    </div>
    <p class="admin-muted" data-media-status></p>
  </div>
</section>
<script defer src="/assets/news-rich-editor.js?v={ASSET_VERSION}"></script>
</body></html>"""


def audit_admin_action(
    repo: Repository,
    env: dict[str, str],
    action: str,
    module: str,
    target_uid: str = "",
    summary: str = "",
    detail: dict[str, Any] | None = None,
    status: str = "success",
) -> None:
    if "operation_logs" not in repo.table_names():
        return
    try:
        auth = current_auth(repo, env)
        user = auth.get("user") or {}
        actor_uid = text_only(user.get("uid"), 160).strip()
        actor_name = text_only(user.get("display_name") or user.get("username") or actor_uid or "system", 160).strip()
        payload = {
            "uid": stable_uid("op", f"{time.time_ns()}:{actor_uid}:{module}:{action}:{target_uid}"),
            "actor_uid": actor_uid,
            "actor_name": actor_name,
            "action": text_only(action, 60).strip() or "system",
            "module": text_only(module, 120).strip(),
            "target_uid": text_only(target_uid, 220).strip(),
            "summary": text_only(summary, 1200).strip(),
            "detail_json": json.dumps(detail or {}, ensure_ascii=False, default=str),
            "status": status if status in {"success", "warning", "failed"} else "success",
        }
        repo.save("operation_logs", payload)
    except Exception:
        # 审计失败不能阻断真实业务操作；生产环境可由外层日志系统继续捕获。
        return


def append_query_params(url: str, params: dict[str, Any]) -> str:
    clean = {key: str(value) for key, value in params.items() if value not in (None, "")}
    if not clean:
        return url
    return f"{url}{'&' if '?' in url else '?'}{urlencode(clean)}"


def admin_list_page_args(query: dict[str, str]) -> tuple[int, int]:
    page = max(1, int_value(query.get("page"), 1))
    per_page = max(20, min(int_value(query.get("per_page"), 80), 200))
    return page, per_page


def admin_paginate_rows(rows: list[dict[str, Any]], query: dict[str, str]) -> tuple[list[dict[str, Any]], int, int, int]:
    page, per_page = admin_list_page_args(query)
    total = len(rows)
    start = (page - 1) * per_page
    return rows[start:start + per_page], page, per_page, total


def admin_pager(table: str, query: dict[str, str], page: int, per_page: int, total_rows: int) -> str:
    total_pages = max(1, (total_rows + per_page - 1) // per_page)
    params = {key: value for key, value in query.items() if key not in {"page", "per_page"} and value}
    per_options = options(["40", "80", "120", "200"], str(per_page))
    prev_link = admin_page_link(table, params, page - 1, per_page, "上一页") if page > 1 else '<span class="button light is-disabled">上一页</span>'
    next_link = admin_page_link(table, params, page + 1, per_page, "下一页") if page < total_pages else '<span class="button light is-disabled">下一页</span>'
    return f"""<nav class="admin-pager" aria-label="后台列表分页">
      {prev_link}
      <span class="admin-muted">第 {page} / {total_pages} 页，共 {total_rows} 条</span>
      {next_link}
      <form method="get" action="/admin/table/{esc(table)}">
        {"".join(f'<input type="hidden" name="{esc(key)}" value="{esc(value)}">' for key, value in params.items())}
        <label>每页<select name="per_page">{per_options}</select></label>
        <button class="button light" type="submit">应用</button>
      </form>
    </nav>"""


def admin_page_link(table: str, params: dict[str, str], page: int, per_page: int, label: str) -> str:
    query = urlencode({**params, "page": str(page), "per_page": str(per_page)})
    return f'<a class="button light" href="/admin/table/{esc(table)}?{esc(query)}">{esc(label)}</a>'


def admin_batch_result_notice(query: dict[str, str]) -> str:
    if query.get("batch_selected") is None and query.get("batch_updated") is None:
        return ""
    selected = text_only(query.get("batch_selected"), 20) or "0"
    updated = text_only(query.get("batch_updated"), 20) or "0"
    deleted = text_only(query.get("batch_deleted"), 20) or "0"
    skipped = text_only(query.get("batch_skipped"), 20) or "0"
    return f'<p class="admin-operation-notice">批量操作完成：选中 {esc(selected)} 条，更新 {esc(updated)} 条，删除 {esc(deleted)} 条，跳过 {esc(skipped)} 条。</p>'


def admin_quick_update_redirect(repo: Repository, env: dict[str, str], table: str, body: bytes, handler: Any) -> ResponseTuple:
    data = _form(body)
    target_uid = text_only(data.get("uid") or data.get("key"), 200).strip()
    location = handler(repo, body)
    audit_admin_action(repo, env, "quick_update", table, target_uid, f"快速修改 {TABLE_MAP.get(table).label if table in TABLE_MAP else table}", {"return_to": location})
    return redirect(location)


def admin_route(repo: Repository, method: str, path: str, query: dict[str, str], body: bytes, env: dict[str, str]) -> ResponseTuple:
    if path == "/admin/setup":
        return admin_setup_route(repo, method, query, body, env)
    if path == "/admin/login":
        return admin_login_route(repo, method, query, body, env)
    if path == "/admin/logout":
        return redirect_with_cookie("/admin/login", clear_auth_cookie())
    guard = admin_auth_response(repo, path, env)
    if guard:
        return guard
    if path == "/admin":
        counts = repo.counts()
        cards = "".join(f'<a class="admin-card" href="/admin/table/{esc(table.name)}"><strong>{esc(table.label)}</strong><span>{counts.get(table.name, 0)} 条</span></a>' for table in TABLES if has_permission(repo, env, table.name, "can_view"))
        breadcrumbs = [("后台", "/admin")]
        export_link = '<a class="button" href="/admin/export">导入与导出</a>' if has_permission(repo, env, "export", "can_export") else ""
        permission_link = '<a class="button ghost" href="/admin/permissions">权限管理</a>' if has_permission(repo, env, "auth_permissions", "can_view") else ""
        return html_response(admin_layout(repo, "后台", f'<section class="admin-hero"><h1>内容管理</h1><p>管理导航、按钮、教师照片、团队、论文、项目、专利、学生、动态、课程、权限和导出。</p><p>{export_link}{permission_link}</p></section><section class="admin-grid">{cards}</section>', env, breadcrumbs))
    if path == "/admin/tools/news-rich-editor":
        if not has_permission(repo, env, "news", "can_edit"):
            return html_response(admin_denied_html(repo, env, "富文本工具", "当前账号没有编辑动态正文的权限。"), 403)
        return html_response(news_rich_editor_tool_page())
    if path == "/admin/permissions":
        if not has_permission(repo, env, "auth_permissions", "can_view"):
            return html_response(admin_denied_html(repo, env, "权限管理", "当前账号没有查看权限管理的权限。"), 403)
        return html_response(admin_layout(repo, "权限管理", admin_permissions_home(repo, env), env, [("后台", "/admin"), ("权限管理", "/admin/permissions")]))
    if path in {"/admin/i18n-dictionary", "/admin/i18n-dictionary/save"}:
        if not has_permission(repo, env, "translation_cache", "can_view"):
            return html_response(admin_denied_html(repo, env, "手动中英词典", "当前账号没有查看翻译工具的权限。"), 403)
        breadcrumbs = [("后台", "/admin"), ("翻译缓存", "/admin/table/translation_cache"), ("手动中英词典", "/admin/i18n-dictionary")]
        if method == "POST":
            if not has_permission(repo, env, "translation_cache", "can_edit"):
                return html_response(admin_layout(repo, "手动中英词典", admin_denied_panel("手动中英词典", "当前账号没有编辑词典文件的权限。"), env, breadcrumbs), 403)
            result = i18n_dictionary_save_local_from_body(body, env)
            audit_admin_action(repo, env, "dictionary_save", "translation_cache", "", "保存手动中英词典文件", result)
            return redirect(f"/admin/i18n-dictionary?saved={quote(str(result.get('saved') or 'local'))}&entries={result.get('entries', 0)}")
        return html_response(admin_layout(repo, "手动中英词典", admin_i18n_dictionary_page(query, env), env, breadcrumbs))
    if path == "/admin/export":
        if not has_permission(repo, env, "export", "can_export"):
            return html_response(admin_denied_html(repo, env, "导入与导出", "当前账号没有导出或恢复网站数据的权限。"), 403)
        breadcrumbs = [("后台", "/admin"), ("导入与导出", "/admin/export")]
        if method == "POST":
            if not has_permission(repo, env, "export", "can_edit"):
                return html_response(admin_layout(repo, "导入与导出", admin_denied_panel("导入恢复", "当前账号没有恢复导入数据的权限。"), env, breadcrumbs), 403)
            result = import_restore_payload(repo, body, env)
            audit_admin_action(repo, env, "import_restore", "export", "", "导入恢复网站数据", result, "warning" if result.get("errors") else "success")
            suffix = urlencode({key: str(value) for key, value in result.items() if value not in (None, "")})
            return redirect(f"/admin/export?{suffix}")
        return html_response(admin_layout(repo, "导入与导出", admin_import_export_page(repo, query, env), env, breadcrumbs))
    parts = [part for part in path.split("/") if part]
    if len(parts) >= 3 and parts[1] == "table":
        table = parts[2]
        if table not in TABLE_MAP:
            return html_response(admin_layout(repo, "未知表", "<p>未知表。</p>", env, [("后台", "/admin"), ("未知表", path)]), 404)
        meta = TABLE_MAP[table]
        table_breadcrumbs = [("后台", "/admin"), (meta.label, f"/admin/table/{table}")]
        if not has_permission(repo, env, table, "can_view"):
            return html_response(admin_layout(repo, meta.label, admin_denied_panel(meta.label, "当前账号没有查看该功能页的权限。"), env, table_breadcrumbs), 403)
        if method == "POST":
            action = "can_edit"
            if len(parts) >= 4 and parts[3] == "save":
                form_uid = _form(body).get("uid", "")
                action = "can_edit" if form_uid and repo.get(table, form_uid) else "can_create"
            elif table == "media_assets" and len(parts) >= 4 and parts[3] == "export-used":
                action = "can_export"
            elif len(parts) >= 4 and parts[3] in {"delete"}:
                action = "can_delete"
            elif table == "media_assets" and any(part in {"trash", "clear", "batch"} for part in parts[3:]):
                action = "can_delete" if "delete" in parts[3:] or "clear" in parts[3:] else "can_edit"
            if not has_permission(repo, env, table, action):
                return html_response(admin_layout(repo, meta.label, admin_denied_panel(meta.label, f"当前账号没有{admin_action_label(action)}该类型数据的权限。"), env, table_breadcrumbs), 403)
        if table == "media_assets" and len(parts) >= 4 and parts[3] == "export-used":
            if not has_permission(repo, env, table, "can_export"):
                return html_response(admin_layout(repo, meta.label, admin_denied_panel(meta.label, "当前账号没有导出媒体文件的权限。"), env, table_breadcrumbs), 403)
            mode = "trash" if query.get("scope") == "trash" else "library"
            response, result = media_export_used_response(repo, query, body if method == "POST" else b"", env, mode)
            audit_admin_action(repo, env, "export_media", table, "", "导出媒体文件", result, "warning" if result.get("skipped") else "success")
            return response
        if method == "POST" and table == "media_assets" and len(parts) >= 5 and parts[3] == "trash" and parts[4] == "clear":
            result = media_clear_trash(repo)
            audit_admin_action(repo, env, "delete", table, "trash", "清空媒体回收站", result, "warning" if result.get("skipped") else "success")
            return redirect(append_query_params(f"/admin/table/{table}/trash", {"batch_deleted": result.get("deleted", 0), "batch_skipped": result.get("skipped", 0)}))
        if method == "POST" and table == "media_assets" and len(parts) >= 6 and parts[3] == "trash":
            result = media_apply_action(repo, parts[4], parts[5])
            audit_admin_action(repo, env, parts[5], table, parts[4], f"媒体回收站操作：{parts[5]}", result, "warning" if result.get("skipped") else "success")
            return redirect(append_query_params(f"/admin/table/{table}/trash", {"batch_selected": result.get("selected", 0), "batch_deleted": result.get("deleted", 0), "batch_skipped": result.get("skipped", 0)}))
        if method == "POST" and table == "media_assets" and len(parts) >= 4 and parts[3] == "scan":
            result = media_scan_project_files(repo, env)
            audit_admin_action(repo, env, "scan", table, "", "扫描项目媒体目录", result, "warning" if result.get("unsupported") else "success")
            params = {
                "scan_done": 1,
                "scan_unsupported": result.get("unsupported", 0),
                "scan_scanned": result.get("scanned", 0),
                "scan_added": result.get("added", 0),
                "scan_updated": result.get("updated", 0),
                "scan_skipped": result.get("skipped", 0),
            }
            return redirect(f"/admin/table/{table}?{urlencode({key: str(value) for key, value in params.items() if value})}")
        if method == "POST" and table == "media_assets" and len(parts) >= 5:
            media_key = parts[3]
            action = parts[4]
            result = media_apply_action(repo, media_key, action)
            audit_admin_action(repo, env, action, table, media_key, f"媒体操作：{action}", result, "warning" if result.get("skipped") else "success")
            return redirect(append_query_params(f"/admin/table/{table}", {"batch_selected": result.get("selected", 0), "batch_deleted": result.get("deleted", 0), "batch_skipped": result.get("skipped", 0)}))
        if method == "POST" and table == "media_assets" and len(parts) >= 4 and parts[3] == "batch":
            location, result = media_batch_update(repo, body)
            audit_admin_action(repo, env, "batch_update", table, "", "批量修改媒体库", result, "warning" if result.get("skipped") else "success")
            return redirect(location)
        if method == "POST" and table == "global_settings" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, global_settings_quick_update)
        if method == "POST" and table == "site_settings" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, site_settings_quick_update)
        if method == "POST" and table == "navigation_items" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, navigation_quick_update)
        if method == "POST" and table == "profiles" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, profile_quick_update)
        if method == "POST" and table == "research_interests" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, research_interest_quick_update)
        if method == "POST" and table == "publications" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, publication_quick_update)
        if method == "POST" and table == "projects" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, project_quick_update)
        if method == "POST" and table == "patents" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, patent_quick_update)
        if method == "POST" and table == "students" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, student_quick_update)
        if method == "POST" and table == "student_category_displays" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, student_category_quick_update)
        if method == "POST" and table == "news" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, news_quick_update)
        if method == "POST" and table == "courses" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, course_quick_update)
        if method == "POST" and table == "messages" and len(parts) >= 4 and parts[3] == "quick-update":
            return admin_quick_update_redirect(repo, env, table, body, message_quick_update)
        if method == "POST" and len(parts) >= 4 and parts[3] == "batch-update":
            location, result = admin_batch_update(repo, table, body)
            audit_admin_action(repo, env, "batch_update", table, "", f"批量修改 {meta.label}", result, "warning" if result.get("skipped") else "success")
            return redirect(location)
        if method == "POST" and table == "translation_cache" and len(parts) >= 4 and parts[3] == "scan":
            result = translation_scan_database(repo, env)
            audit_admin_action(repo, env, "scan", table, "", "扫描数据库提取翻译缓存", result)
            return redirect(f"/admin/table/translation_cache?scanned={result.get('created', 0)}&updated={result.get('updated', 0)}&dedicated={result.get('dedicated', 0)}&deleted={result.get('deleted', 0)}")
        if method == "POST" and table == "translation_cache" and len(parts) >= 4 and parts[3] == "auto-translate":
            result = translation_auto_translate(repo, body, env)
            audit_admin_action(repo, env, "auto_translate", table, "", "自动翻译缓存", result, "warning" if result.get("failed") else "success")
            return redirect(f"/admin/table/translation_cache?translated={result.get('translated', 0)}&failed={result.get('failed', 0)}&provider={quote(str(result.get('provider') or ''))}&scope={quote(str(result.get('scope') or ''))}&selected={result.get('selected', 0)}")
        if method == "POST" and table == "translation_cache" and len(parts) >= 4 and parts[3] == "inline":
            inline_result = translation_inline_update(repo, body)
            audit_admin_action(repo, env, "inline_update", table, text_only(_form(body).get("uid"), 200), "手动保存/确认翻译缓存", {"result": inline_result})
            return redirect("/admin/table/translation_cache")
        if method == "POST" and table == "translation_cache" and len(parts) >= 5 and parts[3] == "delete":
            deleted = translation_delete_cache(repo, parts[4])
            audit_admin_action(repo, env, "delete", table, parts[4], "删除翻译缓存", {"deleted": deleted})
            return redirect("/admin/table/translation_cache")
        if method == "POST" and len(parts) >= 4 and parts[3] == "save":
            data = _form(body)
            normalized = normalize_admin_data(meta, data)
            if table == "auth_users":
                existing = repo.get(table, normalized.get("uid") or data.get("uid") or normalized.get("id"))
                new_password = str(data.get("new_password") or "")
                if new_password:
                    normalized["password_hash"] = hash_password(new_password)
                    normalized["must_change_password"] = 0 if not normalized.get("must_change_password") else normalized.get("must_change_password")
                elif existing:
                    normalized["password_hash"] = existing.get("password_hash", "")
                else:
                    normalized["password_hash"] = hash_password(stable_uid("temp-password", str(time.time_ns())))
                    normalized["must_change_password"] = 1
                if not normalized.get("role_uid"):
                    normalized["role_uid"] = "role-visitor"
                if not normalized.get("status"):
                    normalized["status"] = "active"
            if table == "auth_permissions":
                role_uid = text_only(normalized.get("role_uid"), 160).strip()
                module = text_only(normalized.get("module"), 160).strip()
                if role_uid and module:
                    normalized["uid"] = stable_uid("perm", f"{role_uid}:{module}")
            if table == "translation_cache":
                existing = repo.get(table, normalized.get("uid") or data.get("uid") or normalized.get("id"))
                if existing:
                    for readonly_name in ("source_ref_key", "source_hash", "source_refs", "source_text"):
                        normalized[readonly_name] = existing.get(readonly_name, normalized.get(readonly_name, ""))
            saved = repo.save(table, normalized)
            if table == "publications" and not text_only(data.get("sort_order"), 40).strip():
                saved["sort_order"] = int_value(saved.get("id"), int_value(saved.get("sort_order"), 0))
                saved = repo.save(table, saved)
            if table == "projects" and not text_only(data.get("sort_order"), 40).strip():
                saved["sort_order"] = int_value(saved.get("id"), int_value(saved.get("sort_order"), 0))
                saved = repo.save(table, saved)
            if table == "patents" and not text_only(data.get("sort_order"), 40).strip():
                saved["sort_order"] = int_value(saved.get("id"), int_value(saved.get("sort_order"), 0))
                saved = repo.save(table, saved)
            if table == "students" and not text_only(data.get("sort_order"), 40).strip():
                saved["sort_order"] = int_value(saved.get("id"), int_value(saved.get("sort_order"), 0))
                saved = repo.save(table, saved)
            if table == "news" and not text_only(data.get("sort_order"), 40).strip():
                saved["sort_order"] = int_value(saved.get("id"), int_value(saved.get("sort_order"), 0))
                saved = repo.save(table, saved)
            if table == "courses" and not text_only(data.get("sort_order"), 40).strip():
                saved["sort_order"] = int_value(saved.get("id"), int_value(saved.get("sort_order"), 0))
                saved = repo.save(table, saved)
            audit_admin_action(repo, env, "save", table, text_only(saved.get("uid") or saved.get("id"), 200), f"保存 {meta.label}", {"action": data.get("_action") or "save"})
            if data.get("_action") == "save_continue":
                saved_key = saved.get("uid") or saved.get("id") or normalized.get("uid")
                return redirect(f"/admin/table/{table}/{saved_key}")
            return redirect(f"/admin/table/{table}")
        if len(parts) == 4 and parts[3] == "new":
            if not has_permission(repo, env, table, "can_create"):
                return html_response(admin_layout(repo, f"新增 {meta.label}", admin_denied_panel(meta.label, "当前账号没有新增该类型数据的权限。"), env, table_breadcrumbs), 403)
            breadcrumbs = table_breadcrumbs + [("新增", f"/admin/table/{table}/new")]
            return html_response(admin_layout(repo, f"新增 {meta.label}", admin_form(meta, admin_new_row(repo, table), repo), env, breadcrumbs))
        if table == "media_assets" and len(parts) == 4 and parts[3] == "trash":
            breadcrumbs = table_breadcrumbs + [("回收站", f"/admin/table/{table}/trash")]
            return html_response(admin_layout(repo, "媒体回收站", admin_media_table(repo, query, env, "trash"), env, breadcrumbs))
        if len(parts) == 4:
            if not has_permission(repo, env, table, "can_edit"):
                return html_response(admin_layout(repo, f"编辑 {meta.label}", admin_denied_panel(meta.label, "当前账号没有编辑该类型数据的权限。"), env, table_breadcrumbs), 403)
            row = repo.get(table, parts[3]) or {}
            label = admin_row_title(meta, row) if row else "未找到"
            breadcrumbs = table_breadcrumbs + [(label, f"/admin/table/{table}/{parts[3]}")]
            return html_response(admin_layout(repo, f"编辑 {meta.label}", admin_form(meta, row, repo), env, breadcrumbs))
        if table == "media_assets":
            content = admin_media_table(repo, query, env, "library")
        elif table == "site_settings":
            rows = site_settings_filter_rows(repo.list(table, Query(limit=100, order_by="id", descending=False)), query.get("q", ""))
            content = admin_site_settings_table(meta, rows, query, env)
        elif table == "global_settings":
            rows = global_settings_filter_rows(repo.list(table, Query(limit=100, order_by="id", descending=False)), query.get("q", ""))
            content = admin_global_settings_table(meta, rows, query, env)
        elif table == "navigation_items":
            nav_order_by, nav_descending = navigation_sort_args(query.get("sort", "sort_asc"))
            nav_filters = {key: query.get(key, "") for key in ("kind", "location", "enabled")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=nav_filters, limit=1000, order_by=nav_order_by, descending=nav_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=False))
            content = admin_navigation_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "profiles":
            profile_order_by, profile_descending = profile_sort_args(query.get("sort", "sort_asc"))
            profile_filters = {key: query.get(key, "") for key in ("role", "title", "organization", "lab", "is_active", "is_featured")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=profile_filters, limit=1000, order_by=profile_order_by, descending=profile_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=False))
            content = admin_profiles_table(meta, rows, query, all_rows, env) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "research_interests":
            research_order_by, research_descending = research_interest_sort_args(query.get("sort", "sort_asc"))
            research_filters = {"visibility": query.get("visibility", "")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=research_filters, limit=1000, order_by=research_order_by, descending=research_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=False))
            content = admin_research_interests_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "publications":
            publication_order_by, publication_descending = publication_admin_sort_args(query.get("sort", "year_desc"))
            publication_filters = {key: query.get(key, "") for key in ("year", "venue", "publication_type", "author_role", "index_type", "visibility", "is_featured")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=publication_filters, limit=1000, order_by=publication_order_by, descending=publication_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="year", descending=True))
            display_style = str(active_global(repo).get("publication_display_style") or "gbt")
            content = admin_publications_table(meta, rows, query, all_rows, display_style) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "projects":
            project_order_by, project_descending = project_admin_sort_args(query.get("sort", "sort_desc"))
            project_filters = {key: query.get(key, "") for key in ("source", "fund_name", "status", "visibility", "is_featured")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=project_filters, limit=1000, order_by=project_order_by, descending=project_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=True))
            content = admin_projects_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "patents":
            patent_order_by, patent_descending = patent_admin_sort_args(query.get("sort", "sort_desc"))
            patent_filters = {key: query.get(key, "") for key in ("country", "patent_type", "legal_status", "visibility", "is_featured")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=patent_filters, limit=1000, order_by=patent_order_by, descending=patent_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=True))
            content = admin_patents_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "students":
            student_order_by, student_descending = student_admin_sort_args(query.get("sort", "sort_desc"))
            student_filters = {key: query.get(key, "") for key in ("degree", "category", "grade", "status", "visibility", "is_featured")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=student_filters, limit=1000, order_by=student_order_by, descending=student_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=True))
            content = admin_students_table(meta, rows, query, all_rows, env) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "student_category_displays":
            category_order_by, category_descending = student_category_sort_args(query.get("sort", "order_asc"))
            category_filters = {"enabled": query.get("enabled", "")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=category_filters, limit=1000, order_by=category_order_by, descending=category_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="display_order", descending=False))
            student_rows = repo.list("students", Query(limit=1000, order_by="sort_order", descending=True))
            content = admin_student_categories_table(meta, rows, query, all_rows, student_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "news":
            news_order_by, news_descending = news_admin_sort_args(query.get("sort", "published_desc"))
            news_filters = {key: query.get(key, "") for key in ("category", "content_format", "visibility", "is_featured", "allow_comments")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=news_filters, limit=1000, order_by=news_order_by, descending=news_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="published_at", descending=True))
            content = admin_news_table(meta, rows, query, all_rows, env) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "courses":
            course_order_by, course_descending = course_admin_sort_args(query.get("sort", "sort_desc"))
            course_filters = {key: query.get(key, "") for key in ("semester", "audience", "material_visibility", "visibility", "is_featured")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=course_filters, limit=1000, order_by=course_order_by, descending=course_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=True))
            content = admin_courses_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "messages":
            message_order_by, message_descending = message_admin_sort_args(query.get("sort", "updated_desc"))
            message_filters = {key: query.get(key, "") for key in ("message_type", "status", "visibility")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=message_filters, limit=1000, order_by=message_order_by, descending=message_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="updated_at", descending=True))
            content = admin_messages_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "translation_cache":
            content = admin_translation_cache_table(meta, repo, query, env)
        elif table == "auth_roles":
            auth_order_by, auth_descending = auth_role_sort_args(query.get("sort", "level_desc"))
            auth_filters = {key: query.get(key, "") for key in ("is_active", "is_system")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=auth_filters, limit=1000, order_by=auth_order_by, descending=auth_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=False))
            content = admin_auth_roles_table(meta, rows, query, all_rows, repo) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "auth_users":
            auth_order_by, auth_descending = auth_user_sort_args(query.get("sort", "updated_desc"))
            auth_filters = {key: query.get(key, "") for key in ("role_uid", "status", "visibility")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=auth_filters, limit=1000, order_by=auth_order_by, descending=auth_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="updated_at", descending=True))
            content = admin_auth_users_table(meta, rows, query, all_rows, repo) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "auth_permissions":
            auth_order_by, auth_descending = auth_permission_sort_args(query.get("sort", "module_asc"))
            auth_filters = {key: query.get(key, "") for key in ("role_uid", "module", "can_view", "can_create", "can_edit", "can_delete", "can_export")}
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=auth_filters, limit=1000, order_by=auth_order_by, descending=auth_descending))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=False))
            content = admin_auth_permissions_table(meta, rows, query, all_rows, repo) + admin_pager(table, query, page, per_page, total_rows)
        elif table == "operation_logs":
            log_filters = {key: query.get(key, "") for key in ("action", "module", "status")}
            log_order = "created_at" if query.get("sort") != "module" else "module"
            log_desc = query.get("sort") != "module"
            rows_full = repo.list(table, Query(q=query.get("q", ""), filters=log_filters, limit=1000, order_by=log_order, descending=log_desc))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            all_rows = repo.list(table, Query(limit=1000, order_by="created_at", descending=True))
            content = admin_operation_logs_table(meta, rows, query, all_rows) + admin_pager(table, query, page, per_page, total_rows)
        else:
            rows_full = repo.list(table, Query(q=query.get("q", ""), limit=1000, order_by="updated_at", descending=True))
            rows, page, per_page, total_rows = admin_paginate_rows(rows_full, query)
            content = admin_table(meta, rows, query.get("q", "")) + admin_pager(table, query, page, per_page, total_rows)
        return html_response(admin_layout(repo, meta.label, content, env, table_breadcrumbs))
    return html_response(admin_layout(repo, "后台", "<p>后台路径不存在。</p>", env, [("后台", "/admin"), ("路径不存在", path)]), 404)


def layout(repo: Repository, title: str, content: str, env: dict[str, str], site: dict[str, Any] | None = None) -> str:
    site = site or active_site(repo)
    lang = current_lang(env)
    site_name = front_value(repo, env, "site_settings", site, "site_name", 200) or localized_site_name(site, lang)
    brand_html = site_brand_html(site, site_name, env)
    nav_html = "".join(f'<a href="{esc(lang_url(str(item.get("path") or "/"), env))}">{esc(front_value(repo, env, "navigation_items", item, "title", 120) or localized_nav_title(item, lang))}</a>' for item in nav(repo, "header", env))
    lang_switch = language_switch(env)
    auth = current_auth(repo, env)
    user = auth.get("user") if auth else {}
    user_label = text_only((user or {}).get("display_name") or (user or {}).get("username"), 80).strip()
    user_badge = f'<span class="front-user-badge" title="{esc(user_label)}">{esc(user_label)}</span>' if user_label else ""
    auth_link = f'<a class="login-link" href="/logout">{esc(t(env, "logout"))}</a>' if auth else f'<a class="login-link" href="{esc(lang_url("/login", env))}">{esc(t(env, "login"))}</a>'
    admin_link = f'<a class="admin-link" href="/admin">{esc(t(env, "admin"))}</a>' if auth and has_permission(repo, env, "admin", "can_view") else ""
    is_admin_page = "admin-shell" in content
    body_class = ' class="admin-page"' if is_admin_page else ""
    menu_label = "Menu" if lang == "en" else "菜单"
    page_title = site_name if not title else f"{title} - {site_name}"
    seo_description = front_value(repo, env, "site_settings", site, "seo_description", 500)
    footer_html = site_footer_html(repo, env, site, site_name)
    seo_links = seo_language_links(env)
    media_links = site_media_head_links(site, site_name, page_title, seo_description, env)
    robots_meta = '<meta name="robots" content="noindex,nofollow,noarchive">' if is_admin_page or str(env.get("_PATH") or "") in {"/login", "/register"} else ""
    return f"""<!doctype html>
<html lang="{"en" if lang == "en" else "zh-CN"}"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(page_title)}</title><meta name="description" content="{esc(seo_description)}">
{robots_meta}
{seo_links}
{media_links}
<link rel="stylesheet" href="/assets/site.css?v={ASSET_VERSION}"><script defer src="/assets/site.js?v={ASSET_VERSION}"></script></head>
<body{body_class}><header class="site-header"><a class="brand" href="{esc(lang_url("/", env))}">{brand_html}</a><button class="front-nav-toggle" type="button" data-front-nav-toggle aria-controls="front-site-nav" aria-expanded="false">{esc(menu_label)}</button><nav id="front-site-nav" class="site-nav">{nav_html}</nav><div class="header-actions">{lang_switch}{user_badge}{auth_link}{admin_link}</div></header>
<main>{content}</main><footer class="site-footer">{footer_html}</footer><button class="back-to-top" type="button" data-back-to-top aria-label="{esc('Back to top' if lang == 'en' else '回到顶部')}" title="{esc('Back to top' if lang == 'en' else '回到顶部')}">↑</button></body></html>"""


def site_media_key(site: dict[str, Any], field: str, fallback: str = "default/site-logo.png") -> str:
    return text_only(site.get(field), 300).strip() or fallback


def site_media_src(site: dict[str, Any], field: str, env: dict[str, str], fallback: str = "default/site-logo.png") -> str:
    key = site_media_key(site, field, fallback)
    return versioned_media_url(media_url(key, env.get("PUBLIC_MEDIA_BASE_URL", "")))


def versioned_media_url(url: str) -> str:
    if not url or url.startswith(("http://", "https://")):
        return url
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}v={ASSET_VERSION}"


def site_brand_html(site: dict[str, Any], site_name: str, env: dict[str, str]) -> str:
    logo_src = site_media_src(site, "logo_key", env)
    logo = f'<img class="brand-logo" src="{esc(logo_src)}" alt="" loading="eager" decoding="async">' if logo_src else ""
    return f'{logo}<span class="brand-text">{esc(site_name)}</span>'


def site_media_head_links(site: dict[str, Any], site_name: str, page_title: str, seo_description: str, env: dict[str, str]) -> str:
    favicon = site_media_src(site, "favicon_key", env)
    og_image = site_media_src(site, "og_image_key", env)
    canonical = site_absolute_url(str(env.get("SITE_URL") or ""), str(env.get("_PATH") or "/"))
    og_image_abs = site_absolute_url(str(env.get("SITE_URL") or ""), og_image) if og_image else ""
    lines = [
        f'<meta property="og:site_name" content="{esc(site_name)}">',
        f'<meta property="og:title" content="{esc(page_title)}">',
        f'<meta property="og:description" content="{esc(seo_description)}">',
        f'<meta property="og:type" content="website">',
        f'<meta property="og:url" content="{esc(canonical)}">',
        '<meta name="twitter:card" content="summary_large_image">',
    ]
    if favicon:
        lines.extend([
            f'<link rel="icon" href="{esc(favicon)}">',
            f'<link rel="apple-touch-icon" href="{esc(favicon)}">',
        ])
    if og_image_abs:
        lines.extend([
            f'<meta property="og:image" content="{esc(og_image_abs)}">',
            f'<meta name="twitter:image" content="{esc(og_image_abs)}">',
        ])
    return "\n".join(lines)


def site_footer_html(repo: Repository, env: dict[str, str], site: dict[str, Any], fallback: str) -> str:
    raw = str(site.get("footer_text") or "").strip()
    if current_lang(env) == "en":
        translated = front_value(repo, env, "site_settings", site, "footer_text", 3000).strip()
        if translated and translated != text_only(raw, 3000).strip():
            raw = translated
    raw = raw or fallback
    if "<" in raw and ">" in raw:
        return f'<div class="site-footer-html">{render_limited_html(raw)}</div>'
    return f'<div class="site-footer-text">{paragraphs(raw)}</div>'


def admin_layout(repo: Repository, title: str, content: str, env: dict[str, str], breadcrumbs: list[tuple[str, str]] | None = None) -> str:
    auth = current_auth(repo, env)
    user = auth.get("user") or {}
    role = auth.get("role") or {}
    trail_items = breadcrumbs or [("后台", "/admin")]
    active_path = admin_active_path(trail_items)
    primary_links = [admin_sidebar_link("/admin", "后台首页", "首", active_path)]
    if has_permission(repo, env, "export", "can_export"):
        primary_links.append(admin_sidebar_link("/admin/export", "导入与导出", "导", active_path))
    if has_permission(repo, env, "auth_permissions", "can_view"):
        primary_links.append(admin_sidebar_link("/admin/permissions", "权限管理", "权", active_path))
    table_groups = [
        ("站点配置", ("site_settings", "global_settings", "navigation_items")),
        ("人员与互动", ("profiles", "students", "student_category_displays", "messages")),
        ("科研成果", ("research_interests", "publications", "projects", "patents")),
        ("内容与媒体", ("news", "courses", "media_assets", "translation_cache", "autofetch_logs")),
        ("账号安全", ("auth_roles", "auth_users", "auth_permissions", "operation_logs")),
    ]
    grouped_tables = {name for _label, names in table_groups for name in names}
    sections = [admin_sidebar_section("常用操作", "".join(primary_links))]
    for label, names in table_groups:
        section = admin_table_sidebar_section(repo, env, label, names, active_path)
        if section:
            sections.append(section)
    other_tables = tuple(table.name for table in TABLES if table.name not in grouped_tables)
    other_section = admin_table_sidebar_section(repo, env, "其他数据", other_tables, active_path)
    if other_section:
        sections.append(other_section)
    account = ""
    if user:
        account = f"""<div class="admin-account">
          <span class="admin-account-avatar">{esc(admin_user_initial(user))}</span>
          <div><strong>{esc(user.get("display_name") or user.get("username"))}</strong><span>{esc(role.get("name") or "")}</span></div>
          <a href="/admin/logout">退出</a>
        </div>"""
    side = f"""<button class="admin-sidebar-toggle" type="button" data-admin-nav-toggle aria-controls="admin-sidebar-panel" aria-expanded="false">后台导航</button>
    <aside class="admin-sidebar" id="admin-sidebar-panel">
      <div class="admin-sidebar-brand"><span>TS</span><div><strong>管理后台</strong><small>Teacher Site</small></div></div>
      <nav class="admin-sidebar-nav" aria-label="后台导航">{"".join(sections)}</nav>
      {account}
    </aside>"""
    trail = admin_breadcrumbs(trail_items)
    return layout(repo, title, f'<div class="admin-shell">{side}<section class="admin-main">{trail}{content}</section></div>', env)


def admin_active_path(breadcrumbs: list[tuple[str, str]]) -> str:
    if len(breadcrumbs) > 1 and breadcrumbs[1][1].startswith("/admin/table/"):
        return breadcrumbs[1][1]
    if breadcrumbs:
        return breadcrumbs[-1][1]
    return "/admin"


def admin_sidebar_section(label: str, links: str) -> str:
    if not links:
        return ""
    return f'<section class="admin-sidebar-section"><h2>{esc(label)}</h2>{links}</section>'


def admin_table_sidebar_section(repo: Repository, env: dict[str, str], label: str, names: tuple[str, ...], active_path: str) -> str:
    links = []
    for name in names:
        meta = TABLE_MAP.get(name)
        if meta and has_permission(repo, env, name, "can_view"):
            links.append(admin_sidebar_link(f"/admin/table/{esc(name)}", meta.label, admin_nav_icon(name), active_path))
    return admin_sidebar_section(label, "".join(links))


def admin_sidebar_link(href: str, label: str, icon: str, active_path: str) -> str:
    active = admin_sidebar_active(href, active_path)
    return f'<a class="admin-sidebar-link{" is-active" if active else ""}" href="{esc(href)}"><span class="admin-nav-icon">{esc(icon)}</span><span>{esc(label)}</span></a>'


def admin_sidebar_active(href: str, active_path: str) -> bool:
    if href == "/admin":
        return active_path == "/admin"
    return active_path == href or active_path.startswith(href + "/")


def admin_nav_icon(table_or_key: str) -> str:
    return {
        "site_settings": "站",
        "global_settings": "全",
        "navigation_items": "航",
        "profiles": "师",
        "students": "生",
        "student_category_displays": "组",
        "messages": "信",
        "research_interests": "向",
        "publications": "文",
        "projects": "项",
        "patents": "专",
        "news": "动",
        "courses": "课",
        "media_assets": "媒",
        "translation_cache": "译",
        "autofetch_logs": "取",
        "operation_logs": "审",
        "auth_roles": "角",
        "auth_users": "户",
        "auth_permissions": "限",
    }.get(table_or_key, "表")


def admin_user_initial(user: dict[str, Any]) -> str:
    name = text_only(user.get("display_name") or user.get("username"), 80).strip()
    return (name[:1] or "管").upper()


def admin_setup_route(repo: Repository, method: str, query: dict[str, str], body: bytes, env: dict[str, str]) -> ResponseTuple:
    if auth_users_exist(repo):
        return redirect("/admin/login")
    error = ""
    if method == "POST":
        data = _form(body)
        username = text_only(data.get("username"), 80).strip()
        password = str(data.get("password") or "")
        password_confirm = str(data.get("password_confirm") or "")
        if not username:
            error = "请填写登录账号。"
        elif len(password) < 10:
            error = "密码至少需要 10 位。"
        elif password != password_confirm:
            error = "两次输入的密码不一致。"
        else:
            user = {
                "uid": stable_uid("user", username.lower()),
                "username": username,
                "password_hash": hash_password(password),
                "display_name": text_only(data.get("display_name") or username, 80),
                "email": text_only(data.get("email"), 160),
                "role_uid": "role-super-admin",
                "status": "active",
                "must_change_password": 0,
                "last_login_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "visibility": "owner",
            }
            saved = repo.save("auth_users", user)
            return redirect_with_cookie("/admin", auth_cookie(repo, env, saved))
    return html_response(admin_auth_page(repo, env, "初始化高级管理员", f"""
      <form class="auth-panel" method="post" action="/admin/setup">
        <h1>初始化高级管理员</h1>
        <p>当前数据库没有有效账号。请创建第一个高级管理员；系统不会预置默认弱密码。</p>
        {f'<p class="auth-error">{esc(error)}</p>' if error else ''}
        <label>登录账号<input name="username" autocomplete="username" required></label>
        <label>显示名称<input name="display_name" autocomplete="name"></label>
        <label>邮箱<input name="email" type="email" autocomplete="email"></label>
        <label>密码<input name="password" type="password" autocomplete="new-password" required minlength="10"></label>
        <label>确认密码<input name="password_confirm" type="password" autocomplete="new-password" required minlength="10"></label>
        <button type="submit">创建并登录</button>
      </form>"""))


def admin_login_route(repo: Repository, method: str, query: dict[str, str], body: bytes, env: dict[str, str]) -> ResponseTuple:
    if not auth_users_exist(repo):
        return redirect("/admin/setup")
    next_url = safe_admin_next(query.get("next") or "/admin")
    return redirect(lang_url(f"/login?next={quote(next_url)}", env))


def front_login_route(repo: Repository, method: str, query: dict[str, str], body: bytes, env: dict[str, str]) -> ResponseTuple:
    if not auth_users_exist(repo):
        if login_next_targets_admin(query.get("next")):
            return redirect(lang_url("/admin/setup", env))
        return html_response(layout(repo, t(env, "login"), f'<section class="notice"><h1>{esc(t(env, "login_unavailable"))}</h1><p>{esc(t(env, "admin_not_initialized"))}</p></section>', env), 404)
    error = ""
    next_url = safe_local_next(query.get("next") or lang_url("/", env))
    if method == "POST":
        data = _form(body)
        username = text_only(data.get("username"), 80).strip().lower()
        password = str(data.get("password") or "")
        wait_seconds = login_rate_wait_seconds(username, env)
        if wait_seconds > 0:
            minutes = max(1, (wait_seconds + 59) // 60)
            error = t(env, "login_rate_limited").format(minutes=minutes)
        else:
            user = next((row for row in repo.list("auth_users", Query(limit=1000)) if str(row.get("username") or "").lower() == username), {})
            if not user or str(user.get("status") or "active") != "active" or not verify_password(password, user.get("password_hash")):
                record_login_failure(username, env)
                error = t(env, "login_failed")
            else:
                clear_login_failures(username, env)
                user["last_login_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
                saved = repo.save("auth_users", user)
                return redirect_with_cookie(next_url, auth_cookie(repo, env, saved))
    register_hint = ""
    if public_registration_allowed(repo):
        register_hint = f'<div class="auth-secondary auth-register-action"><span>{esc(t(env, "no_account"))}</span><a class="button ghost" href="{esc(lang_url("/register", env))}">{esc(t(env, "register"))}</a></div>'
    else:
        register_hint = f'<p class="auth-secondary">{esc(t(env, "registration_closed_hint"))}</p>'
    return html_response(layout(repo, t(env, "login"), f"""
      <section class="auth-page">
        <form class="auth-panel" method="post" action="{esc(lang_url(f"/login?next={quote(next_url)}", env))}">
          <h1>{esc(t(env, "login"))}</h1>
          <p>{esc(t(env, "login_scope_hint"))}</p>
          {f'<p class="auth-error">{esc(error)}</p>' if error else ''}
          <label>{esc(t(env, "username"))}<input name="username" autocomplete="username" placeholder="{esc(t(env, "username_placeholder"))}" required></label>
          <label>{esc(t(env, "password"))}<input name="password" type="password" autocomplete="current-password" placeholder="{esc(t(env, "password_placeholder"))}" required></label>
          <button type="submit">{esc(t(env, "login"))}</button>
          {register_hint}
        </form>
      </section>""", env))


def front_register_route(repo: Repository, method: str, query: dict[str, str], body: bytes, env: dict[str, str]) -> ResponseTuple:
    if not auth_users_exist(repo):
        return html_response(layout(repo, t(env, "register"), f'<section class="notice"><h1>{esc(t(env, "registration_unavailable"))}</h1><p>{esc(t(env, "admin_not_initialized"))}</p></section>', env), 404)
    if not public_registration_allowed(repo):
        return html_response(layout(repo, t(env, "register"), f"""
          <section class="auth-page">
            <div class="auth-panel">
              <h1>{esc(t(env, "registration_closed"))}</h1>
              <p>{esc(t(env, "registration_closed_message"))}</p>
              <a class="button ghost" href="{esc(lang_url("/login", env))}">{esc(t(env, "back_login"))}</a>
            </div>
          </section>""", env), 403)
    error = ""
    if method == "POST":
        data = _form(body)
        username = text_only(data.get("username"), 80).strip()
        username_key = username.lower()
        password = str(data.get("password") or "")
        password_confirm = str(data.get("password_confirm") or "")
        existing = next((row for row in repo.list("auth_users", Query(limit=1000)) if str(row.get("username") or "").lower() == username_key), {})
        if not username:
            error = t(env, "username_required")
        elif existing:
            error = t(env, "username_taken")
        elif len(password) < 10:
            error = t(env, "password_too_short")
        elif password != password_confirm:
            error = t(env, "password_mismatch")
        else:
            user = {
                "uid": stable_uid("user", username_key),
                "username": username,
                "password_hash": hash_password(password),
                "display_name": text_only(data.get("display_name") or username, 80),
                "email": text_only(data.get("email"), 160),
                "role_uid": "role-visitor",
                "status": "active",
                "must_change_password": 0,
                "last_login_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                "visibility": "authenticated",
            }
            saved = repo.save("auth_users", user)
            return redirect_with_cookie(lang_url("/", env), auth_cookie(repo, env, saved))
    return html_response(layout(repo, t(env, "register"), f"""
      <section class="auth-page">
        <form class="auth-panel" method="post" action="{esc(lang_url("/register", env))}">
          <h1>{esc(t(env, "register"))}</h1>
          <p>{esc(t(env, "register_scope_hint"))}</p>
          {f'<p class="auth-error">{esc(error)}</p>' if error else ''}
          <label>{esc(t(env, "username"))}<input name="username" autocomplete="username" placeholder="{esc(t(env, "register_username_placeholder"))}" required></label>
          <label>{esc(t(env, "display_name"))}<input name="display_name" autocomplete="name" placeholder="{esc(t(env, "display_name_placeholder"))}"></label>
          <label>{esc(t(env, "email"))}<input name="email" type="email" autocomplete="email" placeholder="{esc(t(env, "email_placeholder"))}"></label>
          <label>{esc(t(env, "password"))}<input name="password" type="password" autocomplete="new-password" placeholder="{esc(t(env, "new_password_placeholder"))}" required minlength="10"></label>
          <label>{esc(t(env, "password_confirm"))}<input name="password_confirm" type="password" autocomplete="new-password" placeholder="{esc(t(env, "password_confirm_placeholder"))}" required minlength="10"></label>
          <button type="submit">{esc(t(env, "register"))}</button>
          <p class="auth-secondary"><a href="{esc(lang_url("/login", env))}">{esc(t(env, "back_login"))}</a></p>
        </form>
      </section>""", env))


def safe_admin_next(value: Any) -> str:
    target = str(value or "/admin")
    return target if target.startswith("/admin") and not target.startswith("//") else "/admin"


def safe_local_next(value: Any) -> str:
    target = str(value or "/")
    if target.startswith("/") and not target.startswith("//"):
        return target
    return "/"


def login_next_targets_admin(value: Any) -> bool:
    target = safe_local_next(value)
    parsed = urlparse(target)
    return parsed.path == "/admin" or parsed.path.startswith("/admin/")


def public_registration_allowed(repo: Repository) -> bool:
    return truthy(active_global(repo).get("allow_public_registration"), default=False)


def admin_auth_page(repo: Repository, env: dict[str, str], title: str, content: str) -> str:
    return layout(repo, title, f'<section class="auth-page">{content}</section>', env)


def admin_permissions_home(repo: Repository, env: dict[str, str]) -> str:
    roles = repo.list("auth_roles", Query(limit=1000, order_by="sort_order", descending=False))
    users = repo.list("auth_users", Query(limit=1000, order_by="updated_at", descending=True))
    perms = repo.list("auth_permissions", Query(limit=1000))
    role_cards = []
    for role in roles:
        role_uid = str(role.get("uid") or "")
        role_users = sum(1 for user in users if str(user.get("role_uid") or "") == role_uid)
        enabled_perms = sum(1 for perm in perms if str(perm.get("role_uid") or "") == role_uid and any(truthy(perm.get(flag), default=False) for flag in ("can_view", "can_create", "can_edit", "can_delete", "can_export")))
        role_cards.append(f"""<article class="permission-role-card">
          <strong>{esc(role.get("name") or role_uid)}</strong>
          <span>层级 {esc(role.get("level") or 0)} · {role_users} 个账号 · {enabled_perms} 条授权</span>
          <small>{esc(role.get("description") or "")}</small>
        </article>""")
    return f"""<section class="admin-card permission-home">
      <div class="admin-card-head"><div><h1>权限管理</h1><p class="admin-muted">角色决定后台模块访问与编辑权限；前台内容仍沿用原有“可见范围”字段，并按登录账号角色层级放行。</p></div></div>
      <div class="permission-actions">
        <a class="button" href="/admin/table/auth_users">用户账号</a>
        <a class="button ghost" href="/admin/table/auth_roles">权限角色</a>
        <a class="button ghost" href="/admin/table/auth_permissions">角色权限</a>
      </div>
      <div class="permission-role-grid">{''.join(role_cards) or empty(env)}</div>
    </section>"""


def admin_denied_panel(title: str, message: str) -> str:
    return f'<section class="admin-card admin-denied"><div><h1>{esc(title)}</h1><p>{esc(message)}</p><a class="button ghost" href="/admin">返回后台首页</a></div></section>'


def admin_denied_html(repo: Repository, env: dict[str, str], title: str, message: str) -> str:
    return admin_layout(repo, title, admin_denied_panel(title, message), env, [("后台", "/admin"), (title, env.get("_PATH") or "/admin")])


def admin_action_label(action: str) -> str:
    return {"can_view": "查看", "can_create": "新增", "can_edit": "编辑", "can_delete": "删除", "can_export": "导出"}.get(action, "操作")


def admin_breadcrumbs(items: list[tuple[str, str]]) -> str:
    links = []
    for index, (label, href) in enumerate(items):
        if index == len(items) - 1:
            links.append(f'<span aria-current="page">{esc(label)}</span>')
        else:
            links.append(f'<a href="{esc(href)}">{esc(label)}</a>')
    return f'<nav class="admin-breadcrumbs" aria-label="当前链路路径">{"<span>/</span>".join(links)}</nav>'


def admin_row_title(meta: Table, row: dict[str, Any]) -> str:
    return str(row.get(meta.title_field) or row.get("title") or row.get("name") or row.get("uid") or row.get("id") or "记录")


def admin_new_row(repo: Repository, table: str) -> dict[str, Any]:
    if table == "site_settings":
        return {
            "uid": "site-default",
            "is_active": 1,
            "site_name": "教师个人与团队网站",
            "site_name_en": "Teacher and Research Group",
            "hero_title": "",
            "hero_subtitle": "",
            "logo_key": "default/site-logo.png",
            "favicon_key": "default/site-logo.png",
            "og_image_key": "default/site-logo.png",
            "homepage_profile_uid": "profile-main-teacher",
            "homepage_publication_limit": 5,
            "homepage_news_limit": 4,
        }
    if table == "global_settings":
        return {
            "uid": "global-default",
            "allow_public_registration": 0,
            "allow_anonymous_messages": 1,
            "upload_max_size_mb": 10,
            "upload_allowed_extensions": ".jpg,.jpeg,.png,.webp,.svg,.pdf,.doc,.docx,.xls,.xlsx,.csv",
            "media_trash_retention_days": 30,
            "news_pdf_engine": "native",
            "news_pdf_allow_download": 1,
            "translation_provider": "auto",
            "translation_providers": TRANSLATION_FALLBACK_PROVIDERS,
            "mymemory_email": "",
            "microsoft_translator_endpoint": MICROSOFT_TRANSLATOR_DEFAULT_ENDPOINT,
            "translation_batch_size": 10,
            "translation_worker_count": 4,
            "translation_timeout_seconds": 12,
            "publication_metadata_provider": "manual",
            "publication_metadata_providers": "crossref,openalex,semanticscholar",
            "publication_display_style": "gbt",
            "patent_metadata_providers": "patentsview,epo_ops",
            "publication_suggestion_cache_seconds": 30,
            "profile_suggestion_cache_seconds": 30,
            "project_suggestion_cache_seconds": 30,
            "patent_suggestion_cache_seconds": 30,
            "student_suggestion_cache_seconds": 30,
            "news_suggestion_cache_seconds": 30,
            "course_suggestion_cache_seconds": 30,
        }
    if table == "publications":
        return {"uid": next_publication_uid(repo), "visibility": "public", "is_featured": 0}
    if table == "projects":
        return {"uid": next_table_uid(repo, "projects", "project"), "visibility": "public", "is_featured": 0, "status": "在研"}
    if table == "patents":
        return {"uid": next_table_uid(repo, "patents", "patent"), "visibility": "public", "is_featured": 0, "country": "中国", "legal_status": "申请中"}
    if table == "students":
        return {"uid": next_table_uid(repo, "students", "student"), "visibility": "public", "contact_visibility": "public", "is_featured": 0, "status": "在读"}
    if table == "student_category_displays":
        return {"uid": next_table_uid(repo, "student_category_displays", "cat"), "enabled": 1, "display_order": next_display_order(repo, "student_category_displays")}
    if table == "news":
        uid = next_table_uid(repo, "news", "news")
        return {"uid": uid, "slug": uid, "content_format": "plain", "allow_comments": 1, "visibility": "public", "is_featured": 0, "published_at": time.strftime("%Y-%m-%d"), "sort_order": next_sort_order(repo, "news")}
    if table == "courses":
        return {"uid": next_table_uid(repo, "courses", "course"), "material_visibility": "public", "visibility": "public", "is_featured": 0, "sort_order": next_sort_order(repo, "courses")}
    if table == "messages":
        return {"uid": next_table_uid(repo, "messages", "msg"), "message_type": "other", "status": "new", "visibility": "staff"}
    if table == "translation_cache":
        return {"uid": next_table_uid(repo, "translation_cache", "tr"), "source_lang": "zh", "target_lang": "en", "provider": "manual", "status": "pending", "is_manual": 1, "is_current": 1}
    if table == "auth_roles":
        return {"uid": next_table_uid(repo, "auth_roles", "role"), "level": 20, "visibility_scopes": "public,authenticated", "is_system": 0, "is_active": 1, "sort_order": next_sort_order(repo, "auth_roles")}
    if table == "auth_users":
        return {"uid": next_table_uid(repo, "auth_users", "user"), "role_uid": "role-visitor", "status": "active", "must_change_password": 1, "visibility": "authenticated"}
    if table == "auth_permissions":
        return {"uid": next_table_uid(repo, "auth_permissions", "perm"), "role_uid": "role-visitor", "module": "admin", "can_view": 0, "can_create": 0, "can_edit": 0, "can_delete": 0, "can_export": 0, "sort_order": next_sort_order(repo, "auth_permissions")}
    return {}


def next_publication_uid(repo: Repository) -> str:
    return next_table_uid(repo, "publications", "pub")


def next_table_uid(repo: Repository, table: str, prefix: str) -> str:
    for attempt in range(20):
        uid = stable_uid(prefix, f"{time.time_ns()}-{attempt}")
        if not repo.get(table, uid):
            return uid
    return f"{prefix}-{time.time_ns()}"


def next_display_order(repo: Repository, table: str) -> int:
    rows = repo.list(table, Query(limit=1000, order_by="display_order", descending=True))
    if not rows:
        return 1
    return int_value(rows[0].get("display_order"), 0) + 1


def next_sort_order(repo: Repository, table: str) -> int:
    rows = repo.list(table, Query(limit=1000, order_by="sort_order", descending=True))
    if not rows:
        return 1
    return int_value(rows[0].get("sort_order"), 0) + 1


def admin_table(meta: Table, rows: list[dict[str, Any]], q: str) -> str:
    head = "".join(f"<th>{esc(field)}</th>" for field in meta.list_fields)
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        cells = "".join(f"<td>{esc(row.get(field))}</td>" for field in meta.list_fields)
        body.append(f'<tr>{cells}<td><a href="/admin/table/{esc(meta.name)}/{esc(key)}">编辑</a></td></tr>')
    return f"""<section class="admin-card"><div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/{esc(meta.name)}/new">新增</a></div>
    <form class="filters"><input name="q" value="{esc(q)}" placeholder="搜索"><button>搜索</button></form>
    <div class="table-wrap"><table><thead><tr>{head}<th>操作</th></tr></thead><tbody>{''.join(body) or '<tr><td>暂无数据</td></tr>'}</tbody></table></div></section>"""


def admin_import_export_page(repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    platform = text_only(env.get("PLATFORM") or "ubuntu", 40) or "ubuntu"
    is_cloudflare = platform == "cloudflare"
    scanned = query.get("scan") == "1"
    counts = repo.counts() if scanned else {}
    note = import_export_notice(query)
    group_cards = []
    for group_key, group_label, tables in EXPORT_TABLE_GROUPS:
        table_links = "".join(export_table_chip(table, counts.get(table) if scanned else None) for table in tables)
        group_tables = ",".join(tables)
        group_count = f'{sum(counts.get(table, 0) for table in tables)} 条记录' if scanned else "未扫描"
        group_cards.append(f"""<article class="export-group-card">
          <div><strong>{esc(group_label)}</strong><span>{esc(group_count)}</span></div>
          <p>{table_links}</p>
          <div class="export-actions">
            <a class="button ghost" href="/api/export/site.json?tables={esc(group_tables)}">JSON</a>
            <a class="button light" href="/api/export/main.xlsx?tables={esc(group_tables)}" {'aria-disabled="true" tabindex="-1"' if is_cloudflare else ''}>Excel</a>
          </div>
        </article>""")
    import_controls = import_restore_form(repo, is_cloudflare)
    platform_note = "Cloudflare Worker 环境适合轻量 JSON 导出；批量恢复导入、文件上传解析和 Excel 生成已灰色处理，建议在 Ubuntu/本地环境执行。" if is_cloudflare else "当前为本地/Ubuntu 运行模式，可执行 JSON 导出、CSV/Excel 导出，以及上传 JSON/CSV/Excel 或粘贴 JSON 恢复导入。"
    main_count = str(sum(counts.get(table, 0) for table in EXPORT_MAIN_TABLES)) if scanned else "未扫描"
    media_count = str(counts.get("media_assets", 0)) if scanned else "未扫描"
    translation_count = str(counts.get("translation_cache", 0)) if scanned else "未扫描"
    scan_hint = "已手动扫描当前数据量。" if scanned else "为减少数据库读取，当前不自动统计记录数。"
    return f"""<section class="admin-card import-export-card">
      <div class="admin-card-head"><div><h1>导入与导出</h1><p class="admin-muted">用于 Cloudflare Worker 与 Ubuntu 之间迁移教师、论文、项目、专利、学生、动态、课程、留言、媒体索引和翻译缓存等主要数据。</p></div></div>
      {note}
      <div class="platform-capability"><strong>运行平台：{esc(platform)}</strong><span>{esc(platform_note)}</span></div>
      <div class="export-scan-row"><span>{esc(scan_hint)}</span><a class="button ghost" href="/admin/export?scan=1">扫描当前数据量</a></div>
      <div class="export-overview">
        <div><strong>{len(EXPORT_MAIN_TABLES)}</strong><span>主要数据表</span></div>
        <div><strong>{esc(main_count)}</strong><span>主要记录数</span></div>
        <div><strong>{esc(media_count)}</strong><span>媒体索引</span></div>
        <div><strong>{esc(translation_count)}</strong><span>翻译缓存</span></div>
      </div>
      <section class="export-section">
        <h2>集成导出</h2>
        <div class="export-action-grid">
          <a class="export-action-card" href="/api/export/main.json"><strong>主要数据 JSON</strong><span>推荐迁移格式，不包含用户密码哈希和权限账号。</span></a>
          <a class="export-action-card" href="/api/export/site.json"><strong>整站 JSON</strong><span>包含所有表，适合高级备份；请妥善保管权限账号数据。</span></a>
          <a class="export-action-card{' is-disabled' if is_cloudflare else ''}" href="/api/export/main.xlsx" {'aria-disabled="true" tabindex="-1"' if is_cloudflare else ''}><strong>主要数据 Excel</strong><span>适合人工检查和小规模整理；Cloudflare 上禁用。</span></a>
          <a class="export-action-card" href="/api/export/i18n-dictionary.json"><strong>手动词典 JSON</strong><span>只导出根目录/R2 中的中英互译词典，不包含翻译缓存。</span></a>
        </div>
      </section>
      <section class="export-section">
        <h2>分组导出</h2>
        <div class="export-group-grid">{"".join(group_cards)}</div>
      </section>
      <section class="export-section">
        <h2>分表导出</h2>
        <div class="export-table-grid">{"".join(export_table_row(table, counts.get(table) if scanned else None) for table in EXPORT_MAIN_TABLES)}</div>
      </section>
      {import_controls}
    </section>"""


def import_export_notice(query: dict[str, str]) -> str:
    if query.get("imported"):
        backup = query.get("backup", "")
        backup_text = f" 自动备份：{backup}。" if backup else ""
        return f'<p class="auth-success">已导入 {esc(query.get("imported"))} 条记录，跳过 {esc(query.get("skipped", "0"))} 条。{esc(backup_text)}</p>'
    if query.get("dry_run"):
        return f'<p class="auth-success">导入预演完成：涉及 {esc(query.get("tables", "0"))} 个表，计划写入 {esc(query.get("planned", "0"))} 条，新建 {esc(query.get("create", "0"))} 条，更新 {esc(query.get("update", "0"))} 条，替换模式将删除 {esc(query.get("delete", "0"))} 条，跳过 {esc(query.get("skipped", "0"))} 条。确认无误后请再次提交“确认导入”。</p>'
    if query.get("error"):
        return f'<p class="auth-error">{esc(query.get("error"))}</p>'
    return ""


def export_count_text(count: int | None) -> str:
    return str(count) if count is not None else "未扫描"


def export_table_chip(table: str, count: int | None) -> str:
    meta = TABLE_MAP.get(table)
    return f'<span class="export-table-chip">{esc((meta.label if meta else table))}<b>{esc(export_count_text(count))}</b></span>'


def export_table_row(table: str, count: int | None) -> str:
    meta = TABLE_MAP[table]
    return f"""<article class="export-table-row">
      <div><strong>{esc(meta.label)}</strong><span>{esc(table)} · {esc(export_count_text(count))}</span></div>
      <p>{esc(", ".join(meta.field_names[:6]))}</p>
      <div><a class="button ghost" href="/api/export/table.json?table={esc(table)}">JSON</a><a class="button light" href="/api/export/table.csv?table={esc(table)}">CSV</a></div>
    </article>"""


def import_restore_form(repo: Repository, is_cloudflare: bool) -> str:
    disabled = ' disabled aria-disabled="true"' if is_cloudflare else ""
    group_checks = []
    for _group_key, group_label, tables in EXPORT_TABLE_GROUPS:
        checks = "".join(f'<label class="restore-check"><input type="checkbox" name="tables" value="{esc(table)}" checked{disabled}><span>{esc(TABLE_MAP[table].label)}</span><small>{esc(table)}</small></label>' for table in tables if table in EXPORT_MAIN_TABLES)
        group_checks.append(f'<fieldset><legend>{esc(group_label)}</legend><div>{checks}</div></fieldset>')
    table_options = "".join(f'<option value="{esc(table)}">{esc(TABLE_MAP[table].label)} / {esc(table)}</option>' for table in EXPORT_MAIN_TABLES)
    disabled_note = '<p class="restore-note">Cloudflare Worker 环境暂不执行批量恢复导入和上传解析，避免单次请求写入过多 D1 数据。请先导出 JSON 后在 Ubuntu/本地环境恢复。</p>' if is_cloudflare else '<p class="restore-note">推荐上传从本中心导出的 JSON、CSV 或 Excel；未上传文件时，才读取下方粘贴的 JSON。</p>'
    return f"""<section class="export-section import-section">
      <div class="restore-section-head"><div><h2>恢复导入</h2>{disabled_note}</div><span>文件不进入媒体库 · 只在提交时解析</span></div>
      <form method="post" action="/admin/export" class="import-restore-form" enctype="multipart/form-data">
        <section class="restore-step-card restore-scope-card">
          <div class="restore-step-head"><b>1</b><div><strong>选择恢复范围</strong><span>按分组勾选需要写入的数据表</span></div></div>
          <div class="import-table-checks">{"".join(group_checks)}</div>
        </section>
        <section class="restore-step-card restore-source-card">
          <div class="restore-step-head"><b>2</b><div><strong>选择数据来源</strong><span>优先上传文件，也可粘贴 JSON 备用</span></div></div>
          <div class="restore-format-hints"><span>JSON 多表</span><span>CSV 单表</span><span>Excel 工作表</span></div>
          <div class="import-options-grid">
            <label class="import-file-field"><span>上传文件</span><input type="file" name="restore_file" accept=".json,.csv,.xlsx,application/json,text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"{disabled}><small class="field-help">支持 .json、.csv、.xlsx；文件只在提交时解析。</small></label>
            <label><span>文件类型</span><select name="source_type"{disabled}><option value="auto">自动识别</option><option value="json">JSON</option><option value="csv">CSV</option><option value="xlsx">Excel xlsx</option></select><small class="field-help">文件扩展名不明确时，可手动指定。</small></label>
            <label><span>CSV/单表目标</span><select name="target_table"{disabled}><option value="">自动/不指定</option>{table_options}</select><small class="field-help">CSV 或单工作表 Excel 建议指定目标表。</small></label>
            <label class="import-json-field"><span>粘贴 JSON 备用</span><textarea name="payload" rows="5" placeholder='粘贴 {{"tables":{{"profiles":[...]}}}} 或 {{"profiles":[...]}}'{disabled}></textarea><small class="field-help">未上传文件时读取这里；不会恢复权限账号表。</small></label>
          </div>
        </section>
        <section class="restore-step-card restore-mode-card">
          <div class="restore-step-head"><b>3</b><div><strong>确认恢复方式</strong><span>优先使用合并更新，替换需谨慎</span></div></div>
          <label><span>导入模式</span><select name="mode"{disabled}><option value="merge">合并更新</option><option value="replace">替换所选表</option></select><small class="field-help">合并按 uid/id 覆盖同名记录；替换会先清空所选表。</small></label>
          <label class="restore-confirm-check"><input type="checkbox" name="confirm_restore" value="1"{disabled}><span>我已完成预演并确认本次恢复范围，允许系统先自动备份再写入数据。</span></label>
          <div class="restore-risk-box"><strong>恢复前建议</strong><span>先点击“预演导入”查看新建、更新、删除数量；确认导入前系统会自动保存本次涉及表的 JSON 备份。</span></div>
        </section>
        <div class="restore-submit-bar"><span>预演不会写库；确认导入需勾选确认项，并会先自动备份。</span><button class="button light" type="submit" name="restore_action" value="dry_run"{disabled}>预演导入</button><button type="submit" name="restore_action" value="restore"{disabled}>确认导入</button></div>
      </form>
    </section>"""


def export_api_route(repo: Repository, path: str, query: dict[str, str], env: dict[str, str]) -> ResponseTuple:
    if path.endswith("/i18n-dictionary.json"):
        payload = i18n_dictionary_current_payload(env)
        return binary_response(payload["content"], "application/json; charset=utf-8", I18N_DICTIONARY_FILENAME)
    if path.endswith("/main.json"):
        tables = export_selected_tables(query.get("tables"), default=EXPORT_MAIN_TABLES)
        return download_json_response(export_payload(repo, tables), "teacher-site-main.json")
    if path.endswith("/site.json"):
        tables = export_selected_tables(query.get("tables"), default=tuple(repo.table_names()))
        return download_json_response(export_payload(repo, tables), "teacher-site-site.json")
    if path.endswith("/table.json"):
        table = text_only(query.get("table"), 80).strip()
        if table not in TABLE_MAP:
            return json_response({"ok": False, "message": "未知数据表。"}, 404)
        return download_json_response(export_payload(repo, (table,)), f"teacher-site-{table}.json")
    if path.endswith("/table.csv"):
        table = text_only(query.get("table"), 80).strip()
        if table not in TABLE_MAP:
            return json_response({"ok": False, "message": "未知数据表。"}, 404)
        rows = repo.list(table, Query(limit=1000))
        return binary_response(csv_bytes(rows, TABLE_MAP[table].field_names), "text/csv; charset=utf-8", f"teacher-site-{table}.csv")
    if path.endswith("/main.xlsx"):
        if env.get("PLATFORM") == "cloudflare":
            return json_response({"ok": False, "message": "Cloudflare Worker 环境不生成 Excel。"}, 501)
        tables = export_selected_tables(query.get("tables"), default=EXPORT_MAIN_TABLES)
        try:
            return binary_response(excel_bytes(ExportViewRepository(repo, tables)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "teacher-site-main.xlsx")
        except Exception as exc:
            return json_response({"ok": False, "message": f"Excel 导出不可用：{exc}"}, 500)
    return json_response({"ok": False, "message": "未知导出接口。"}, 404)


def export_selected_tables(raw: Any, default: tuple[str, ...]) -> tuple[str, ...]:
    selected = [item.strip() for item in text_only(raw, 2000).split(",") if item.strip()]
    valid = tuple(table for table in selected if table in TABLE_MAP)
    return valid or tuple(table for table in default if table in TABLE_MAP)


def export_payload(repo: Repository, tables: tuple[str, ...]) -> dict[str, Any]:
    payload = export_json(ExportViewRepository(repo, tables))
    payload["tables"] = {table: payload["tables"].get(table, []) for table in tables}
    payload["table_labels"] = {table: TABLE_MAP[table].label for table in tables}
    payload["export_scope"] = "selected"
    payload["exported_at"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    return payload


class ExportViewRepository:
    def __init__(self, repo: Repository, tables: tuple[str, ...]):
        self.repo = repo
        self.tables = tuple(table for table in tables if table in TABLE_MAP)

    def list(self, table: str, query: Query | None = None) -> list[dict[str, Any]]:
        if table not in self.tables:
            return []
        return self.repo.list(table, query)


def download_json_response(data: Any, filename: str) -> ResponseTuple:
    return binary_response(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"), "application/json; charset=utf-8", filename)


def binary_response(content: bytes, content_type: str, filename: str, status: int = 200) -> ResponseTuple:
    return status, security_headers() + [("content-type", content_type), ("content-disposition", f'attachment; filename="{filename}"')], content


def import_restore_payload(repo: Repository, body: bytes, env: dict[str, str]) -> dict[str, Any]:
    if env.get("PLATFORM") == "cloudflare":
        return {"error": "Cloudflare Worker 环境暂不执行批量恢复导入。"}
    if len(body) > 12 * 1024 * 1024:
        return {"error": "导入文件超过 12MB，请拆分为分组或分表文件后恢复。"}
    data = import_restore_form_data(body, env)
    mode = text_only(import_form_value(data, "mode", "merge"), 40).strip() or "merge"
    selected = [table for table in import_form_values(data, "tables") if table in EXPORT_MAIN_TABLES]
    if not selected:
        return {"error": "请选择至少一个需要恢复的数据表。"}
    source_type = text_only(import_form_value(data, "source_type", "auto"), 40).strip().lower() or "auto"
    target_table = text_only(import_form_value(data, "target_table", ""), 80).strip()
    submit_action = text_only(import_form_value(data, "restore_action", "dry_run"), 40).strip()
    confirmed = truthy(import_form_value(data, "confirm_restore", ""), default=False)
    upload = import_form_file(data, "restore_file")
    raw_payload = import_form_value(data, "payload", "")
    parsed_result = parse_import_restore_source(upload, raw_payload, source_type, target_table, selected)
    if parsed_result.get("error"):
        return {"error": str(parsed_result["error"])}
    tables_data = parsed_result.get("tables")
    if not isinstance(tables_data, dict):
        return {"error": "导入内容中没有可识别的数据表。"}
    restore_tables = [table for table in selected if table in tables_data]
    if not restore_tables:
        return {"error": "导入内容没有匹配已勾选的数据表。"}
    if mode not in {"merge", "replace"}:
        mode = "merge"
    preview = analyze_import_restore(repo, tables_data, restore_tables, mode)
    if submit_action != "restore" or not confirmed:
        return {
            "dry_run": 1,
            "planned": preview["planned"],
            "create": preview["create"],
            "update": preview["update"],
            "delete": preview["delete"],
            "skipped": preview["skipped"],
            "tables": len(restore_tables),
        }
    backup_path = create_restore_backup(repo, tuple(restore_tables))
    if not backup_path:
        return {"error": "自动备份失败，已取消恢复导入。请检查 exports/backups 写入权限。"}
    imported = 0
    skipped = 0
    for table in restore_tables:
        rows = tables_data.get(table)
        if not isinstance(rows, list):
            skipped += 1
            continue
        if mode == "replace":
            for old in repo.list(table, Query(limit=1000)):
                old_key = str(old.get("uid") or old.get("id") or "").strip()
                if old_key:
                    repo.delete(table, old_key)
        meta = TABLE_MAP[table]
        for row in rows[:1000]:
            if not isinstance(row, dict):
                skipped += 1
                continue
            clean = {field.name: row.get(field.name, "") for field in meta.fields if field.name in row}
            if not clean.get("uid"):
                seed = import_row_uid_seed(row)
                if seed:
                    clean["uid"] = stable_uid(table.rstrip("s"), seed)
            if not clean.get("uid"):
                skipped += 1
                continue
            repo.save(table, normalize_import_row(meta, clean))
            imported += 1
    return {"imported": imported, "skipped": skipped, "backup": backup_path.name if backup_path else ""}


def analyze_import_restore(repo: Repository, tables_data: dict[str, Any], restore_tables: list[str], mode: str) -> dict[str, int]:
    planned = 0
    create = 0
    update = 0
    delete = 0
    skipped = 0
    for table in restore_tables:
        rows = tables_data.get(table)
        if not isinstance(rows, list):
            skipped += 1
            continue
        if mode == "replace":
            delete += len(repo.list(table, Query(limit=1000)))
        meta = TABLE_MAP[table]
        for row in rows[:1000]:
            if not isinstance(row, dict):
                skipped += 1
                continue
            clean = {field.name: row.get(field.name, "") for field in meta.fields if field.name in row}
            uid = str(clean.get("uid") or "").strip()
            if not uid:
                seed = import_row_uid_seed(row)
                uid = stable_uid(table.rstrip("s"), seed) if seed else ""
            if not uid:
                skipped += 1
                continue
            planned += 1
            if repo.get(table, uid):
                update += 1
            else:
                create += 1
    return {"planned": planned, "create": create, "update": update, "delete": delete, "skipped": skipped}


def create_restore_backup(repo: Repository, tables: tuple[str, ...]) -> Path | None:
    if not tables:
        return None
    backup_dir = Path("exports") / "backups"
    filename = f"restore-backup-{time.strftime('%Y%m%d-%H%M%S')}.json"
    payload = export_payload(repo, tables)
    payload["backup_reason"] = "before-import-restore"
    try:
        backup_dir.mkdir(parents=True, exist_ok=True)
        target = backup_dir / filename
        target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return target
    except OSError:
        return None


def import_restore_form_data(body: bytes, env: dict[str, str]) -> dict[str, Any]:
    content_type = env.get("_CONTENT_TYPE", "")
    if content_type.startswith("multipart/form-data"):
        return parse_multipart(body, content_type)
    return _form_multi(body)


def import_form_values(data: dict[str, Any], name: str) -> list[str]:
    value = data.get(name)
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item) for item in value if not isinstance(item, dict)]
    if isinstance(value, dict):
        return []
    return [str(value)]


def import_form_value(data: dict[str, Any], name: str, default: str = "") -> str:
    values = import_form_values(data, name)
    return values[-1] if values else default


def import_form_file(data: dict[str, Any], name: str) -> dict[str, Any] | None:
    value = data.get(name)
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        for item in reversed(value):
            if isinstance(item, dict):
                return item
    return None


def parse_import_restore_source(upload: dict[str, Any] | None, raw_payload: str, source_type: str, target_table: str, selected: list[str]) -> dict[str, Any]:
    filename = str((upload or {}).get("filename") or "")
    content = (upload or {}).get("content") if upload else None
    if isinstance(content, bytes) and content:
        detected = detect_import_source_type(filename, source_type)
        if detected == "json":
            return parse_import_json(content.decode("utf-8-sig", "ignore"))
        if detected == "csv":
            table = target_table if target_table in selected else (selected[0] if len(selected) == 1 else "")
            if table not in selected:
                return {"error": "CSV 恢复需要在“CSV/单表目标”中指定一个已勾选的数据表。"}
            return {"tables": {table: parse_import_csv(content)}}
        if detected == "xlsx":
            return parse_import_xlsx(content, target_table if target_table in selected else "", selected)
        return {"error": "无法识别上传文件类型，请手动选择 JSON、CSV 或 Excel xlsx。"}
    if not raw_payload.strip():
        return {"error": "请上传恢复文件，或粘贴 JSON 数据。"}
    if source_type not in {"auto", "json"}:
        return {"error": "粘贴输入当前仅支持 JSON；CSV/Excel 请使用文件上传。"}
    return parse_import_json(raw_payload)


def detect_import_source_type(filename: str, source_type: str) -> str:
    if source_type in {"json", "csv", "xlsx"}:
        return source_type
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":
        return "json"
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    return ""


def parse_import_json(text: str) -> dict[str, Any]:
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        return {"error": f"JSON 解析失败：{exc}"}
    tables_data = parsed.get("tables") if isinstance(parsed, dict) and isinstance(parsed.get("tables"), dict) else parsed
    if not isinstance(tables_data, dict):
        return {"error": "JSON 中没有可识别的数据表。"}
    return {"tables": {table: rows for table, rows in tables_data.items() if table in EXPORT_MAIN_TABLES and isinstance(rows, list)}}


def parse_import_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", "ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [{str(key or "").strip(): value for key, value in row.items() if key} for row in reader]


def parse_import_xlsx(content: bytes, target_table: str, selected: list[str]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return {"error": f"Excel 恢复不可用：{exc}"}
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return {"error": f"Excel 解析失败：{exc}"}
    tables: dict[str, list[dict[str, Any]]] = {}
    try:
        for sheet in workbook.worksheets:
            table = sheet.title if sheet.title in EXPORT_MAIN_TABLES else ""
            if not table and target_table and len(workbook.worksheets) == 1:
                table = target_table
            if table not in selected:
                continue
            rows_iter = sheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                continue
            fields = [str(value or "").strip() for value in header]
            rows = []
            for values in rows_iter:
                row = {field: values[index] if index < len(values) else "" for index, field in enumerate(fields) if field}
                if any(value not in (None, "") for value in row.values()):
                    rows.append(row)
            tables[table] = rows
    finally:
        workbook.close()
    if not tables:
        return {"error": "Excel 中没有匹配已勾选数据表的工作表。"}
    return {"tables": tables}


def import_row_uid_seed(row: dict[str, Any]) -> str:
    for key in ("uid", "id", "slug", "title", "name", "name_en", "object_key", "email", "doi", "patent_number", "application_number"):
        value = str(row.get(key) or "").strip()
        if value:
            return value
    return ""


def normalize_import_row(meta: Table, row: dict[str, Any]) -> dict[str, Any]:
    data = {key: "" if value is None else str(value) for key, value in row.items()}
    if meta.name in {"news", "projects", "patents", "students", "global_settings", "translation_cache"}:
        data = normalize_admin_data(meta, data)
    return data


def admin_auth_roles_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], repo: Repository) -> str:
    users = repo.list("auth_users", Query(limit=1000))
    perms = repo.list("auth_permissions", Query(limit=1000))
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        active = truthy(row.get("is_active"), default=True)
        system = truthy(row.get("is_system"), default=False)
        user_count = sum(1 for item in users if str(item.get("role_uid") or "") == str(key))
        perm_count = sum(1 for item in perms if str(item.get("role_uid") or "") == str(key) and any(truthy(item.get(flag), default=False) for flag in ("can_view", "can_create", "can_edit", "can_delete", "can_export")))
        body.append(f"""<article class="auth-admin-row auth-role-row{' is-disabled' if not active else ''}">
          {admin_batch_select("auth_roles", key)}
          <div class="auth-admin-main"><strong>{esc(row.get("name") or key)}</strong><span>{esc(key)} · 层级 {esc(row.get("level") or 0)}</span><small>{esc(text_only(row.get("description"), 180) or "未填写说明")}</small></div>
          <div class="auth-admin-scope"><span>{esc(row.get("visibility_scopes") or "public")}</span><small>前台可见范围</small></div>
          <div class="auth-admin-state">
            {auth_badge("启用" if active else "停用", active)}
            {auth_badge("系统角色" if system else "自定义", True)}
            <small>{user_count} 个账号 · {perm_count} 条授权</small>
          </div>
          <div class="auth-admin-actions"><a class="button ghost" href="/admin/table/auth_roles/{esc(key)}">编辑</a><a class="button light" href="/admin/table/auth_permissions?role_uid={esc(key)}" target="_blank" rel="noreferrer">权限</a></div>
        </article>""")
    filters = f"""<form class="filters auth-admin-search" method="get" action="/admin/table/auth_roles">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索角色名称、UID、说明、可见范围">
      <select name="is_active">{navigation_pair_options([("", "启用状态"), ("1", "启用"), ("0", "停用")], query.get("is_active", ""))}</select>
      <select name="is_system">{navigation_pair_options([("", "角色来源"), ("1", "系统"), ("0", "自定义")], query.get("is_system", ""))}</select>
      <select name="sort">{navigation_pair_options(auth_role_sort_options(), query.get("sort", "level_desc"))}</select>
      <button>筛选</button><a class="button ghost" href="/admin/table/auth_roles">重置</a>
    </form>"""
    return f"""<section class="admin-card auth-admin-card compact-admin-card">
      <div class="admin-card-head"><div><h1>{esc(meta.label)}</h1><p class="admin-muted">角色决定前台可见范围和后台授权归属；系统角色建议保留，新增自定义角色时再单独配置权限。</p></div><a class="button" href="/admin/table/auth_roles/new">新增</a></div>
      {admin_list_tools(filters, admin_batch_toolbar("auth_roles", meta, query, all_rows))}
      <div class="auth-admin-head auth-role-head"><span></span><span>角色</span><span>可见范围</span><span>状态摘要</span><span>操作</span></div>
      <div class="auth-admin-list">{"".join(body) or empty()}</div>
    </section>"""


def admin_auth_users_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], repo: Repository) -> str:
    roles = repo.list("auth_roles", Query(limit=1000, order_by="sort_order", descending=False))
    role_map = {str(role.get("uid") or ""): role for role in roles}
    role_choices = [("", "角色")] + [(str(role.get("uid") or ""), text_only(role.get("name") or role.get("uid"), 80)) for role in roles]
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        status = text_only(row.get("status"), 40).strip() or "active"
        active = status == "active"
        role = role_map.get(str(row.get("role_uid") or ""), {})
        body.append(f"""<article class="auth-admin-row auth-user-row{' is-disabled' if not active else ''}">
          {admin_batch_select("auth_users", key)}
          <div class="auth-admin-main"><strong>{esc(row.get("username") or key)}</strong><span>{esc(row.get("display_name") or "未填写显示名称")} · {esc(row.get("email") or "未填写邮箱")}</span><small>{esc(key)}</small></div>
          <div class="auth-admin-role"><strong>{esc(role.get("name") or row.get("role_uid") or "未分配")}</strong><span>{esc(row.get("role_uid") or "")}</span></div>
          <div class="auth-admin-state">{auth_badge("启用" if active else "停用", active)}{auth_badge(f'可见 {row.get("visibility") or "authenticated"}', True)}<small>最后登录：{esc(row.get("last_login_at") or "暂无")}</small></div>
          <div class="auth-admin-actions"><a class="button ghost" href="/admin/table/auth_users/{esc(key)}">编辑</a></div>
        </article>""")
    filters = f"""<form class="filters auth-admin-search" method="get" action="/admin/table/auth_users">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索账号、显示名称、邮箱、角色 UID">
      <select name="role_uid">{navigation_pair_options(role_choices, query.get("role_uid", ""))}</select>
      <select name="status">{navigation_pair_options([("", "状态"), ("active", "启用"), ("disabled", "停用")], query.get("status", ""))}</select>
      <select name="visibility">{navigation_pair_options([("", "账号可见"), ("authenticated", "authenticated"), ("owner", "owner"), ("staff", "staff"), ("public", "public")], query.get("visibility", ""))}</select>
      <select name="sort">{navigation_pair_options(auth_user_sort_options(), query.get("sort", "updated_desc"))}</select>
      <button>筛选</button><a class="button ghost" href="/admin/table/auth_users">重置</a>
    </form>"""
    return f"""<section class="admin-card auth-admin-card compact-admin-card">
      <div class="admin-card-head"><div><h1>{esc(meta.label)}</h1><p class="admin-muted">账号负责登录身份；后台权限来自角色权限，密码只可重设，不显示原文。</p></div><a class="button" href="/admin/table/auth_users/new">新增</a></div>
      {admin_list_tools(filters, admin_batch_toolbar("auth_users", meta, query, all_rows))}
      <div class="auth-admin-head auth-user-head"><span></span><span>账号</span><span>角色</span><span>状态</span><span>操作</span></div>
      <div class="auth-admin-list">{"".join(body) or empty()}</div>
    </section>"""


def admin_auth_permissions_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], repo: Repository) -> str:
    roles = repo.list("auth_roles", Query(limit=1000, order_by="sort_order", descending=False))
    role_map = {str(role.get("uid") or ""): role for role in roles}
    role_choices = [("", "角色")] + [(str(role.get("uid") or ""), text_only(role.get("name") or role.get("uid"), 80)) for role in roles]
    module_choices = [("", "模块")] + [(module, module_label(module)) for module in admin_modules()]
    flag_choices = [("", "不筛选"), ("1", "允许"), ("0", "禁止")]
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        role = role_map.get(str(row.get("role_uid") or ""), {})
        flags = [permission_flag(row, "can_view", "查"), permission_flag(row, "can_create", "增"), permission_flag(row, "can_edit", "改"), permission_flag(row, "can_delete", "删"), permission_flag(row, "can_export", "导")]
        body.append(f"""<article class="auth-admin-row auth-permission-row">
          {admin_batch_select("auth_permissions", key)}
          <div class="auth-admin-main"><strong>{esc(module_label(str(row.get("module") or "")))}</strong><span>{esc(row.get("module") or "")}</span><small>{esc(key)}</small></div>
          <div class="auth-admin-role"><strong>{esc(role.get("name") or row.get("role_uid") or "未分配")}</strong><span>{esc(row.get("role_uid") or "")}</span></div>
          <div class="auth-permission-flags">{"".join(flags)}</div>
          <div class="auth-admin-actions"><a class="button ghost" href="/admin/table/auth_permissions/{esc(key)}">编辑</a></div>
        </article>""")
    filters = f"""<form class="filters auth-admin-search" method="get" action="/admin/table/auth_permissions">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索角色、模块、UID">
      <select name="role_uid">{navigation_pair_options(role_choices, query.get("role_uid", ""))}</select>
      <select name="module">{navigation_pair_options(module_choices, query.get("module", ""))}</select>
      <select name="can_view">{navigation_pair_options([("", "查看")] + flag_choices[1:], query.get("can_view", ""))}</select>
      <select name="can_edit">{navigation_pair_options([("", "编辑")] + flag_choices[1:], query.get("can_edit", ""))}</select>
      <select name="can_delete">{navigation_pair_options([("", "删除")] + flag_choices[1:], query.get("can_delete", ""))}</select>
      <select name="sort">{navigation_pair_options(auth_permission_sort_options(), query.get("sort", "module_asc"))}</select>
      <button>筛选</button><a class="button ghost" href="/admin/table/auth_permissions">重置</a>
    </form>"""
    return f"""<section class="admin-card auth-admin-card compact-admin-card">
      <div class="admin-card-head"><div><h1>{esc(meta.label)}</h1><p class="admin-muted">角色权限控制后台模块的查看、新增、编辑、删除和导出。危险权限建议逐项确认后再批量修改。</p></div><a class="button" href="/admin/table/auth_permissions/new">新增</a></div>
      {admin_list_tools(filters, admin_batch_toolbar("auth_permissions", meta, query, all_rows))}
      <div class="auth-admin-head auth-permission-head"><span></span><span>模块</span><span>角色</span><span>权限</span><span>操作</span></div>
      <div class="auth-admin-list">{"".join(body) or empty()}</div>
    </section>"""


def admin_operation_logs_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    actions = sorted({text_only(row.get("action"), 80).strip() for row in all_rows if text_only(row.get("action"), 80).strip()})
    modules = sorted({text_only(row.get("module"), 120).strip() for row in all_rows if text_only(row.get("module"), 120).strip()})
    statuses = [("", "全部状态"), ("success", "成功"), ("warning", "警告"), ("failed", "失败")]
    filters = f"""<form class="filters operation-log-search" method="get" action="/admin/table/operation_logs">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索操作者、模块、对象、摘要、详情">
      <select name="action"><option value="">全部操作</option>{options(actions, query.get("action", ""))}</select>
      <select name="module"><option value="">全部模块</option>{options(modules, query.get("module", ""))}</select>
      <select name="status">{navigation_pair_options(statuses, query.get("status", ""))}</select>
      <select name="sort">{navigation_pair_options([("created_desc", "最近操作"), ("module", "按模块")], query.get("sort", "created_desc"))}</select>
      <button>筛选</button><a class="button ghost" href="/admin/table/operation_logs">重置</a>
    </form>"""
    body = []
    for row in rows:
        status = text_only(row.get("status"), 40).strip() or "success"
        detail = text_only(row.get("detail_json"), 3000).strip()
        target = text_only(row.get("target_uid"), 220).strip()
        module = text_only(row.get("module"), 120).strip()
        target_link = f'<a href="/admin/table/{esc(module)}/{esc(target)}" target="_blank" rel="noreferrer">{esc(target)}</a>' if module in TABLE_MAP and target else esc(target or "整表/系统")
        body.append(f"""<article class="operation-log-row status-{esc(status)}">
          <div class="operation-log-time"><strong>{esc(row.get("created_at") or row.get("updated_at") or "")}</strong><small>{esc(row.get("uid") or "")}</small></div>
          <div class="operation-log-actor"><strong>{esc(row.get("actor_name") or "system")}</strong><small>{esc(row.get("actor_uid") or "")}</small></div>
          <div class="operation-log-module"><strong>{esc(module_label(module) if module else "系统")}</strong><span>{esc(row.get("action") or "")}</span></div>
          <div class="operation-log-target">{target_link}</div>
          <div class="operation-log-summary"><span>{esc(row.get("summary") or "")}</span>{f'<small title="{esc(detail)}">{esc(text_only(detail, 260))}</small>' if detail else ""}</div>
          <div class="operation-log-status">{auth_badge({"success": "成功", "warning": "警告", "failed": "失败"}.get(status, status), status != "failed")}</div>
        </article>""")
    return f"""<section class="admin-card auth-admin-card operation-log-card compact-admin-card">
      <div class="admin-card-head"><div><h1>{esc(meta.label)}</h1><p class="admin-muted">记录后台保存、快速修改、批量修改、导入恢复、媒体操作和翻译任务。建议只授予高级管理员查看。</p></div></div>
      {admin_list_tools(filters, "")}
      <div class="operation-log-head"><span>时间</span><span>操作者</span><span>模块/操作</span><span>对象</span><span>摘要/详情</span><span>状态</span></div>
      <div class="operation-log-list">{"".join(body) or empty()}</div>
    </section>"""


def auth_badge(label: str, ok: bool) -> str:
    klass = "status-available" if ok else "status-missing"
    return f'<span class="admin-status-badge {klass}">{esc(label)}</span>'


def permission_flag(row: dict[str, Any], key: str, label: str) -> str:
    enabled = truthy(row.get(key), default=False)
    return f'<span class="permission-flag{" is-on" if enabled else ""}" title="{esc(key)}">{esc(label)}</span>'


def auth_role_sort_options() -> list[tuple[str, str]]:
    return [("level_desc", "层级高到低"), ("sort_asc", "排序小到大"), ("name_asc", "名称 A-Z"), ("updated_desc", "最近更新")]


def auth_role_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_asc": ("sort_order", False),
        "name_asc": ("name", False),
        "updated_desc": ("updated_at", True),
    }.get(value, ("level", True))


def auth_user_sort_options() -> list[tuple[str, str]]:
    return [("updated_desc", "最近更新"), ("login_desc", "最近登录"), ("username_asc", "账号 A-Z"), ("role_asc", "角色 A-Z")]


def auth_user_sort_args(value: str) -> tuple[str, bool]:
    return {
        "login_desc": ("last_login_at", True),
        "username_asc": ("username", False),
        "role_asc": ("role_uid", False),
    }.get(value, ("updated_at", True))


def auth_permission_sort_options() -> list[tuple[str, str]]:
    return [("module_asc", "模块 A-Z"), ("role_asc", "角色 A-Z"), ("sort_asc", "排序小到大"), ("updated_desc", "最近更新")]


def auth_permission_sort_args(value: str) -> tuple[str, bool]:
    return {
        "role_asc": ("role_uid", False),
        "sort_asc": ("sort_order", False),
        "updated_desc": ("updated_at", True),
    }.get(value, ("module", False))


def admin_site_settings_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], env: dict[str, str]) -> str:
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        active_badge = "已启用" if truthy(row.get("is_active"), default=True) else "停用"
        active_class = "status-available" if truthy(row.get("is_active"), default=True) else "status-missing"
        body.append(f"""<article class="site-admin-row">
          <div class="site-admin-title">
            <div>
              <strong>{esc(row.get("site_name") or key or "site-default")}</strong>
              <span>{esc(row.get("site_name_en") or "英文站名未设置")} · {esc(key or "")}</span>
            </div>
            <div class="site-admin-actions">
              <span class="admin-status-badge {esc(active_class)}">{esc(active_badge)}</span>
              <a class="button ghost" href="/admin/table/site_settings/{esc(key)}">编辑全部</a>
              <a class="button light" href="/" target="_blank" rel="noreferrer">查看首页</a>
              <a class="button light" href="/admin/table/media_assets" target="_blank" rel="noreferrer">媒体库</a>
            </div>
          </div>
          <div class="site-setting-line"><span>基础名称</span><p>{esc(site_settings_name_summary(row))}</p></div>
          <div class="site-setting-line"><span>首页主视觉</span><p><b>{esc(row.get("hero_title") or "首页标题未设置")}</b><small title="{esc(row.get("hero_subtitle") or "")}">{esc(text_only(row.get("hero_subtitle") or "首页简介未设置", 220))}</small></p></div>
          <div class="site-setting-line"><span>品牌媒体</span><p>{site_settings_media_summary(row)}</p></div>
          <div class="site-setting-line"><span>SEO</span><p>{esc(site_settings_seo_summary(row))}</p></div>
          <div class="site-setting-line"><span>首页内容</span><p>{esc(site_settings_homepage_summary(row))}</p></div>
          <div class="site-setting-line"><span>页脚</span><p>{esc(text_only(row.get("footer_text") or "页脚文本未设置", 220))}</p></div>
          <form class="site-inline-form site-setting-line" method="post" action="/admin/table/site_settings/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <span>常用修改</span>
            <div class="site-control-grid">
              {admin_switch_control("is_active", row.get("is_active"), "启用")}
              <label class="site-quick-field site-quick-wide"><span>中文站名</span><input name="site_name" value="{esc(row.get("site_name") or "")}"></label>
              <label class="site-quick-field site-quick-wide"><span>英文站名</span><input name="site_name_en" value="{esc(row.get("site_name_en") or "")}"></label>
              <label class="site-quick-field site-quick-wide"><span>首页教师 UID</span><input name="homepage_profile_uid" value="{esc(row.get("homepage_profile_uid") or "")}"></label>
              <label class="site-quick-field"><span>首页论文</span><input type="number" name="homepage_publication_limit" value="{esc(row.get("homepage_publication_limit") or 5)}"></label>
              <label class="site-quick-field"><span>首页动态</span><input type="number" name="homepage_news_limit" value="{esc(row.get("homepage_news_limit") or 4)}"></label>
              <button class="button light" type="submit">应用</button>
            </div>
          </form>
        </article>""")
    first_row = rows[0] if rows else {}
    return f"""<section class="admin-card site-admin-card compact-admin-card">
      <div class="admin-card-head"><div><h1>{esc(meta.label)}</h1><p class="admin-muted">站点名称、首页主视觉、品牌媒体、SEO 与首页展示数量集中维护；常用项可在概览中直接修改。</p></div><a class="button" href="/admin/table/site_settings/new">新增</a></div>
      <form class="filters site-admin-search" method="get" action="/admin/table/site_settings">
        <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索站名、首页标题、SEO、首页教师 UID">
        <button>搜索</button><a class="button ghost" href="/admin/table/site_settings">重置</a>
      </form>
      <div class="site-admin-summary">
        {site_settings_summary_card("站点名称", site_settings_name_summary(first_row) if rows else "暂无设置")}
        {site_settings_summary_card("首页展示", site_settings_homepage_summary(first_row) if rows else "暂无设置")}
        {site_settings_summary_card("SEO", site_settings_seo_summary(first_row) if rows else "暂无设置")}
        {site_settings_summary_card("品牌媒体", site_settings_media_summary_text(first_row) if rows else "暂无设置")}
      </div>
      <div class="site-admin-list">
        {"".join(body) or '<p class="empty">暂无站点设置。</p>'}
      </div>
    </section>"""


def site_settings_summary_card(title: str, value: str) -> str:
    return f'<div class="site-summary-card"><span>{esc(title)}</span><strong title="{esc(value)}">{esc(value)}</strong></div>'


def site_settings_name_summary(row: dict[str, Any]) -> str:
    if not row:
        return "暂无设置"
    cn = text_only(row.get("site_name") or "中文站名未设置", 160)
    en = text_only(row.get("site_name_en") or "英文站名未设置", 160)
    return f"{cn} / {en}"


def site_settings_homepage_summary(row: dict[str, Any]) -> str:
    if not row:
        return "暂无设置"
    teacher = text_only(row.get("homepage_profile_uid") or "未设置首页教师", 120)
    pubs = int_value(row.get("homepage_publication_limit"), 5)
    news = int_value(row.get("homepage_news_limit"), 4)
    return f"首页教师 {teacher} / 论文 {pubs} 条 / 动态 {news} 条"


def site_settings_media_summary_text(row: dict[str, Any]) -> str:
    if not row:
        return "暂无设置"
    items = []
    for label, name in (("Logo", "logo_key"), ("Favicon", "favicon_key"), ("分享图", "og_image_key")):
        value = text_only(row.get(name), 140).strip()
        items.append(f"{label}: {value or '未设置'}")
    return " / ".join(items)


def site_settings_media_summary(row: dict[str, Any]) -> str:
    chips = []
    for label, name in (("Logo", "logo_key"), ("Favicon", "favicon_key"), ("分享图", "og_image_key")):
        value = text_only(row.get(name), 200).strip()
        klass = "site-media-chip is-empty" if not value else "site-media-chip"
        chips.append(f'<span class="{klass}" title="{esc(value or "未设置")}"><b>{esc(label)}</b>{esc(value or "未设置")}</span>')
    return "".join(chips)


def site_settings_seo_summary(row: dict[str, Any]) -> str:
    if not row:
        return "暂无设置"
    title = text_only(row.get("seo_title") or "SEO 标题未设置", 120)
    desc = text_only(row.get("seo_description") or "SEO 描述未设置", 160)
    keywords = text_only(row.get("seo_keywords") or "关键词未设置", 120)
    return f"{title} / {desc} / {keywords}"


def site_settings_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid") or "site-default"
    row = repo.get("site_settings", key) or active_site(repo) or {"uid": key}
    row["is_active"] = 1 if truthy(data.get("is_active"), default=False) else 0
    for name, limit in (("site_name", 160), ("site_name_en", 160), ("homepage_profile_uid", 160)):
        row[name] = text_only(data.get(name), limit).strip()
    for name, default in (("homepage_publication_limit", 5), ("homepage_news_limit", 4)):
        row[name] = max(0, int_value(data.get(name), default))
    repo.save("site_settings", row)
    return "/admin/table/site_settings"


def site_settings_filter_rows(rows: list[dict[str, Any]], q: str) -> list[dict[str, Any]]:
    needle = text_only(q, 200).strip().casefold()
    if not needle:
        return rows
    return [row for row in rows if needle in " ".join(text_only(value, 12000) for value in row.values()).casefold()]


def admin_global_settings_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], env: dict[str, str]) -> str:
    body = []
    provider_labels = translation_provider_labels()
    for row in rows:
        row = global_settings_translation_defaults(row)
        key = row.get("uid") or row.get("id")
        translation_provider = text_only(row.get("translation_provider"), 80).strip() or "auto"
        pub_style = text_only(row.get("publication_display_style"), 40).strip() or "gbt"
        enabled_sources = global_settings_enabled_summary(row)
        usable_sources = translation_usable_provider_summary_text(row, env)
        body.append(f"""<article class="global-admin-row">
          <div class="global-admin-title">
            <div><strong>{esc(key or "global-default")}</strong><span>{esc(row.get("notify_email") or "通知邮箱未设置")}</span></div>
            <div class="global-admin-actions">
              <a class="button ghost" href="/admin/table/global_settings/{esc(key)}">编辑全部</a>
              <a class="button light" href="/admin/table/translation_cache" target="_blank" rel="noreferrer">翻译缓存</a>
              <a class="button light" href="/admin/table/media_assets" target="_blank" rel="noreferrer">媒体库</a>
            </div>
          </div>
          <div class="global-setting-line"><span>访问与留言</span><p>{esc(global_settings_access_summary(row))}</p></div>
          <div class="global-setting-line"><span>上传与媒体</span><p>{esc(global_settings_upload_summary(row))}</p></div>
          <div class="global-setting-line"><span>翻译服务</span><p><b>{esc(provider_labels.get(translation_provider, translation_provider))}</b><small title="{esc(enabled_sources)}">启用：{esc(enabled_sources)}</small><small title="{esc(usable_sources)}">当前可用：{esc(usable_sources)}</small></p></div>
          <div class="global-setting-line"><span>学术平台</span><p><b>论文显示：{esc(pub_style.upper() if pub_style != "source" else "原始引用")}</b><small>{esc(global_settings_platform_summary(row))}</small></p></div>
          <div class="global-setting-line"><span>缓存与性能</span><p><b>{esc(global_settings_cache_summary(row))}</b><small>翻译批量 {esc(row.get("translation_batch_size") or "10")} / 并发 {esc(row.get("translation_worker_count") or "4")} / 超时 {esc(row.get("translation_timeout_seconds") or "12")} 秒</small></p></div>
          <form class="global-inline-form global-setting-line" method="post" action="/admin/table/global_settings/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <span>常用修改</span>
            <div class="global-control-grid">
              {admin_switch_control("allow_public_registration", row.get("allow_public_registration"), "注册")}
              {admin_switch_control("allow_anonymous_messages", row.get("allow_anonymous_messages"), "匿名留言")}
              {admin_switch_control("news_pdf_allow_download", row.get("news_pdf_allow_download"), "PDF下载")}
              <label class="global-quick-field"><span>上传 MB</span><input type="number" name="upload_max_size_mb" value="{esc(row.get("upload_max_size_mb") or 10)}"></label>
              <label class="global-quick-field"><span>回收站天</span><input type="number" name="media_trash_retention_days" value="{esc(row.get("media_trash_retention_days") or 30)}"></label>
              <label class="global-quick-field"><span>翻译源</span><select name="translation_provider">{translation_provider_options(row, translation_provider)}</select></label>
              <label class="global-quick-field"><span>论文格式</span><select name="publication_display_style">{navigation_pair_options(publication_display_style_pairs(), pub_style)}</select></label>
              <button class="button light" type="submit">应用</button>
            </div>
          </form>
        </article>""")
    return f"""<section class="admin-card global-admin-card compact-admin-card">
      <div class="admin-card-head"><div><h1>{esc(meta.label)}</h1><p class="admin-muted">常用运行参数可在概览中快速修改；密钥、平台列表和缓存细项请进入编辑页维护。</p></div><a class="button" href="/admin/table/global_settings/new">新增</a></div>
      <form class="filters global-admin-search" method="get" action="/admin/table/global_settings">
        <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索 uid、翻译源、平台、通知邮箱">
        <button>搜索</button><a class="button ghost" href="/admin/table/global_settings">重置</a>
      </form>
      <div class="global-admin-summary">
        {global_settings_summary_card("访问与留言", global_settings_access_summary(rows[0]) if rows else "暂无设置")}
        {global_settings_summary_card("上传与媒体", global_settings_upload_summary(rows[0]) if rows else "暂无设置")}
        {global_settings_summary_card("翻译源", translation_usable_provider_summary_text(rows[0], env) if rows else "暂无设置")}
        {global_settings_summary_card("缓存性能", global_settings_cache_summary(rows[0]) if rows else "暂无设置")}
      </div>
      <div class="global-admin-list">
        {"".join(body) or '<p class="empty">暂无通用设置。</p>'}
      </div>
    </section>"""


def global_settings_summary_card(title: str, value: str) -> str:
    return f'<div class="global-summary-card"><span>{esc(title)}</span><strong title="{esc(value)}">{esc(value)}</strong></div>'


def global_settings_access_summary(row: dict[str, Any]) -> str:
    registration = "开放注册" if truthy(row.get("allow_public_registration"), default=False) else "关闭注册"
    messages = "允许匿名留言" if truthy(row.get("allow_anonymous_messages"), default=True) else "仅实名/后台留言"
    return f"{registration} / {messages}"


def global_settings_upload_summary(row: dict[str, Any]) -> str:
    size = int_value(row.get("upload_max_size_mb"), 10)
    trash_days = int_value(row.get("media_trash_retention_days"), 30)
    extensions = text_only(row.get("upload_allowed_extensions"), 260).strip()
    return f"上传 {size} MB / 回收站 {trash_days} 天 / {extensions or '扩展名未设置'}"


def global_settings_enabled_summary(row: dict[str, Any]) -> str:
    providers = parse_translation_providers(row)
    labels = translation_provider_labels()
    return " / ".join(labels.get(provider, provider) for provider in providers) or "未启用"


def global_settings_platform_summary(row: dict[str, Any]) -> str:
    pubs = ", ".join(parse_platforms(row.get("publication_metadata_providers") or row.get("publication_metadata_provider")))
    pats = ", ".join(parse_patent_platforms(row.get("patent_metadata_providers")))
    return f"论文: {pubs or '未设'} / 专利: {pats or '未设'}"


def global_settings_cache_summary(row: dict[str, Any]) -> str:
    values = [
        int_value(row.get("publication_suggestion_cache_seconds"), 30),
        int_value(row.get("profile_suggestion_cache_seconds"), 30),
        int_value(row.get("project_suggestion_cache_seconds"), 30),
        int_value(row.get("patent_suggestion_cache_seconds"), 30),
        int_value(row.get("student_suggestion_cache_seconds"), 30),
        int_value(row.get("news_suggestion_cache_seconds"), 30),
        int_value(row.get("course_suggestion_cache_seconds"), 30),
    ]
    unique = sorted(set(values))
    return f"填法提示缓存 {unique[0]} 秒" if len(unique) == 1 else f"填法提示缓存 {min(unique)}-{max(unique)} 秒"


def admin_switch_control(name: str, value: Any, label: str) -> str:
    checked = " checked" if truthy(value, default=False) else ""
    return f'<label class="admin-switch"><input type="hidden" name="{esc(name)}" value="0"><input type="checkbox" name="{esc(name)}" value="1"{checked}><span></span><em>{esc(label)}</em></label>'


def publication_display_style_pairs() -> list[tuple[str, str]]:
    return [("gbt", "GB/T"), ("elsevier", "Elsevier"), ("apa", "APA"), ("ieee", "IEEE"), ("bibtex", "BibTeX"), ("source", "原始引用")]


def global_settings_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid") or "global-default"
    row = repo.get("global_settings", key) or active_global(repo) or {"uid": key}
    for name in ("allow_public_registration", "allow_anonymous_messages", "news_pdf_allow_download"):
        row[name] = 1 if truthy(data.get(name), default=False) else 0
    for name, default in (("upload_max_size_mb", 10), ("media_trash_retention_days", 30)):
        row[name] = max(0, int_value(data.get(name), default))
    provider = text_only(data.get("translation_provider"), 80).strip()
    if provider in translation_provider_labels():
        row["translation_provider"] = provider
    display_style = text_only(data.get("publication_display_style"), 40).strip()
    if display_style in {value for value, _label in publication_display_style_pairs()}:
        row["publication_display_style"] = display_style
    repo.save("global_settings", global_settings_translation_defaults(row))
    return "/admin/table/global_settings"


def global_settings_filter_rows(rows: list[dict[str, Any]], q: str) -> list[dict[str, Any]]:
    needle = text_only(q, 200).strip().casefold()
    if not needle:
        return rows
    return [row for row in rows if needle in " ".join(text_only(value, 12000) for value in row.values()).casefold()]


def global_settings_translation_defaults(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row or {})
    if not text_only(normalized.get("translation_provider"), 80).strip():
        normalized["translation_provider"] = "auto"
    providers_text = text_only(normalized.get("translation_providers"), 500).strip()
    if not providers_text or providers_text.lower() == "manual":
        normalized["translation_providers"] = default_translation_provider_list(normalized)
    if not text_only(normalized.get("microsoft_translator_endpoint"), 500).strip():
        normalized["microsoft_translator_endpoint"] = MICROSOFT_TRANSLATOR_DEFAULT_ENDPOINT
    if not text_only(normalized.get("mymemory_email"), 200).strip():
        notify = text_only(normalized.get("notify_email"), 200).strip()
        if is_email_like(notify):
            normalized["mymemory_email"] = notify
    if not int_value(normalized.get("translation_batch_size"), 0):
        normalized["translation_batch_size"] = 10
    if not int_value(normalized.get("translation_worker_count"), 0):
        normalized["translation_worker_count"] = 4
    if not int_value(normalized.get("translation_timeout_seconds"), 0):
        normalized["translation_timeout_seconds"] = 12
    return normalized


def default_translation_provider_list(settings: dict[str, Any]) -> str:
    providers = ["auto", "mymemory"]
    if text_only(settings.get("libretranslate_url"), 500).strip():
        providers.append("libretranslate")
    if text_only(settings.get("deepl_api_key"), 500).strip():
        providers.append("deepl_free")
    if text_only(settings.get("google_translate_api_key"), 500).strip():
        providers.append("google_translate")
    if text_only(settings.get("microsoft_translator_key"), 500).strip():
        providers.append("microsoft_translator")
    providers.append("argos_local")
    return ",".join(dict.fromkeys(providers))


def is_email_like(value: Any) -> bool:
    text = text_only(value, 200).strip()
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", text))


def admin_translation_cache_table(meta: Table, repo: Repository, query: dict[str, str], env: dict[str, str]) -> str:
    groups = translation_requirement_groups(repo, env)
    filtered = filter_translation_requirements(groups, query)
    summary = translation_requirement_summary(groups)
    quality_summary = translation_quality_summary(groups)
    display_rows, page, per_page, total_rows = admin_paginate_rows(filtered, query)
    scan_note = translation_action_note(query)
    settings = active_global(repo)
    provider_options = translation_provider_options(settings, query.get("provider", ""))
    scope_options = navigation_pair_options(translation_scope_options(), query.get("scope", "priority"))
    provider_summary = translation_provider_summary(settings, env)
    dictionary_source = text_only(env.get("_I18N_DICTIONARY_SOURCE") or ("local" if i18n_dictionary_entries(env) else "none"), 40).strip()
    rows = []
    for item in display_rows:
        cache = item.get("cache") or {}
        status = text_only(item.get("status"), 40)
        english = text_only(item.get("english_text"), 12000)
        quality_flags = translation_quality_flags(item)
        quality_html = "".join(f'<span class="translation-quality-badge">{esc(label)}</span>' for _key, label in quality_flags)
        cache_key = cache.get("uid") or translation_cache_uid("shared", str(item.get("source_hash")))
        cache_edit = f'/admin/table/translation_cache/{esc(cache_key)}'
        source_links = translation_source_links(item)
        refs_count = len(item.get("refs") or [])
        form_id = f"translation-form-{safe_slug(str(cache_key))}"
        confirmed = status in {"cached", "dedicated"} or text_only(cache.get("status"), 40).strip() in {"success", "reviewed"}
        confirm_action = "unconfirm" if confirmed else "confirm"
        confirm_label = "取消确认" if confirmed else "确认"
        dictionary_locked = status == "dictionary"
        textarea_attrs = ' disabled title="该译文来自手动词典，请到词典编辑页修改。"' if dictionary_locked else f' title="{esc(english)}"'
        action_html = (
            f'<a class="button light" href="/admin/i18n-dictionary?q={quote(str(item.get("source_text") or ""))}" target="_blank" rel="noreferrer">改词典</a>'
            if dictionary_locked
            else f"""<button class="button secondary translation-confirm-toggle" type="submit" form="{esc(form_id)}" name="_translation_action" value="{esc(confirm_action)}">{esc(confirm_label)}</button>
            <button class="button light" type="submit" form="{esc(form_id)}" name="_translation_action" value="save">保存</button>
            <a class="button ghost" href="{esc(cache_edit)}" target="_blank" rel="noreferrer">详情</a>
            <form method="post" action="/admin/table/translation_cache/delete/{esc(cache_key)}"><button class="button danger" type="submit" data-confirm="确定删除这条翻译缓存任务吗？原始内容不会被删除，之后可重新扫描生成。">删除</button></form>"""
        )
        rows.append(f"""<article class="translation-admin-row status-{esc(status)}">
          <label class="translation-select-cell" title="选择后可只翻译勾选缓存"><input type="checkbox" name="selected" value="{esc(cache_key)}" form="translation-auto-form"></label>
          <div class="translation-source-cell">
            <strong>{esc(item.get("source_title") or item.get("table_label"))}</strong>
            <span>{source_links}</span>
            <small>{esc(item.get("source_hash"))}{f" / {refs_count} 个来源" if refs_count > 1 else ""}</small>
          </div>
          <div class="translation-text-cell" title="{esc(item.get("source_text"))}">{esc(item.get("source_text"))}</div>
          <div class="translation-status-cell">{translation_status_badge(status)}{quality_html}<small>{esc(item.get("source_hash"))}</small></div>
          <form id="{esc(form_id)}" class="translation-inline-form" method="post" action="/admin/table/translation_cache/inline">
            <input type="hidden" name="uid" value="{esc(cache_key)}">
            <input type="hidden" name="source_hash" value="{esc(item.get("source_hash"))}">
            <input type="hidden" name="source_text" value="{esc(item.get("source_text"))}">
            <input type="hidden" name="source_refs" value="{esc(item.get("source_refs"))}">
            <textarea name="translated_text" rows="2"{textarea_attrs} placeholder="填写或修正英文译文">{esc(english)}</textarea>
          </form>
          <div class="translation-actions-cell">
            {action_html}
          </div>
        </article>""")
    return f"""<section class="admin-card compact-admin-card translation-admin-card">
      <div class="translation-sticky-tools">
      <div class="translation-toolbar-row">
        <strong class="translation-toolbar-title">{esc(meta.label)}</strong>
          <form method="post" action="/admin/table/translation_cache/scan"><button class="button secondary" type="submit">扫描数据库</button></form>
          <a class="button light" href="/admin/i18n-dictionary">编辑词典</a>
          <a class="button ghost" href="/api/export/i18n-dictionary.json">导出词典</a>
          <form id="translation-auto-form" class="translation-auto-form" method="post" action="/admin/table/translation_cache/auto-translate">
            <select name="provider">{provider_options}</select>
            <select name="scope">{scope_options}</select>
            <button class="button secondary" type="submit">自动翻译</button>
            <span id="translation-selected-count" class="admin-muted translation-selected-note" title="未勾选缓存时，将按当前范围自动处理候选项。">未选=按范围</span>
          </form>
        <div class="translation-summary">
          {provider_summary}
          <span title="前台固定文案和手动译文优先读取该词典来源">词典 <strong>{esc(dictionary_source)}</strong></span>
          <span>全 <strong>{summary.get("total", 0)}</strong></span>
          <span>缺 <strong>{summary.get("missing", 0)}</strong></span>
          <span>缓存 <strong>{summary.get("cached", 0)}</strong></span>
          <span>词典 <strong>{summary.get("dictionary", 0)}</strong></span>
          <span>待 <strong>{summary.get("pending", 0)}</strong></span>
          <span>未确 <strong>{summary.get("unconfirmed", 0)}</strong></span>
          <span>变 <strong>{summary.get("stale", 0)}</strong></span>
          <span>错 <strong>{summary.get("failed", 0)}</strong></span>
          <span>异 <strong>{quality_summary.get("anomaly", 0)}</strong></span>
          <span>专 <strong>{summary.get("dedicated", 0)}</strong></span>
          {scan_note}
        </div>
      </div>
      <div id="translation-progress" class="translation-progress" hidden>
        <div class="translation-progress-head"><strong>自动翻译进度</strong><span id="translation-progress-text">准备开始</span><span id="translation-progress-sources" class="translation-progress-sources" hidden></span><button id="translation-progress-stop" class="button ghost" type="button">停止</button></div>
        <progress id="translation-progress-bar" value="0" max="1"></progress>
        <p id="translation-progress-note" class="admin-muted">翻译过程以小批量后台请求执行，不会阻塞其他页面编辑和保存。</p>
      </div>
      {translation_filter_form(query)}
      </div>
      <div class="translation-admin-list">
        <div class="translation-admin-head"><label class="translation-select-cell"><input id="translation-select-all" type="checkbox" aria-label="全选当前列表"></label><span>源域/来源</span><span>原文</span><span>状态</span><span>英文内容</span><span>操作</span></div>
        {"".join(rows) or '<p class="empty">暂无缓存任务。点击“扫描数据库”提取前台字段。</p>'}
      </div>
      {admin_pager("translation_cache", query, page, per_page, total_rows)}
    </section>"""


def translation_action_note(query: dict[str, str]) -> str:
    parts = []
    if query.get("scanned") is not None:
        parts.append(f"新增 {text_only(query.get('scanned'), 20)}")
        parts.append(f"更新 {text_only(query.get('updated'), 20)}")
        parts.append(f"专属字段跳过 {text_only(query.get('dedicated'), 20)}")
        if query.get("deleted") is not None:
            parts.append(f"清理失效 {text_only(query.get('deleted'), 20)}")
    if query.get("translated") is not None:
        provider = text_only(query.get("provider"), 80).strip()
        parts.append(f"{provider or 'provider'} 翻译 {text_only(query.get('translated'), 20)}")
        parts.append(f"失败 {text_only(query.get('failed'), 20)}")
        if query.get("selected"):
            parts.append(f"指定 {text_only(query.get('selected'), 20)}")
        scope = translation_scope_label(query.get("scope", ""))
        if scope:
            parts.append(scope)
    if query.get("dictionary_synced") is not None:
        parts.append(f"词典条目 {text_only(query.get('dictionary_synced'), 20)}")
        parts.append(f"新增 {text_only(query.get('dictionary_created'), 20)}")
        parts.append(f"更新 {text_only(query.get('dictionary_updated'), 20)}")
    if query.get("dictionary_saved") is not None:
        parts.append(f"词典保存到 {text_only(query.get('dictionary_saved'), 80)}")
        parts.append(f"条目 {text_only(query.get('dictionary_entries'), 20)}")
    return f'<span class="admin-muted">{"；".join(parts)}</span>' if parts else ""


def admin_i18n_dictionary_page(query: dict[str, str], env: dict[str, str]) -> str:
    entries = i18n_dictionary_entries(env)
    rows = [
        {
            "key": key,
            "zh": text_only(entry.get("zh"), 12000).strip(),
            "en": text_only(entry.get("en"), 12000).strip(),
            "context": text_only(entry.get("context"), 300).strip(),
        }
        for key, entry in entries.items()
        if isinstance(entry, dict)
    ]
    q = text_only(query.get("q"), 300).strip().casefold()
    if q:
        rows = [row for row in rows if q in " ".join([row["key"], row["zh"], row["en"], row["context"]]).casefold()]
    sort = text_only(query.get("sort"), 40).strip() or "key"
    if sort == "zh":
        rows.sort(key=lambda row: row["zh"])
    elif sort == "en":
        rows.sort(key=lambda row: row["en"])
    elif sort == "context":
        rows.sort(key=lambda row: (row["context"], row["key"]))
    else:
        rows.sort(key=lambda row: row["key"])
    display_rows, page, per_page, total_rows = admin_paginate_rows(rows, query)
    source = text_only(env.get("_I18N_DICTIONARY_SOURCE") or ("local" if entries else "none"), 40)
    note = ""
    if query.get("saved"):
        note = f'<p class="auth-success">词典已保存到 {esc(query.get("saved"))}，共 {esc(query.get("entries", "0"))} 条。</p>'
    rendered_rows = []
    for index, row in enumerate(display_rows):
        rendered_rows.append(f"""<article class="dictionary-admin-row">
          <label class="dictionary-delete-cell" title="勾选后保存时删除此条"><input type="checkbox" name="delete__{index}" value="1"></label>
          <input name="key__{index}" value="{esc(row["key"])}" placeholder="key">
          <textarea name="zh__{index}" rows="2" placeholder="中文原文">{esc(row["zh"])}</textarea>
          <textarea name="en__{index}" rows="2" placeholder="英文译文">{esc(row["en"])}</textarea>
          <input name="context__{index}" value="{esc(row["context"])}" placeholder="用途/备注">
          <input type="hidden" name="old_key__{index}" value="{esc(row["key"])}">
        </article>""")
    start_blank = len(display_rows)
    for offset in range(5):
        index = start_blank + offset
        rendered_rows.append(f"""<article class="dictionary-admin-row dictionary-new-row">
          <span class="dictionary-delete-cell admin-muted">新</span>
          <input name="key__{index}" placeholder="new-key">
          <textarea name="zh__{index}" rows="2" placeholder="中文原文"></textarea>
          <textarea name="en__{index}" rows="2" placeholder="英文译文"></textarea>
          <input name="context__{index}" placeholder="ui / common / 学科词汇">
          <input type="hidden" name="old_key__{index}" value="">
        </article>""")
    entry_count = len(display_rows) + 5
    return f"""<section class="admin-card compact-admin-card dictionary-admin-card">
      <div class="translation-sticky-tools dictionary-sticky-tools">
        <div class="translation-toolbar-row">
          <strong class="translation-toolbar-title">手动中英词典</strong>
          <a class="button ghost" href="/admin/table/translation_cache">返回翻译缓存</a>
          <a class="button light" href="/api/export/i18n-dictionary.json">导出词典</a>
          <span class="admin-muted">来源：{esc(source)} · 当前 {len(entries)} 条</span>
        </div>
        <form class="filters translation-admin-search dictionary-admin-search" method="get" action="/admin/i18n-dictionary">
          <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索 key、中文、英文、用途">
          <select class="translation-filter-sort" name="sort">{navigation_pair_options([("key", "按 key"), ("zh", "按中文"), ("en", "按英文"), ("context", "按用途")], sort)}</select>
          <select class="translation-filter-page" name="per_page">{options(["40", "80", "120", "200"], str(query.get("per_page") or "80"))}</select>
          <button>搜索</button><a class="button ghost" href="/admin/i18n-dictionary">重置</a>
        </form>
      </div>
      {note}
      <form class="dictionary-edit-form" method="post" action="/admin/i18n-dictionary/save">
        <input type="hidden" name="entry_count" value="{entry_count}">
        <div class="dictionary-admin-list">
          <div class="dictionary-admin-head"><span>删</span><span>Key</span><span>中文原文</span><span>英文译文</span><span>用途</span></div>
          {"".join(rendered_rows) or '<p class="empty">暂无词典条目。</p>'}
        </div>
        <div class="form-actions sticky-actions">
          <button type="submit">保存词典文件</button>
          <a class="button ghost" href="/admin/table/translation_cache">返回</a>
        </div>
      </form>
      {admin_pager("i18n-dictionary", query, page, per_page, total_rows)}
    </section>"""


def translation_provider_options(settings: dict[str, Any], selected: str = "") -> str:
    settings = global_settings_translation_defaults(settings)
    providers = parse_translation_providers(settings)
    selected = selected or text_only(settings.get("translation_provider"), 80).strip() or (providers[0] if providers else "manual")
    if selected not in providers and providers:
        selected = providers[0]
    labels = translation_provider_labels()
    return "".join(f'<option value="{esc(provider)}"{" selected" if provider == selected else ""}>{esc(labels.get(provider, provider))}</option>' for provider in providers)


def translation_provider_labels() -> dict[str, str]:
    return {
        "manual": "手动",
        "auto": "自动选择",
        "libretranslate": "LibreTranslate",
        "deepl_free": "DeepL Free",
        "google_translate": "Google Translate",
        "microsoft_translator": "Microsoft Translator",
        "mymemory": "MyMemory",
        "argos_local": "Argos Local",
    }


def translation_provider_summary(settings: dict[str, Any], env: dict[str, str] | None = None) -> str:
    env = env or {}
    settings = global_settings_translation_defaults(settings)
    labels = translation_provider_labels()
    default_provider = text_only(settings.get("translation_provider"), 80).strip() or "auto"
    enabled = translation_usable_providers(settings, env)
    enabled_labels = [labels.get(item, item) for item in enabled]
    default_label = labels.get(default_provider, default_provider)
    detail = " -> ".join(enabled_labels) if default_provider == "auto" and enabled_labels else " / ".join(enabled_labels)
    if not detail:
        detail = "暂无可用源"
    return f'<span class="translation-provider-summary" id="translation-provider-summary"><strong>可用源</strong>{esc(default_label)}{f"：{esc(detail)}" if detail else ""}</span>'


def translation_usable_provider_summary_text(settings: dict[str, Any], env: dict[str, str] | None = None) -> str:
    env = env or {}
    settings = global_settings_translation_defaults(settings)
    providers = translation_usable_providers(settings, env)
    return translation_provider_ids_summary(providers) or "暂无可用源"


def translation_provider_ids_summary(providers: list[str]) -> str:
    labels = translation_provider_labels()
    cleaned = []
    for provider in providers:
        item = text_only(provider, 80).strip()
        if item and item not in cleaned:
            cleaned.append(item)
    return " / ".join(labels.get(provider, provider) for provider in cleaned)


def translation_running_provider_ids(provider: str, settings: dict[str, Any], env: dict[str, str]) -> list[str]:
    provider = text_only(provider, 80).strip() or "auto"
    if provider == "auto":
        return translation_usable_providers(settings, env)
    return [provider] if translation_provider_is_usable(provider, settings, env) else []


def translation_usable_providers(settings: dict[str, Any], env: dict[str, str] | None = None) -> list[str]:
    env = env or {}
    settings = global_settings_translation_defaults(settings)
    usable = []
    for provider in parse_translation_providers(settings):
        if provider in {"auto", "manual"}:
            continue
        if translation_provider_is_usable(provider, settings, env):
            usable.append(provider)
    return usable


def translation_provider_is_usable(provider: str, settings: dict[str, Any], env: dict[str, str]) -> bool:
    settings = global_settings_translation_defaults(settings)
    if provider == "libretranslate":
        return bool(text_only(settings.get("libretranslate_url"), 500).strip())
    if provider == "deepl_free":
        return bool(text_only(settings.get("deepl_api_key"), 500).strip())
    if provider == "google_translate":
        return bool(text_only(settings.get("google_translate_api_key"), 500).strip())
    if provider == "microsoft_translator":
        return bool(text_only(settings.get("microsoft_translator_key"), 500).strip())
    if provider == "mymemory":
        return True
    if provider == "argos_local":
        return env.get("PLATFORM") != "cloudflare" and argos_translate_available()
    return False


def argos_translate_available() -> bool:
    try:
        import argostranslate.translate  # noqa: F401
    except ImportError:
        return False
    return True


def translation_scope_options() -> list[tuple[str, str]]:
    return [
        ("priority", "待填写优先"),
        ("missing", "仅待填写"),
        ("unconfirmed", "仅未确认/需复核"),
    ]


def translation_scope_label(scope: Any) -> str:
    labels = dict(translation_scope_options())
    return labels.get(text_only(scope, 40).strip(), "")


def parse_translation_providers(settings: dict[str, Any]) -> list[str]:
    raw = text_only(settings.get("translation_providers") or settings.get("translation_provider") or default_translation_provider_list(settings), 500)
    if raw.strip().lower() in {"", "manual"}:
        raw = default_translation_provider_list(settings)
    aliases = {
        "libre": "libretranslate", "libre_translate": "libretranslate",
        "deepl": "deepl_free", "deepl_api_free": "deepl_free",
        "google": "google_translate", "google_cloud": "google_translate",
        "bing": "microsoft_translator", "microsoft": "microsoft_translator", "azure": "microsoft_translator",
        "argos": "argos_local", "manual": "manual", "auto": "auto",
    }
    providers = []
    for part in raw.replace("；", ",").replace(";", ",").replace("\n", ",").split(","):
        item = aliases.get(part.strip().lower(), part.strip().lower())
        if item in {"manual", "auto", "libretranslate", "deepl_free", "google_translate", "microsoft_translator", "mymemory", "argos_local"} and item not in providers:
            providers.append(item)
    return providers or ["auto", "libretranslate", "deepl_free", "google_translate", "microsoft_translator", "mymemory", "argos_local"]


def translation_cache_groups(repo: Repository) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in repo.list("translation_cache", Query(limit=1000)):
        if text_only(row.get("target_lang"), 20).strip() not in {"en", "EN", "english", "English"}:
            continue
        source = text_only(row.get("source_text"), 12000).strip()
        if not source:
            continue
        source_hash = text_only(row.get("source_hash"), 80).strip() or translation_source_hash(source)
        item = grouped.setdefault(source_hash, {
            "source_hash": source_hash,
            "source_text": source,
            "english_text": "",
            "status": "pending",
            "cache": row,
            "refs": [],
            "source_refs": "",
        })
        item["refs"].extend(translation_parse_source_refs(row.get("source_refs")))
        if not item.get("english_text") and row.get("translated_text"):
            item["english_text"] = text_only(row.get("translated_text"), 12000).strip()
            item["cache"] = row
        item["status"] = translation_preferred_status(str(item.get("status") or ""), text_only(row.get("status"), 40).strip(), bool(row.get("translated_text")))
        if row.get("uid") and (not item.get("cache") or str(row.get("uid")).startswith("tr-")):
            item["cache"] = row
    for item in grouped.values():
        refs = translation_unique_refs(item.get("refs") or [])
        item["refs"] = refs
        item["source_refs"] = translation_refs_json(refs)
        item["source_title"] = translation_source_title(refs)
    return list(grouped.values())


def translation_requirement_groups(repo: Repository, env: dict[str, str] | None = None) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for item in frontend_translation_requirements(repo, env):
        source_hash = str(item.get("source_hash") or "")
        if not source_hash:
            continue
        cache = dict(item.get("cache") or {})
        translated = text_only(cache.get("translated_text") or item.get("english_text"), 12000).strip()
        status = translation_requirement_group_status(str(item.get("status") or ""), translated, cache)
        group = grouped.setdefault(
            source_hash,
            {
                "source_hash": source_hash,
                "source_text": item.get("source_text") or "",
                "english_text": "",
                "status": status,
                "cache": cache,
                "refs": [],
                "source_refs": "",
                "table": item.get("table"),
                "table_label": item.get("table_label"),
                "field": item.get("field"),
                "field_label": item.get("field_label"),
                "row_title": item.get("row_title"),
                "ref_key": item.get("ref_key"),
            },
        )
        group["refs"].append(translation_requirement_ref(item))
        group["status"] = translation_requirement_group_preferred(str(group.get("status") or ""), status)
        if translated and (not group.get("english_text") or status in {"cached", "dedicated", "unconfirmed"}):
            group["english_text"] = translated
        if cache and (not group.get("cache") or cache.get("translated_text") or str(cache.get("uid") or "").startswith("tr-")):
            group["cache"] = cache
    for item in grouped.values():
        refs = translation_unique_refs(item.get("refs") or [])
        item["refs"] = refs
        item["source_refs"] = translation_refs_json(refs)
        item["source_title"] = translation_source_title(refs)
    return list(grouped.values())


def translation_requirement_group_status(status: str, translated: str, cache: dict[str, Any]) -> str:
    cache_status = text_only(cache.get("status"), 40).strip()
    if status in {"dedicated", "dictionary"}:
        return status
    if cache_status in {"success", "reviewed", "cached"}:
        return "cached"
    if cache_status == "failed":
        return "failed"
    if translated and status in {"cached", "pending", "missing"}:
        return "unconfirmed"
    return status or "missing"


def translation_requirement_group_preferred(current: str, next_status: str) -> str:
    rank = {
        "missing": 0,
        "pending": 1,
        "failed": 2,
        "stale": 3,
        "unconfirmed": 4,
        "cached": 5,
        "dictionary": 6,
        "dedicated": 7,
    }
    return next_status if rank.get(next_status, 9) < rank.get(current, 9) else current


def translation_preferred_status(current: str, next_status: str, has_text: bool) -> str:
    if next_status in {"failed"}:
        return "failed" if current not in {"cached", "success", "reviewed"} else current
    if next_status in {"success", "reviewed"}:
        return "cached"
    if has_text and next_status in {"pending", ""} and current not in {"cached"}:
        return "unconfirmed"
    if next_status == "stale" and current not in {"cached"}:
        return "stale"
    return current or next_status or "pending"


def translation_source_title(refs: list[dict[str, Any]]) -> str:
    if not refs:
        return "手动缓存"
    first = refs[0]
    title = f"{first.get('table_label') or first.get('table')}"
    if first.get("field_label"):
        title += f" / {first.get('field_label')}"
    return title


def translation_parse_source_refs(value: Any) -> list[dict[str, Any]]:
    text = text_only(value, 12000).strip()
    if not text:
        return []
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return [dict(item) for item in data if isinstance(item, dict)]
    except (ValueError, TypeError):
        pass
    refs = []
    for line in text.splitlines() or [text]:
        line = line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split(" / ")]
        refs.append({"table_label": parts[0] if parts else "来源", "row_title": parts[1] if len(parts) > 1 else "", "field_label": parts[2] if len(parts) > 2 else "", "front_href": parts[3] if len(parts) > 3 else ""})
    return refs


def translation_unique_refs(refs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen = set()
    unique = []
    for ref in refs:
        key = (ref.get("table"), ref.get("row_key"), ref.get("field"), ref.get("source_ref_key"), ref.get("admin_href"))
        if key in seen:
            continue
        seen.add(key)
        unique.append(ref)
    return unique


def translation_refs_json(refs: list[dict[str, Any]]) -> str:
    return json.dumps(refs[:100], ensure_ascii=False, separators=(",", ":"))


def translation_source_links(item: dict[str, Any]) -> str:
    refs = item.get("refs") or []
    links = []
    for ref in refs[:3]:
        label = " / ".join(str(ref.get(key) or "") for key in ("table_label", "row_title", "field_label") if ref.get(key)) or "来源"
        admin_href = ref.get("admin_href") or ""
        front_href = ref.get("front_href") or ""
        if admin_href:
            links.append(f'<a href="{esc(safe_href(str(admin_href)))}" target="_blank" rel="noreferrer">{esc(label)}</a>')
        else:
            links.append(f'<span>{esc(label)}</span>')
        if front_href:
            links.append(f'<a class="translation-front-link" href="{esc(safe_href(str(front_href)))}" target="_blank" rel="noreferrer">前台</a>')
    if len(refs) > 3:
        links.append(f'<span class="admin-muted">等 {len(refs)} 个来源</span>')
    return "".join(links) or '<span class="admin-muted">手动缓存</span>'


def frontend_translation_requirements(repo: Repository, env: dict[str, str] | None = None) -> list[dict[str, Any]]:
    env = env or {}
    cache_rows = repo.list("translation_cache", Query(limit=1000))
    by_exact: dict[tuple[str, str], dict[str, Any]] = {}
    by_ref: dict[str, dict[str, Any]] = {}
    by_hash: dict[str, dict[str, Any]] = {}
    for cache in cache_rows:
        if text_only(cache.get("target_lang"), 20).strip() not in {"en", "EN", "english", "English"}:
            continue
        if not truthy(cache.get("is_current"), default=True):
            continue
        ref_key = text_only(cache.get("source_ref_key"), 300).strip()
        source_hash = text_only(cache.get("source_hash"), 80).strip()
        if ref_key:
            by_ref[ref_key] = cache
        if source_hash:
            by_hash[source_hash] = cache
        if ref_key and source_hash:
            by_exact[(ref_key, source_hash)] = cache
    requirements: list[dict[str, Any]] = []
    for table, fields in FRONTEND_TRANSLATION_FIELDS.items():
        meta = TABLE_MAP.get(table)
        if not meta:
            continue
        for row in frontend_translation_rows(repo, table):
            row_key = text_only(row.get("uid") or row.get("slug") or row.get("id"), 200).strip()
            for field in fields:
                source = text_only(row.get(field), 12000).strip()
                if not source:
                    continue
                english_field = ENGLISH_FIELD_OVERRIDES.get((table, field), "")
                direct = text_only(row.get(english_field), 12000).strip() if english_field else ""
                ref_key = translation_ref_key(table, row, field)
                source_hash = translation_source_hash(source)
                cache = by_exact.get((ref_key, source_hash)) or by_hash.get(source_hash)
                stale_cache = by_ref.get(ref_key) if not cache else None
                cached_text = text_only((cache or {}).get("translated_text"), 12000).strip()
                dictionary_text = i18n_dictionary_lookup_source(env, source, "en")
                if direct:
                    status = "dedicated"
                    english_text = direct
                elif dictionary_text:
                    status = "dictionary"
                    english_text = dictionary_text
                elif cache and cached_text:
                    status = "cached"
                    english_text = cached_text
                elif cache:
                    status = "pending"
                    english_text = ""
                elif stale_cache:
                    status = "stale"
                    english_text = text_only(stale_cache.get("translated_text"), 12000).strip()
                else:
                    status = "missing"
                    english_text = ""
                requirements.append(
                    {
                        "table": table,
                        "table_label": meta.label,
                        "field": field,
                        "field_label": translation_field_label(meta, field),
                        "row_key": row_key,
                        "row_title": admin_row_title(meta, row),
                        "source_text": source,
                        "english_field": english_field,
                        "english_text": english_text,
                        "ref_key": ref_key,
                        "source_hash": source_hash,
                        "status": status,
                        "cache": cache or stale_cache or {},
                        "front_href": translation_front_href(table, row),
                        "admin_href": f"/admin/table/{table}/{row_key}",
                    }
                )
    return requirements


def frontend_translation_rows(repo: Repository, table: str) -> list[dict[str, Any]]:
    if table == "navigation_items":
        return repo.list(table, Query(filters={"enabled": 1}, public_only=True, limit=1000, order_by="sort_order"))
    if table == "student_category_displays":
        return repo.list(table, Query(filters={"enabled": 1}, limit=1000, order_by="display_order"))
    if table == "site_settings":
        row = active_site(repo)
        return [row] if row else []
    return repo.list(table, Query(public_only=True, limit=1000))


def translation_field_label(meta: Table, field_name: str) -> str:
    for field in meta.fields:
        if field.name == field_name:
            return field.label
    return field_name


def translation_front_href(table: str, row: dict[str, Any]) -> str:
    pattern = FRONTEND_TABLE_URLS.get(table, "/")
    uid = text_only(row.get("uid") or row.get("id"), 160).strip()
    slug = safe_slug(str(row.get("slug") or row.get("title") or uid))
    return pattern.format(uid=uid, slug=slug)


def filter_translation_requirements(rows: list[dict[str, Any]], query: dict[str, str]) -> list[dict[str, Any]]:
    q = text_only(query.get("q"), 300).strip().casefold()
    table = text_only(query.get("source_table"), 80).strip()
    status = text_only(query.get("status"), 40).strip()
    quality = text_only(query.get("quality"), 40).strip()
    result = rows
    if q:
        result = [row for row in result if q in " ".join([str(row.get(key, "")) for key in ("table_label", "field_label", "row_title", "source_text", "english_text", "ref_key", "source_refs")] + [json.dumps(row.get("refs") or [], ensure_ascii=False)]).casefold()]
    if table:
        result = [row for row in result if row.get("table") == table or any(ref.get("table") == table for ref in (row.get("refs") or []))]
    if status:
        result = [row for row in result if row.get("status") == status]
    if quality:
        result = [row for row in result if any(flag_key == quality or quality == "anomaly" for flag_key, _label in translation_quality_flags(row))]
    sort = query.get("sort", "status")
    if sort == "table":
        result.sort(key=lambda row: (str(row.get("table")), str(row.get("row_title")), str(row.get("field"))))
    elif sort == "field":
        result.sort(key=lambda row: (str(row.get("field_label")), str(row.get("table")), str(row.get("row_title"))))
    else:
        rank = {"missing": 0, "pending": 1, "failed": 2, "stale": 3, "cached": 4, "dictionary": 5, "dedicated": 6}
        result.sort(key=lambda row: (rank.get(str(row.get("status")), 9), str(row.get("table")), str(row.get("row_title"))))
    return result


def translation_requirement_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary = {"total": len(rows), "dedicated": 0, "dictionary": 0, "cached": 0, "pending": 0, "unconfirmed": 0, "stale": 0, "missing": 0, "failed": 0}
    for row in rows:
        status = str(row.get("status") or "")
        if status in summary:
            summary[status] += 1
    return summary


def contains_cjk(value: Any) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", text_only(value, 12000)))


def translation_quality_flags(row: dict[str, Any]) -> list[tuple[str, str]]:
    source = text_only(row.get("source_text"), 12000).strip()
    english = text_only(row.get("english_text"), 12000).strip()
    status = text_only(row.get("status"), 40).strip()
    flags: list[tuple[str, str]] = []
    if status in {"dictionary", "dedicated"}:
        return flags
    if status == "failed":
        flags.append(("failed", "翻译失败"))
    if status != "dedicated" and not english:
        flags.append(("missing", "空译文"))
    if source and english and source.casefold() == english.casefold():
        flags.append(("same", "原译相同"))
    if source and contains_cjk(source) and english and contains_cjk(english):
        flags.append(("cjk", "仍含中文"))
    letters = [char for char in english if char.isalpha()]
    if len(letters) >= 12 and sum(1 for char in letters if char.isupper()) / max(1, len(letters)) > 0.85:
        flags.append(("caps", "疑似全大写"))
    return flags


def translation_quality_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary = {"anomaly": 0, "missing": 0, "failed": 0, "same": 0, "cjk": 0, "caps": 0}
    for row in rows:
        flags = translation_quality_flags(row)
        if flags:
            summary["anomaly"] += 1
        for key, _label in flags:
            summary[key] = summary.get(key, 0) + 1
    return summary


def translation_filter_form(query: dict[str, str]) -> str:
    tables = [(name, TABLE_MAP[name].label) for name in FRONTEND_TRANSLATION_FIELDS if name in TABLE_MAP]
    statuses = [("", "全部状态"), ("missing", "缺缓存"), ("pending", "待填写"), ("unconfirmed", "未确认/需复核"), ("failed", "失败"), ("stale", "原文已变"), ("cached", "已缓存"), ("dictionary", "词典命中"), ("dedicated", "专属英文")]
    quality = [("", "异常检查"), ("anomaly", "全部异常"), ("missing", "空译文"), ("failed", "翻译失败"), ("same", "原译相同"), ("cjk", "仍含中文"), ("caps", "疑似全大写")]
    sorts = [("status", "按状态优先"), ("table", "按来源模块"), ("field", "按字段")]
    table_options = '<option value="">全部模块</option>' + "".join(f'<option value="{esc(name)}"{" selected" if query.get("source_table") == name else ""}>{esc(label)}</option>' for name, label in tables)
    return f"""<form class="filters translation-admin-search" method="get" action="/admin/table/translation_cache">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索来源、原文、英文、ref key">
      <select class="translation-filter-source" name="source_table">{table_options}</select>
      <select class="translation-filter-status" name="status">{navigation_pair_options(statuses, query.get("status", ""))}</select>
      <select class="translation-filter-quality" name="quality">{navigation_pair_options(quality, query.get("quality", ""))}</select>
      <select class="translation-filter-sort" name="sort">{navigation_pair_options(sorts, query.get("sort", "status"))}</select>
      <select class="translation-filter-page" name="per_page">{options(["40", "80", "120", "200"], str(query.get("per_page") or "80"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/translation_cache">重置</a>
    </form>"""


def translation_status_badge(status: str) -> str:
    labels = {"dedicated": "专属英文", "dictionary": "词典命中", "cached": "已缓存", "pending": "待填写", "unconfirmed": "未确认", "failed": "失败", "stale": "原文已变", "missing": "缺缓存"}
    return f'<span class="translation-status-badge translation-status-{esc(status)}">{esc(labels.get(status, status or "未知"))}</span>'


def translation_scan_database(repo: Repository, env: dict[str, str] | None = None) -> dict[str, int]:
    result = {"created": 0, "updated": 0, "dedicated": 0, "deleted": 0}
    rows = translation_scan_cache_rows(repo, result, env)
    current_hashes = {row["source_hash"] for row in rows if row.get("source_hash")}
    existing = {text_only(row.get("source_hash"), 80).strip(): row for row in repo.list("translation_cache", Query(limit=1000)) if text_only(row.get("source_hash"), 80).strip()}
    for row in rows:
        old = existing.get(row["source_hash"])
        if old:
            merged = {**old, "source_ref_key": row["source_ref_key"], "source_refs": row["source_refs"], "source_text": row["source_text"], "source_hash": row["source_hash"], "target_lang": "en", "is_current": 1}
            if not text_only(merged.get("translated_text"), 12000).strip():
                merged["status"] = "pending"
            repo.save("translation_cache", merged)
            result["updated"] += 1
        else:
            repo.save("translation_cache", row)
            result["created"] += 1
    for old in repo.list("translation_cache", Query(limit=1000)):
        old_hash = text_only(old.get("source_hash"), 80).strip()
        if not old_hash or old_hash in current_hashes:
            continue
        key = text_only(old.get("uid") or old.get("id"), 200).strip()
        if key and repo.delete("translation_cache", key):
            result["deleted"] += 1
    return result


def translation_scan_cache_rows(repo: Repository, result: dict[str, int] | None = None, env: dict[str, str] | None = None) -> list[dict[str, Any]]:
    result = result if result is not None else {"created": 0, "updated": 0, "dedicated": 0}
    grouped: dict[str, dict[str, Any]] = {}
    for item in frontend_translation_requirements(repo, env):
        if item.get("status") in {"dedicated", "dictionary"}:
            result[str(item.get("status"))] = result.get(str(item.get("status")), 0) + 1
            continue
        source_hash = str(item.get("source_hash") or "")
        ref = translation_requirement_ref(item)
        seed_english = text_only(item.get("english_text"), 12000).strip()
        group = grouped.setdefault(
            source_hash,
            {
                "uid": translation_cache_uid("shared", source_hash),
                "source_hash": source_hash,
                "source_ref_key": f"shared:{source_hash}",
                "source_text": item.get("source_text") or "",
                "source_lang": "zh",
                "target_lang": "en",
                "translated_text": seed_english,
                "provider": "dictionary" if seed_english and item.get("table") == "dictionary" else "manual",
                "status": "success" if seed_english else "pending",
                "is_manual": 1,
                "is_current": 1,
                "_refs": [],
                "error_message": "",
            },
        )
        group["_refs"].append(ref)
    rows = []
    for group in grouped.values():
        refs = translation_unique_refs(group.pop("_refs", []))
        group["source_refs"] = translation_refs_json(refs)
        rows.append(group)
    return rows


def i18n_dictionary_document(entries: dict[str, dict[str, str]]) -> dict[str, Any]:
    clean_entries: dict[str, dict[str, str]] = {}
    for key, entry in entries.items():
        clean_key = text_only(key, 180).strip()
        if not clean_key:
            continue
        zh = text_only(entry.get("zh"), 12000).strip()
        en = text_only(entry.get("en"), 12000).strip()
        context = text_only(entry.get("context"), 300).strip()
        if not zh and not en:
            continue
        clean_entries[clean_key] = {"zh": zh, "en": en, "context": context}
    return {
        "version": 1,
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "description": "Manual Chinese-English dictionary for frontend fixed text and reusable content translations. Cloudflare reads R2 i18n/i18n_dictionary.json first, then falls back to this bundled file.",
        "entries": dict(sorted(clean_entries.items(), key=lambda item: item[0])),
    }


def i18n_dictionary_payload_from_entries(entries: dict[str, dict[str, str]]) -> dict[str, Any]:
    payload = i18n_dictionary_document(entries)
    content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8") + b"\n"
    return {"content": content, "entries": len(payload.get("entries") or {}), "key": I18N_DICTIONARY_R2_KEY, "payload": payload}


def i18n_dictionary_current_payload(env: dict[str, str] | None = None) -> dict[str, Any]:
    return i18n_dictionary_payload_from_entries(i18n_dictionary_entries(env))


def i18n_dictionary_update_payload(body: bytes, env: dict[str, str] | None = None) -> dict[str, Any]:
    form = _form(body)
    entries = {key: dict(value) for key, value in i18n_dictionary_entries(env).items() if isinstance(value, dict)}
    total = int_value(form.get("entry_count"), 0)
    for index in range(total):
        old_key = text_only(form.get(f"old_key__{index}"), 180).strip()
        key = text_only(form.get(f"key__{index}"), 180).strip()
        zh = text_only(form.get(f"zh__{index}"), 12000).strip()
        en = text_only(form.get(f"en__{index}"), 12000).strip()
        context = text_only(form.get(f"context__{index}"), 300).strip()
        delete = truthy(form.get(f"delete__{index}"), default=False)
        if old_key:
            entries.pop(old_key, None)
        if delete or not key or (not zh and not en):
            continue
        entries[key] = {"zh": zh, "en": en, "context": context}
    return i18n_dictionary_payload_from_entries(entries)


def translation_dictionary_payload(repo: Repository, env: dict[str, str] | None = None) -> dict[str, Any]:
    return i18n_dictionary_current_payload(env)


def translation_dictionary_save_payload(repo: Repository, env: dict[str, str] | None = None) -> dict[str, Any]:
    return i18n_dictionary_current_payload(env)


def translation_dictionary_save_local(repo: Repository, env: dict[str, str] | None = None) -> dict[str, Any]:
    payload = i18n_dictionary_current_payload(env)
    root_path = Path(I18N_DICTIONARY_FILENAME)
    root_path.write_bytes(payload["content"])
    I18N_DICTIONARY_CACHE.update({"path": "", "mtime": -1.0, "data": None})
    return {"saved": "local", "path": str(root_path), "entries": payload.get("entries", 0)}


def i18n_dictionary_save_local_from_body(body: bytes, env: dict[str, str] | None = None) -> dict[str, Any]:
    payload = i18n_dictionary_update_payload(body, env)
    root_path = Path(I18N_DICTIONARY_FILENAME)
    root_path.write_bytes(payload["content"])
    I18N_DICTIONARY_CACHE.update({"path": "", "mtime": -1.0, "data": None})
    return {"saved": "local", "path": str(root_path), "entries": payload.get("entries", 0)}


def translation_sync_missing(repo: Repository) -> int:
    return translation_scan_database(repo).get("created", 0)


def translation_requirement_ref(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "table": item.get("table"),
        "table_label": item.get("table_label"),
        "field": item.get("field"),
        "field_label": item.get("field_label"),
        "row_key": item.get("row_key"),
        "row_title": item.get("row_title"),
        "source_ref_key": item.get("ref_key"),
        "front_href": item.get("front_href"),
        "admin_href": item.get("admin_href"),
    }


def translation_inline_update(repo: Repository, body: bytes) -> dict[str, Any]:
    data = _form(body)
    source_hash = text_only(data.get("source_hash"), 80).strip()
    translated = text_only(data.get("translated_text"), 12000).strip()
    action = text_only(data.get("_translation_action"), 40).strip()
    uid = text_only(data.get("uid"), 200).strip() or translation_cache_uid("shared", source_hash)
    base = repo.get("translation_cache", uid) or {
        "uid": uid,
        "source_hash": source_hash,
        "source_ref_key": f"shared:{source_hash}",
        "source_text": text_only(data.get("source_text"), 12000).strip(),
        "source_lang": "zh",
        "target_lang": "en",
        "provider": "manual",
        "is_current": 1,
    }
    if action == "unconfirm":
        next_status = "pending"
        next_is_manual = 0
        next_provider = text_only(base.get("provider"), 80).strip() or "manual"
    elif action == "confirm" and translated:
        next_status = "success"
        next_is_manual = 1
        next_provider = "manual"
    else:
        next_status = "pending"
        next_is_manual = 1 if action == "save" else int_value(base.get("is_manual"), 0)
        next_provider = "manual" if action == "save" else text_only(base.get("provider"), 80).strip() or "manual"
    changes = {
        **base,
        "translated_text": translated,
        "source_hash": source_hash or base.get("source_hash"),
        "source_text": text_only(data.get("source_text"), 12000).strip() or base.get("source_text") or "",
        "source_refs": text_only(data.get("source_refs"), 12000).strip() or base.get("source_refs") or "",
        "target_lang": "en",
        "provider": next_provider,
        "is_manual": next_is_manual,
        "is_current": 1,
        "status": next_status,
        "error_message": "",
    }
    repo.save("translation_cache", changes)
    if source_hash:
        for row in repo.list("translation_cache", Query(limit=1000)):
            if text_only(row.get("source_hash"), 80).strip() == source_hash and str(row.get("uid")) != str(changes.get("uid")):
                row.update({key: changes[key] for key in ("translated_text", "provider", "is_manual", "is_current", "status", "error_message")})
                repo.save("translation_cache", row)
    return changes


def translation_inline_payload(row: dict[str, Any]) -> dict[str, Any]:
    status = text_only(row.get("status"), 40).strip() or "pending"
    display_status = "cached" if status in {"success", "reviewed"} else "unconfirmed" if text_only(row.get("translated_text"), 12000).strip() else status
    confirmed = status in {"success", "reviewed"}
    return {
        "ok": True,
        "uid": translation_cache_row_key(row),
        "source_hash": text_only(row.get("source_hash"), 80).strip(),
        "status": status,
        "display_status": display_status,
        "translated_text": text_only(row.get("translated_text"), 12000),
        "confirmed": confirmed,
        "action_value": "unconfirm" if confirmed else "confirm",
        "action_label": "取消确认" if confirmed else "确认",
    }


def translation_delete_cache(repo: Repository, uid_or_id: str) -> bool:
    row = repo.get("translation_cache", uid_or_id)
    if not row:
        return False
    source_hash = text_only(row.get("source_hash"), 80).strip()
    deleted = repo.delete("translation_cache", uid_or_id)
    if source_hash:
        for other in repo.list("translation_cache", Query(limit=1000)):
            if text_only(other.get("source_hash"), 80).strip() == source_hash:
                repo.delete("translation_cache", other.get("uid") or other.get("id") or "")
    return deleted


def translation_auto_translate(repo: Repository, body: bytes, env: dict[str, str]) -> dict[str, Any]:
    form = _form(body)
    form_multi = _form_multi(body)
    settings = active_global(repo)
    provider = text_only(form.get("provider") or settings.get("translation_provider") or "libretranslate", 80).strip()
    if provider == "manual":
        return {"provider": provider, "translated": 0, "failed": 0}
    scope = text_only(form.get("scope") or "priority", 40).strip()
    if scope not in {key for key, _label in translation_scope_options()}:
        scope = "priority"
    selected = {text_only(item, 200).strip() for item in form_multi.get("selected", []) if text_only(item, 200).strip()}
    batch_size = int_value(settings.get("translation_batch_size"), 10)
    batch_size = max(1, min(batch_size, 5 if env.get("PLATFORM") == "cloudflare" else 50))
    worker_count = int_value(settings.get("translation_worker_count"), 4)
    worker_count = 1 if env.get("PLATFORM") == "cloudflare" else max(1, min(worker_count, 8))
    candidates = translation_auto_candidates(repo, scope, selected)[:batch_size]
    translated, failed, _items = translation_translate_rows(repo, candidates, provider, settings, env, worker_count)
    return {"provider": provider, "translated": translated, "failed": failed, "scope": scope, "selected": len(selected)}


def translation_job_state(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    raw = text_only(settings.get("translation_job_state"), 20000).strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except (TypeError, ValueError):
        return {}
    return data if isinstance(data, dict) else {}


def translation_save_job_state(repo: Repository, state: dict[str, Any]) -> dict[str, Any]:
    settings = active_global(repo)
    uid = text_only(settings.get("uid"), 120).strip() or "global-default"
    saved = {
        **settings,
        "uid": uid,
        "translation_job_state": json.dumps(state, ensure_ascii=False, separators=(",", ":")),
    }
    repo.save("global_settings", saved)
    return state


def translation_job_public_state(repo: Repository, env: dict[str, str], state: dict[str, Any] | None = None) -> dict[str, Any]:
    settings = active_global(repo)
    state = dict(state if state is not None else translation_job_state(repo))
    if not state:
        state = {"status": "idle", "done": 0, "total": 0, "translated_total": 0, "failed_total": 0, "processed": []}
    state.setdefault("status", "idle")
    state.setdefault("done", 0)
    state.setdefault("total", 0)
    state.setdefault("translated_total", 0)
    state.setdefault("failed_total", 0)
    state.setdefault("processed", [])
    state["available_provider_summary"] = translation_usable_provider_summary_text(settings, env)
    state["available_providers"] = translation_usable_providers(settings, env)
    active_providers = [text_only(item, 80).strip() for item in (state.get("active_providers") or []) if text_only(item, 80).strip()]
    if not active_providers and state.get("status") == "running":
        active_providers = translation_running_provider_ids(text_only(state.get("provider"), 80).strip() or "auto", settings, env)
    last_step_providers = [text_only(item, 80).strip() for item in (state.get("last_step_providers") or []) if text_only(item, 80).strip()]
    current_providers = last_step_providers or active_providers
    state["active_providers"] = active_providers
    state["last_step_providers"] = last_step_providers
    state["active_provider_summary"] = translation_provider_ids_summary(current_providers)
    state["last_step_provider_summary"] = translation_provider_ids_summary(last_step_providers)
    state["provider_label"] = translation_provider_labels().get(text_only(state.get("provider"), 80).strip(), text_only(state.get("provider"), 80).strip() or "自动选择")
    state["scope_label"] = translation_scope_label(state.get("scope"))
    state["running"] = state.get("status") == "running"
    return state


def translation_job_status_payload(repo: Repository, env: dict[str, str]) -> dict[str, Any]:
    return {"ok": True, "state": translation_job_public_state(repo, env)}


def translation_job_start(repo: Repository, body: bytes, env: dict[str, str]) -> dict[str, Any]:
    form = _form(body)
    form_multi = _form_multi(body)
    settings = active_global(repo)
    provider = text_only(form.get("provider") or settings.get("translation_provider") or "auto", 80).strip() or "auto"
    scope = text_only(form.get("scope") or "priority", 40).strip()
    if scope not in {key for key, _label in translation_scope_options()}:
        scope = "priority"
    if provider == "manual":
        state = {"status": "idle", "message": "手动模式不会自动翻译。", "updated_at": int(time.time())}
        translation_save_job_state(repo, state)
        return {"ok": False, "done": True, "message": state["message"], "state": translation_job_public_state(repo, env, state)}
    if provider != "auto" and not translation_provider_is_usable(provider, settings, env):
        label = translation_provider_labels().get(provider, provider)
        state = {"status": "idle", "message": f"{label} 当前不可用，请先在通用设置中配置。", "updated_at": int(time.time())}
        translation_save_job_state(repo, state)
        return {"ok": False, "done": True, "message": state["message"], "state": translation_job_public_state(repo, env, state)}
    if provider == "auto" and not translation_usable_providers(settings, env):
        state = {"status": "idle", "message": "当前没有可用自动翻译源，请先在通用设置中配置。", "updated_at": int(time.time())}
        translation_save_job_state(repo, state)
        return {"ok": False, "done": True, "message": state["message"], "state": translation_job_public_state(repo, env, state)}
    translation_scan_database(repo, env)
    selected = [text_only(item, 200).strip() for item in form_multi.get("selected", []) if text_only(item, 200).strip()][:500]
    candidate_count = len(translation_auto_candidates(repo, scope, set(selected)))
    now = int(time.time())
    nonce = time.time_ns()
    state = {
        "job_id": stable_uid("trjob", f"{nonce}:{provider}:{scope}:{','.join(selected)}"),
        "status": "running" if candidate_count else "completed",
        "provider": provider,
        "active_providers": translation_running_provider_ids(provider, settings, env),
        "last_step_providers": [],
        "scope": scope,
        "selected": selected,
        "processed": [],
        "total": candidate_count,
        "done": 0,
        "remaining": candidate_count,
        "translated_total": 0,
        "failed_total": 0,
        "stop_requested": False,
        "message": "自动翻译任务已启动。" if candidate_count else "没有需要自动翻译的缓存。",
        "started_at": now,
        "updated_at": now,
    }
    translation_save_job_state(repo, state)
    public_state = translation_job_public_state(repo, env, state)
    return {"ok": True, "done": candidate_count == 0, "message": state["message"], "state": public_state, **public_state}


def translation_job_stop(repo: Repository, env: dict[str, str]) -> dict[str, Any]:
    state = translation_job_state(repo)
    now = int(time.time())
    if state.get("status") == "running":
        state["status"] = "stopped"
        state["stop_requested"] = True
        state["message"] = "已请求停止自动翻译。"
    else:
        state.setdefault("status", "stopped")
        state["message"] = "当前没有正在运行的自动翻译任务。"
    state["updated_at"] = now
    translation_save_job_state(repo, state)
    public_state = translation_job_public_state(repo, env, state)
    return {"ok": True, "done": True, "message": state["message"], "state": public_state, **public_state}


def translation_auto_translate_step(repo: Repository, body: bytes, env: dict[str, str]) -> dict[str, Any]:
    form = _form(body)
    form_multi = _form_multi(body)
    settings = active_global(repo)
    job_id = text_only(form.get("job_id"), 120).strip()
    job_state = translation_job_state(repo) if job_id else {}
    job_mode = bool(job_id)
    if job_mode:
        if text_only(job_state.get("job_id"), 120).strip() != job_id:
            public_state = translation_job_public_state(repo, env, job_state)
            return {"ok": False, "done": True, "message": "已有新的自动翻译任务，此旧任务已停止。", "state": public_state, **public_state}
        if job_state.get("status") != "running" or truthy(job_state.get("stop_requested"), default=False):
            job_state["status"] = "stopped" if job_state.get("status") != "completed" else "completed"
            job_state["updated_at"] = int(time.time())
            translation_save_job_state(repo, job_state)
            public_state = translation_job_public_state(repo, env, job_state)
            return {"ok": True, "done": True, "message": text_only(job_state.get("message"), 300) or "自动翻译已停止。", "state": public_state, **public_state}
        provider = text_only(job_state.get("provider") or settings.get("translation_provider") or "auto", 80).strip()
        scope = text_only(job_state.get("scope") or "priority", 40).strip()
        selected = {text_only(item, 200).strip() for item in (job_state.get("selected") or []) if text_only(item, 200).strip()}
        processed = {text_only(item, 200).strip() for item in (job_state.get("processed") or []) if text_only(item, 200).strip()}
    else:
        provider = text_only(form.get("provider") or settings.get("translation_provider") or "libretranslate", 80).strip()
        scope = text_only(form.get("scope") or "priority", 40).strip()
        selected = {text_only(item, 200).strip() for item in form_multi.get("selected", []) if text_only(item, 200).strip()}
        processed = {text_only(item, 200).strip() for item in form_multi.get("processed", []) if text_only(item, 200).strip()}
    if scope not in {key for key, _label in translation_scope_options()}:
        scope = "priority"
    if provider == "manual":
        return {"ok": False, "done": True, "message": "手动模式不会自动翻译。", "provider": provider, "scope": scope, "total": 0, "done_count": 0, "remaining": 0, "translated": 0, "failed": 0, "processed": sorted(processed)}
    all_candidates = translation_auto_candidates(repo, scope, selected)
    candidate_keys = [translation_cache_row_key(row) for row in all_candidates]
    candidate_key_set = {key for key in candidate_keys if key}
    processed_known = {key for key in processed if key in candidate_key_set}
    pending_candidates = [row for row in all_candidates if translation_cache_row_key(row) not in processed_known]
    configured_batch = max(1, min(int_value(settings.get("translation_batch_size"), 10), 5 if env.get("PLATFORM") == "cloudflare" else 50))
    requested_step = int_value(form.get("step_size"), 5 if env.get("PLATFORM") == "cloudflare" else max(4, int_value(settings.get("translation_worker_count"), 4)))
    step_cap = 5 if env.get("PLATFORM") == "cloudflare" else 12
    step_size = max(1, min(requested_step, configured_batch, step_cap))
    candidates = pending_candidates[:step_size]
    worker_count = int_value(settings.get("translation_worker_count"), 4)
    worker_count = 1 if env.get("PLATFORM") == "cloudflare" else max(1, min(worker_count, 8, len(candidates) or 1))
    translated, failed, items = translation_translate_rows(repo, candidates, provider, settings, env, worker_count)
    step_providers = []
    for item in items:
        item_provider = text_only(item.get("provider"), 80).strip()
        if item_provider and item_provider not in step_providers:
            step_providers.append(item_provider)
    active_providers = translation_running_provider_ids(provider, settings, env)
    next_processed = set(processed_known)
    next_processed.update(item["uid"] for item in items if item.get("uid"))
    total = len(candidate_keys)
    done_count = min(total, len(next_processed))
    remaining = max(0, total - done_count)
    message = "没有需要自动翻译的缓存。" if total == 0 else f"已处理 {done_count}/{total}，成功 {translated}，失败 {failed}。"
    public_state: dict[str, Any] = {}
    if job_mode:
        current_state = translation_job_state(repo)
        stopped = text_only(current_state.get("job_id"), 120).strip() == job_id and (current_state.get("status") == "stopped" or truthy(current_state.get("stop_requested"), default=False))
        job_state.update({
            "processed": sorted(next_processed),
            "total": total,
            "done": done_count,
            "remaining": remaining,
            "translated_total": int_value(job_state.get("translated_total"), 0) + translated,
            "failed_total": int_value(job_state.get("failed_total"), 0) + failed,
            "active_providers": active_providers,
            "last_step_providers": step_providers,
            "status": "stopped" if stopped else "completed" if remaining == 0 else "running",
            "stop_requested": bool(stopped),
            "message": "自动翻译已停止。" if stopped else message,
            "updated_at": int(time.time()),
        })
        translation_save_job_state(repo, job_state)
        public_state = translation_job_public_state(repo, env, job_state)
    return {
        "ok": True,
        "done": remaining == 0 or public_state.get("status") in {"stopped", "completed"},
        "message": public_state.get("message") or message,
        "provider": provider,
        "scope": scope,
        "total": total,
        "done_count": done_count,
        "remaining": remaining,
        "translated": translated,
        "failed": failed,
        "active_providers": active_providers,
        "last_step_providers": step_providers,
        "active_provider_summary": translation_provider_ids_summary(step_providers or active_providers),
        "selected": len(selected),
        "processed": sorted(next_processed),
        "items": items,
        "state": public_state,
        **public_state,
    }


def translation_translate_rows(repo: Repository, candidates: list[dict[str, Any]], provider: str, settings: dict[str, Any], env: dict[str, str], worker_count: int) -> tuple[int, int, list[dict[str, Any]]]:
    translated = 0
    failed = 0
    tasks = translation_translate_tasks(candidates, env)
    providers = translation_task_providers(provider, settings, env)

    def run(task_index_task: tuple[int, dict[str, Any]]) -> list[tuple[dict[str, Any], str, str, str]]:
        task_index, task = task_index_task
        preferred = translation_preferred_task_provider(providers, task_index, provider, env)
        return translation_run_task(task, provider, preferred, settings, env)

    if worker_count > 1 and len(tasks) > 1:
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=min(worker_count, len(tasks))) as executor:
            task_results = list(executor.map(run, enumerate(tasks)))
    else:
        task_results = [run(item) for item in enumerate(tasks)]
    results = [item for group in task_results for item in group]
    items = []
    for row, value, used_provider, error in results:
        key = translation_cache_row_key(row)
        if value:
            row["translated_text"] = value
            row["provider"] = used_provider or provider
            row["status"] = "pending"
            row["is_manual"] = 0
            row["is_current"] = 1
            row["error_message"] = ""
            translated += 1
            items.append({"uid": key, "ok": True, "status": "pending", "display_status": "unconfirmed", "translated_text": value, "provider": row["provider"], "message": ""})
        else:
            row["provider"] = used_provider or provider
            row["status"] = "failed"
            row["error_message"] = error or "翻译服务未返回译文"
            failed += 1
            items.append({"uid": key, "ok": False, "status": "failed", "display_status": "failed", "translated_text": text_only(row.get("translated_text"), 12000).strip(), "provider": row["provider"], "message": row["error_message"]})
        repo.save("translation_cache", row)
    return translated, failed, items


def translation_translate_tasks(candidates: list[dict[str, Any]], env: dict[str, str]) -> list[dict[str, Any]]:
    buckets: dict[str, list[dict[str, Any]]] = {}
    for index, row in enumerate(candidates):
        key = translation_row_origin_key(row) or f"single:{index}:{translation_cache_row_key(row)}"
        buckets.setdefault(key, []).append(row)
    tasks: list[dict[str, Any]] = []
    for rows in buckets.values():
        chunk: list[dict[str, Any]] = []
        chunk_chars = 0
        for row in rows:
            source = text_only(row.get("source_text"), 12000).strip()
            projected = chunk_chars + len(source) + 80
            if chunk and (len(chunk) >= TRANSLATION_BUNDLE_MAX_ITEMS or projected > TRANSLATION_BUNDLE_MAX_CHARS):
                tasks.append({"rows": chunk, "bundled": len(chunk) > 1})
                chunk = []
                chunk_chars = 0
            chunk.append(row)
            chunk_chars += len(source) + 80
        if chunk:
            tasks.append({"rows": chunk, "bundled": len(chunk) > 1})
    return tasks


def translation_row_origin_key(row: dict[str, Any]) -> str:
    refs = translation_parse_source_refs(row.get("source_refs"))
    for ref in refs:
        table = text_only(ref.get("table"), 120).strip()
        row_key = text_only(ref.get("row_key"), 200).strip()
        if table and row_key:
            return f"{table}:{row_key}"
    ref_key = text_only(row.get("source_ref_key"), 300).strip()
    if ref_key and not ref_key.startswith("shared:"):
        parts = ref_key.split(":")
        if len(parts) >= 2:
            return ":".join(parts[:2])
    return ""


def translation_task_providers(provider: str, settings: dict[str, Any], env: dict[str, str]) -> list[str]:
    if provider == "auto":
        providers = translation_usable_providers(settings, env)
        if not providers:
            raise RuntimeError("当前没有可用自动翻译源，请先在通用设置中配置。")
        return providers
    return [provider]


def translation_preferred_task_provider(providers: list[str], task_index: int, provider: str, env: dict[str, str]) -> str:
    if provider == "auto" and providers:
        return providers[task_index % len(providers)]
    return providers[0] if providers else provider


def translation_run_task(task: dict[str, Any], provider: str, preferred: str, settings: dict[str, Any], env: dict[str, str]) -> list[tuple[dict[str, Any], str, str, str]]:
    rows = list(task.get("rows") or [])
    if task.get("bundled") and len(rows) > 1:
        payload = translation_bundle_payload(rows)
        try:
            value, used_provider = translate_text_with_provider_fallback(provider, payload, settings, env, preferred)
            parsed = translation_parse_bundle(value, len(rows))
            if len(parsed) == len(rows) and all(text_only(item, 12000).strip() for item in parsed):
                return [(row, parsed[index], used_provider, "") for index, row in enumerate(rows)]
        except Exception:
            pass
    results: list[tuple[dict[str, Any], str, str, str]] = []
    for row in rows:
        try:
            value, used_provider = translate_text_with_provider_fallback(provider, row.get("source_text"), settings, env, preferred)
            results.append((row, value, used_provider, ""))
        except Exception as exc:
            results.append((row, "", preferred or provider, str(exc)[:500]))
    return results


def translate_text_with_provider_fallback(provider: str, value: Any, settings: dict[str, Any], env: dict[str, str], preferred: str = "") -> tuple[str, str]:
    if provider != "auto":
        return translate_text_provider(provider, value, settings, env), provider
    providers = translation_usable_providers(settings, env)
    if preferred in providers:
        providers = [preferred] + [item for item in providers if item != preferred]
    errors = []
    for item in providers:
        try:
            return translate_text_provider(item, value, settings, env), item
        except Exception as exc:
            errors.append(f"{item}: {exc}")
    raise RuntimeError("自动选择翻译失败：" + "；".join(errors[:5]))


def translation_bundle_payload(rows: list[dict[str, Any]]) -> str:
    parts = []
    for index, row in enumerate(rows, 1):
        marker = translation_bundle_marker(index)
        parts.append(f"{marker}\n{text_only(row.get('source_text'), 12000).strip()}")
    return "\n\n".join(parts)


def translation_bundle_marker(index: int) -> str:
    return f"[[[{TRANSLATION_BUNDLE_MARKER}_{index:03d}]]]"


def translation_parse_bundle(value: Any, expected: int) -> list[str]:
    text = text_only(value, 12000).strip()
    if not text:
        return []
    pattern = re.compile(r"\[\[\[" + re.escape(TRANSLATION_BUNDLE_MARKER) + r"_(\d{3})\]\]\]\s*(.*?)(?=\s*\[\[\[" + re.escape(TRANSLATION_BUNDLE_MARKER) + r"_\d{3}\]\]\]|\s*$)", re.S)
    found = {int(match.group(1)): match.group(2).strip() for match in pattern.finditer(text)}
    if len(found) == expected:
        return [found.get(index, "") for index in range(1, expected + 1)]
    chunks = re.split(r"\n\s*\n", text)
    chunks = [chunk.strip() for chunk in chunks if chunk.strip()]
    return chunks if len(chunks) == expected else []


def translation_auto_candidates(repo: Repository, scope: str = "priority", selected: set[str] | None = None) -> list[dict[str, Any]]:
    selected = selected or set()
    rows = []
    for row in repo.list("translation_cache", Query(limit=1000)):
        key = text_only(row.get("uid") or row.get("id"), 200).strip()
        if selected and key not in selected:
            continue
        if not translation_row_can_auto_translate(row):
            continue
        bucket = translation_auto_bucket(row)
        if scope == "missing" and bucket != "missing":
            continue
        if scope == "unconfirmed" and bucket != "unconfirmed":
            continue
        rows.append(row)
    rank = {"missing": 0, "unconfirmed": 1}
    rows.sort(key=lambda row: (
        rank.get(translation_auto_bucket(row), 9),
        str(row.get("source_ref_key") or row.get("uid") or ""),
    ))
    return rows


def translation_cache_row_key(row: dict[str, Any]) -> str:
    return text_only(row.get("uid") or row.get("id"), 200).strip()


def translation_row_can_auto_translate(row: dict[str, Any]) -> bool:
    status = text_only(row.get("status"), 40).strip()
    return (
        text_only(row.get("target_lang"), 20).strip() in {"en", "EN", "english", "English"}
        and truthy(row.get("is_current"), default=True)
        and text_only(row.get("source_text"), 12000).strip()
        and status not in {"success", "reviewed", "cached", "dedicated"}
    )


def translation_auto_bucket(row: dict[str, Any]) -> str:
    translated = text_only(row.get("translated_text"), 12000).strip()
    return "unconfirmed" if translated else "missing"


def translate_text_provider(provider: str, value: Any, settings: dict[str, Any], env: dict[str, str]) -> str:
    text = text_only(value, 12000).strip()
    if not text:
        return ""
    if provider == "auto":
        errors = []
        providers = translation_usable_providers(settings, env)
        if not providers:
            raise RuntimeError("当前没有可用自动翻译源，请先在通用设置中配置。")
        for item in providers:
            try:
                return translate_text_provider(item, text, settings, env)
            except Exception as exc:
                errors.append(f"{item}: {exc}")
        raise RuntimeError("自动选择翻译失败：" + "；".join(errors[:5]))
    if provider == "libretranslate":
        return translate_libretranslate(text, settings, env)
    if provider == "deepl_free":
        return translate_deepl_free(text, settings, env)
    if provider == "google_translate":
        return translate_google_translate(text, settings, env)
    if provider == "microsoft_translator":
        return translate_microsoft_translator(text, settings, env)
    if provider == "mymemory":
        return translate_mymemory(text, settings)
    if provider == "argos_local":
        if env.get("PLATFORM") == "cloudflare":
            raise RuntimeError("Argos Local 不适用于 Cloudflare Worker，请在 Ubuntu/本地环境使用。")
        return translate_argos_local(text)
    raise RuntimeError("未知翻译服务")


def translate_libretranslate(text: str, settings: dict[str, Any], env: dict[str, str]) -> str:
    base = text_only(settings.get("libretranslate_url"), 500).strip().rstrip("/")
    if not base:
        raise RuntimeError("请在通用设置中配置 LibreTranslate 地址，例如自托管 /translate 服务。")
    payload = {"q": text, "source": "zh", "target": "en", "format": "text"}
    api_key = text_only(settings.get("libretranslate_api_key"), 500).strip()
    if api_key:
        payload["api_key"] = api_key
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    timeout = max(3, min(int_value(settings.get("translation_timeout_seconds"), 12), 25 if env.get("PLATFORM") == "cloudflare" else 60))
    request = Request(base + "/translate", data=data, headers={"Content-Type": "application/json", "Accept": "application/json", "User-Agent": "teacher-site/0.1"})
    try:
        with urlopen(request, timeout=timeout) as response:
            result = json.loads(response.read(500_000).decode("utf-8", "ignore"))
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc
    translated = text_only(result.get("translatedText"), 12000).strip()
    if not translated:
        raise RuntimeError(text_only(result.get("error"), 500).strip() or "LibreTranslate 未返回译文")
    return translated


def translate_mymemory(text: str, settings: dict[str, Any]) -> str:
    params = {"q": text, "langpair": "zh-CN|en"}
    email = text_only(settings.get("mymemory_email"), 200).strip()
    if email:
        params["de"] = email
    data = http_json("https://api.mymemory.translated.net/get?" + urlencode(params))
    translated = text_only((data.get("responseData") or {}).get("translatedText"), 12000).strip()
    if not translated:
        raise RuntimeError(text_only(data.get("responseDetails"), 500).strip() or "MyMemory 未返回译文")
    return translated


def translate_deepl_free(text: str, settings: dict[str, Any], env: dict[str, str]) -> str:
    key = text_only(settings.get("deepl_api_key"), 500).strip()
    if not key:
        raise RuntimeError("请在通用设置中配置 DeepL API Key。")
    timeout = max(3, min(int_value(settings.get("translation_timeout_seconds"), 12), 25 if env.get("PLATFORM") == "cloudflare" else 60))
    data = urlencode({"text": text, "source_lang": "ZH", "target_lang": "EN"}).encode("utf-8")
    request = Request(
        "https://api-free.deepl.com/v2/translate",
        data=data,
        headers={"Authorization": f"DeepL-Auth-Key {key}", "Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json", "User-Agent": "teacher-site/0.1"},
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read(500_000).decode("utf-8", "ignore"))
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc
    translations = payload.get("translations") or []
    translated = text_only((translations[0] if translations else {}).get("text"), 12000).strip()
    if not translated:
        raise RuntimeError("DeepL Free 未返回译文")
    return translated


def translate_google_translate(text: str, settings: dict[str, Any], env: dict[str, str]) -> str:
    key = text_only(settings.get("google_translate_api_key"), 500).strip()
    if not key:
        raise RuntimeError("请在通用设置中配置 Google Translate API Key。")
    timeout = max(3, min(int_value(settings.get("translation_timeout_seconds"), 12), 25 if env.get("PLATFORM") == "cloudflare" else 60))
    data = urlencode({"q": text, "source": "zh-CN", "target": "en", "format": "text", "key": key}).encode("utf-8")
    request = Request("https://translation.googleapis.com/language/translate/v2", data=data, headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json", "User-Agent": "teacher-site/0.1"})
    try:
        with urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read(500_000).decode("utf-8", "ignore"))
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc
    translations = ((payload.get("data") or {}).get("translations") or [])
    translated = text_only((translations[0] if translations else {}).get("translatedText"), 12000).strip()
    if not translated:
        raise RuntimeError("Google Translate 未返回译文")
    return translated


def translate_microsoft_translator(text: str, settings: dict[str, Any], env: dict[str, str]) -> str:
    key = text_only(settings.get("microsoft_translator_key"), 500).strip()
    if not key:
        raise RuntimeError("请在通用设置中配置 Microsoft Translator Key。")
    endpoint = text_only(settings.get("microsoft_translator_endpoint"), 500).strip().rstrip("/") or "https://api.cognitive.microsofttranslator.com"
    region = text_only(settings.get("microsoft_translator_region"), 100).strip()
    timeout = max(3, min(int_value(settings.get("translation_timeout_seconds"), 12), 25 if env.get("PLATFORM") == "cloudflare" else 60))
    headers = {
        "Ocp-Apim-Subscription-Key": key,
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "teacher-site/0.1",
    }
    if region:
        headers["Ocp-Apim-Subscription-Region"] = region
    request = Request(
        endpoint + "/translate?api-version=3.0&from=zh-Hans&to=en",
        data=json.dumps([{"Text": text}], ensure_ascii=False).encode("utf-8"),
        headers=headers,
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read(500_000).decode("utf-8", "ignore"))
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc
    translations = ((payload[0] if isinstance(payload, list) and payload else {}).get("translations") or [])
    translated = text_only((translations[0] if translations else {}).get("text"), 12000).strip()
    if not translated:
        raise RuntimeError("Microsoft Translator 未返回译文")
    return translated


def translate_argos_local(text: str) -> str:
    try:
        from argostranslate import translate
    except ImportError as exc:
        raise RuntimeError("当前 Python 环境未安装 argostranslate 或未安装中英模型。") from exc
    translated = text_only(translate.translate(text, "zh", "en"), 12000).strip()
    if not translated:
        raise RuntimeError("Argos Local 未返回译文，请确认已安装 zh -> en 模型。")
    return translated


def admin_navigation_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        enabled = str(row.get("enabled") or "0") in {"1", "true", "True", "yes", "on"}
        toggle_label = "停" if enabled else "启"
        icon = navigation_icon_preview(row.get("icon"))
        title_en = text_only(row.get("title_en"), 80).strip()
        uid_line = text_only(key, 80)
        body.append(f"""<tr class="nav-admin-row{' is-disabled' if not enabled else ''}">
          <td>{admin_batch_select("navigation_items", key)}</td>
          <td class="nav-title-cell"><strong>{esc(row.get("title"))}</strong><small>{esc(uid_line)}</small></td>
          <td class="nav-en-cell">{esc(title_en or "未设置")}</td>
          <td class="nav-icon-cell">{icon}</td>
          <td>{esc(row.get("kind"))}</td>
          <td><span class="admin-status-badge status-{'active' if enabled else 'trash'}">{'启用' if enabled else '停用'}</span></td>
          <td>{esc(row.get("location"))}</td>
          <td class="nav-path-cell"><span>{esc(row.get("path") or row.get("url_name") or row.get("fragment"))}</span></td>
          <td>
            <form class="nav-inline-form" method="post" action="/admin/table/navigation_items/quick-update">
              <input type="hidden" name="uid" value="{esc(key)}">
              <input type="number" name="sort_order" value="{esc(row.get("sort_order"))}" title="排序值，数字越小越靠前">
              <button class="button light" type="submit" name="_nav_action" value="save_sort">应用</button>
              <button class="button {'danger' if enabled else 'secondary'}" type="submit" name="_nav_action" value="toggle_enabled">{toggle_label}</button>
            </form>
          </td>
          <td><a class="button ghost" href="/admin/table/navigation_items/{esc(key)}">编辑</a></td>
        </tr>""")
    return f"""<section class="admin-card nav-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/navigation_items/new">新增</a></div>
      {admin_list_tools(navigation_filter_form(query, all_rows), admin_batch_toolbar("navigation_items", meta, query, all_rows))}
      <div class="table-wrap"><table class="nav-admin-table"><thead><tr><th>选</th><th>标题</th><th>英文标题</th><th>图标</th><th>类型</th><th>状态</th><th>位置</th><th>目标</th><th>快速修改</th><th>操作</th></tr></thead><tbody>{''.join(body) or '<tr><td colspan="10">暂无数据</td></tr>'}</tbody></table></div>
    </section>"""


def navigation_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    kinds = sorted({text_only(row.get("kind"), 40).strip() for row in rows if text_only(row.get("kind"), 40).strip()})
    locations = sorted({text_only(row.get("location"), 80).strip() for row in rows if text_only(row.get("location"), 80).strip()})
    return f"""<form class="filters nav-admin-search" method="get" action="/admin/table/navigation_items">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索标题、英文标题、路径、位置、图标">
      <select name="kind"><option value="">全部类型</option>{options(kinds, query.get("kind", ""))}</select>
      <select name="location"><option value="">全部位置</option>{options(locations, query.get("location", ""))}</select>
      <select name="enabled">{navigation_pair_options([("", "全部状态"), ("1", "启用"), ("0", "停用")], query.get("enabled", ""))}</select>
      <select name="sort">{navigation_pair_options(navigation_sort_pairs(), query.get("sort", "sort_asc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/navigation_items">重置</a>
    </form>"""


def navigation_pair_options(pairs: list[tuple[str, str]], selected: str) -> str:
    return "".join(f'<option value="{esc(value)}"{" selected" if str(value) == str(selected) else ""}>{esc(label)}</option>' for value, label in pairs)


def navigation_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_asc", "排序从小到大"),
        ("sort_desc", "排序从大到小"),
        ("title_asc", "中文标题 A-Z"),
        ("title_en_asc", "英文标题 A-Z"),
        ("location_asc", "位置"),
        ("kind_asc", "类型"),
        ("enabled_desc", "启用优先"),
    ]


def navigation_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_asc": ("sort_order", False),
        "sort_desc": ("sort_order", True),
        "title_asc": ("title", False),
        "title_en_asc": ("title_en", False),
        "location_asc": ("location", False),
        "kind_asc": ("kind", False),
        "enabled_desc": ("enabled", True),
    }.get(value, ("sort_order", False))


def navigation_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("navigation_items", key)
    if not row:
        return "/admin/table/navigation_items"
    action = data.get("_nav_action", "save_sort")
    if action == "toggle_enabled":
        enabled = str(row.get("enabled") or "0") in {"1", "true", "True", "yes", "on"}
        row["enabled"] = 0 if enabled else 1
    else:
        row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("sort_order"), 0))
    repo.save("navigation_items", row)
    return "/admin/table/navigation_items"


def navigation_icon_preview(value: Any) -> str:
    text = text_only(value, 300).strip()
    if not text:
        return '<span class="nav-icon-preview empty">无</span>'
    if text.startswith(("http://", "https://", "/media/", "media/", "public/media/")) or text.lower().endswith((".svg", ".png", ".jpg", ".jpeg", ".webp", ".gif")):
        src = media_url(normalize_media_key(text)) if not text.startswith(("http://", "https://")) else safe_href(text)
        return f'<span class="nav-icon-preview"><img src="{esc(src)}" alt="" loading="lazy"></span>'
    return f'<span class="nav-icon-preview">{esc(text[:2])}</span>'


def admin_profiles_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], env: dict[str, str]) -> str:
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        active = str(row.get("is_active") or "0") in {"1", "true", "True", "yes", "on"}
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        contact = " / ".join(item for item in [text_only(row.get("email"), 120), text_only(row.get("phone"), 80), text_only(row.get("office"), 80)] if item) or "未填写"
        body.append(f"""<article class="profile-admin-row{' is-disabled' if not active else ''}">
          {admin_batch_select("profiles", key)}
          <figure class="profile-admin-avatar-cell">{image_tag(row.get("avatar_key"), row.get("name") or "", "profile-admin-avatar", env.get("PUBLIC_MEDIA_BASE_URL", ""))}</figure>
          <div class="profile-admin-person">
            <strong>{esc(row.get("name"))}</strong>
            <span>{esc(row.get("name_en") or "英文名未设置")}</span>
            <small>{esc(key)}</small>
          </div>
          <div class="profile-admin-identity">
            <span>{esc(row.get("role") or "未设角色")}</span>
            <span>{esc(row.get("title") or "未设职称")}</span>
            <small>{esc(" / ".join(item for item in [text_only(row.get("organization"), 80), text_only(row.get("lab"), 80)] if item) or "单位/团队未设置")}</small>
          </div>
          <div class="profile-admin-contact"><span>{esc(contact)}</span><small>{esc(profile_admin_links(row))}</small></div>
          <div class="profile-admin-flags">
            <span class="admin-status-badge status-{'active' if active else 'trash'}">{'启用' if active else '停用'}</span>
            <span class="admin-status-badge status-{'active' if featured else 'trash'}">{'首页' if featured else '普通'}</span>
            <small>{esc(row.get("visibility") or "public")}</small>
          </div>
          <form class="profile-inline-form" method="post" action="/admin/table/profiles/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="profile-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order"))}" title="排序值，数字越小越靠前"></label>
            <button class="button light" type="submit" name="_profile_action" value="save_inline">应用</button>
          </form>
          <div class="profile-admin-actions">
            <a class="button ghost" href="/admin/table/profiles/{esc(key)}">编辑</a>
            <form method="post" action="/admin/table/profiles/quick-update"><input type="hidden" name="uid" value="{esc(key)}"><button class="button {'danger' if active else 'secondary'}" type="submit" name="_profile_action" value="toggle_active">{'停用' if active else '启用'}</button></form>
            <form method="post" action="/admin/table/profiles/quick-update"><input type="hidden" name="uid" value="{esc(key)}"><button class="button light" type="submit" name="_profile_action" value="toggle_featured">{'取消首页' if featured else '设首页'}</button></form>
          </div>
        </article>""")
    return f"""<section class="admin-card profile-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/profiles/new">新增</a></div>
      {admin_list_tools(profile_filter_form(query, all_rows), admin_batch_toolbar("profiles", meta, query, all_rows))}
      <div class="profile-admin-list">
        <div class="profile-admin-head"><span>选</span><span>照片</span><span>姓名</span><span>身份</span><span>联系与链接</span><span>状态</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无团队成员。</p>'}
      </div>
    </section>"""


def profile_admin_links(row: dict[str, Any]) -> str:
    pairs = [
        ("主页", row.get("personal_homepage")),
        ("Scholar", row.get("google_scholar")),
        ("ORCID", row.get("orcid")),
        ("DBLP", row.get("dblp")),
        ("GitHub", row.get("github")),
        ("CNKI", row.get("cnki")),
    ]
    labels = [label for label, value in pairs if text_only(value, 200).strip()]
    return " · ".join(labels) if labels else "链接未填写"


def profile_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    roles = sorted({text_only(row.get("role"), 80).strip() for row in rows if text_only(row.get("role"), 80).strip()})
    titles = sorted({text_only(row.get("title"), 80).strip() for row in rows if text_only(row.get("title"), 80).strip()})
    orgs = sorted({text_only(row.get("organization"), 80).strip() for row in rows if text_only(row.get("organization"), 80).strip()})
    labs = sorted({text_only(row.get("lab"), 80).strip() for row in rows if text_only(row.get("lab"), 80).strip()})
    return f"""<form class="filters profile-admin-search" method="get" action="/admin/table/profiles">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索姓名、英文名、角色、职称、单位、邮箱、简介">
      <select name="role"><option value="">全部角色</option>{options(roles, query.get("role", ""))}</select>
      <select name="title"><option value="">全部职称</option>{options(titles, query.get("title", ""))}</select>
      <select name="organization"><option value="">全部单位</option>{options(orgs, query.get("organization", ""))}</select>
      <select name="lab"><option value="">全部团队</option>{options(labs, query.get("lab", ""))}</select>
      <select name="is_active">{navigation_pair_options([("", "全部状态"), ("1", "启用"), ("0", "停用")], query.get("is_active", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "首页不限"), ("1", "首页展示"), ("0", "非首页")], query.get("is_featured", ""))}</select>
      <select name="sort">{navigation_pair_options(profile_sort_pairs(), query.get("sort", "sort_asc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/profiles">重置</a>
    </form>"""


def profile_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_asc", "排序从小到大"),
        ("sort_desc", "排序从大到小"),
        ("name_asc", "姓名 A-Z"),
        ("role_asc", "角色"),
        ("title_asc", "职称"),
        ("org_asc", "单位"),
        ("active_desc", "启用优先"),
        ("featured_desc", "首页优先"),
    ]


def profile_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_asc": ("sort_order", False),
        "sort_desc": ("sort_order", True),
        "name_asc": ("name", False),
        "role_asc": ("role", False),
        "title_asc": ("title", False),
        "org_asc": ("organization", False),
        "active_desc": ("is_active", True),
        "featured_desc": ("is_featured", True),
    }.get(value, ("sort_order", False))


def profile_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("profiles", key)
    if not row:
        return "/admin/table/profiles"
    action = data.get("_profile_action", "save_inline")
    if action == "toggle_active":
        active = str(row.get("is_active") or "0") in {"1", "true", "True", "yes", "on"}
        row["is_active"] = 0 if active else 1
    elif action == "toggle_featured":
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        row["is_featured"] = 0 if featured else 1
    else:
        row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("sort_order"), 0))
    repo.save("profiles", row)
    return "/admin/table/profiles"


def admin_research_interests_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    body = []
    visibility_choices = research_interest_visibility_choices(meta)
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        is_public = visibility == "public"
        description = text_only(row.get("description"), 500).strip()
        body.append(f"""<article class="research-admin-row{' is-disabled' if not is_public else ''}">
          {admin_batch_select("research_interests", key)}
          <div class="research-admin-title">
            <strong>{esc(row.get("name"))}</strong>
            <span>{esc(row.get("name_en") or "英文名称未设置")}</span>
            <small>{esc(key)}</small>
          </div>
          <div class="research-admin-description" title="{esc(description)}">{esc(description or "描述未填写")}</div>
          <div class="research-admin-flags">
            <span class="admin-status-badge status-{'active' if is_public else 'trash'}">{esc(visibility)}</span>
            <small>排序 {esc(row.get("sort_order"))}</small>
          </div>
          <form class="research-inline-form" method="post" action="/admin/table/research_interests/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="research-quick-field"><span>中文名</span><input name="name" value="{esc(row.get("name"))}" title="中文方向名称"></label>
            <label class="research-quick-field"><span>英文名</span><input name="name_en" value="{esc(row.get("name_en"))}" title="英文方向名称"></label>
            <label class="research-quick-field"><span>可见</span><select name="visibility" title="可见范围">{options(list(visibility_choices), visibility)}</select></label>
            <label class="research-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order"))}" title="排序值，数字越小越靠前"></label>
            <button class="button light" type="submit" name="_research_action" value="save_inline">应用</button>
          </form>
          <div class="research-admin-actions">
            <a class="button ghost" href="/admin/table/research_interests/{esc(key)}">编辑</a>
          </div>
        </article>""")
    return f"""<section class="admin-card research-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/research_interests/new">新增</a></div>
      {admin_list_tools(research_interest_filter_form(query, all_rows, meta), admin_batch_toolbar("research_interests", meta, query, all_rows))}
      <div class="research-admin-list">
        <div class="research-admin-head"><span>选</span><span>方向名称</span><span>描述摘要</span><span>状态</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无研究方向。</p>'}
      </div>
    </section>"""


def research_interest_filter_form(query: dict[str, str], rows: list[dict[str, Any]], meta: Table) -> str:
    visibility_values = sorted({text_only(row.get("visibility"), 40).strip() for row in rows if text_only(row.get("visibility"), 40).strip()})
    if not visibility_values:
        visibility_values = list(research_interest_visibility_choices(meta))
    return f"""<form class="filters research-admin-search" method="get" action="/admin/table/research_interests">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索中文名称、英文名称、描述">
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="sort">{navigation_pair_options(research_interest_sort_pairs(), query.get("sort", "sort_asc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/research_interests">重置</a>
    </form>"""


def research_interest_visibility_choices(meta: Table) -> tuple[str, ...]:
    for field in meta.fields:
        if field.name == "visibility" and field.choices:
            return field.choices
    return ("public", "authenticated", "staff", "owner", "hidden")


def research_interest_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_asc", "排序从小到大"),
        ("sort_desc", "排序从大到小"),
        ("name_asc", "中文名称 A-Z"),
        ("name_en_asc", "英文名称 A-Z"),
        ("visibility_asc", "可见范围"),
    ]


def research_interest_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_asc": ("sort_order", False),
        "sort_desc": ("sort_order", True),
        "name_asc": ("name", False),
        "name_en_asc": ("name_en", False),
        "visibility_asc": ("visibility", False),
    }.get(value, ("sort_order", False))


def research_interest_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("research_interests", key)
    if not row:
        return "/admin/table/research_interests"
    name = text_only(data.get("name"), 300).strip()
    row["name"] = name or row.get("name")
    row["name_en"] = text_only(data.get("name_en"), 300).strip()
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("sort_order"), 0))
    repo.save("research_interests", row)
    return "/admin/table/research_interests"


def admin_publications_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], display_style: str) -> str:
    body = []
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        authors = text_only(row.get("authors"), 300).strip()
        tags = publication_admin_tags(row)
        front_href = f'/publications?q={quote(text_only(row.get("title"), 160).strip())}'
        title_text = text_only(row.get("title"), 500).strip()
        doi_href = doi_url(row.get("doi"))
        title_node = f'<a href="{esc(doi_href)}" target="_blank" rel="noreferrer" title="打开 DOI：{esc(row.get("doi"))}">{esc(title_text or "未命名论文")}</a>' if doi_href else f'<strong>{esc(title_text or "未命名论文")}</strong>'
        body.append(f"""<article class="publication-admin-row{' is-disabled' if visibility != 'public' else ''}">
          {admin_batch_select("publications", key)}
          <div class="publication-admin-authors" title="{esc(authors)}">{esc(authors or "作者未填写")}</div>
          <div class="publication-admin-title" title="{esc(title_text)}">
            {title_node}
            <small>{esc(key)}</small>
          </div>
          <div class="publication-admin-venue">
            <strong>{esc(row.get("venue") or "期刊/会议未填写")}</strong>
            <span>{esc(publication_issue_line(row))}</span>
            <small>{esc(row.get("doi") or "DOI 未填写")}</small>
          </div>
          <div class="publication-admin-tags">
            {tags}
            <span class="admin-status-badge status-{'active' if visibility == 'public' else 'trash'}">{esc(visibility)}</span>
            <span class="admin-status-badge status-{'active' if featured else 'trash'}">{'代表作' if featured else '普通'}</span>
          </div>
          <form class="publication-inline-form" method="post" action="/admin/table/publications/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="publication-quick-field"><span>代表</span><select name="is_featured" title="是否代表作">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if featured else "0")}</select></label>
            <label class="publication-quick-field"><span>可见</span><select name="visibility" title="可见范围">{options(list(visibility_choices), visibility)}</select></label>
            <label class="publication-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order"))}" title="排序值"></label>
            <button class="button light" type="submit" name="_publication_action" value="save_inline">应用</button>
          </form>
          <div class="publication-admin-actions">
            <a class="button ghost" href="/admin/table/publications/{esc(key)}">编辑</a>
            {f'<a class="button light" href="{esc(doi_href)}" target="_blank" rel="noreferrer">DOI</a>' if doi_href else ""}
            <a class="button light" href="{esc(front_href)}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card publication-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/publications/new">新增</a></div>
      {admin_list_tools(publication_admin_filter_form(query, all_rows), admin_batch_toolbar("publications", meta, query, all_rows))}
      <div class="publication-admin-list">
        <div class="publication-admin-head"><span>选</span><span>作者</span><span>标题</span><span>期刊</span><span>标签</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无论文。</p>'}
      </div>
    </section>"""


def publication_admin_source(row: dict[str, Any]) -> str:
    parts = []
    for key in ("venue", "year", "volume", "issue", "pages"):
        value = text_only(row.get(key), 120).strip()
        if value:
            parts.append(value)
    return esc(" / ".join(parts) or "来源信息未填写")


def publication_issue_line(row: dict[str, Any]) -> str:
    details = []
    for label, key in [("年", "year"), ("卷", "volume"), ("期", "issue"), ("页", "pages")]:
        value = text_only(row.get(key), 80).strip()
        if value:
            details.append(f"{label} {value}")
    return " / ".join(details) or "出版信息未填写"


def doi_url(value: Any) -> str:
    text = text_only(value, 200).strip()
    if not text:
        return ""
    text = text.removeprefix("https://doi.org/").removeprefix("http://doi.org/").removeprefix("doi:")
    return safe_href(f"https://doi.org/{text}")


def publication_admin_tags(row: dict[str, Any]) -> str:
    values = []
    custom = text_only(row.get("display_tags"), 500).strip()
    if custom:
        chunks = custom.replace("；", ";").replace("，", ",").replace("\n", ",").replace(";", ",").split(",")
        values.extend(part.strip() for part in chunks if part.strip())
    else:
        values.extend(text_only(row.get(key), 80).strip() for key in ("publication_type", "index_type") if text_only(row.get(key), 80).strip())
    return "".join(f'<span class="publication-admin-tag">{esc(value[:40])}</span>' for value in values[:5]) or '<span class="publication-admin-tag muted">无标签</span>'


def publication_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    years = sorted({text_only(row.get("year"), 20).strip() for row in rows if text_only(row.get("year"), 20).strip()}, key=lambda value: int_value(value, 0), reverse=True)
    venues = sorted({text_only(row.get("venue"), 160).strip() for row in rows if text_only(row.get("venue"), 160).strip()})
    types = sorted({text_only(row.get("publication_type"), 80).strip() for row in rows if text_only(row.get("publication_type"), 80).strip()})
    roles = sorted({text_only(row.get("author_role"), 80).strip() for row in rows if text_only(row.get("author_role"), 80).strip()})
    indexes = sorted({text_only(row.get("index_type"), 80).strip() for row in rows if text_only(row.get("index_type"), 80).strip()})
    visibility_values = sorted({text_only(row.get("visibility"), 40).strip() for row in rows if text_only(row.get("visibility"), 40).strip()})
    return f"""<form class="filters publication-admin-search" method="get" action="/admin/table/publications">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索题名、作者、期刊会议、DOI、标签、摘要">
      <select name="year"><option value="">全部年份</option>{options(years, query.get("year", ""))}</select>
      <select name="venue"><option value="">全部期刊会议</option>{options(venues, query.get("venue", ""))}</select>
      <select name="publication_type"><option value="">全部类型</option>{options(types, query.get("publication_type", ""))}</select>
      <select name="author_role"><option value="">作者角色</option>{options(roles, query.get("author_role", ""))}</select>
      <select name="index_type"><option value="">全部收录</option>{options(indexes, query.get("index_type", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "代表不限"), ("1", "代表作"), ("0", "普通")], query.get("is_featured", ""))}</select>
      <select name="sort">{navigation_pair_options(publication_admin_sort_pairs(), query.get("sort", "year_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/publications">重置</a>
    </form>"""


def publication_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("year_desc", "年份从新到旧"),
        ("year_asc", "年份从旧到新"),
        ("sort_asc", "排序从小到大"),
        ("sort_desc", "排序从大到小"),
        ("title_asc", "题名 A-Z"),
        ("venue_asc", "期刊会议"),
        ("featured_desc", "代表作优先"),
    ]


def publication_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "year_desc": ("year", True),
        "year_asc": ("year", False),
        "sort_asc": ("sort_order", False),
        "sort_desc": ("sort_order", True),
        "title_asc": ("title", False),
        "venue_asc": ("venue", False),
        "featured_desc": ("is_featured", True),
    }.get(value, ("year", True))


def table_field_choices(meta: Table, name: str) -> tuple[str, ...]:
    for field in meta.fields:
        if field.name == name and field.choices:
            return field.choices
    return ()


BATCH_UPDATE_FIELDS: dict[str, tuple[str, ...]] = {
    "navigation_items": ("enabled", "visibility", "location", "kind"),
    "profiles": ("is_active", "is_featured", "visibility", "contact_visibility", "role", "title", "organization", "lab"),
    "research_interests": ("visibility",),
    "publications": ("is_featured", "visibility", "pdf_visibility", "publication_type", "author_role", "index_type"),
    "projects": ("is_featured", "visibility", "status", "source", "fund_name"),
    "patents": ("is_featured", "visibility", "legal_status", "patent_type", "country"),
    "students": ("is_featured", "visibility", "contact_visibility", "degree", "category", "grade", "status"),
    "student_category_displays": ("enabled",),
    "news": ("is_featured", "visibility", "allow_comments", "category", "content_format"),
    "courses": ("is_featured", "visibility", "material_visibility", "semester", "audience"),
    "messages": ("status", "visibility", "message_type"),
    "auth_roles": ("is_active", "visibility_scopes"),
    "auth_users": ("role_uid", "status", "visibility", "must_change_password"),
    "auth_permissions": ("can_view", "can_create", "can_edit", "can_delete", "can_export"),
}


def admin_list_tools(filter_html: str, batch_html: str) -> str:
    return f'<div class="admin-list-tools">{filter_html}{batch_html}</div>'


def admin_batch_select(table: str, key: Any) -> str:
    return f'<label class="admin-batch-select-cell" title="选择此条记录"><input type="checkbox" name="selected" value="{esc(key)}" form="{esc(table)}-batch-form" data-batch-table="{esc(table)}"><span class="sr-only">选择</span></label>'


def admin_batch_toolbar(table: str, meta: Table, query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    field_controls = admin_batch_field_controls(table, meta, all_rows)
    notice = admin_batch_result_notice(query)
    sort_name = "display_order" if table == "student_category_displays" else "sort_order"
    if sort_name not in meta.field_names:
        sort_controls = ""
    else:
        sort_controls = f"""
        <label class="admin-batch-field admin-batch-number"><span>排序起始</span><input type="number" name="_sort_start" placeholder="例：100"></label>
        <label class="admin-batch-field admin-batch-number"><span>排序步长</span><input type="number" name="_sort_step" value="1" placeholder="例：1"></label>"""
    return f"""{notice}<form id="{esc(table)}-batch-form" class="admin-batch-toolbar" data-batch-table="{esc(table)}" method="post" action="/admin/table/{esc(table)}/batch-update">
      <input type="hidden" name="return_to" value="{esc(admin_table_return_to(table, query))}">
      <label class="admin-batch-select-all"><input type="checkbox" data-batch-select-all="{esc(table)}"> 全选</label>
      {field_controls}{sort_controls}
      <button type="submit">应用到选中</button>
      <span class="admin-muted" data-batch-count="{esc(table)}">已选 0 条</span>
      <small class="admin-batch-help">示例：筛选“非公开”后全选，将可见范围统一改为 public；填写“排序起始 100、步长 10”可按当前列表顺序重排。留空的配置不会改动。</small>
    </form>"""


def admin_batch_field_controls(table: str, meta: Table, all_rows: list[dict[str, Any]]) -> str:
    by_name = {field.name: field for field in meta.fields}
    controls = []
    for name in BATCH_UPDATE_FIELDS.get(table, ()):
        field = by_name.get(name)
        if not field:
            continue
        label = field.label.replace("可见范围", "可见").replace("联系方式可见性", "联系可见").replace("资料可见性", "资料可见").replace("首页展示", "首页")
        if field.kind == "bool":
            choices = [("", f"{label}不变"), ("1", "是/启用"), ("0", "否/停用")]
            controls.append(f'<label class="admin-batch-field"><span>{esc(label)}</span><select name="{esc(name)}__set">{navigation_pair_options(choices, "")}</select></label>')
            continue
        values = list(field.choices) if field.choices else admin_batch_values(all_rows, name)
        if values:
            controls.append(f'<label class="admin-batch-field"><span>{esc(label)}</span><select name="{esc(name)}__set"><option value="">{esc(label)}不变</option>{options(values, "")}</select></label>')
        else:
            controls.append(f'<label class="admin-batch-field admin-batch-text"><span>{esc(label)}</span><input name="{esc(name)}__set" placeholder="留空不改"></label>')
    return "".join(controls)


def admin_batch_values(rows: list[dict[str, Any]], field: str) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        value = text_only(row.get(field), 180).strip()
        key = value.casefold()
        if value and key not in seen:
            values.append(value)
            seen.add(key)
    return values[:80]


def admin_table_return_to(table: str, query: dict[str, str]) -> str:
    params = {key: value for key, value in query.items() if value}
    suffix = f"?{urlencode(params)}" if params else ""
    return f"/admin/table/{table}{suffix}"


def admin_batch_update(repo: Repository, table: str, body: bytes) -> tuple[str, dict[str, Any]]:
    if table not in TABLE_MAP:
        return "/admin", {"selected": 0, "updated": 0, "deleted": 0, "skipped": 0}
    data = _form_multi(body)
    selected = [text_only(item, 180).strip() for item in data.get("selected", []) if text_only(item, 180).strip()][:500]
    return_to = text_only((data.get("return_to") or [f"/admin/table/{table}"])[-1], 500).strip()
    if not return_to.startswith(f"/admin/table/{table}"):
        return_to = f"/admin/table/{table}"
    if not selected:
        return append_query_params(return_to, {"batch_selected": 0, "batch_updated": 0, "batch_deleted": 0, "batch_skipped": 0}), {"selected": 0, "updated": 0, "deleted": 0, "skipped": 0}
    meta = TABLE_MAP[table]
    allowed = set(BATCH_UPDATE_FIELDS.get(table, ()))
    changes: dict[str, Any] = {}
    for field in meta.fields:
        if field.name not in allowed:
            continue
        raw_values = data.get(f"{field.name}__set", [])
        value = text_only(raw_values[-1], 500).strip() if raw_values else ""
        if value == "":
            continue
        if field.kind == "bool":
            changes[field.name] = 1 if value in {"1", "true", "True", "yes", "on"} else 0
        elif field.kind == "number":
            changes[field.name] = int_value(value, 0)
        else:
            changes[field.name] = value
    sort_name = "display_order" if table == "student_category_displays" else "sort_order"
    sort_start = text_only((data.get("_sort_start") or [""])[-1], 40).strip()
    sort_step = int_value((data.get("_sort_step") or ["1"])[-1], 1)
    sort_step = sort_step if sort_step != 0 else 1
    updated = 0
    skipped = 0
    for index, key in enumerate(selected):
        row = repo.get(table, key)
        if not row:
            skipped += 1
            continue
        next_row = {**row, **changes}
        if sort_start and sort_name in meta.field_names:
            next_row[sort_name] = int_value(sort_start, 0) + index * sort_step
        if next_row != row:
            repo.save(table, next_row)
            updated += 1
        else:
            skipped += 1
    result = {"selected": len(selected), "updated": updated, "deleted": 0, "skipped": skipped, "changes": list(changes)}
    return append_query_params(return_to, {"batch_selected": len(selected), "batch_updated": updated, "batch_deleted": 0, "batch_skipped": skipped}), result


def publication_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("publications", key)
    if not row:
        return "/admin/table/publications"
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["is_featured"] = 1 if str(data.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("sort_order"), 0))
    repo.save("publications", row)
    return "/admin/table/publications"


def admin_projects_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    body = []
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        name = text_only(row.get("name"), 400).strip()
        status = text_only(row.get("status"), 80).strip() or "未设置"
        status_value = "" if status == "未设置" else status
        front_href = f'/projects?q={quote(name)}'
        body.append(f"""<article class="project-admin-row{' is-disabled' if visibility != 'public' else ''}">
          {admin_batch_select("projects", key)}
          <div class="project-admin-main" title="{esc(name)}">
            <strong>{esc(name or "未命名项目")}</strong>
            <small>{esc(key)}</small>
          </div>
          <div class="project-admin-fund">
            <strong>{esc(row.get("source") or "来源未填写")}</strong>
            <span>{esc(row.get("fund_name") or "基金/计划未填写")}</span>
            <small>{esc(row.get("project_number") or "项目编号未填写")}</small>
          </div>
          <div class="project-admin-people">
            <span>负责人：{esc(row.get("principal") or "未填写")}</span>
            <small title="{esc(row.get("members") or "")}">成员：{esc(text_only(row.get("members"), 180) or "未填写")}</small>
          </div>
          <div class="project-admin-state">
            <span class="publication-admin-tag">{esc(status)}</span>
            <small>{esc(project_period(row))}</small>
            <small>{esc(project_amount(row))}</small>
          </div>
          <form class="project-inline-form" method="post" action="/admin/table/projects/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="project-quick-field"><span>状态</span><input name="status" value="{esc(status_value)}" data-project-suggest="status" autocomplete="off"></label>
            <label class="project-quick-field"><span>首页</span><select name="is_featured">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if featured else "0")}</select></label>
            <label class="project-quick-field"><span>可见</span><select name="visibility">{options(list(visibility_choices), visibility)}</select></label>
            <label class="project-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order") or row.get("id"))}" title="排序值，留空保存时使用记录 ID"></label>
            <button class="button light" type="submit" name="_project_action" value="save_inline">应用</button>
          </form>
          <div class="project-admin-actions">
            <a class="button ghost" href="/admin/table/projects/{esc(key)}">编辑</a>
            <a class="button light" href="{esc(front_href)}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card project-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/projects/new">新增</a></div>
      {admin_list_tools(project_admin_filter_form(query, all_rows), admin_batch_toolbar("projects", meta, query, all_rows))}
      <div class="project-admin-list">
        <div class="project-admin-head"><span>选</span><span>项目</span><span>来源/基金</span><span>人员</span><span>状态/周期</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无项目。</p>'}
      </div>
    </section>"""


def project_period(row: dict[str, Any]) -> str:
    start = text_only(row.get("start_date"), 40).strip()
    end = text_only(row.get("end_date"), 40).strip()
    return f"{start or '未填开始'} 至 {end or '未填结束'}"


def project_amount(row: dict[str, Any]) -> str:
    amount = text_only(row.get("amount"), 80).strip()
    return f"金额：{amount}" if amount else "金额未填写"


def project_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    sources = sorted({text_only(row.get("source"), 160).strip() for row in rows if text_only(row.get("source"), 160).strip()})
    funds = sorted({text_only(row.get("fund_name"), 180).strip() for row in rows if text_only(row.get("fund_name"), 180).strip()})
    statuses = sorted({text_only(row.get("status"), 80).strip() for row in rows if text_only(row.get("status"), 80).strip()})
    visibility_values = sorted({text_only(row.get("visibility"), 40).strip() for row in rows if text_only(row.get("visibility"), 40).strip()})
    return f"""<form class="filters project-admin-search" method="get" action="/admin/table/projects">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索项目名称、来源、基金、编号、负责人、成员">
      <select name="source"><option value="">全部来源</option>{options(sources, query.get("source", ""))}</select>
      <select name="fund_name"><option value="">全部基金/计划</option>{options(funds, query.get("fund_name", ""))}</select>
      <select name="status"><option value="">全部状态</option>{options(statuses, query.get("status", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "首页不限"), ("1", "首页展示"), ("0", "普通")], query.get("is_featured", ""))}</select>
      <select name="sort">{navigation_pair_options(project_admin_sort_pairs(), query.get("sort", "sort_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/projects">重置</a>
    </form>"""


def project_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_desc", "排序从大到小"),
        ("sort_asc", "排序从小到大"),
        ("start_desc", "开始时间从新到旧"),
        ("end_desc", "结束时间从新到旧"),
        ("name_asc", "项目名称 A-Z"),
        ("status_asc", "状态"),
        ("featured_desc", "首页展示优先"),
    ]


def project_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_desc": ("sort_order", True),
        "sort_asc": ("sort_order", False),
        "start_desc": ("start_date", True),
        "end_desc": ("end_date", True),
        "name_asc": ("name", False),
        "status_asc": ("status", False),
        "featured_desc": ("is_featured", True),
    }.get(value, ("sort_order", True))


def project_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("projects", key)
    if not row:
        return "/admin/table/projects"
    row["status"] = text_only(data.get("status"), 80).strip() or row.get("status") or ""
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["is_featured"] = 1 if str(data.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("id"), int_value(row.get("sort_order"), 0)))
    repo.save("projects", row)
    return "/admin/table/projects"


def admin_patents_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    body = []
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        name = text_only(row.get("name"), 400).strip()
        legal_status = text_only(row.get("legal_status"), 80).strip() or "未设置"
        front_href = f'/patents?q={quote(name)}'
        body.append(f"""<article class="patent-admin-row{' is-disabled' if visibility != 'public' else ''}">
          {admin_batch_select("patents", key)}
          <div class="patent-admin-main" title="{esc(name)}">
            <strong>{esc(name or "未命名专利/软著")}</strong>
            <small>{esc(key)}</small>
          </div>
          <div class="patent-admin-people">
            <span title="{esc(row.get("inventors") or "")}">发明人/作者：{esc(text_only(row.get("inventors"), 160) or "未填写")}</span>
            <small title="{esc(row.get("owner") or "")}">权利人：{esc(text_only(row.get("owner"), 160) or "未填写")}</small>
          </div>
          <div class="patent-admin-numbers">
            <span title="{esc(row.get("application_number") or "")}">申请：{esc(row.get("application_number") or "未填写")} <em>{esc(row.get("application_date") or "")}</em></span>
            <small title="{esc(row.get("grant_number") or "")}">授权：{esc(row.get("grant_number") or "未填写")} <em>{esc(row.get("grant_date") or "")}</em></small>
          </div>
          <div class="patent-admin-status">
            <span class="publication-admin-tag">{esc(row.get("patent_type") or "类型未填写")}</span>
            <span class="publication-admin-tag">{esc(row.get("country") or "国别未填写")}</span>
            <span class="publication-admin-tag">{esc(legal_status)}</span>
          </div>
          <form class="patent-inline-form" method="post" action="/admin/table/patents/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="patent-quick-field"><span>首页</span><select name="is_featured">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if featured else "0")}</select></label>
            <label class="patent-quick-field"><span>可见</span><select name="visibility">{options(list(visibility_choices), visibility)}</select></label>
            <label class="patent-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order") or row.get("id"))}" title="排序值，留空保存时使用记录 ID"></label>
            <button class="button light" type="submit" name="_patent_action" value="save_inline">应用</button>
          </form>
          <div class="patent-admin-actions">
            <a class="button ghost" href="/admin/table/patents/{esc(key)}">编辑</a>
            <a class="button light" href="{esc(front_href)}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card patent-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/patents/new">新增</a></div>
      {admin_list_tools(patent_admin_filter_form(query, all_rows), admin_batch_toolbar("patents", meta, query, all_rows))}
      <div class="patent-admin-list">
        <div class="patent-admin-head"><span>选</span><span>名称</span><span>人员</span><span>编号</span><span>状态</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无专利与软著。</p>'}
      </div>
    </section>"""


def patent_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    countries = sorted({text_only(row.get("country"), 80).strip() for row in rows if text_only(row.get("country"), 80).strip()})
    types = sorted({text_only(row.get("patent_type"), 120).strip() for row in rows if text_only(row.get("patent_type"), 120).strip()})
    statuses = sorted({text_only(row.get("legal_status"), 80).strip() for row in rows if text_only(row.get("legal_status"), 80).strip()})
    visibility_values = sorted({text_only(row.get("visibility"), 40).strip() for row in rows if text_only(row.get("visibility"), 40).strip()})
    return f"""<form class="filters patent-admin-search" method="get" action="/admin/table/patents">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索名称、申请号、授权号、发明人、权利人">
      <select name="country"><option value="">全部国别</option>{options(countries, query.get("country", ""))}</select>
      <select name="patent_type"><option value="">全部类型</option>{options(types, query.get("patent_type", ""))}</select>
      <select name="legal_status"><option value="">全部状态</option>{options(statuses, query.get("legal_status", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "首页不限"), ("1", "首页展示"), ("0", "普通")], query.get("is_featured", ""))}</select>
      <select name="sort">{navigation_pair_options(patent_admin_sort_pairs(), query.get("sort", "sort_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/patents">重置</a>
    </form>"""


def patent_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_desc", "排序从大到小"),
        ("sort_asc", "排序从小到大"),
        ("application_desc", "申请时间从新到旧"),
        ("grant_desc", "授权时间从新到旧"),
        ("name_asc", "名称 A-Z"),
        ("status_asc", "状态"),
        ("featured_desc", "首页展示优先"),
    ]


def patent_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_desc": ("sort_order", True),
        "sort_asc": ("sort_order", False),
        "application_desc": ("application_date", True),
        "grant_desc": ("grant_date", True),
        "name_asc": ("name", False),
        "status_asc": ("legal_status", False),
        "featured_desc": ("is_featured", True),
    }.get(value, ("sort_order", True))


def patent_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("patents", key)
    if not row:
        return "/admin/table/patents"
    row["legal_status"] = text_only(data.get("legal_status"), 80).strip() or row.get("legal_status") or ""
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["is_featured"] = 1 if str(data.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("id"), int_value(row.get("sort_order"), 0)))
    repo.save("patents", row)
    return "/admin/table/patents"


def admin_students_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], env: dict[str, str]) -> str:
    body = []
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        name = text_only(row.get("name"), 120).strip()
        name_en = text_only(row.get("name_en"), 160).strip()
        display_name = name or name_en or "未命名学生"
        student_number = text_only(row.get("student_id"), 80).strip()
        degree = text_only(row.get("degree"), 80).strip()
        category = text_only(row.get("category"), 80).strip()
        grade = text_only(row.get("grade"), 80).strip()
        status = text_only(row.get("status"), 80).strip() or "未设置"
        direction = text_only(row.get("direction"), 220).strip()
        destination = text_only(row.get("destination"), 220).strip()
        contact = " / ".join(item for item in [text_only(row.get("email"), 140), text_only(row.get("homepage"), 220)] if item) or "未填写"
        front_href = f'/students?q={quote(display_name)}'
        period = student_period(row)
        body.append(f"""<article class="student-admin-row{' is-disabled' if visibility != 'public' else ''}">
          {admin_batch_select("students", key)}
          <div class="student-admin-avatar-cell">{image_tag(row.get("avatar_key"), display_name, "student-admin-avatar", env.get("PUBLIC_MEDIA_BASE_URL", ""))}</div>
          <div class="student-admin-person">
            <strong>{esc(display_name)}</strong>
            <span>{esc(name_en or "英文名未设置")}</span>
            <small>{esc(key)}{f" / {esc(student_number)}" if student_number else ""}</small>
          </div>
          <div class="student-admin-study">
            <span>{esc(" / ".join(item for item in [degree, category, grade] if item) or "层次/分组/年级未填写")}</span>
            <span class="publication-admin-tag">{esc(status)}</span>
            <small>{esc(period)}</small>
          </div>
          <div class="student-admin-direction">
            <strong title="{esc(direction)}">{esc(direction or "研究方向未填写")}</strong>
            <small title="{esc(destination)}">去向：{esc(destination or "未填写")}</small>
          </div>
          <div class="student-admin-contact" title="{esc(contact)}"><span>{esc(contact)}</span><small>{esc(row.get("contact_visibility") or "public")}</small></div>
          <form class="student-inline-form" method="post" action="/admin/table/students/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="student-quick-field"><span>首页</span><select name="is_featured">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if featured else "0")}</select></label>
            <label class="student-quick-field"><span>可见</span><select name="visibility">{options(list(visibility_choices), visibility)}</select></label>
            <label class="student-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order") or row.get("id"))}" title="排序值，留空保存时使用记录 ID"></label>
            <button class="button light" type="submit" name="_student_action" value="save_inline">应用</button>
          </form>
          <div class="student-admin-actions">
            <a class="button ghost" href="/admin/table/students/{esc(key)}">编辑</a>
            <a class="button light" href="{esc(front_href)}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card student-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/students/new">新增</a></div>
      {admin_list_tools(student_admin_filter_form(query, all_rows), admin_batch_toolbar("students", meta, query, all_rows))}
      <div class="student-admin-list">
        <div class="student-admin-head"><span>选</span><span>照片</span><span>学生</span><span>学籍</span><span>方向/去向</span><span>联系</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无学生。</p>'}
      </div>
    </section>"""


def student_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    degrees = student_admin_values(rows, "degree")
    categories = student_admin_values(rows, "category")
    grades = student_admin_values(rows, "grade")
    statuses = student_admin_values(rows, "status")
    visibility_values = student_admin_values(rows, "visibility")
    return f"""<form class="filters student-admin-search" method="get" action="/admin/table/students">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索姓名、英文名、学号、层次、年级、方向、状态、邮箱、去向">
      <select name="degree"><option value="">全部层次</option>{options(degrees, query.get("degree", ""))}</select>
      <select name="category"><option value="">全部分组</option>{options(categories, query.get("category", ""))}</select>
      <select name="grade"><option value="">全部年级</option>{options(grades, query.get("grade", ""))}</select>
      <select name="status"><option value="">全部状态</option>{options(statuses, query.get("status", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "首页不限"), ("1", "首页展示"), ("0", "普通")], query.get("is_featured", ""))}</select>
      <select name="sort">{navigation_pair_options(student_admin_sort_pairs(), query.get("sort", "sort_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/students">重置</a>
    </form>"""


def student_admin_values(rows: list[dict[str, Any]], field: str) -> list[str]:
    return sorted({text_only(row.get(field), 160).strip() for row in rows if text_only(row.get(field), 160).strip()})


def student_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_desc", "排序从大到小"),
        ("sort_asc", "排序从小到大"),
        ("grade_desc", "年级从高到低"),
        ("grade_asc", "年级从低到高"),
        ("degree_asc", "层次 A-Z"),
        ("degree_desc", "层次 Z-A"),
        ("category_asc", "分组 A-Z"),
        ("category_desc", "分组 Z-A"),
        ("status_asc", "状态 A-Z"),
        ("status_desc", "状态 Z-A"),
        ("enroll_desc", "入学时间从新到旧"),
        ("enroll_asc", "入学时间从旧到新"),
        ("graduate_desc", "毕业时间从新到旧"),
        ("graduate_asc", "毕业时间从旧到新"),
        ("name_asc", "姓名 A-Z"),
        ("name_desc", "姓名 Z-A"),
        ("featured_desc", "首页展示优先"),
    ]


def student_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_desc": ("sort_order", True),
        "sort_asc": ("sort_order", False),
        "grade_desc": ("grade", True),
        "grade_asc": ("grade", False),
        "degree_asc": ("degree", False),
        "degree_desc": ("degree", True),
        "category_asc": ("category", False),
        "category_desc": ("category", True),
        "status_asc": ("status", False),
        "status_desc": ("status", True),
        "enroll_desc": ("enrollment_date", True),
        "enroll_asc": ("enrollment_date", False),
        "graduate_desc": ("graduation_date", True),
        "graduate_asc": ("graduation_date", False),
        "name_asc": ("name", False),
        "name_desc": ("name", True),
        "featured_desc": ("is_featured", True),
    }.get(value, ("sort_order", True))


def student_period(row: dict[str, Any]) -> str:
    start = text_only(row.get("enrollment_date"), 40).strip()
    end = text_only(row.get("graduation_date"), 40).strip()
    if start or end:
        return f"{start or '未填入学'} 至 {end or '未填毕业'}"
    return "入学/毕业时间未填写"


def student_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("students", key)
    if not row:
        return "/admin/table/students"
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["is_featured"] = 1 if str(data.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("id"), int_value(row.get("sort_order"), 0)))
    repo.save("students", row)
    return "/admin/table/students"


def admin_student_categories_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], student_rows: list[dict[str, Any]]) -> str:
    body = []
    for row in rows:
        key = row.get("uid") or row.get("id")
        enabled = str(row.get("enabled") or "0") in {"1", "true", "True", "yes", "on"}
        usage = student_category_usage(row, student_rows)
        keyword_text = text_only(row.get("keywords"), 500).strip()
        body.append(f"""<article class="student-category-admin-row{' is-disabled' if not enabled else ''}">
          {admin_batch_select("student_category_displays", key)}
          <div class="student-category-admin-main">
            <strong>{esc(row.get("label") or "未命名分组")}</strong>
            <span>{esc(row.get("label_en") or "英文标签未设置")}</span>
            <small>{esc(key)}</small>
          </div>
          <div class="student-category-admin-key">
            <span>{esc(row.get("key") or "key 未设置")}</span>
            <small title="{esc(keyword_text)}">{esc(keyword_text or "关键词未填写")}</small>
          </div>
          <div class="student-category-admin-usage" title="{esc(usage['title'])}">
            <span>{esc(usage['count_label'])}</span>
            <small>{esc(usage['sample_label'])}</small>
          </div>
          <div class="student-category-admin-state">
            <span class="admin-status-badge status-{'active' if enabled else 'trash'}">{'启用' if enabled else '停用'}</span>
            <small>排序 {esc(row.get("display_order"))}</small>
          </div>
          <form class="student-category-inline-form" method="post" action="/admin/table/student_category_displays/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="student-category-quick-field"><span>启用</span><select name="enabled">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if enabled else "0")}</select></label>
            <label class="student-category-quick-field"><span>排序</span><input type="number" name="display_order" value="{esc(row.get("display_order") or row.get("id"))}" title="前台分组显示顺序"></label>
            <button class="button light" type="submit" name="_student_category_action" value="save_inline">应用</button>
          </form>
          <div class="student-category-admin-actions">
            <a class="button ghost" href="/admin/table/student_category_displays/{esc(key)}">编辑</a>
            <a class="button light" href="/students?category={esc(quote(row.get("label") or row.get("key") or ""))}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card student-category-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/student_category_displays/new">新增</a></div>
      {admin_list_tools(student_category_filter_form(query, all_rows), admin_batch_toolbar("student_category_displays", meta, query, all_rows))}
      <div class="student-category-admin-list">
        <div class="student-category-admin-head"><span>选</span><span>分组标签</span><span>Key/关键词</span><span>使用情况</span><span>状态</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无学生分组。</p>'}
      </div>
    </section>"""


def student_category_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    return f"""<form class="filters student-category-admin-search" method="get" action="/admin/table/student_category_displays">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索分组 key、中文标签、英文标签、匹配关键词">
      <select name="enabled">{navigation_pair_options([("", "全部状态"), ("1", "启用"), ("0", "停用")], query.get("enabled", ""))}</select>
      <select name="sort">{navigation_pair_options(student_category_sort_pairs(), query.get("sort", "order_asc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/student_category_displays">重置</a>
    </form>"""


def student_category_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("order_asc", "排序从小到大"),
        ("order_desc", "排序从大到小"),
        ("label_asc", "中文标签 A-Z"),
        ("label_desc", "中文标签 Z-A"),
        ("key_asc", "Key A-Z"),
        ("key_desc", "Key Z-A"),
        ("enabled_desc", "启用优先"),
    ]


def student_category_sort_args(value: str) -> tuple[str, bool]:
    return {
        "order_asc": ("display_order", False),
        "order_desc": ("display_order", True),
        "label_asc": ("label", False),
        "label_desc": ("label", True),
        "key_asc": ("key", False),
        "key_desc": ("key", True),
        "enabled_desc": ("enabled", True),
    }.get(value, ("display_order", False))


def student_category_usage(category: dict[str, Any], students: list[dict[str, Any]]) -> dict[str, str]:
    tokens = [text_only(category.get("key"), 80), text_only(category.get("label"), 120), text_only(category.get("label_en"), 120)]
    tokens.extend(split_publication_tag_text(text_only(category.get("keywords"), 500)))
    normalized = {token.casefold() for token in tokens if token.strip()}
    matches = []
    for student in students:
        value = text_only(student.get("category"), 120).strip().casefold()
        if value and any(value == token or token in value for token in normalized):
            matches.append(student)
    names = [text_only(student.get("name"), 80).strip() or text_only(student.get("name_en"), 120).strip() for student in matches]
    names = [name for name in names if name]
    sample = "、".join(names[:4])
    if len(names) > 4:
        sample += f" 等 {len(names)} 人"
    return {
        "count_label": f"{len(matches)} 名学生",
        "sample_label": sample or "暂无匹配学生",
        "title": "、".join(names) or "暂无匹配学生",
    }


def student_category_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("student_category_displays", key)
    if not row:
        return "/admin/table/student_category_displays"
    row["enabled"] = 1 if str(data.get("enabled") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["display_order"] = int_value(data.get("display_order"), int_value(row.get("id"), int_value(row.get("display_order"), 0)))
    repo.save("student_category_displays", row)
    return "/admin/table/student_category_displays"


def admin_news_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]], env: dict[str, str]) -> str:
    body = []
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        comments = str(row.get("allow_comments") or "0") in {"1", "true", "True", "yes", "on"}
        title = text_only(row.get("title"), 240).strip()
        content = text_only(row.get("content"), 600).strip()
        front_href = f'/news/{safe_slug(str(row.get("slug") or title or key))}'
        related = news_related_summary(row)
        body.append(f"""<article class="news-admin-row{' is-disabled' if visibility != 'public' else ''}">
          {admin_batch_select("news", key)}
          <div class="news-admin-cover">{image_tag(row.get("cover_key"), title or "动态", "news-admin-thumb", env.get("PUBLIC_MEDIA_BASE_URL", "")) if row.get("cover_key") else '<span class="news-admin-thumb placeholder">N</span>'}</div>
          <div class="news-admin-main">
            <strong title="{esc(title)}">{esc(title or "未命名动态")}</strong>
            <small>{esc(key)} / {esc(row.get("slug") or "slug 未设置")}</small>
            <span title="{esc(content)}">{esc(content or "正文未填写")}</span>
          </div>
          <div class="news-admin-meta">
            <span class="publication-admin-tag">{esc(row.get("category") or "分类未填")}</span>
            <span>{esc(row.get("published_at") or "发布时间未填")}</span>
            <small>{esc(row.get("content_format") or "plain")}</small>
          </div>
          <div class="news-admin-related" title="{esc(related)}">{esc(related)}</div>
          <div class="news-admin-state">
            <span class="admin-status-badge status-{'active' if visibility == 'public' else 'trash'}">{esc(visibility)}</span>
            <span class="admin-status-badge status-{'active' if featured else 'trash'}">{'首页' if featured else '普通'}</span>
            <small>{'允许评论' if comments else '关闭评论'}</small>
          </div>
          <form class="news-inline-form" method="post" action="/admin/table/news/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="news-quick-field"><span>首页</span><select name="is_featured">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if featured else "0")}</select></label>
            <label class="news-quick-field"><span>可见</span><select name="visibility">{options(list(visibility_choices), visibility)}</select></label>
            <label class="news-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order") or row.get("id"))}" title="排序值，留空保存时使用记录 ID"></label>
            <button class="button light" type="submit" name="_news_action" value="save_inline">应用</button>
          </form>
          <div class="news-admin-actions">
            <a class="button ghost" href="/admin/table/news/{esc(key)}">编辑</a>
            <a class="button light" href="{esc(front_href)}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card news-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/news/new">新增</a></div>
      {admin_list_tools(news_admin_filter_form(query, all_rows), admin_batch_toolbar("news", meta, query, all_rows))}
      <div class="news-admin-list">
        <div class="news-admin-head"><span>选</span><span>封面</span><span>标题/摘要</span><span>分类/时间</span><span>关联</span><span>状态</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无动态。</p>'}
      </div>
    </section>"""


def news_related_summary(row: dict[str, Any]) -> str:
    pairs = [
        ("论文", row.get("related_publication_uid")),
        ("项目", row.get("related_project_uid")),
        ("学生", row.get("related_student_uid")),
    ]
    values = [f"{label}:{text_only(value, 120)}" for label, value in pairs if text_only(value, 120).strip()]
    return " / ".join(values) if values else "未关联"


def news_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    categories = student_admin_values(rows, "category")
    formats = student_admin_values(rows, "content_format")
    visibility_values = student_admin_values(rows, "visibility")
    return f"""<form class="filters news-admin-search" method="get" action="/admin/table/news">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索标题、分类、正文、slug、关联论文/项目/学生 UID">
      <select name="category"><option value="">全部分类</option>{options(categories, query.get("category", ""))}</select>
      <select name="content_format"><option value="">全部格式</option>{options(formats, query.get("content_format", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "首页不限"), ("1", "首页展示"), ("0", "普通")], query.get("is_featured", ""))}</select>
      <select name="allow_comments">{navigation_pair_options([("", "评论不限"), ("1", "允许评论"), ("0", "关闭评论")], query.get("allow_comments", ""))}</select>
      <select name="sort">{navigation_pair_options(news_admin_sort_pairs(), query.get("sort", "published_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/news">重置</a>
    </form>"""


def news_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("published_desc", "发布时间从新到旧"),
        ("published_asc", "发布时间从旧到新"),
        ("sort_desc", "排序从大到小"),
        ("sort_asc", "排序从小到大"),
        ("title_asc", "标题 A-Z"),
        ("title_desc", "标题 Z-A"),
        ("category_asc", "分类 A-Z"),
        ("featured_desc", "首页展示优先"),
        ("updated_desc", "最近更新优先"),
    ]


def news_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "published_desc": ("published_at", True),
        "published_asc": ("published_at", False),
        "sort_desc": ("sort_order", True),
        "sort_asc": ("sort_order", False),
        "title_asc": ("title", False),
        "title_desc": ("title", True),
        "category_asc": ("category", False),
        "featured_desc": ("is_featured", True),
        "updated_desc": ("updated_at", True),
    }.get(value, ("published_at", True))


def news_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("news", key)
    if not row:
        return "/admin/table/news"
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["is_featured"] = 1 if str(data.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("id"), int_value(row.get("sort_order"), 0)))
    repo.save("news", row)
    return "/admin/table/news"


def admin_courses_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    body = []
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        visibility = text_only(row.get("visibility"), 40).strip() or "public"
        featured = str(row.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"}
        name = text_only(row.get("name"), 240).strip()
        summary = text_only(row.get("summary"), 500).strip()
        resources = course_resource_summary(row)
        front_href = f'/courses?q={quote(name)}'
        body.append(f"""<article class="course-admin-row{' is-disabled' if visibility != 'public' else ''}">
          {admin_batch_select("courses", key)}
          <div class="course-admin-main">
            <strong title="{esc(name)}">{esc(name or "未命名课程")}</strong>
            <small>{esc(key)}</small>
            <span title="{esc(summary)}">{esc(summary or "课程简介未填写")}</span>
          </div>
          <div class="course-admin-meta">
            <span class="publication-admin-tag">{esc(row.get("semester") or "学期未填")}</span>
            <span>{esc(row.get("audience") or "授课对象未填")}</span>
          </div>
          <div class="course-admin-resources" title="{esc(resources)}">{esc(resources)}</div>
          <div class="course-admin-state">
            <span class="admin-status-badge status-{'active' if visibility == 'public' else 'trash'}">{esc(visibility)}</span>
            <span class="admin-status-badge status-{'active' if featured else 'trash'}">{'首页' if featured else '普通'}</span>
            <small>资料：{esc(row.get("material_visibility") or "public")}</small>
          </div>
          <form class="course-inline-form" method="post" action="/admin/table/courses/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="course-quick-field"><span>首页</span><select name="is_featured">{navigation_pair_options([("1", "是"), ("0", "否")], "1" if featured else "0")}</select></label>
            <label class="course-quick-field"><span>可见</span><select name="visibility">{options(list(visibility_choices), visibility)}</select></label>
            <label class="course-quick-field"><span>排序</span><input type="number" name="sort_order" value="{esc(row.get("sort_order") or row.get("id"))}" title="排序值，留空保存时使用记录 ID"></label>
            <button class="button light" type="submit" name="_course_action" value="save_inline">应用</button>
          </form>
          <div class="course-admin-actions">
            <a class="button ghost" href="/admin/table/courses/{esc(key)}">编辑</a>
            <a class="button light" href="{esc(front_href)}" target="_blank" rel="noreferrer">前台</a>
          </div>
        </article>""")
    return f"""<section class="admin-card course-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/courses/new">新增</a></div>
      {admin_list_tools(course_admin_filter_form(query, all_rows), admin_batch_toolbar("courses", meta, query, all_rows))}
      <div class="course-admin-list">
        <div class="course-admin-head"><span>选</span><span>课程/简介</span><span>学期/对象</span><span>资料</span><span>状态</span><span>快速修改</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无课程。</p>'}
      </div>
    </section>"""


def course_resource_summary(row: dict[str, Any]) -> str:
    parts = []
    if text_only(row.get("syllabus_key"), 240).strip():
        parts.append(f"大纲:{text_only(row.get('syllabus_key'), 120)}")
    if text_only(row.get("material_key"), 240).strip():
        parts.append(f"课件:{text_only(row.get('material_key'), 120)}")
    references = text_only(row.get("references_text"), 180).strip()
    if references:
        parts.append(f"参考:{references}")
    return " / ".join(parts) if parts else "资料未配置"


def course_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    semesters = student_admin_values(rows, "semester")
    audiences = student_admin_values(rows, "audience")
    material_visibility = student_admin_values(rows, "material_visibility")
    visibility_values = student_admin_values(rows, "visibility")
    return f"""<form class="filters course-admin-search" method="get" action="/admin/table/courses">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索课程名称、学期、授课对象、简介、资料 key、参考资料">
      <select name="semester"><option value="">全部学期</option>{options(semesters, query.get("semester", ""))}</select>
      <select name="audience"><option value="">全部对象</option>{options(audiences, query.get("audience", ""))}</select>
      <select name="material_visibility"><option value="">全部资料可见性</option>{options(material_visibility, query.get("material_visibility", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="is_featured">{navigation_pair_options([("", "首页不限"), ("1", "首页展示"), ("0", "普通")], query.get("is_featured", ""))}</select>
      <select name="sort">{navigation_pair_options(course_admin_sort_pairs(), query.get("sort", "sort_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/courses">重置</a>
    </form>"""


def course_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("sort_desc", "排序从大到小"),
        ("sort_asc", "排序从小到大"),
        ("semester_desc", "学期从新到旧"),
        ("semester_asc", "学期从旧到新"),
        ("name_asc", "课程名称 A-Z"),
        ("name_desc", "课程名称 Z-A"),
        ("audience_asc", "授课对象 A-Z"),
        ("featured_desc", "首页展示优先"),
        ("updated_desc", "最近更新优先"),
    ]


def course_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "sort_desc": ("sort_order", True),
        "sort_asc": ("sort_order", False),
        "semester_desc": ("semester", True),
        "semester_asc": ("semester", False),
        "name_asc": ("name", False),
        "name_desc": ("name", True),
        "audience_asc": ("audience", False),
        "featured_desc": ("is_featured", True),
        "updated_desc": ("updated_at", True),
    }.get(value, ("sort_order", True))


def course_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("courses", key)
    if not row:
        return "/admin/table/courses"
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "public"
    row["is_featured"] = 1 if str(data.get("is_featured") or "0") in {"1", "true", "True", "yes", "on"} else 0
    row["sort_order"] = int_value(data.get("sort_order"), int_value(row.get("id"), int_value(row.get("sort_order"), 0)))
    repo.save("courses", row)
    return "/admin/table/courses"


def admin_messages_table(meta: Table, rows: list[dict[str, Any]], query: dict[str, str], all_rows: list[dict[str, Any]]) -> str:
    body = []
    status_choices = message_status_pairs()
    visibility_choices = table_field_choices(meta, "visibility")
    for row in rows:
        key = row.get("uid") or row.get("id")
        name = text_only(row.get("name"), 120).strip()
        email = text_only(row.get("email"), 180).strip()
        subject = text_only(row.get("subject"), 260).strip()
        content = text_only(row.get("content"), 800).strip()
        message_type = text_only(row.get("message_type"), 60).strip() or "other"
        status = text_only(row.get("status"), 40).strip() or "new"
        visibility = text_only(row.get("visibility"), 40).strip() or "staff"
        attachment = text_only(row.get("attachment_key"), 220).strip()
        mail_action = f'<a class="button light" href="mailto:{esc(email)}?subject={quote(subject)}">邮件</a>' if email else ""
        attachment_action = f'<a class="button light" href="{esc(media_url(attachment))}" target="_blank" rel="noreferrer">附件</a>' if attachment else ""
        attachment_text = attachment or "无附件"
        body.append(f"""<article class="message-admin-row{' is-disabled' if status == 'archived' else ''}">
          {admin_batch_select("messages", key)}
          <div class="message-admin-contact">
            <strong title="{esc(name)}">{esc(name or "未留姓名")}</strong>
            <span title="{esc(email)}">{esc(email or "邮箱未填")}</span>
            <small>{esc(key)}</small>
          </div>
          <div class="message-admin-main">
            <strong title="{esc(subject)}">{esc(subject or "无主题留言")}</strong>
            <span title="{esc(content)}">{esc(content or "留言正文未填写")}</span>
          </div>
          <div class="message-admin-type">
            <span class="publication-admin-tag">{esc(message_type_label(message_type))}</span>
            <span class="admin-status-badge status-{esc('active' if status in {'new', 'read', 'replied'} else 'trash')}">{esc(message_status_label(status))}</span>
          </div>
          <div class="message-admin-attachment" title="{esc(attachment_text)}">{esc(attachment_text)}</div>
          <div class="message-admin-visibility"><span>{esc(visibility)}</span></div>
          <form class="message-inline-form" method="post" action="/admin/table/messages/quick-update">
            <input type="hidden" name="uid" value="{esc(key)}">
            <label class="message-quick-field"><span>状态</span><select name="status">{navigation_pair_options(status_choices, status)}</select></label>
            <label class="message-quick-field"><span>可见</span><select name="visibility">{options(list(visibility_choices), visibility)}</select></label>
            <button class="button light" type="submit" name="_message_action" value="save_inline">应用</button>
          </form>
          <div class="message-admin-actions">
            <a class="button ghost" href="/admin/table/messages/{esc(key)}">编辑</a>
            {mail_action}
            {attachment_action}
          </div>
        </article>""")
    return f"""<section class="admin-card message-admin-card compact-admin-card">
      <div class="admin-card-head"><h1>{esc(meta.label)}</h1><a class="button" href="/admin/table/messages/new">新增</a></div>
      {admin_list_tools(message_admin_filter_form(query, all_rows), admin_batch_toolbar("messages", meta, query, all_rows))}
      <div class="message-admin-list">
        <div class="message-admin-head"><span>选</span><span>联系人</span><span>主题/内容</span><span>类型/状态</span><span>附件</span><span>可见</span><span>快速处理</span><span>操作</span></div>
        {"".join(body) or '<p class="empty">暂无留言。</p>'}
      </div>
    </section>"""


def message_admin_filter_form(query: dict[str, str], rows: list[dict[str, Any]]) -> str:
    types = student_admin_values(rows, "message_type")
    statuses = student_admin_values(rows, "status")
    visibility_values = student_admin_values(rows, "visibility")
    return f"""<form class="filters message-admin-search" method="get" action="/admin/table/messages">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索姓名、邮箱、类型、主题、正文、附件 key">
      <select name="message_type"><option value="">全部类型</option>{navigation_pair_options([(value, message_type_label(value)) for value in types], query.get("message_type", ""))}</select>
      <select name="status"><option value="">全部状态</option>{navigation_pair_options([(value, message_status_label(value)) for value in statuses], query.get("status", ""))}</select>
      <select name="visibility"><option value="">全部可见范围</option>{options(visibility_values, query.get("visibility", ""))}</select>
      <select name="sort">{navigation_pair_options(message_admin_sort_pairs(), query.get("sort", "updated_desc"))}</select>
      <button>搜索</button><a class="button ghost" href="/admin/table/messages">重置</a>
    </form>"""


def message_admin_sort_pairs() -> list[tuple[str, str]]:
    return [
        ("updated_desc", "最近更新优先"),
        ("updated_asc", "最早更新优先"),
        ("status_asc", "状态 A-Z"),
        ("type_asc", "类型 A-Z"),
        ("subject_asc", "主题 A-Z"),
        ("name_asc", "姓名 A-Z"),
    ]


def message_admin_sort_args(value: str) -> tuple[str, bool]:
    return {
        "updated_desc": ("updated_at", True),
        "updated_asc": ("updated_at", False),
        "status_asc": ("status", False),
        "type_asc": ("message_type", False),
        "subject_asc": ("subject", False),
        "name_asc": ("name", False),
    }.get(value, ("updated_at", True))


def message_type_label(value: str) -> str:
    labels = {
        "recruiting": "招生",
        "cooperation": "合作",
        "paper": "论文",
        "project": "项目",
        "course": "课程",
        "other": "其他",
    }
    return labels.get(value, value or "其他")


def message_status_pairs() -> list[tuple[str, str]]:
    return [
        ("new", "新留言"),
        ("read", "已读"),
        ("replied", "已回复"),
        ("archived", "已归档"),
    ]


def message_status_label(value: str) -> str:
    return dict(message_status_pairs()).get(value, value or "新留言")


def message_quick_update(repo: Repository, body: bytes) -> str:
    data = _form(body)
    key = data.get("uid", "")
    row = repo.get("messages", key)
    if not row:
        return "/admin/table/messages"
    row["status"] = text_only(data.get("status"), 40).strip() or row.get("status") or "new"
    row["visibility"] = text_only(data.get("visibility"), 40).strip() or row.get("visibility") or "staff"
    repo.save("messages", row)
    return "/admin/table/messages"


def admin_media_table(repo: Repository, query: dict[str, str], env: dict[str, str], mode: str = "library") -> str:
    media_auto_empty_trash(repo)
    is_trash = mode == "trash"
    all_rows = repo.list("media_assets", Query(filters={"status": "trash" if is_trash else "active"}, limit=1000, order_by="updated_at", descending=True))
    order_by, descending = media_sort_args(query.get("sort", "updated_desc"))
    filters = {"status": "trash" if is_trash else "active"}
    if query.get("category"):
        filters["category"] = query["category"]
    if query.get("mime_type"):
        filters["mime_type"] = query["mime_type"]
    if query.get("storage_kind"):
        filters["storage_kind"] = query["storage_kind"]
    page = max(1, int_value(query.get("page"), 1))
    per_page = max(20, min(int_value(query.get("per_page"), 80), 200))
    queried_rows = repo.list("media_assets", Query(q=query.get("q", ""), filters=filters, limit=1000, order_by=order_by, descending=descending))
    usage_map = media_usage_map(repo)
    queried_rows = media_filter_by_file_state(queried_rows, query.get("file_state", ""), usage_map, env)
    total_rows = len(queried_rows)
    start = (page - 1) * per_page
    rows = queried_rows[start:start + per_page]
    items = []
    total_size = sum(int_value(row.get("size")) for row in queried_rows)
    active_rows = repo.list("media_assets", Query(filters={"status": "active"}, limit=1000))
    active_count = sum(1 for row in active_rows if str(row.get("status") or "active") == "active" and media_file_exists(row, env))
    missing_count = sum(1 for row in active_rows if str(row.get("status") or "active") == "active" and not media_file_exists(row, env))
    trash_count = sum(1 for row in repo.list("media_assets", Query(filters={"status": "trash"}, limit=1000)) if str(row.get("status") or "active") == "trash")
    for row in rows:
        key = str(row.get("object_key") or "")
        usage = usage_map.get(normalize_media_key(key), [])
        usage_html = "".join(admin_media_usage_link(item) for item in usage) or '<span class="admin-muted">未发现引用</span>'
        status = str(row.get("status") or "active")
        exists = media_file_exists(row, env)
        row_state = media_row_state(row, env)
        row_classes = " ".join(filter(None, ["media-row", "is-trash" if status == "trash" else "", "is-missing" if not exists else ""]))
        items.append(f"""<article class="{esc(row_classes)}">
          <label class="media-select" title="选择此媒体"><input type="checkbox" name="selected" value="{esc(row.get("uid") or row.get("id"))}" form="media-batch-form"><span class="sr-only">选择</span></label>
          <a class="media-thumb{' is-missing' if not exists else ''}" href="{esc(media_url(key, env.get("PUBLIC_MEDIA_BASE_URL", "")))}" target="_blank" rel="noreferrer">{media_preview(row, env)}</a>
          <div class="media-info">
            <div class="media-title-line"><h2>{esc(row.get("title") or key or "未命名媒体")}</h2>{admin_media_status_badge(row_state)}</div>
            <p class="media-key">{esc(key)}</p>
          </div>
          <dl class="media-facts media-meta">
            {admin_fact("分类", row.get("category"))}
            {admin_fact("存储", media_storage_label(row))}
            {admin_fact("类型", row.get("mime_type"))}
            {admin_media_size_fact(key, row.get("size"))}
            {admin_fact("校验", text_only(row.get("checksum"), 18) if row.get("checksum") else "")}
          </dl>
          <div class="media-usage"><div>{usage_html}</div></div>
          <div class="media-actions">
            <a class="button ghost" href="/admin/table/media_assets/{esc(row.get("uid") or row.get("id"))}">编辑</a>
            {media_action_button(row, "restore", "恢复", trash_context=is_trash) if status == "trash" else media_action_button(row, "trash", "回收站")}
            {media_action_button(row, "delete", "删除", danger=True, trash_context=is_trash)}
          </div>
        </article>""")
    title = "媒体回收站" if is_trash else "媒体库"
    subtitle = f"共 {total_rows} 个{'回收站' if is_trash else '媒体'}记录，当前筛选容量 {format_bytes(total_size)}。"
    top_action = (
        '<a class="button ghost" href="/admin/table/media_assets">返回媒体库</a>'
        if is_trash else
        f'<a class="button ghost" href="/admin/table/media_assets/trash">回收站 {trash_count}</a>'
    )
    scan_action = "" if is_trash else '<form method="post" action="/admin/table/media_assets/scan"><button class="button light" type="submit">扫描项目媒体</button></form>'
    export_used_action = f'<a class="button light" href="{esc(media_export_used_url(query, mode))}" title="快捷导出当前筛选范围内被使用的媒体文件">导出媒体</a>'
    delete_scope_label = "媒体文件和媒体库记录"
    clear_trash = (
        f'<form method="post" action="/admin/table/media_assets/trash/clear" '
        f'data-confirm="确定清空回收站中的{delete_scope_label}吗？">'
        f'<button class="button danger" type="submit">一键清空回收站</button></form>'
    ) if is_trash and total_rows else ""
    return f"""<section class="admin-card media-admin-card">
      <div class="admin-card-head">
        <div><h1>{title}</h1><p class="admin-muted">{subtitle} 可用 {active_count} / 缺失 {missing_count} / 回收站 {trash_count}。</p></div>
        <div class="admin-card-head-actions">{top_action}{scan_action}{export_used_action}{clear_trash}<a class="button" href="/admin/table/media_assets/new">新增</a></div>
      </div>
      {media_scan_result_notice(query)}
      {media_capacity_panel(queried_rows)}
      <div class="media-sticky-tools">
      {admin_batch_result_notice(query)}
      {media_filter_form(query, all_rows, mode)}
      <form id="media-batch-form" class="media-batch-toolbar" method="post" action="/admin/table/media_assets/batch" data-delete-scope="{esc(delete_scope_label)}">
        <input type="hidden" name="return_to" value="{esc('/admin/table/media_assets/trash' if is_trash else '/admin/table/media_assets')}">
        <input type="hidden" name="scope" value="{esc(mode)}">
        <label><input type="checkbox" id="admin-media-select-all"> 全选</label>
        <select name="batch_action">
          <option value="update">批量修改</option>
          {'<option value="restore">恢复可用</option>' if is_trash else '<option value="trash">移到回收站</option>'}
          <option value="delete">彻底删除{'记录' if env.get("PLATFORM") == "cloudflare" else '文件'}</option>
        </select>
        <select name="batch_status">
          <option value="">状态不变</option>
          <option value="active">可用</option>
          <option value="trash">回收站</option>
        </select>
        <input name="batch_category" placeholder="例：profile / icon / news-cover，留空不改">
        <button type="submit">应用到选中</button>
        <button class="button light" type="submit" formaction="/admin/table/media_assets/export-used" formmethod="post" title="导出所有选中的本地媒体文件">导出选中媒体</button>
        <span id="admin-media-selected-count" class="admin-muted">已选 0 个</span>
        <small class="media-batch-help">示例：筛选“文件缺失”后全选，可批量移到回收站或彻底删除；统一分类只在选择“批量修改”时生效。</small>
      </form>
      </div>
      <div class="media-admin-list"><div class="media-list-head"><span>选择</span><span>预览</span><span>媒体文件</span><span>文件信息</span><span>使用位置</span><span>操作</span></div>{"".join(items) or '<p class="empty">暂无媒体文件。</p>'}</div>
      {media_pager(query, mode, page, per_page, total_rows)}
    </section>"""


def media_capacity_panel(rows: list[dict[str, Any]]) -> str:
    active_count = sum(1 for row in rows if str(row.get("status") or "active") != "trash")
    trash_count = sum(1 for row in rows if str(row.get("status") or "active") == "trash")
    recorded_total = sum(int_value(row.get("size")) for row in rows)
    return f"""<section class="media-capacity-panel" aria-label="媒体容量概览">
      <div><span>媒体记录</span><strong id="media-capacity-count">{len(rows)}</strong><small>可用 {active_count} / 回收站 {trash_count}</small></div>
      <div><span>媒体容量</span><strong id="media-capacity-total">{esc(format_bytes(recorded_total))}</strong><small id="media-capacity-note">正在等待检测</small></div>
      <div><span>磁盘容量</span><strong id="media-disk-total">待检测</strong><small id="media-disk-free">点击刷新或等待自动检测</small></div>
      <button type="button" id="media-refresh-stats">刷新检测</button>
    </section>"""


def media_scan_result_notice(query: dict[str, str]) -> str:
    if "scan_done" not in query and "scan_added" not in query and "scan_unsupported" not in query:
        return ""
    if query.get("scan_unsupported"):
        return '<p class="admin-result warning">当前运行环境不支持扫描项目文件系统；Cloudflare Worker 生产环境建议通过部署资源或 R2 管理媒体。</p>'
    added = int_value(query.get("scan_added"))
    updated = int_value(query.get("scan_updated"))
    skipped = int_value(query.get("scan_skipped"))
    scanned = int_value(query.get("scan_scanned"))
    return (
        f'<p class="admin-result success">已扫描 {scanned} 个媒体文件，新增登记 {added} 个，'
        f'更新元数据 {updated} 个，跳过 {skipped} 个。</p>'
    )


def media_export_used_url(query: dict[str, str], mode: str) -> str:
    params = {
        key: query.get(key, "")
        for key in ("q", "category", "mime_type", "storage_kind", "file_state", "sort")
        if query.get(key)
    }
    if mode == "trash":
        params["scope"] = "trash"
    suffix = f"?{urlencode(params)}" if params else ""
    return f"/admin/table/media_assets/export-used{suffix}"


def media_filter_form(query: dict[str, str], rows: list[dict[str, Any]], mode: str) -> str:
    action = "/admin/table/media_assets/trash" if mode == "trash" else "/admin/table/media_assets"
    categories = sorted({text_only(row.get("category"), 80).strip() for row in rows if text_only(row.get("category"), 80).strip()})
    mime_types = sorted({text_only(row.get("mime_type"), 120).strip() for row in rows if text_only(row.get("mime_type"), 120).strip()})
    storage_kinds = [("", "全部存储"), ("static", "静态包"), ("local", "本地"), ("r2", "R2"), ("external", "外链")]
    file_states = [("", "全部状态"), ("available", "可用"), ("missing", "文件缺失"), ("used", "有引用"), ("unused", "未引用")]
    return f"""<form class="filters admin-media-search" method="get" action="{action}">
      <input name="q" value="{esc(query.get("q", ""))}" placeholder="搜索标题、对象 key、分类、MIME、状态">
      <select name="category"><option value="">全部分类</option>{options(categories, query.get("category", ""))}</select>
      <select name="mime_type"><option value="">全部类型</option>{options(mime_types, query.get("mime_type", ""))}</select>
      <select name="storage_kind">{navigation_pair_options(storage_kinds, query.get("storage_kind", ""))}</select>
      <select name="file_state">{navigation_pair_options(file_states, query.get("file_state", ""))}</select>
      <select name="sort">{media_sort_options(query.get("sort", "updated_desc"))}</select>
      <select name="per_page">{options(["40", "80", "120", "200"], str(query.get("per_page") or "80"))}</select>
      <button>搜索</button><a class="button ghost" href="{action}">重置</a>
    </form>"""


def media_sort_options(selected: str) -> str:
    pairs = [
        ("updated_desc", "最近更新"),
        ("created_desc", "最近添加"),
        ("title_asc", "标题 A-Z"),
        ("key_asc", "对象 key"),
        ("size_desc", "大小从大到小"),
        ("size_asc", "大小从小到大"),
        ("category_asc", "分类"),
        ("type_asc", "MIME 类型"),
    ]
    return "".join(f'<option value="{esc(value)}"{" selected" if value == selected else ""}>{esc(label)}</option>' for value, label in pairs)


def media_sort_args(value: str) -> tuple[str, bool]:
    return {
        "updated_desc": ("updated_at", True),
        "created_desc": ("created_at", True),
        "title_asc": ("title", False),
        "key_asc": ("object_key", False),
        "size_desc": ("size", True),
        "size_asc": ("size", False),
        "category_asc": ("category", False),
        "type_asc": ("mime_type", False),
    }.get(value, ("updated_at", True))


def media_filter_by_file_state(rows: list[dict[str, Any]], file_state: str, usage_map: dict[str, list[dict[str, str]]], env: dict[str, str]) -> list[dict[str, Any]]:
    state = text_only(file_state, 40).strip()
    if not state:
        return rows
    filtered = []
    for row in rows:
        key = normalize_media_key(str(row.get("object_key") or ""))
        exists = media_file_exists(row, env)
        used = bool(usage_map.get(key))
        if state == "available" and media_row_state(row, env) != "available":
            continue
        if state == "missing" and exists:
            continue
        if state == "used" and not used:
            continue
        if state == "unused" and used:
            continue
        filtered.append(row)
    return filtered


def media_pager(query: dict[str, str], mode: str, page: int, per_page: int, total_rows: int) -> str:
    total_pages = max(1, (total_rows + per_page - 1) // per_page)
    if total_pages <= 1:
        return ""
    base = "/admin/table/media_assets/trash" if mode == "trash" else "/admin/table/media_assets"
    params = {key: value for key, value in query.items() if key != "page" and value}
    prev_link = media_page_link(base, params, page - 1, "上一页") if page > 1 else '<span class="button light is-disabled">上一页</span>'
    next_link = media_page_link(base, params, page + 1, "下一页") if page < total_pages else '<span class="button light is-disabled">下一页</span>'
    return f'<nav class="media-pager">{prev_link}<span class="admin-muted">第 {page} / {total_pages} 页，共 {total_rows} 条</span>{next_link}</nav>'


def media_page_link(base: str, params: dict[str, str], page: int, label: str) -> str:
    from urllib.parse import urlencode

    query = urlencode({**params, "page": str(page)})
    return f'<a class="button light" href="{esc(base + ("?" + query if query else ""))}">{esc(label)}</a>'


def admin_media_size_fact(key: str, value: Any) -> str:
    size = int_value(value)
    label = format_bytes(size) if size else "待检测"
    return f'<dt>大小</dt><dd class="media-size-value" data-media-key="{esc(normalize_media_key(key))}" data-media-size="{size}">{esc(label)}</dd>'


def media_options_payload(repo: Repository, env: dict[str, str], query: dict[str, str]) -> dict[str, Any]:
    q = text_only(query.get("q"), 120)
    page = max(1, int_value(query.get("page"), 1))
    per_page = max(20, min(120, int_value(query.get("per_page"), 60)))
    all_rows = repo.list("media_assets", Query(q=q, filters={"status": "active"}, limit=1000, order_by="updated_at", descending=True))
    rows = all_rows[(page - 1) * per_page: page * per_page]
    items = []
    for row in rows:
        key = normalize_media_key(str(row.get("object_key") or ""))
        if not key:
            continue
        mime = text_only(row.get("mime_type"), 120)
        items.append({
            "uid": row.get("uid") or row.get("id"),
            "key": key,
            "title": text_only(row.get("title") or key, 160),
            "category": text_only(row.get("category"), 80),
            "storage_kind": media_storage_kind(row),
            "storage_label": media_storage_label(row),
            "mime_type": mime,
            "size": int_value(row.get("size")),
            "url": media_url(key, env.get("PUBLIC_MEDIA_BASE_URL", "")),
            "is_image": mime.startswith("image/") or key.lower().endswith((".svg", ".png", ".jpg", ".jpeg", ".gif", ".webp")),
        })
    return {
        "items": items,
        "folders": media_folder_options(repo),
        "page": page,
        "per_page": per_page,
        "total": len(all_rows),
        "has_more": page * per_page < len(all_rows),
    }


def media_folder_options(repo: Repository) -> list[str]:
    folders: set[str] = set()
    for row in repo.list("media_assets", Query(limit=1000)):
        key = normalize_media_key(str(row.get("object_key") or ""))
        if "/" in key:
            folders.add(safe_media_folder("/".join(key.split("/")[:-1])))
    for root in (Path("public") / "media", Path("media")):
        if not root.exists():
            continue
        try:
            root_resolved = root.resolve()
            for item in root.rglob("*"):
                if not item.is_dir():
                    continue
                relative = item.resolve().relative_to(root_resolved).as_posix()
                if relative and len(relative.split("/")) <= 3:
                    folders.add(safe_media_folder(relative))
        except OSError:
            continue
    defaults = ["icons", "profile", "students", "news", "publications", "projects", "patents", "courses", "messages", "site"]
    ordered = [item for item in defaults if item in folders or item in {"icons", "profile", "students"}]
    ordered.extend(sorted(item for item in folders if item not in ordered))
    return ordered[:120]


def upload_allowed_extensions(repo: Repository) -> set[str]:
    settings = active_global(repo)
    raw = text_only(settings.get("upload_allowed_extensions"), 1000).strip()
    if not raw:
        raw = ".jpg,.jpeg,.png,.webp,.svg,.pdf,.doc,.docx,.xls,.xlsx,.csv"
    allowed = set()
    for item in re.split(r"[,;\s]+", raw):
        ext = item.strip().lower()
        if not ext:
            continue
        if not ext.startswith("."):
            ext = "." + ext
        if re.fullmatch(r"\.[a-z0-9]{1,12}", ext):
            allowed.add(ext)
    return allowed or {".jpg", ".jpeg", ".png", ".webp", ".pdf", ".csv"}


def upload_max_bytes(repo: Repository) -> int:
    settings = active_global(repo)
    configured_mb = int_value(settings.get("upload_max_size_mb"), 10)
    # Keep the application-level cap aligned with the lightweight stdlib server guard.
    safe_mb = max(1, min(configured_mb, 10))
    return safe_mb * 1024 * 1024


def inspect_upload_file(filename: str, content: bytes, allowed_extensions: set[str]) -> dict[str, Any]:
    ext = Path(filename).suffix.lower()
    if ext not in allowed_extensions:
        return {"ok": False, "message": f"不允许上传 {ext or '无扩展名'} 文件，请在通用设置中维护允许扩展名。"}
    detected = detect_upload_mime(content, ext)
    if not detected.get("ok"):
        return detected
    valid_exts = detected.get("extensions") or {ext}
    if ext not in valid_exts:
        return {"ok": False, "message": f"文件内容与扩展名不匹配：当前为 {ext}，检测到 {', '.join(sorted(valid_exts))}。"}
    return {"ok": True, "mime": detected.get("mime") or mimetypes.guess_type(filename)[0] or "application/octet-stream"}


def detect_upload_mime(content: bytes, ext: str) -> dict[str, Any]:
    head = content[:4096]
    if head.startswith(b"\x89PNG\r\n\x1a\n"):
        return {"ok": True, "mime": "image/png", "extensions": {".png"}}
    if head.startswith(b"\xff\xd8\xff"):
        return {"ok": True, "mime": "image/jpeg", "extensions": {".jpg", ".jpeg"}}
    if head.startswith(b"GIF87a") or head.startswith(b"GIF89a"):
        return {"ok": True, "mime": "image/gif", "extensions": {".gif"}}
    if head.startswith(b"RIFF") and b"WEBP" in head[:16]:
        return {"ok": True, "mime": "image/webp", "extensions": {".webp"}}
    if head.startswith(b"%PDF-"):
        return {"ok": True, "mime": "application/pdf", "extensions": {".pdf"}}
    if head.startswith(b"PK\x03\x04") and ext in {".docx", ".xlsx", ".pptx", ".zip"}:
        return {"ok": True, "mime": mimetypes.guess_type(f"file{ext}")[0] or "application/zip", "extensions": {ext}}
    if head.startswith(b"\xd0\xcf\x11\xe0") and ext in {".doc", ".xls", ".ppt"}:
        return {"ok": True, "mime": mimetypes.guess_type(f"file{ext}")[0] or "application/octet-stream", "extensions": {ext}}
    if ext == ".svg":
        return validate_svg_upload(content)
    if ext in {".csv", ".txt", ".md", ".json", ".yaml", ".yml"}:
        if b"\x00" in head:
            return {"ok": False, "message": "文本文件中检测到二进制内容，已拒绝上传。"}
        return {"ok": True, "mime": mimetypes.guess_type(f"file{ext}")[0] or "text/plain", "extensions": {ext}}
    return {"ok": False, "message": "无法识别文件内容，请确认文件未损坏且扩展名正确。"}


def validate_svg_upload(content: bytes) -> dict[str, Any]:
    if len(content) > 1024 * 1024:
        return {"ok": False, "message": "SVG 文件超过 1MB，请压缩或转为 PNG/WebP。"}
    text = content[:1024 * 1024].decode("utf-8", "ignore").lower()
    if "<svg" not in text:
        return {"ok": False, "message": "SVG 文件缺少 <svg> 根元素。"}
    dangerous = ("<script", "javascript:", "data:text/html", "onload=", "onclick=", "onerror=", "<foreignobject", "<iframe", "<object", "<embed")
    if any(token in text for token in dangerous):
        return {"ok": False, "message": "SVG 中包含脚本、事件或嵌入对象，已拒绝上传。"}
    return {"ok": True, "mime": "image/svg+xml", "extensions": {".svg"}}


def prepare_media_upload(repo: Repository, body: bytes, env: dict[str, str], storage_kind: str = "local", key_prefix: str = "") -> dict[str, Any]:
    max_bytes = upload_max_bytes(repo)
    if len(body) > max_bytes + 1024 * 1024:
        return {"ok": False, "message": f"文件超过 {max_bytes // 1024 // 1024}MB，请压缩后再上传。"}
    form = parse_multipart(body, env.get("_CONTENT_TYPE", ""))
    upload = form.get("file")
    if not isinstance(upload, dict) or not upload.get("content"):
        return {"ok": False, "message": "未收到上传文件。"}
    filename = safe_media_filename(str(form.get("file_name") or upload.get("filename") or "media.bin"))
    folder = safe_media_folder(str(form.get("folder") or "icons"))
    content = upload.get("content") or b""
    if not isinstance(content, bytes):
        return {"ok": False, "message": "上传内容无效。"}
    if len(content) > max_bytes:
        return {"ok": False, "message": f"文件超过 {max_bytes // 1024 // 1024}MB。"}
    inspected = inspect_upload_file(filename, content, upload_allowed_extensions(repo))
    if not inspected.get("ok"):
        return {"ok": False, "message": str(inspected.get("message") or "文件校验失败。")}
    storage = normalize_media_storage_kind(storage_kind, "local")
    if storage == "r2":
        key = unique_media_object_key(folder, filename, key_prefix or "uploads")
        path = None
    else:
        key, path = unique_media_path(folder, filename, env)
        storage = "local"
    mime = str(inspected.get("mime") or upload.get("content_type") or mimetypes.guess_type(filename)[0] or "application/octet-stream")
    row = media_record_data(key, text_only(form.get("title") or Path(filename).stem, 160), folder, mime, len(content), storage)
    return {"ok": True, "key": key, "path": path, "content": content, "mime": mime, "row": row, "url": media_url(key, env.get("PUBLIC_MEDIA_BASE_URL", ""))}


def media_upload_payload(repo: Repository, body: bytes, env: dict[str, str]) -> dict[str, Any]:
    if env.get("PLATFORM") == "cloudflare":
        return {"ok": False, "message": "Cloudflare Worker 环境请通过 R2 写入接口保存媒体；当前请求未进入 R2 适配器。"}
    prepared = prepare_media_upload(repo, body, env, "local")
    if not prepared.get("ok"):
        return prepared
    path = prepared.get("path")
    content = prepared.get("content") or b""
    if not isinstance(path, Path):
        return {"ok": False, "message": "文件保存路径无效。"}
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)
    except OSError:
        return {"ok": False, "message": "文件保存失败。"}
    row = repo.save("media_assets", prepared["row"])
    return {"ok": True, "key": prepared["key"], "url": prepared["url"], "item": row}


def prepare_media_crop(repo: Repository, body: bytes, env: dict[str, str], storage_kind: str = "local", key_prefix: str = "") -> dict[str, Any]:
    data = _form(body)
    image_data = data.get("image_data", "")
    if not image_data.startswith("data:image/") or ";base64," not in image_data:
        return {"ok": False, "message": "裁剪图片数据无效。"}
    meta, encoded = image_data.split(";base64,", 1)
    try:
        content = base64.b64decode(encoded, validate=True)
    except (ValueError, TypeError):
        return {"ok": False, "message": "裁剪图片解码失败。"}
    if len(content) > 10 * 1024 * 1024:
        return {"ok": False, "message": "裁剪结果超过 10MB。"}
    image_type = meta.removeprefix("data:image/").split("+", 1)[0].lower()
    ext = {"jpeg": "jpg", "jpg": "jpg", "png": "png", "webp": "webp"}.get(image_type, "png")
    folder = safe_media_folder(data.get("folder") or "icons")
    requested = safe_media_filename(data.get("file_name") or f"icon-{int(time.time())}.{ext}")
    requested_ext = Path(requested).suffix.lower().lstrip(".")
    if requested_ext == "jpeg":
        requested_ext = "jpg"
    if requested_ext != ext:
        requested = f"{Path(requested).stem}.{ext}"
    replace_current = truthy(data.get("replace_current"), default=False)
    replace_key = normalize_media_key(text_only(data.get("replace_key"), 800).strip())
    replace_uid = text_only(data.get("replace_uid"), 200).strip()
    if replace_current and replace_key:
        row = repo.get("media_assets", replace_uid) if replace_uid else None
        if not row:
            row = next((item for item in repo.list("media_assets", Query(limit=1000)) if normalize_media_key(str(item.get("object_key") or "")) == replace_key), None)
        if not row:
            return {"ok": False, "message": "未找到要替换的媒体记录。"}
        storage = media_storage_kind(row)
        if normalize_media_storage_kind(storage_kind, "local") == "r2":
            if storage != "r2" and not r2_preferred_key(replace_key):
                return {"ok": False, "message": "Cloudflare 只能替换 R2/uploads 媒体；静态资源请随代码重新部署。"}
            target = None
            row["storage_kind"] = "r2"
        else:
            if storage == "static":
                return {"ok": False, "message": "静态包媒体不允许在线覆盖，请上传为新媒体或随代码重新部署。"}
            target = media_target_path_for_key(replace_key, env)
            if not target:
                return {"ok": False, "message": "替换路径无效。"}
            row["storage_kind"] = "local"
        mime = "image/jpeg" if ext in {"jpg", "jpeg"} else f"image/{ext}"
        row["size"] = len(content)
        row["mime_type"] = mime
        row["checksum"] = hashlib.sha1(content).hexdigest()[:12]
        if data.get("title"):
            row["title"] = text_only(data.get("title"), 160)
        return {"ok": True, "replaced": True, "key": replace_key, "path": target, "content": content, "mime": mime, "row": row, "url": media_url(replace_key, env.get("PUBLIC_MEDIA_BASE_URL", ""))}
    storage = normalize_media_storage_kind(storage_kind, "local")
    if storage == "r2":
        key = unique_media_object_key(folder, requested, key_prefix or "uploads")
        path = None
    else:
        key, path = unique_media_path(folder, requested, env)
        storage = "local"
    mime = "image/jpeg" if ext in {"jpg", "jpeg"} else f"image/{ext}"
    row = media_record_data(key, text_only(data.get("title") or Path(requested).stem, 160), folder, mime, len(content), storage)
    return {"ok": True, "key": key, "path": path, "content": content, "mime": mime, "row": row, "url": media_url(key, env.get("PUBLIC_MEDIA_BASE_URL", ""))}


def media_crop_payload(repo: Repository, body: bytes, env: dict[str, str]) -> dict[str, Any]:
    if env.get("PLATFORM") == "cloudflare":
        return {"ok": False, "message": "Cloudflare Worker 环境请通过 R2 写入接口保存裁剪结果；当前请求未进入 R2 适配器。"}
    prepared = prepare_media_crop(repo, body, env, "local")
    if not prepared.get("ok"):
        return prepared
    path = prepared.get("path")
    content = prepared.get("content") or b""
    if not isinstance(path, Path):
        return {"ok": False, "message": "裁剪图片保存路径无效。"}
    if prepared.get("replaced"):
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(content)
        except OSError:
            return {"ok": False, "message": "替换源文件失败。"}
        saved = repo.save("media_assets", prepared["row"])
        return {"ok": True, "replaced": True, "key": prepared["key"], "url": prepared["url"], "item": saved}
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)
    except OSError:
        return {"ok": False, "message": "裁剪图片保存失败。"}
    row = repo.save("media_assets", prepared["row"])
    return {"ok": True, "key": prepared["key"], "url": prepared["url"], "item": row}


def publication_parse_payload(body: bytes) -> dict[str, Any]:
    data = _form(body)
    raw = text_only(data.get("citation"), 12000).strip()
    fmt = text_only(data.get("format"), 40).strip().lower() or "auto"
    if not raw:
        return {"ok": False, "message": "请先粘贴 BibTeX、GB/T、IEEE 或普通引文文本。"}
    fields = parse_publication_reference(raw, fmt)
    return {"ok": True, "fields": fields, "message": f"已解析 {len(fields)} 个字段" if fields else "未识别到可自动填写的字段"}


def publication_duplicates_payload(repo: Repository, query: dict[str, str]) -> dict[str, Any]:
    uid = text_only(query.get("uid"), 160).strip()
    title = normalize_lookup_text(query.get("title"))
    doi = normalize_doi(query.get("doi"))
    matches = []
    for row in repo.list("publications", Query(limit=1000)):
        key = str(row.get("uid") or row.get("id") or "")
        if uid and key == uid:
            continue
        reasons = []
        row_title = normalize_lookup_text(row.get("title"))
        row_doi = normalize_doi(row.get("doi"))
        if doi and row_doi and doi == row_doi:
            reasons.append("DOI 相同")
        if title and row_title and (title == row_title or title in row_title or row_title in title):
            reasons.append("标题相近")
        if reasons:
            matches.append({
                "uid": key,
                "title": row.get("title") or "",
                "authors": text_only(row.get("authors"), 200),
                "year": row.get("year") or "",
                "doi": row.get("doi") or "",
                "edit_url": f"/admin/table/publications/{key}",
                "reasons": reasons,
            })
    return {"ok": True, "matches": matches[:10]}


def publication_lookup_payload(repo: Repository, body: bytes) -> dict[str, Any]:
    data = _form(body)
    lookup_text = text_only(data.get("lookup_text"), 800).strip()
    lookup_doi = find_doi(lookup_text) if lookup_text else ""
    doi = normalize_doi(lookup_doi or data.get("doi"))
    title = text_only("" if lookup_doi else lookup_text, 600).strip() or text_only(data.get("title"), 600).strip()
    settings = active_global(repo)
    selected = parse_platforms(data.get("platforms") or settings.get("publication_metadata_providers") or settings.get("publication_metadata_provider"))
    if not title and not doi:
        return {"ok": False, "message": "请先填写标题或 DOI。", "results": []}
    results = []
    for platform in selected[:3]:
        try:
            result = lookup_publication_platform(platform, title, doi)
        except Exception as exc:
            result = {"platform": platform, "ok": False, "message": f"查验失败：{exc}"}
        results.append(result)
    fields: dict[str, Any] = {}
    for result in results:
        if result.get("ok") and result.get("fields"):
            for key, value in result["fields"].items():
                if value not in (None, "") and not fields.get(key):
                    fields[key] = value
    return {"ok": bool(fields), "fields": fields, "results": results, "platforms": selected}


def publication_citations_payload(body: bytes) -> dict[str, Any]:
    data = _form(body)
    row = {field.name: data.get(field.name, "") for field in TABLE_MAP["publications"].fields}
    citations = generated_publication_citations(row)
    return {"ok": True, "fields": {"citation_gbt": citations["gbt"], "citation_elsevier": citations["elsevier"], "citation_apa": citations["apa"], "citation_ieee": citations["ieee"], "bibtex": citations["bibtex"]}}


def project_duplicates_payload(repo: Repository, query: dict[str, str]) -> dict[str, Any]:
    uid = text_only(query.get("uid"), 160).strip()
    name = normalize_lookup_text(query.get("name"))
    number = normalize_identifier(query.get("project_number"))
    matches = []
    for row in repo.list("projects", Query(limit=1000)):
        key = str(row.get("uid") or row.get("id") or "")
        if uid and key == uid:
            continue
        reasons = []
        row_name = normalize_lookup_text(row.get("name"))
        row_number = normalize_identifier(row.get("project_number"))
        if number and row_number and number == row_number:
            reasons.append("项目编号相同")
        if name and row_name and (name == row_name or name in row_name or row_name in name):
            reasons.append("项目名称相近")
        if reasons:
            matches.append({
                "uid": key,
                "title": row.get("name") or "",
                "summary": text_only(row.get("fund_name") or row.get("source"), 200),
                "edit_url": f"/admin/table/projects/{key}",
                "reasons": reasons,
            })
    return {"ok": True, "matches": matches[:10]}


def patent_duplicates_payload(repo: Repository, query: dict[str, str]) -> dict[str, Any]:
    uid = text_only(query.get("uid"), 160).strip()
    name = normalize_lookup_text(query.get("name"))
    application_number = normalize_identifier(query.get("application_number"))
    grant_number = normalize_identifier(query.get("grant_number"))
    matches = []
    for row in repo.list("patents", Query(limit=1000)):
        key = str(row.get("uid") or row.get("id") or "")
        if uid and key == uid:
            continue
        reasons = []
        row_name = normalize_lookup_text(row.get("name"))
        row_application = normalize_identifier(row.get("application_number"))
        row_grant = normalize_identifier(row.get("grant_number"))
        if application_number and row_application and application_number == row_application:
            reasons.append("申请号相同")
        if grant_number and row_grant and grant_number == row_grant:
            reasons.append("授权号相同")
        if name and row_name and (name == row_name or name in row_name or row_name in name):
            reasons.append("名称相近")
        if reasons:
            matches.append({
                "uid": key,
                "title": row.get("name") or "",
                "summary": text_only(row.get("patent_type") or row.get("legal_status"), 200),
                "edit_url": f"/admin/table/patents/{key}",
                "reasons": reasons,
            })
    return {"ok": True, "matches": matches[:10]}


def patent_lookup_payload(repo: Repository, body: bytes) -> dict[str, Any]:
    data = _form(body)
    settings = active_global(repo)
    lookup_text = text_only(data.get("lookup_text"), 800).strip()
    identifier = normalize_identifier(lookup_text) or normalize_identifier(data.get("grant_number")) or normalize_identifier(data.get("application_number"))
    title = "" if identifier and identifier == normalize_identifier(lookup_text) else lookup_text
    title = title or text_only(data.get("name"), 600).strip()
    selected = parse_patent_platforms(data.get("platforms") or settings.get("patent_metadata_providers"))
    if not title and not identifier:
        return {"ok": False, "message": "请先填写专利名称、申请号或授权号。", "results": []}
    results = []
    for platform in selected[:3]:
        try:
            result = lookup_patent_platform(platform, title, identifier, settings)
        except Exception as exc:
            result = {"platform": platform, "ok": False, "message": f"查验失败：{exc}"}
        results.append(result)
    fields: dict[str, Any] = {}
    for result in results:
        if result.get("ok") and result.get("fields"):
            for key, value in result["fields"].items():
                if value not in (None, "") and not fields.get(key):
                    fields[key] = value
    return {"ok": bool(fields), "fields": fields, "results": results, "platforms": selected}


def publication_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("publication_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = PUBLICATION_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(PUBLICATION_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(PUBLICATION_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("publications", Query(limit=1000, order_by="id", descending=True))
    payload = {
        "ok": True,
        "fields": {
            "venue": publication_unique_values(rows, "venue"),
            "publication_type": publication_unique_values(rows, "publication_type"),
            "index_type": publication_unique_values(rows, "index_type"),
            "display_tags": publication_split_values(rows, "display_tags"),
        },
    }
    PUBLICATION_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def profile_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("profile_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = PROFILE_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(PROFILE_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(PROFILE_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("profiles", Query(limit=1000, order_by="sort_order", descending=False))
    payload = {
        "ok": True,
        "fields": {
            "role": publication_unique_values(rows, "role"),
            "title": publication_unique_values(rows, "title"),
            "organization": publication_unique_values(rows, "organization"),
            "lab": publication_unique_values(rows, "lab"),
            "office": publication_unique_values(rows, "office"),
        },
    }
    PROFILE_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def publication_unique_values(rows: list[dict[str, Any]], field: str) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        value = text_only(row.get(field), 240).strip()
        key = value.casefold()
        if value and key not in seen:
            values.append(value)
            seen.add(key)
    return values[:80]


def publication_split_values(rows: list[dict[str, Any]], field: str) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        raw = text_only(row.get(field), 800).strip()
        for value in split_publication_tag_text(raw):
            key = value.casefold()
            if key not in seen:
                values.append(value)
                seen.add(key)
    return values[:120]


def split_publication_tag_text(raw: str) -> list[str]:
    text = raw.replace("；", ";").replace("，", ",").replace("\n", ",").replace(";", ",")
    return [part.strip() for part in text.split(",") if part.strip()]


def project_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("project_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = PROJECT_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(PROJECT_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(PROJECT_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("projects", Query(limit=1000, order_by="id", descending=True))
    payload = {
        "ok": True,
        "fields": {
            "source": publication_unique_values(rows, "source"),
            "fund_name": publication_unique_values(rows, "fund_name"),
            "status": publication_unique_values(rows, "status"),
            "principal": publication_unique_values(rows, "principal"),
            "members": project_split_values(rows, "members"),
        },
    }
    PROJECT_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def project_split_values(rows: list[dict[str, Any]], field: str) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in rows:
        raw = text_only(row.get(field), 800).strip()
        for value in split_publication_tag_text(raw):
            key = value.casefold()
            if key not in seen:
                values.append(value)
                seen.add(key)
    return values[:120]


def patent_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("patent_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = PATENT_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(PATENT_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(PATENT_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("patents", Query(limit=1000, order_by="id", descending=True))
    payload = {
        "ok": True,
        "fields": {
            "country": publication_unique_values(rows, "country"),
            "patent_type": publication_unique_values(rows, "patent_type"),
            "legal_status": publication_unique_values(rows, "legal_status"),
            "owner": publication_unique_values(rows, "owner"),
            "inventors": project_split_values(rows, "inventors"),
        },
    }
    PATENT_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def student_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("student_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = STUDENT_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(STUDENT_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(STUDENT_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("students", Query(limit=1000, order_by="id", descending=True))
    payload = {
        "ok": True,
        "fields": {
            "degree": publication_unique_values(rows, "degree"),
            "category": publication_unique_values(rows, "category"),
            "grade": publication_unique_values(rows, "grade"),
            "status": publication_unique_values(rows, "status"),
            "direction": publication_unique_values(rows, "direction"),
            "destination": publication_unique_values(rows, "destination"),
        },
    }
    STUDENT_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def news_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("news_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = NEWS_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(NEWS_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(NEWS_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("news", Query(limit=1000, order_by="published_at", descending=True))
    payload = {"ok": True, "fields": {"category": publication_unique_values(rows, "category")}}
    NEWS_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def course_suggestions_payload(repo: Repository) -> dict[str, Any]:
    settings = active_global(repo)
    ttl = max(0, int_value(settings.get("course_suggestion_cache_seconds"), 30))
    now = time.time()
    cached = COURSE_SUGGESTION_CACHE.get("payload")
    if cached is not None and ttl > 0 and now - float(COURSE_SUGGESTION_CACHE.get("ts") or 0) < ttl and int(COURSE_SUGGESTION_CACHE.get("ttl") or ttl) == ttl:
        return {**cached, "cached": True, "ttl": ttl}
    rows = repo.list("courses", Query(limit=1000, order_by="sort_order", descending=True))
    payload = {
        "ok": True,
        "fields": {
            "semester": publication_unique_values(rows, "semester"),
            "audience": publication_unique_values(rows, "audience"),
        },
    }
    COURSE_SUGGESTION_CACHE.update({"ts": now, "ttl": ttl, "payload": payload})
    return {**payload, "cached": False, "ttl": ttl}


def parse_publication_reference(raw: str, fmt: str = "auto") -> dict[str, Any]:
    text = raw.strip()
    fmt = fmt if fmt in {"auto", "bibtex", "ieee", "elsevier", "gb"} else "auto"
    if fmt == "bibtex" or text.startswith("@"):
        fields = parse_bibtex_reference(text)
    elif fmt == "ieee":
        fields = parse_ieee_reference(text)
    elif fmt == "elsevier":
        fields = parse_elsevier_reference(text)
    elif fmt == "gb":
        fields = parse_gb_reference(text)
    else:
        fields = parse_auto_publication_reference(text)
    doi = find_doi(text)
    if doi and not fields.get("doi"):
        fields["doi"] = doi
    return fields


def parse_bibtex_reference(text: str) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    for key, value in re_find_bib_fields(text).items():
        mapped = {
            "title": "title", "author": "authors", "journal": "venue", "booktitle": "venue",
            "year": "year", "volume": "volume", "number": "issue", "pages": "pages",
            "doi": "doi", "url": "url", "abstract": "abstract", "keywords": "keywords",
        }.get(key.lower())
        if mapped:
            fields[mapped] = value.replace(" and ", "; ") if mapped == "authors" else value
    fields["bibtex"] = text
    lowered = text.lower()
    if "booktitle" in lowered:
        fields["publication_type"] = "会议论文"
    elif "journal" in lowered:
        fields["publication_type"] = "期刊论文"
    return fields


def re_find_bib_fields(text: str) -> dict[str, str]:
    import re

    found: dict[str, str] = {}
    pattern = re.compile(r"(\w+)\s*=\s*(\{(?:[^{}]|\{[^{}]*\})*\}|\"[^\"]*\"|[^,\n]+)", re.S)
    for key, value in pattern.findall(text):
        clean = value.strip().strip(",").strip()
        if clean.startswith("{") and clean.endswith("}"):
            clean = clean[1:-1]
        if clean.startswith('"') and clean.endswith('"'):
            clean = clean[1:-1]
        found[key.lower()] = " ".join(clean.split())
    return found


def parse_auto_publication_reference(text: str) -> dict[str, Any]:
    lowered = text.lower()
    if "[j]" in lowered or "[c]" in lowered or "知网" in lowered:
        return parse_gb_reference(text)
    if '"' in text or "vol." in lowered or "no." in lowered or "pp." in lowered:
        return parse_ieee_reference(text)
    if re_search(r"\(\s*(19|20)\d{2}\s*\)", text):
        return parse_elsevier_reference(text)
    return parse_plain_publication_reference(text)


def parse_plain_publication_reference(text: str) -> dict[str, Any]:
    import re

    fields: dict[str, Any] = {}
    cleaned = " ".join(text.replace("\n", " ").split())
    year_match = re.search(r"\b(19|20)\d{2}\b", cleaned)
    if year_match:
        fields["year"] = year_match.group(0)
    parts = [part.strip(" .") for part in re.split(r"\.\s+|\[J\]|\[C\]", cleaned) if part.strip(" .")]
    if len(parts) >= 2:
        fields["authors"] = parts[0].replace(",", ";")
        fields["title"] = parts[1]
    elif cleaned:
        fields["title"] = cleaned[:500]
    if len(parts) >= 3:
        venue = parts[2]
        if year_match:
            venue = venue.split(year_match.group(0))[0].strip(" ,")
        fields["venue"] = venue[:300]
    lowered = text.lower()
    if "[c]" in lowered or "conference" in lowered:
        fields["publication_type"] = "会议论文"
    elif "[j]" in lowered or "journal" in lowered:
        fields["publication_type"] = "期刊论文"
    return fields


def parse_ieee_reference(text: str) -> dict[str, Any]:
    import re

    fields: dict[str, Any] = {}
    cleaned = " ".join(text.replace("\n", " ").split())
    cleaned = re.sub(r"^\s*\[\d+\]\s*", "", cleaned)
    title_match = re.search(r'"([^"]{5,500})"', cleaned)
    if title_match:
        fields["title"] = title_match.group(1).strip(" ,.")
        before = cleaned[:title_match.start()].strip(" ,.")
        after = cleaned[title_match.end():].strip(" ,.")
        if before:
            fields["authors"] = before.replace(", and ", "; ").replace(" and ", "; ")
        venue = after.split(", vol.", 1)[0].split(", no.", 1)[0].split(", pp.", 1)[0].strip(" ,.")
        if venue:
            fields["venue"] = venue
    year = re.search(r"\b(19|20)\d{2}\b", cleaned)
    if year:
        fields["year"] = year.group(0)
    for target, pattern in [("volume", r"\bvol\.\s*([^,]+)"), ("issue", r"\bno\.\s*([^,]+)"), ("pages", r"\bpp\.\s*([^,]+)")]:
        match = re.search(pattern, cleaned, flags=re.I)
        if match:
            fields[target] = match.group(1).strip(" .")
    fields.setdefault("publication_type", "期刊论文")
    return fields


def parse_elsevier_reference(text: str) -> dict[str, Any]:
    import re

    fields = parse_plain_publication_reference(text)
    cleaned = " ".join(text.replace("\n", " ").split())
    match = re.match(r"(.+?)\.\s+(.+?)\.\s+(.+?)\s+(\d+)\s*\((\d{4})\)\s*([0-9A-Za-z\-–—]+)", cleaned)
    if match:
        fields.update({
            "authors": match.group(1).replace(",", ";"),
            "title": match.group(2).strip(" ."),
            "venue": match.group(3).strip(" ."),
            "volume": match.group(4),
            "year": match.group(5),
            "pages": match.group(6).replace("–", "-").replace("—", "-"),
        })
    fields.setdefault("publication_type", "期刊论文")
    return fields


def parse_gb_reference(text: str) -> dict[str, Any]:
    import re

    fields: dict[str, Any] = {}
    cleaned = " ".join(text.replace("\n", " ").split())
    match = re.match(r"(.+?)\.\s*(.+?)\[(J|C|D|M)\]\.\s*(.+?)(?:[,，]\s*((?:19|20)\d{2}))?(?:[,，]\s*([^:：]+))?(?:[:：]\s*([0-9A-Za-z\-–—]+))?\s*\.?$", cleaned, flags=re.I)
    if match:
        fields["authors"] = "; ".join(part.strip() for part in re.split(r"[;,，；]", match.group(1)) if part.strip())
        fields["title"] = match.group(2).strip(" .")
        kind = match.group(3).upper()
        fields["venue"] = match.group(4).strip(" ,，.")
        if match.group(5):
            fields["year"] = match.group(5)
        issue_text = (match.group(6) or "").strip()
        if issue_text:
            vol_issue = re.match(r"([^()（）]+)[(（]([^()（）]+)[)）]", issue_text)
            if vol_issue:
                fields["volume"] = vol_issue.group(1).strip()
                fields["issue"] = vol_issue.group(2).strip()
            else:
                fields["issue"] = issue_text
        if match.group(7):
            fields["pages"] = match.group(7).replace("–", "-").replace("—", "-")
        fields["publication_type"] = "会议论文" if kind == "C" else "期刊论文" if kind == "J" else ""
    else:
        fields = parse_plain_publication_reference(text)
    return fields


def re_search(pattern: str, text: str) -> bool:
    import re

    return bool(re.search(pattern, text, flags=re.I))


def find_doi(text: str) -> str:
    import re

    match = re.search(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", text, flags=re.I)
    return match.group(0).rstrip(".,;") if match else ""


def normalize_doi(value: Any) -> str:
    text = text_only(value, 200).strip().lower()
    text = text.removeprefix("https://doi.org/").removeprefix("http://doi.org/").removeprefix("doi:")
    return text.strip()


def normalize_lookup_text(value: Any) -> str:
    import re

    return re.sub(r"\s+", " ", text_only(value, 800).casefold()).strip(" .,:;")


def normalize_identifier(value: Any) -> str:
    import re

    text = text_only(value, 300).strip().upper()
    text = text.removeprefix("HTTPS://PATENTS.GOOGLE.COM/PATENT/")
    text = text.removeprefix("HTTP://PATENTS.GOOGLE.COM/PATENT/")
    text = text.split("?", 1)[0].split("#", 1)[0]
    return re.sub(r"[^A-Z0-9]", "", text)


def parse_platforms(value: Any) -> list[str]:
    raw = text_only(value, 500).strip().lower()
    if not raw or raw == "manual":
        raw = "crossref,openalex,semanticscholar"
    aliases = {"semantic": "semanticscholar", "semantic_scholar": "semanticscholar"}
    platforms = []
    for part in raw.replace("；", ",").replace(";", ",").replace("\n", ",").split(","):
        item = aliases.get(part.strip(), part.strip())
        if item in {"crossref", "openalex", "semanticscholar"} and item not in platforms:
            platforms.append(item)
    return platforms or ["crossref", "openalex", "semanticscholar"]


def parse_patent_platforms(value: Any) -> list[str]:
    raw = text_only(value, 500).strip().lower()
    if not raw or raw == "manual":
        raw = "patentsview,epo_ops"
    aliases = {"epo": "epo_ops", "ops": "epo_ops", "patents_view": "patentsview"}
    platforms = []
    for part in raw.replace("；", ",").replace(";", ",").replace("\n", ",").split(","):
        item = aliases.get(part.strip(), part.strip())
        if item in {"patentsview", "epo_ops"} and item not in platforms:
            platforms.append(item)
    return platforms or ["patentsview", "epo_ops"]


def lookup_publication_platform(platform: str, title: str, doi: str) -> dict[str, Any]:
    if platform == "crossref":
        return lookup_crossref(title, doi)
    if platform == "openalex":
        return lookup_openalex(title, doi)
    if platform == "semanticscholar":
        return lookup_semantic_scholar(title, doi)
    return {"platform": platform, "ok": False, "message": "未知平台"}


def lookup_patent_platform(platform: str, title: str, identifier: str, settings: dict[str, Any]) -> dict[str, Any]:
    if platform == "patentsview":
        return lookup_patentsview(title, identifier, settings)
    if platform == "epo_ops":
        return lookup_epo_ops(title, identifier, settings)
    return {"platform": platform, "ok": False, "message": "未知平台"}


def http_json(url: str) -> dict[str, Any]:
    request = Request(url, headers={"User-Agent": "teacher-site/0.1 (mailto:admin@example.edu)", "Accept": "application/json"})
    try:
        with urlopen(request, timeout=7) as response:
            return json.loads(response.read(2_000_000).decode("utf-8", "ignore"))
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc


def http_json_with_headers(url: str, headers: dict[str, str]) -> dict[str, Any]:
    request_headers = {"User-Agent": "teacher-site/0.1 (mailto:admin@example.edu)", "Accept": "application/json"}
    request_headers.update(headers)
    request = Request(url, headers=request_headers)
    try:
        with urlopen(request, timeout=8) as response:
            return json.loads(response.read(2_000_000).decode("utf-8", "ignore"))
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc


def http_text(request: Request, limit: int = 2_000_000) -> str:
    try:
        with urlopen(request, timeout=8) as response:
            return response.read(limit).decode("utf-8", "ignore")
    except URLError as exc:
        raise RuntimeError(str(exc.reason)) from exc


def lookup_patentsview(title: str, identifier: str, settings: dict[str, Any]) -> dict[str, Any]:
    api_key = text_only(settings.get("patentsview_api_key"), 300).strip()
    if not api_key:
        return {"platform": "patentsview", "ok": False, "message": "PatentsView 需要在通用设置中配置 API Key。"}
    fields = [
        "patent_id", "patent_title", "patent_date", "patent_type", "patent_abstract",
        "inventors.inventor_name_first", "inventors.inventor_name_last",
        "assignees.assignee_organization", "assignees.assignee_first_name", "assignees.assignee_last_name",
        "applications.app_number", "applications.app_date",
    ]
    if identifier:
        query_obj: dict[str, Any] = {"_or": [{"patent_id": identifier}, {"applications.app_number": identifier}]}
    else:
        query_obj = {"_text_all": {"patent_title": title}}
    params = {
        "q": json.dumps(query_obj, ensure_ascii=False, separators=(",", ":")),
        "f": json.dumps(fields, separators=(",", ":")),
        "o": json.dumps({"size": 1}, separators=(",", ":")),
    }
    data = http_json_with_headers("https://search.patentsview.org/api/v1/patent/?" + urlencode(params), {"X-Api-Key": api_key})
    item = first_patentsview_item(data)
    return {"platform": "patentsview", "ok": bool(item), "fields": fields_from_patentsview(item), "message": "PatentsView 已返回结果" if item else "PatentsView 无结果"}


def first_patentsview_item(data: dict[str, Any]) -> dict[str, Any]:
    for key in ("patents", "data", "results"):
        values = data.get(key)
        if isinstance(values, list) and values:
            return values[0] if isinstance(values[0], dict) else {}
    if isinstance(data.get("patent"), dict):
        return data["patent"]
    return {}


def fields_from_patentsview(item: dict[str, Any]) -> dict[str, Any]:
    inventors = []
    for inventor in item.get("inventors") or []:
        name = " ".join(part for part in [inventor.get("inventor_name_first"), inventor.get("inventor_name_last"), inventor.get("name")] if part)
        if name:
            inventors.append(name)
    owners = []
    for assignee in item.get("assignees") or []:
        name = assignee.get("assignee_organization") or " ".join(part for part in [assignee.get("assignee_first_name"), assignee.get("assignee_last_name"), assignee.get("name")] if part)
        if name:
            owners.append(name)
    applications = item.get("applications") or []
    application = applications[0] if applications and isinstance(applications[0], dict) else {}
    return compact_fields({
        "name": item.get("patent_title") or item.get("title"),
        "country": patent_country_from_identifier(item.get("patent_id") or item.get("patent_number")),
        "patent_type": item.get("patent_type") or "发明专利",
        "application_number": application.get("app_number"),
        "application_date": normalize_date_input(application.get("app_date")),
        "grant_number": item.get("patent_id") or item.get("patent_number"),
        "grant_date": normalize_date_input(item.get("patent_date")),
        "inventors": "; ".join(inventors[:20]),
        "owner": "; ".join(owners[:10]),
        "legal_status": "已授权" if item.get("patent_date") else "",
        "summary": item.get("patent_abstract") or item.get("abstract"),
    })


def lookup_epo_ops(title: str, identifier: str, settings: dict[str, Any]) -> dict[str, Any]:
    client_id = text_only(settings.get("epo_ops_client_id"), 300).strip()
    client_secret = text_only(settings.get("epo_ops_client_secret"), 300).strip()
    if not client_id or not client_secret:
        return {"platform": "epo_ops", "ok": False, "message": "EPO OPS 需要在通用设置中配置 Client ID 和 Secret。"}
    token = epo_ops_token(client_id, client_secret)
    if identifier:
        path = f"https://ops.epo.org/3.2/rest-services/published-data/publication/epodoc/{quote(identifier)}/biblio"
    else:
        query = f'ti="{title}"'
        path = "https://ops.epo.org/3.2/rest-services/published-data/search/biblio?" + urlencode({"q": query})
    request = Request(path, headers={"Authorization": f"Bearer {token}", "Accept": "application/xml", "User-Agent": "teacher-site/0.1"})
    text = http_text(request)
    fields = fields_from_epo_ops_xml(text)
    return {"platform": "epo_ops", "ok": bool(fields), "fields": fields, "message": "EPO OPS 已返回结果" if fields else "EPO OPS 无结果"}


def epo_ops_token(client_id: str, client_secret: str) -> str:
    token = base64.b64encode(f"{client_id}:{client_secret}".encode("utf-8")).decode("ascii")
    body = urlencode({"grant_type": "client_credentials"}).encode("utf-8")
    request = Request(
        "https://ops.epo.org/3.2/auth/accesstoken",
        data=body,
        headers={"Authorization": f"Basic {token}", "Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
        method="POST",
    )
    data = json.loads(http_text(request, 200_000) or "{}")
    access = text_only(data.get("access_token"), 1000).strip()
    if not access:
        raise RuntimeError("EPO OPS 未返回 access_token")
    return access


def fields_from_epo_ops_xml(text: str) -> dict[str, Any]:
    import xml.etree.ElementTree as ET

    root = ET.fromstring(text)
    titles = [node.text.strip() for node in root.iter() if node.tag.endswith("invention-title") and node.text and node.text.strip()]
    names = [node.text.strip() for node in root.iter() if node.tag.endswith("name") and node.text and node.text.strip()]
    doc_numbers = [node.text.strip() for node in root.iter() if node.tag.endswith("doc-number") and node.text and node.text.strip()]
    dates = [normalize_date_input(node.text.strip()) for node in root.iter() if node.tag.endswith("date") and node.text and node.text.strip()]
    country = first_xml_text(root, "country")
    kind = first_xml_text(root, "kind")
    return compact_fields({
        "name": first_text(titles),
        "country": country,
        "patent_type": patent_type_from_kind(kind),
        "grant_number": doc_numbers[0] if doc_numbers else "",
        "application_number": doc_numbers[1] if len(doc_numbers) > 1 else "",
        "grant_date": dates[0] if dates else "",
        "application_date": dates[1] if len(dates) > 1 else "",
        "inventors": "; ".join(names[:10]),
        "legal_status": "已公开" if doc_numbers else "",
    })


def first_xml_text(root: Any, suffix: str) -> str:
    for node in root.iter():
        if node.tag.endswith(suffix) and node.text and node.text.strip():
            return node.text.strip()
    return ""


def patent_country_from_identifier(value: Any) -> str:
    text = text_only(value, 40).strip().upper()
    if text.startswith(("US", "EP", "CN", "WO", "JP", "KR")):
        return text[:2]
    if text and text[0].isdigit():
        return "US"
    return ""


def patent_type_from_kind(value: Any) -> str:
    text = text_only(value, 40).strip().upper()
    if text.startswith("U"):
        return "实用新型专利"
    if text.startswith("S"):
        return "外观设计专利"
    return "发明专利" if text else ""


def lookup_crossref(title: str, doi: str) -> dict[str, Any]:
    url = f"https://api.crossref.org/works/{quote(doi)}" if doi else "https://api.crossref.org/works?" + urlencode({"query.title": title, "rows": "1"})
    data = http_json(url)
    item = data.get("message", {})
    if "items" in item:
        item = (item.get("items") or [{}])[0]
    return {"platform": "crossref", "ok": bool(item), "fields": fields_from_crossref(item), "message": "Crossref 已返回结果" if item else "Crossref 无结果"}


def lookup_openalex(title: str, doi: str) -> dict[str, Any]:
    if doi:
        item = http_json(f"https://api.openalex.org/works/https://doi.org/{quote(doi)}")
        if not item.get("id"):
            item = {}
    else:
        data = http_json("https://api.openalex.org/works?" + urlencode({"search": title, "per-page": "1"}))
        item = (data.get("results") or [{}])[0]
    return {"platform": "openalex", "ok": bool(item), "fields": fields_from_openalex(item), "message": "OpenAlex 已返回结果" if item else "OpenAlex 无结果"}


def lookup_semantic_scholar(title: str, doi: str) -> dict[str, Any]:
    fields = "title,authors,venue,year,externalIds,publicationTypes,abstract,url,journal"
    if doi:
        item = http_json(f"https://api.semanticscholar.org/graph/v1/paper/DOI:{quote(doi)}?" + urlencode({"fields": fields}))
        if not item.get("paperId"):
            item = {}
    else:
        data = http_json("https://api.semanticscholar.org/graph/v1/paper/search?" + urlencode({"query": title, "limit": "1", "fields": fields}))
        item = (data.get("data") or [{}])[0]
    return {"platform": "semanticscholar", "ok": bool(item), "fields": fields_from_semantic(item), "message": "Semantic Scholar 已返回结果" if item else "Semantic Scholar 无结果"}


def fields_from_crossref(item: dict[str, Any]) -> dict[str, Any]:
    issued = item.get("issued", {}).get("date-parts", [[]])
    authors = "; ".join(" ".join(part for part in [author.get("given"), author.get("family")] if part) for author in item.get("author", [])[:20])
    return compact_fields({
        "title": first_text(item.get("title")),
        "authors": authors,
        "venue": first_text(item.get("container-title")),
        "year": issued[0][0] if issued and issued[0] else "",
        "volume": item.get("volume"),
        "issue": item.get("issue"),
        "pages": item.get("page"),
        "doi": item.get("DOI"),
        "url": item.get("URL"),
        "publication_type": publication_type_from_source(item.get("type")),
    })


def fields_from_openalex(item: dict[str, Any]) -> dict[str, Any]:
    authors = "; ".join(text_only(authorship.get("author", {}).get("display_name"), 120) for authorship in (item.get("authorships") or [])[:20] if authorship.get("author"))
    primary = item.get("primary_location") or {}
    source = primary.get("source") or {}
    return compact_fields({
        "title": item.get("title") or item.get("display_name"),
        "authors": authors,
        "venue": source.get("display_name"),
        "year": item.get("publication_year"),
        "doi": normalize_doi(item.get("doi")),
        "url": item.get("doi") or item.get("id"),
        "publication_type": publication_type_from_source(item.get("type")),
    })


def fields_from_semantic(item: dict[str, Any]) -> dict[str, Any]:
    external = item.get("externalIds") or {}
    journal = item.get("journal") or {}
    return compact_fields({
        "title": item.get("title"),
        "authors": "; ".join(text_only(author.get("name"), 120) for author in (item.get("authors") or [])[:20]),
        "venue": journal.get("name") or item.get("venue"),
        "year": item.get("year"),
        "volume": journal.get("volume"),
        "pages": journal.get("pages"),
        "doi": external.get("DOI"),
        "url": item.get("url"),
        "abstract": item.get("abstract"),
        "publication_type": first_text(item.get("publicationTypes")),
    })


def publication_type_from_source(value: Any) -> str:
    text = text_only(value, 80).lower()
    if "journal" in text:
        return "期刊论文"
    if "conference" in text or "proceedings" in text:
        return "会议论文"
    if "preprint" in text:
        return "预印本"
    return text_only(value, 80)


def first_text(value: Any) -> str:
    if isinstance(value, list):
        return text_only(value[0] if value else "", 300)
    return text_only(value, 300)


def compact_fields(fields: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in fields.items() if value not in (None, "")}


def save_media_record(repo: Repository, key: str, title: str, category: str, mime: str, size: int, storage_kind: str = "local") -> dict[str, Any]:
    return repo.save("media_assets", media_record_data(key, title, category, mime, size, storage_kind))


def media_record_data(key: str, title: str, category: str, mime: str, size: int, storage_kind: str = "local") -> dict[str, Any]:
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]
    return {
        "uid": stable_uid("media", f"{key}-{digest}"),
        "object_key": key,
        "storage_kind": normalize_media_storage_kind(storage_kind, "local"),
        "title": title or key,
        "category": category or "media",
        "mime_type": mime,
        "size": size,
        "status": "active",
        "checksum": digest,
    }


def normalize_media_storage_kind(value: Any, default: str = "static") -> str:
    text = text_only(value, 40).strip().lower()
    return text if text in {"static", "local", "r2", "external"} else default


def unique_media_path(folder: str, filename: str, env: dict[str, str] | None = None) -> tuple[str, Path]:
    folder = safe_media_folder(folder)
    name = safe_media_filename(filename)
    root = runtime_media_root(env)
    stem = Path(name).stem or "media"
    suffix = Path(name).suffix or ".bin"
    candidate = f"{folder}/{stem}{suffix}"
    index = 1
    while (root / candidate).exists():
        candidate = f"{folder}/{stem}-{index}{suffix}"
        index += 1
    return candidate, root / candidate


def runtime_media_root(env: dict[str, str] | None = None) -> Path:
    configured = (env or {}).get("TEACHER_SITE_MEDIA") or os.environ.get("TEACHER_SITE_MEDIA") or "media"
    return Path(configured).expanduser()


def static_media_root(env: dict[str, str] | None = None) -> Path:
    configured = (env or {}).get("TEACHER_SITE_PUBLIC") or os.environ.get("TEACHER_SITE_PUBLIC") or "public"
    return Path(configured).expanduser() / "media"


def media_candidate_roots(env: dict[str, str] | None = None) -> tuple[Path, Path]:
    runtime = runtime_media_root(env)
    static = static_media_root(env)
    return (runtime, static) if runtime != static else (runtime, runtime)


def unique_media_object_key(folder: str, filename: str, prefix: str = "uploads") -> str:
    folder = safe_media_folder(folder)
    prefix = safe_media_folder(prefix or "uploads")
    name = safe_media_filename(filename)
    stem = Path(name).stem or "media"
    suffix = Path(name).suffix or ".bin"
    stamp = f"{int(time.time() * 1000)}-{hashlib.sha1(f'{folder}/{name}/{time.time_ns()}'.encode('utf-8')).hexdigest()[:8]}"
    return normalize_media_key(f"{prefix}/{folder}/{stem}-{stamp}{suffix}")


def safe_media_folder(value: str) -> str:
    text = "".join(char if char.isalnum() or char in "-_/" else "-" for char in value.strip().replace("\\", "/"))
    parts = [part for part in text.split("/") if part and part not in {".", ".."}]
    return "/".join(parts[:3]) or "icons"


def safe_media_filename(value: str) -> str:
    name = Path(value.replace("\\", "/")).name.strip() or "media.bin"
    stem = "".join(char if char.isalnum() or char in "-_." else "-" for char in Path(name).stem).strip(".-") or "media"
    suffix = Path(name).suffix.lower()
    if not suffix or len(suffix) > 12:
        suffix = ".bin"
    return f"{stem[:80]}{suffix}"


def parse_multipart(body: bytes, content_type: str) -> dict[str, Any]:
    marker = "boundary="
    if marker not in content_type:
        return {}
    boundary = content_type.split(marker, 1)[1].split(";", 1)[0].strip().strip('"')
    if not boundary:
        return {}
    delimiter = ("--" + boundary).encode("utf-8")
    result: dict[str, Any] = {}
    for part in body.split(delimiter):
        part = part.strip(b"\r\n")
        if not part or part == b"--" or b"\r\n\r\n" not in part:
            continue
        raw_headers, content = part.split(b"\r\n\r\n", 1)
        headers = {}
        for line in raw_headers.decode("latin1", "ignore").splitlines():
            key, _, value = line.partition(":")
            headers[key.strip().lower()] = value.strip()
        disposition = headers.get("content-disposition", "")
        params = multipart_params(disposition)
        name = params.get("name")
        if not name:
            continue
        if "filename" in params:
            multipart_add_value(result, name, {"filename": params.get("filename") or "", "content": content.rstrip(b"\r\n"), "content_type": headers.get("content-type", "")})
        else:
            multipart_add_value(result, name, content.decode("utf-8", "ignore").rstrip("\r\n"))
    return result


def multipart_add_value(result: dict[str, Any], name: str, value: Any) -> None:
    if name not in result:
        result[name] = value
        return
    current = result[name]
    if isinstance(current, list):
        current.append(value)
    else:
        result[name] = [current, value]


def multipart_params(value: str) -> dict[str, str]:
    params: dict[str, str] = {}
    for part in value.split(";"):
        key, sep, raw = part.strip().partition("=")
        if sep:
            params[key.strip().lower()] = raw.strip().strip('"')
    return params


def media_trash_retention_days(rows_or_repo: Any) -> int:
    if hasattr(rows_or_repo, "list"):
        settings = active_global(rows_or_repo)
        return max(1, int_value(settings.get("media_trash_retention_days"), 30))
    return 30


def media_auto_empty_trash(repo: Repository) -> None:
    retention_days = media_trash_retention_days(repo)
    cutoff = time.time() - retention_days * 86400
    for row in repo.list("media_assets", Query(limit=1000)):
        if str(row.get("status") or "active") != "trash":
            continue
        changed_at = parse_timestamp(row.get("updated_at") or row.get("created_at"))
        if changed_at and changed_at < cutoff:
            if media_delete_physical_file_for_row(row):
                repo.delete("media_assets", str(row.get("uid") or row.get("id")))


def parse_timestamp(value: Any) -> float | None:
    text = text_only(value, 40).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return time.mktime(time.strptime(text[:19], fmt))
        except ValueError:
            continue
    return None


def media_cache_now() -> float:
    return time.time()


def media_cache_load() -> dict[str, Any]:
    try:
        if MEDIA_STATS_CACHE_PATH.is_file():
            data = json.loads(MEDIA_STATS_CACHE_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
    except (OSError, ValueError, TypeError):
        return {}
    return {}


def media_cache_save(cache: dict[str, Any]) -> None:
    try:
        MEDIA_STATS_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        MEDIA_STATS_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    except OSError:
        return


def media_cached_entry(cache: dict[str, Any], section: str, key: str = "value", now: float | None = None) -> tuple[dict[str, Any] | None, int]:
    now = now or media_cache_now()
    area = cache.get(section)
    entry = area.get(key) if isinstance(area, dict) and section == "files" else area
    if not isinstance(entry, dict):
        return None, 0
    created_at = float(entry.get("time") or 0)
    age = max(0, int(now - created_at))
    value = entry.get("value")
    if age > MEDIA_STATS_CACHE_TTL_SECONDS or not isinstance(value, dict):
        return None, age
    return dict(value), age


def media_cache_put(cache: dict[str, Any], section: str, value: dict[str, Any], key: str = "value", now: float | None = None) -> dict[str, Any]:
    now = now or media_cache_now()
    if section == "files":
        files = cache.get("files")
        if not isinstance(files, dict):
            files = {}
            cache["files"] = files
        files[key] = {"time": now, "value": value}
    else:
        cache[section] = {"time": now, "value": value}
    return cache


def with_media_cache_meta(payload: dict[str, Any], cached: bool, age: int = 0) -> dict[str, Any]:
    return {
        **payload,
        "cached": cached,
        "cache_age_seconds": age,
        "cache_ttl_seconds": MEDIA_STATS_CACHE_TTL_SECONDS,
    }


def media_summary_payload(repo: Repository, env: dict[str, str], refresh: bool = False) -> dict[str, Any]:
    rows = repo.list("media_assets", Query(limit=1000))
    cache = media_cache_load()
    cached_disk, age = (None, 0) if refresh else media_cached_entry(cache, "disk")
    if cached_disk is None:
        disk = disk_usage_payload(env)
        media_cache_save(media_cache_put(cache, "disk", disk))
        cached = False
        age = 0
    else:
        disk = cached_disk
        cached = True
    recorded_total = sum(int_value(row.get("size")) for row in rows)
    return {
        "records": len(rows),
        "active": sum(1 for row in rows if str(row.get("status") or "active") != "trash"),
        "trash": sum(1 for row in rows if str(row.get("status") or "active") == "trash"),
        "recorded_total": recorded_total,
        "recorded_total_label": format_bytes(recorded_total),
        "retention_days": media_trash_retention_days(repo),
        "disk": disk,
        "cached": cached,
        "cache_age_seconds": age,
        "cache_ttl_seconds": MEDIA_STATS_CACHE_TTL_SECONDS,
    }


def media_file_size_payload(key: str, env: dict[str, str], refresh: bool = False) -> dict[str, Any]:
    clean = normalize_media_key(key)
    cache = media_cache_load()
    cached_payload, age = (None, 0) if refresh else media_cached_entry(cache, "files", clean)
    if cached_payload is not None:
        return with_media_cache_meta(cached_payload, True, age)
    path = media_local_path(clean, env)
    if not path:
        payload = {"key": clean, "exists": False, "size": 0, "label": "未找到"}
    else:
        try:
            size = path.stat().st_size
            payload = {"key": clean, "exists": True, "size": size, "label": format_bytes(size)}
        except OSError:
            payload = {"key": clean, "exists": False, "size": 0, "label": "未找到"}
    media_cache_save(media_cache_put(cache, "files", payload, key=clean))
    return with_media_cache_meta(payload, False)


def disk_usage_payload(env: dict[str, str] | None = None) -> dict[str, Any]:
    if (env or {}).get("PLATFORM") == "cloudflare":
        return {"available": False, "path": "cloudflare", "total": 0, "used": 0, "free": 0, "total_label": "不支持", "used_label": "不支持", "free_label": "不支持"}
    target = first_existing_media_root(env) or Path.cwd()
    try:
        usage = shutil.disk_usage(target)
        return {
            "available": True,
            "path": str(target),
            "total": usage.total,
            "used": usage.used,
            "free": usage.free,
            "total_label": format_bytes(usage.total),
            "used_label": format_bytes(usage.used),
            "free_label": format_bytes(usage.free),
        }
    except OSError:
        return {"available": False, "path": str(target), "total": 0, "used": 0, "free": 0, "total_label": "未知", "used_label": "未知", "free_label": "未知"}


def first_existing_media_root(env: dict[str, str] | None = None) -> Path | None:
    for root in media_candidate_roots(env):
        if root.exists():
            return root
    return None


def media_local_path(key: str, env: dict[str, str] | None = None) -> Path | None:
    clean = normalize_media_key(key)
    if not clean or clean.startswith("/") or ".." in Path(clean).parts:
        return None
    for root in media_candidate_roots(env):
        try:
            root_resolved = root.resolve()
            target = (root / clean).resolve()
            if str(target).startswith(str(root_resolved)) and target.is_file():
                return target
        except OSError:
            continue
    return None


def media_target_path_for_key(key: str, env: dict[str, str] | None = None) -> Path | None:
    clean = normalize_media_key(key)
    if not clean or clean.startswith("/") or ".." in Path(clean).parts:
        return None
    existing = media_local_path(clean, env)
    if existing:
        return existing
    try:
        root = runtime_media_root(env).resolve()
        target = (root / clean).resolve()
        if str(target).startswith(str(root)):
            return target
    except OSError:
        return None
    return None


def media_scan_project_files(repo: Repository, env: dict[str, str]) -> dict[str, int | str]:
    if env.get("PLATFORM") == "cloudflare":
        return {"unsupported": 1, "scanned": 0, "added": 0, "updated": 0, "skipped": 0}
    existing_by_key: dict[str, dict[str, Any]] = {}
    for row in repo.list("media_assets", Query(limit=1000)):
        key = normalize_media_key(str(row.get("object_key") or ""))
        if key and key not in existing_by_key:
            existing_by_key[key] = row
    scanned = 0
    added = 0
    updated = 0
    skipped = 0
    seen: set[str] = set()
    for root, storage_kind in media_scan_roots():
        try:
            root_resolved = root.resolve()
        except OSError:
            skipped += 1
            continue
        for item in root.rglob("*"):
            if not item.is_file():
                continue
            suffix = item.suffix.lower()
            if suffix not in MEDIA_SCAN_EXTENSIONS:
                skipped += 1
                continue
            try:
                key = normalize_media_key(item.resolve().relative_to(root_resolved).as_posix())
            except (OSError, ValueError):
                skipped += 1
                continue
            if not key or key in seen:
                skipped += 1
                continue
            seen.add(key)
            scanned += 1
            try:
                stat = item.stat()
            except OSError:
                skipped += 1
                continue
            mime = mimetypes.guess_type(item.name)[0] or media_guess_mime_from_suffix(suffix)
            category = safe_media_folder(key.rsplit("/", 1)[0]) if "/" in key else "media"
            title = text_only(item.stem, 160) or key
            checksum = hashlib.sha1(f"{key}:{stat.st_size}:{int(stat.st_mtime)}".encode("utf-8")).hexdigest()[:12]
            row = existing_by_key.get(key) or media_find_record_by_key(repo, key)
            if row:
                before = dict(row)
                row["size"] = stat.st_size
                row["mime_type"] = row.get("mime_type") or mime
                row["category"] = row.get("category") or category
                row["title"] = row.get("title") or title
                row["checksum"] = row.get("checksum") or checksum
                row["storage_kind"] = normalize_media_storage_kind(row.get("storage_kind"), storage_kind)
                if row != before:
                    repo.save("media_assets", row)
                    updated += 1
                else:
                    skipped += 1
                continue
            saved = save_media_record(repo, key, title, category, mime, stat.st_size, storage_kind)
            saved["checksum"] = checksum
            repo.save("media_assets", saved)
            existing_by_key[key] = saved
            added += 1
    return {"unsupported": 0, "scanned": scanned, "added": added, "updated": updated, "skipped": skipped}


def media_find_record_by_key(repo: Repository, key: str) -> dict[str, Any] | None:
    clean = normalize_media_key(key)
    if not clean:
        return None
    for row in repo.list("media_assets", Query(q=clean, limit=1000)):
        if normalize_media_key(str(row.get("object_key") or "")) == clean:
            return row
    return None


def media_scan_roots() -> list[tuple[Path, str]]:
    roots = []
    for root, storage_kind in ((Path("media"), "local"), (Path("public") / "media", "static")):
        if root.exists() and root.is_dir():
            roots.append((root, storage_kind))
    return roots


def media_guess_mime_from_suffix(suffix: str) -> str:
    if suffix in {".svg"}:
        return "image/svg+xml"
    if suffix in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif", ".bmp", ".ico", ".tif", ".tiff"}:
        return f"image/{suffix.lstrip('.').replace('jpg', 'jpeg')}"
    if suffix in {".mp4", ".webm", ".mov", ".m4v", ".ogv"}:
        return "video/mp4" if suffix in {".mp4", ".m4v", ".mov"} else f"video/{suffix.lstrip('.')}"
    if suffix in {".mp3", ".wav", ".ogg", ".flac", ".m4a", ".aac"}:
        return f"audio/{suffix.lstrip('.')}"
    if suffix == ".pdf":
        return "application/pdf"
    return "application/octet-stream"


def media_delete_physical_file_for_row(row: dict[str, Any] | None) -> bool:
    if not row:
        return False
    key = normalize_media_key(str(row.get("object_key") or ""))
    if not key or media_storage_kind(row) == "external":
        return True
    path = media_local_path(key)
    if not path:
        return True
    if not media_path_inside_managed_root(path):
        return False
    try:
        path.unlink()
        media_cache_forget_file(key)
        return True
    except OSError:
        return False


def media_path_inside_managed_root(path: Path) -> bool:
    try:
        resolved = path.resolve()
    except OSError:
        return False
    for root in media_candidate_roots(env):
        try:
            resolved.relative_to(root.resolve())
            return True
        except (OSError, ValueError):
            continue
    return False


def media_cache_forget_file(key: str) -> None:
    clean = normalize_media_key(key)
    if not clean:
        return
    cache = media_cache_load()
    files = cache.get("files")
    if isinstance(files, dict) and clean in files:
        files.pop(clean, None)
        media_cache_save(cache)


def media_export_used_response(repo: Repository, query: dict[str, str], body: bytes, env: dict[str, str], mode: str = "library") -> tuple[ResponseTuple, dict[str, Any]]:
    if env.get("PLATFORM") == "cloudflare":
        result = {"ok": False, "exported": 0, "skipped": 0, "reason": "cloudflare"}
        return json_response({"ok": False, "message": "Cloudflare Worker 环境暂不能从 Static Assets 文件系统打包媒体文件；如需生产环境导出媒体，请后续接入 R2 对象读取。"}, 501), result
    form = _form_multi(body) if body else {}
    selected = [value for value in form.get("selected", []) if value][:500]
    selected_mode = bool(selected)
    usage_map = media_usage_map(repo)
    raw_rows = media_export_candidate_rows(repo, query, selected, mode)
    entries: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    for row in raw_rows:
        key = normalize_media_key(str(row.get("object_key") or ""))
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        usage = usage_map.get(key, [])
        if not selected_mode and not usage:
            skipped.append({"object_key": key, "reason": "未发现使用位置"})
            continue
        path = media_local_path(key)
        if not path or not path.is_file():
            skipped.append({"object_key": key, "reason": "本地文件缺失"})
            continue
        if not media_path_inside_managed_root(path):
            skipped.append({"object_key": key, "reason": "文件不在受管理媒体目录"})
            continue
        try:
            size = path.stat().st_size
        except OSError:
            skipped.append({"object_key": key, "reason": "无法读取文件信息"})
            continue
        entries.append({
            "object_key": key,
            "title": text_only(row.get("title") or key, 200),
            "category": text_only(row.get("category"), 120),
            "mime_type": text_only(row.get("mime_type"), 160) or mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            "size": size,
            "status": text_only(row.get("status") or "active", 40),
            "path": path,
            "usage": usage,
        })
    result = {"ok": bool(entries), "selected": len(selected), "candidates": len(raw_rows), "exported": len(entries), "skipped": len(skipped), "scope": "selected" if selected_mode else "used"}
    if not entries:
        message = "没有可导出的选中媒体文件。请确认所选记录是本地媒体且文件存在。" if selected_mode else "没有可导出的已使用媒体文件。请先确认媒体仍被引用且本地文件存在。"
        return json_response({"ok": False, "message": message, "skipped": skipped[:50]}, 404), result
    payload = io.BytesIO()
    exported_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    manifest_files = []
    with zipfile.ZipFile(payload, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for entry in entries:
            key = str(entry["object_key"])
            arcname = media_zip_arcname(key)
            archive.write(entry["path"], arcname)
            manifest_files.append({
                "object_key": key,
                "archive_path": arcname,
                "title": entry["title"],
                "category": entry["category"],
                "mime_type": entry["mime_type"],
                "size": entry["size"],
                "status": entry["status"],
                "usage": entry["usage"],
            })
        manifest = {
            "exported_at": exported_at,
            "scope": "selected" if selected_mode else f"{mode}_used",
            "scope_note": "导出所有选中的本地媒体文件；usage 为空表示当前未发现引用。" if selected_mode else "快捷导出当前筛选范围内被使用的媒体文件。",
            "file_count": len(manifest_files),
            "total_bytes": sum(int_value(item.get("size")) for item in manifest_files),
            "skipped": skipped,
            "files": manifest_files,
        }
        archive.writestr("media_manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        archive.writestr("media_manifest.csv", csv_bytes(media_manifest_csv_rows(manifest_files, skipped), ["object_key", "archive_path", "title", "category", "mime_type", "size", "status", "usage", "skip_reason"]))
    filename = f"teacher-site-used-media-{time.strftime('%Y%m%d-%H%M%S')}.zip"
    return binary_response(payload.getvalue(), "application/zip", filename), result


def media_export_candidate_rows(repo: Repository, query: dict[str, str], selected: list[str], mode: str) -> list[dict[str, Any]]:
    if selected:
        return [row for key in selected if (row := repo.get("media_assets", key))]
    filters = {"status": "trash" if mode == "trash" else "active"}
    if query.get("category"):
        filters["category"] = query["category"]
    if query.get("mime_type"):
        filters["mime_type"] = query["mime_type"]
    order_by, descending = media_sort_args(query.get("sort", "updated_desc"))
    return repo.list("media_assets", Query(q=query.get("q", ""), filters=filters, limit=1000, order_by=order_by, descending=descending))


def media_zip_arcname(key: str) -> str:
    clean = normalize_media_key(key)
    parts = [part for part in clean.replace("\\", "/").split("/") if part and part not in {".", ".."}]
    if not parts:
        parts = ["media.bin"]
    return "media/" + "/".join(parts)


def media_manifest_csv_rows(files: list[dict[str, Any]], skipped: list[dict[str, str]]) -> list[dict[str, Any]]:
    rows = []
    for item in files:
        usage = item.get("usage") if isinstance(item.get("usage"), list) else []
        rows.append({
            "object_key": item.get("object_key", ""),
            "archive_path": item.get("archive_path", ""),
            "title": item.get("title", ""),
            "category": item.get("category", ""),
            "mime_type": item.get("mime_type", ""),
            "size": item.get("size", 0),
            "status": item.get("status", ""),
            "usage": "; ".join(f"{use.get('table_label', '')}/{use.get('field_label', '')}/{use.get('title', '')}" for use in usage if isinstance(use, dict)),
            "skip_reason": "",
        })
    for item in skipped:
        rows.append({
            "object_key": item.get("object_key", ""),
            "archive_path": "",
            "title": "",
            "category": "",
            "mime_type": "",
            "size": "",
            "status": "",
            "usage": "",
            "skip_reason": item.get("reason", ""),
        })
    return rows


def media_apply_action(repo: Repository, media_key: str, action: str) -> dict[str, int | str]:
    if action == "delete":
        row = repo.get("media_assets", media_key)
        if media_delete_physical_file_for_row(row):
            deleted = 1 if repo.delete("media_assets", media_key) else 0
            return {"selected": 1, "updated": 0, "deleted": deleted, "skipped": 0 if deleted else 1, "action": action}
        return {"selected": 1, "updated": 0, "deleted": 0, "skipped": 1, "action": action}
    row = repo.get("media_assets", media_key)
    if not row:
        return {"selected": 1, "updated": 0, "deleted": 0, "skipped": 1, "action": action}
    if action == "trash":
        row["status"] = "trash"
    elif action == "restore":
        row["status"] = "active"
    else:
        return {"selected": 1, "updated": 0, "deleted": 0, "skipped": 1, "action": action}
    repo.save("media_assets", row)
    return {"selected": 1, "updated": 1, "deleted": 0, "skipped": 0, "action": action}


def media_clear_trash(repo: Repository) -> dict[str, int | str]:
    deleted = 0
    skipped = 0
    for row in repo.list("media_assets", Query(filters={"status": "trash"}, limit=1000)):
        if media_delete_physical_file_for_row(row) and repo.delete("media_assets", str(row.get("uid") or row.get("id"))):
            deleted += 1
        else:
            skipped += 1
    return {"selected": deleted + skipped, "updated": 0, "deleted": deleted, "skipped": skipped, "action": "delete"}


def media_batch_update(repo: Repository, body: bytes) -> tuple[str, dict[str, int | str]]:
    data = _form_multi(body)
    selected = [value for value in data.get("selected", []) if value]
    action = (data.get("batch_action") or ["update"])[-1]
    category = text_only((data.get("batch_category") or [""])[-1], 120).strip()
    requested_status = (data.get("batch_status") or [""])[-1]
    updated = 0
    deleted = 0
    skipped = 0
    for key in selected[:500]:
        if action == "delete":
            row = repo.get("media_assets", key)
            if media_delete_physical_file_for_row(row) and repo.delete("media_assets", key):
                deleted += 1
            else:
                skipped += 1
            continue
        row = repo.get("media_assets", key)
        if not row:
            skipped += 1
            continue
        before = dict(row)
        if action == "trash":
            row["status"] = "trash"
        elif action == "restore":
            row["status"] = "active"
        elif requested_status in {"active", "trash"}:
            row["status"] = requested_status
        if action == "update" and category:
            row["category"] = category
        if row != before:
            repo.save("media_assets", row)
            updated += 1
        else:
            skipped += 1
    return_to = (data.get("return_to") or ["/admin/table/media_assets"])[-1]
    if return_to not in {"/admin/table/media_assets", "/admin/table/media_assets/trash"}:
        return_to = "/admin/table/media_assets"
    result = {"selected": len(selected[:500]), "updated": updated, "deleted": deleted, "skipped": skipped, "action": action}
    return append_query_params(return_to, {"batch_selected": result["selected"], "batch_updated": updated, "batch_deleted": deleted, "batch_skipped": skipped}), result


def media_usage_map(repo: Repository) -> dict[str, list[dict[str, str]]]:
    usage: dict[str, list[dict[str, str]]] = {}
    for table in TABLES:
        fields = [field for field in table.fields if field.kind == "file"]
        if not fields or table.name == "media_assets":
            continue
        for row in repo.list(table.name, Query(limit=1000)):
            for field in fields:
                key = normalize_media_key(str(row.get(field.name) or ""))
                if not key:
                    continue
                usage.setdefault(key, []).append({
                    "table": table.name,
                    "table_label": table.label,
                    "field_label": field.label,
                    "title": admin_row_title(table, row),
                    "admin_href": f'/admin/table/{table.name}/{row.get("uid") or row.get("id")}',
                    "public_href": public_href_for_row(table.name, row),
                })
        if table.name == "news":
            content_field = next((field for field in table.fields if field.name == "content"), None)
            if not content_field:
                continue
            for row in repo.list(table.name, Query(limit=1000)):
                for key in media_keys_from_html(row.get("content")):
                    usage.setdefault(key, []).append({
                        "table": table.name,
                        "table_label": table.label,
                        "field_label": "正文内容",
                        "title": admin_row_title(table, row),
                        "admin_href": f'/admin/table/{table.name}/{row.get("uid") or row.get("id")}',
                        "public_href": public_href_for_row(table.name, row),
                    })
    return usage


def media_keys_from_html(value: Any) -> list[str]:
    text = "" if value is None else str(value)
    found: list[str] = []
    for match in re.finditer(r"""(?:src|href)=["'](?:/)?(?:public/)?media/([^"'>?#]+)""", text, re.IGNORECASE):
        key = normalize_media_key(match.group(1))
        if key and key not in found:
            found.append(key)
    return found


def normalize_media_key(value: str) -> str:
    text = value.strip()
    for prefix in ("/media/", "media/", "/public/media/", "public/media/"):
        if text.startswith(prefix):
            return text.removeprefix(prefix)
    return text


def admin_media_usage_link(item: dict[str, str]) -> str:
    public = f'<a href="{esc(item["public_href"])}" target="_blank" rel="noreferrer">前台</a>' if item.get("public_href") else ""
    title = f'{item.get("table_label") or ""} · {item.get("field_label") or ""} · {item.get("title") or ""}'
    if item.get("admin_href"):
        title += f' · 后台: {item.get("admin_href")}'
    if item.get("public_href"):
        title += f' · 前台: {item.get("public_href")}'
    return f"""<span class="media-usage-chip" title="{esc(title)}">
      <span>{esc(item.get("table_label"))} · {esc(item.get("field_label"))}</span>
      <a href="{esc(item.get("admin_href"))}" target="_blank" rel="noreferrer" title="{esc(title)}">{esc(text_only(item.get("title"), 24))}</a>
      {public}
    </span>"""


def public_href_for_row(table: str, row: dict[str, Any]) -> str:
    uid = row.get("uid") or row.get("id") or ""
    if table == "profiles":
        return f"/team/{uid}"
    if table == "news":
        return f'/news/{safe_slug(str(row.get("slug") or row.get("title") or uid))}'
    if table == "students":
        return "/students"
    if table == "publications":
        return "/publications"
    if table == "projects":
        return "/projects"
    if table == "patents":
        return "/patents"
    if table == "courses":
        return "/courses"
    if table == "site_settings":
        return "/"
    return ""


def media_file_exists(row: dict[str, Any], env: dict[str, str]) -> bool:
    key = normalize_media_key(str(row.get("object_key") or ""))
    if not key:
        return False
    if key.startswith(("http://", "https://")):
        return True
    storage = media_storage_kind(row)
    if storage == "external":
        return True
    if env.get("PLATFORM") == "cloudflare" or env.get("PUBLIC_MEDIA_BASE_URL"):
        return storage in {"static", "r2"}
    return media_local_path(key, env) is not None


def media_storage_label(row: dict[str, Any] | str) -> str:
    labels = {"static": "静态包", "local": "本地", "r2": "R2", "external": "外链"}
    return labels.get(media_storage_kind(row), "静态包")


def media_row_state(row: dict[str, Any], env: dict[str, str]) -> str:
    status = str(row.get("status") or "active")
    if status == "trash":
        return "trash_missing" if not media_file_exists(row, env) else "trash"
    return "available" if media_file_exists(row, env) else "missing"


def admin_media_status_badge(state: str) -> str:
    labels = {
        "available": "可用",
        "missing": "文件缺失",
        "trash": "回收站",
        "trash_missing": "回收站/缺失",
    }
    klass = "missing" if state in {"missing", "trash_missing"} else state
    return f'<span class="admin-status-badge status-{esc(klass)}">{esc(labels.get(state, state or "未知"))}</span>'


def media_preview(row: dict[str, Any], env: dict[str, str]) -> str:
    key = str(row.get("object_key") or "")
    mime = str(row.get("mime_type") or "").lower()
    if not media_file_exists(row, env):
        return '<span class="media-file-icon media-file-missing">缺失</span>'
    url = media_url(key, env.get("PUBLIC_MEDIA_BASE_URL", ""))
    if mime.startswith("image/") or key.lower().endswith((".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg")):
        return f'<img src="{esc(url)}" alt="{esc(row.get("title") or key)}" loading="lazy" decoding="async">'
    if mime.startswith("video/") or key.lower().endswith((".mp4", ".webm", ".mov")):
        return f'<video src="{esc(url)}" muted preload="metadata"></video>'
    label = "PDF" if "pdf" in mime or key.lower().endswith(".pdf") else ("DOC" if any(key.lower().endswith(ext) for ext in (".doc", ".docx")) else "FILE")
    return f'<span class="media-file-icon">{esc(label)}</span>'


def admin_status_badge(status: str) -> str:
    label = "回收站" if status == "trash" else "可用"
    return f'<span class="admin-status-badge status-{esc(status or "active")}">{label}</span>'


def media_action_button(row: dict[str, Any], action: str, label: str, danger: bool = False, trash_context: bool = False, delete_files: bool = True) -> str:
    key = esc(row.get("uid") or row.get("id"))
    confirm_text = "确定彻底删除该媒体文件和媒体库记录吗？此操作不会经过回收站。" if delete_files else "确定彻底删除该媒体库记录吗？Cloudflare 静态资源文件不会被运行时删除。"
    confirm_attrs = f' data-confirm="{esc(confirm_text)}"' if danger else ""
    klass = "button danger" if danger else "button light"
    action_path = f"/admin/table/media_assets/trash/{key}/{action}" if trash_context else f"/admin/table/media_assets/{key}/{action}"
    return f'<form method="post" action="{action_path}"{confirm_attrs}><button class="{klass}" type="submit">{esc(label)}</button></form>'


def admin_fact(label: str, value: Any) -> str:
    if value in (None, ""):
        return ""
    return f"<dt>{esc(label)}</dt><dd>{esc(value)}</dd>"


def format_bytes(value: Any) -> str:
    size = int_value(value)
    if size <= 0:
        return "0 B"
    units = ["B", "KB", "MB", "GB"]
    amount = float(size)
    unit = units[0]
    for unit in units:
        if amount < 1024 or unit == units[-1]:
            break
        amount /= 1024
    return f"{amount:.1f} {unit}" if unit != "B" else f"{int(amount)} B"


def admin_form(meta: Table, row: dict[str, Any], repo: Repository | None = None) -> str:
    if meta.name == "site_settings":
        return admin_site_settings_form(meta, row)
    if meta.name == "global_settings":
        return admin_global_settings_form(meta, row)
    if meta.name == "navigation_items":
        return admin_navigation_form(meta, row)
    if meta.name == "profiles":
        return admin_profile_form(meta, row)
    if meta.name == "research_interests":
        return admin_research_interest_form(meta, row)
    if meta.name == "projects":
        return admin_project_form(meta, row)
    if meta.name == "patents":
        return admin_patent_form(meta, row, repo)
    if meta.name == "students":
        return admin_student_form(meta, row)
    if meta.name == "student_category_displays":
        return admin_student_category_form(meta, row)
    if meta.name == "news":
        return admin_news_form(meta, row)
    if meta.name == "courses":
        return admin_course_form(meta, row)
    if meta.name == "messages":
        return admin_message_form(meta, row)
    if meta.name == "publications":
        return admin_publication_form(meta, row, repo)
    if meta.name == "translation_cache":
        return admin_translation_cache_form(meta, row)
    if meta.name == "media_assets":
        return admin_media_asset_form(meta, row)
    if meta.name == "auth_users":
        return admin_auth_user_form(meta, row, repo)
    if meta.name == "auth_roles":
        return admin_auth_role_form(meta, row)
    if meta.name == "auth_permissions":
        return admin_auth_permission_form(meta, row, repo)
    fields = []
    for field in meta.fields:
        value = row.get(field.name, "")
        fields.append(admin_field_label(field, value))
    return f'<form class="edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(fields)}{admin_form_actions(meta.name)}</form>'


def admin_auth_user_form(meta: Table, row: dict[str, Any], repo: Repository | None = None) -> str:
    field_map = {field.name: field for field in meta.fields}
    roles = (repo.list("auth_roles", Query(limit=1000, order_by="sort_order", descending=False)) if repo else [])
    role_options = "".join(f'<option value="{esc(role.get("uid"))}" {"selected" if str(row.get("role_uid") or "") == str(role.get("uid") or "") else ""}>{esc(role.get("name") or role.get("uid"))}</option>' for role in roles)
    password_hint = "留空则保留原密码；新增账号时请设置不少于 10 位的密码。"
    sections = [
        ("基础身份", [
            admin_field_label(field_map["uid"], row.get("uid", ""), "账号稳定标识，用于会话和权限关联。"),
            admin_field_label(field_map["username"], row.get("username", ""), "登录账号，建议使用邮箱、工号或便于审计的账号名。"),
            admin_field_label(field_map["display_name"], row.get("display_name", ""), "后台右侧和审计提示中显示的姓名。"),
            admin_field_label(field_map["email"], row.get("email", ""), "用于后续通知、找回或人工核验。"),
        ]),
        ("角色与状态", [
            f'<label><span>角色 UID</span><select name="role_uid">{role_options}</select><small class="field-help">决定后台功能权限和前台可见范围访问层级。</small></label>',
            admin_field_label(field_map["status"], row.get("status", "active"), "停用后该账号不能登录。"),
            admin_field_label(field_map["visibility"], row.get("visibility", "authenticated"), "账号自身记录的可见范围，一般保持 authenticated 或 owner。"),
            admin_field_label(field_map["must_change_password"], row.get("must_change_password", 0), "预留字段：后续可用于强制首次登录修改密码。"),
        ]),
        ("密码安全", [
            f'<label class="auth-wide-field"><span>设置新密码</span><input type="password" name="new_password" autocomplete="new-password" minlength="10" placeholder="留空则不修改当前密码"><small class="field-help">{esc(password_hint)}</small></label>',
            f'<label><span>最后登录</span><input value="{esc(row.get("last_login_at") or "暂无")}" readonly><small class="field-help">系统自动记录，只读。</small></label>',
        ]),
    ]
    hidden = f'<input type="hidden" name="password_hash" value="{esc(row.get("password_hash") or "")}"><input type="hidden" name="last_login_at" value="{esc(row.get("last_login_at") or "")}">'
    return f'<form class="edit-form auth-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{auth_form_sections(sections)}{hidden}{admin_form_actions(meta.name)}</form>'


def admin_auth_role_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    help_map = {
        "uid": "角色稳定标识。系统默认角色建议不要改 UID，以免已有账号失去关联。",
        "name": "后台展示的角色名称，例如高级管理员、普通管理员、员工、访客用户。",
        "level": "层级越高权限越大；100 默认视为超级权限，40 及以上可访问 staff 可见内容。",
        "description": "说明该角色适合谁使用，方便多人维护。",
        "visibility_scopes": "前台可访问的可见范围，多个值用英文逗号分隔，例如 public,authenticated,staff。",
        "is_system": "系统角色标记。建议保留默认角色，新增自定义角色时关闭。",
        "is_active": "关闭后该角色下账号不能通过权限校验。",
        "sort_order": "后台显示排序，数字越小越靠前。",
    }
    sections = [
        ("角色层级", [admin_field_label(field_map[name], row.get(name, ""), help_map.get(name, "")) for name in ("uid", "name", "level", "sort_order")]),
        ("前台可见范围", [admin_field_label(field_map["visibility_scopes"], row.get("visibility_scopes", ""), help_map["visibility_scopes"], textarea_rows=3), admin_field_label(field_map["description"], row.get("description", ""), help_map["description"], textarea_rows=3)]),
        ("系统状态", [admin_field_label(field_map["is_system"], row.get("is_system", 0), help_map["is_system"]), admin_field_label(field_map["is_active"], row.get("is_active", 1), help_map["is_active"])]),
    ]
    return f'<form class="edit-form auth-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{auth_form_sections(sections)}{admin_form_actions(meta.name)}</form>'


def admin_auth_permission_form(meta: Table, row: dict[str, Any], repo: Repository | None = None) -> str:
    field_map = {field.name: field for field in meta.fields}
    roles = (repo.list("auth_roles", Query(limit=1000, order_by="sort_order", descending=False)) if repo else [])
    role_options = "".join(f'<option value="{esc(role.get("uid"))}" {"selected" if str(row.get("role_uid") or "") == str(role.get("uid") or "") else ""}>{esc(role.get("name") or role.get("uid"))}</option>' for role in roles)
    module_options = "".join(f'<option value="{esc(module)}" {"selected" if str(row.get("module") or "") == module else ""}>{esc(module_label(module))}</option>' for module in admin_modules())
    sections = [
        ("规则归属", [
            admin_field_label(field_map["uid"], row.get("uid", ""), "角色权限稳定标识，保存时会按角色和模块自动生成，通常无需手动维护。"),
            f'<label><span>角色 UID</span><select name="role_uid">{role_options}</select><small class="field-help">该条规则属于哪个角色。</small></label>',
            f'<label class="auth-wide-field"><span>功能模块</span><select name="module">{module_options}</select><small class="field-help">控制后台入口、导入与导出或具体数据表的访问。</small></label>',
            admin_field_label(field_map["sort_order"], row.get("sort_order", 0), "权限规则列表排序。"),
        ]),
        ("权限开关", [
            admin_field_label(field_map["can_view"], row.get("can_view", 0), "允许打开概览和详情页。"),
            admin_field_label(field_map["can_create"], row.get("can_create", 0), "允许新增该类型数据。"),
            admin_field_label(field_map["can_edit"], row.get("can_edit", 0), "允许保存修改、批量修改和常用快速修改。"),
            admin_field_label(field_map["can_delete"], row.get("can_delete", 0), "允许删除、清空回收站或执行危险操作。"),
            admin_field_label(field_map["can_export"], row.get("can_export", 0), "允许导出该模块或整站数据。"),
        ]),
    ]
    return f'<form class="edit-form auth-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{auth_form_sections(sections)}{admin_form_actions(meta.name)}</form>'


def auth_form_sections(sections: list[tuple[str, list[str]]]) -> str:
    return "".join(f'<fieldset class="form-section auth-form-section"><legend>{esc(title)}</legend>{"".join(fields)}</fieldset>' for title, fields in sections)


def module_label(module: str) -> str:
    if module in AUTH_MODULE_LABELS:
        return f"{AUTH_MODULE_LABELS[module]} ({module})"
    table = TABLE_MAP.get(module)
    return f"{table.label} ({module})" if table else module


def admin_form_actions(table_name: str) -> str:
    return f"""<div class="edit-form-actions">
      <button type="submit" name="_action" value="save">保存</button>
      <button class="button secondary" type="submit" name="_action" value="save_continue">保存并继续</button>
      <a class="button ghost" href="/admin/table/{esc(table_name)}">返回</a>
    </div>"""


def admin_site_settings_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("site-basic", "基础名称", ["uid", "is_active", "site_name", "site_name_en"]),
        ("site-home", "首页主视觉", ["hero_title", "hero_subtitle", "homepage_profile_uid", "homepage_publication_limit", "homepage_news_limit"]),
        ("site-brand", "品牌与分享媒体", ["logo_key", "favicon_key", "og_image_key"]),
        ("site-seo", "SEO 与检索", ["seo_title", "seo_description", "seo_keywords"]),
        ("site-footer", "页脚", ["footer_text"]),
    ]
    help_map = {
        "uid": "站点设置稳定标识。默认 site-default，建议保持不变，便于 Cloudflare D1 与 Ubuntu 数据迁移。",
        "is_active": "启用后作为当前站点设置。未来多站点配置时，可保留备用配置但仅启用一个主配置。",
        "site_name": "中文前台顶部品牌名、页面标题和页脚默认名称。",
        "site_name_en": "英文模式优先使用的站点名称；留空时英文前台会从翻译缓存或中文名兜底。",
        "hero_title": "首页第一屏主标题，建议保持短句，突出研究方向或团队定位。",
        "hero_subtitle": "首页简介，可作为教师简介缺失时的首页说明，也会进入英文翻译缓存。",
        "homepage_profile_uid": "首页主展示教师或负责人 UID，可在“教师与团队”列表中查看。",
        "homepage_publication_limit": "首页展示论文数量。数值越小越利于 Worker 低 CPU 请求。",
        "homepage_news_limit": "首页展示动态数量。建议保持 3-6 条，避免首页过长。",
        "logo_key": "顶部或未来品牌位使用的 Logo 媒体 key，可选择媒体库、上传或直接输入公开 URL。",
        "favicon_key": "浏览器标签页图标媒体 key，建议使用正方形小图标。",
        "og_image_key": "社交分享图媒体 key，建议使用横向封面图。",
        "seo_title": "搜索引擎页面标题；留空时使用站点名称。",
        "seo_description": "搜索引擎摘要描述，也会写入前台 meta description。",
        "seo_keywords": "SEO 关键词，多个关键词用逗号分隔。",
        "footer_text": "前台底部内容，可填写纯文本或安全 HTML；脚本、事件属性和危险链接会被过滤。",
    }
    nav_items = []
    sections = []
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            extra_class = "site-wide-field" if name in {"hero_subtitle", "seo_description", "seo_keywords", "footer_text"} else ""
            rows = 5 if name == "footer_text" else 3 if name in {"hero_subtitle", "seo_description"} else 2 if name == "seo_keywords" else None
            if field.kind == "bool":
                labels.append(admin_switch_field(field, value, help_map.get(name, "")))
            else:
                labels.append(admin_field_label(field, value, help_map.get(name, ""), extra_class=extra_class, textarea_rows=rows))
        sections.append(f'<fieldset class="form-section site-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form site-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save"><section class="site-edit-sticky"><nav class="site-edit-nav">{"".join(nav_items)}</nav></section>{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_global_settings_form(meta: Table, row: dict[str, Any]) -> str:
    row = global_settings_translation_defaults(row)
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("global-access", "访问与留言", ["uid", "allow_public_registration", "allow_anonymous_messages", "notify_email"]),
        ("global-upload", "上传与媒体", ["upload_max_size_mb", "upload_allowed_extensions", "media_trash_retention_days"]),
        ("global-news-pdf", "动态 PDF", ["news_pdf_engine", "news_pdf_allow_download", "news_pdf_watermark"]),
        ("global-translation", "翻译服务", ["translation_provider", "translation_providers", "translation_batch_size", "translation_worker_count", "translation_timeout_seconds"]),
        ("global-translation-keys", "翻译接口配置", ["libretranslate_url", "libretranslate_api_key", "deepl_api_key", "google_translate_api_key", "microsoft_translator_key", "microsoft_translator_region", "microsoft_translator_endpoint", "mymemory_email"]),
        ("global-academic", "论文与专利平台", ["publication_metadata_provider", "publication_metadata_providers", "publication_display_style", "patent_metadata_providers", "patentsview_api_key", "epo_ops_client_id", "epo_ops_client_secret"]),
        ("global-cache", "缓存与性能", ["publication_suggestion_cache_seconds", "profile_suggestion_cache_seconds", "project_suggestion_cache_seconds", "patent_suggestion_cache_seconds", "student_suggestion_cache_seconds", "news_suggestion_cache_seconds", "course_suggestion_cache_seconds"]),
        ("global-system", "系统状态", ["translation_job_state"]),
    ]
    help_map = {
        "uid": "全站通用设置的稳定标识。通常保持 global-default，便于导出迁移和 Worker D1 同步。",
        "allow_public_registration": "是否允许访客公开注册。当前站点以展示为主，通常保持关闭。",
        "allow_anonymous_messages": "是否允许未登录访客提交联系留言。招生咨询场景通常开启。",
        "notify_email": "后台通知邮箱，可用于未来留言提醒、导入失败提醒等。",
        "upload_max_size_mb": "后台上传单个文件大小上限。为兼容低 CPU 场景，当前程序级安全上限为 10MB。",
        "upload_allowed_extensions": "允许上传的扩展名，多个用逗号分隔，例如 .jpg,.png,.pdf,.docx。",
        "media_trash_retention_days": "媒体进入回收站后的保留天数；打开媒体库时会清理超期记录。",
        "news_pdf_engine": "动态 PDF 预览引擎。native 更轻量；pdfjs 可作为未来增强预览方式。",
        "news_pdf_allow_download": "是否允许访客下载动态或课程中关联的 PDF 文件。",
        "news_pdf_watermark": "PDF 水印文字，留空则不加水印；后续 PDF 处理可复用。",
        "translation_provider": "默认自动翻译服务。auto 会按启用平台和可用配置自动选择。",
        "translation_providers": "启用翻译服务顺序，用逗号分隔。默认使用 auto,mymemory,argos_local；配置了 URL 或 API Key 后再加入 LibreTranslate、DeepL、Google、Microsoft。",
        "libretranslate_url": "LibreTranslate 服务地址。免密公开实例可能有频率限制，自建实例更稳定。",
        "libretranslate_api_key": "LibreTranslate API Key；没有密钥的实例可留空。",
        "deepl_api_key": "DeepL Free/API Key。未填写时 DeepL 不会进入可用源。",
        "google_translate_api_key": "Google Translate API Key。未填写时 Google 不会进入可用源。",
        "microsoft_translator_key": "Bing/Microsoft Translator Key。未填写时 Microsoft 不会进入可用源。",
        "microsoft_translator_region": "Microsoft Translator 资源区域，例如 eastasia、global。",
        "microsoft_translator_endpoint": "Microsoft Translator endpoint，默认使用 Azure 官方翻译 API 地址。",
        "mymemory_email": "MyMemory 可选邮箱，用于服务方额度识别；留空时会优先复用通知邮箱。请使用真实可联系邮箱，不生成虚假邮箱。",
        "translation_batch_size": "每次自动翻译任务处理的最大条数。Worker 中会自动限制为小批量，并尽量把同源字段打包成一次请求。",
        "translation_worker_count": "本地/Ubuntu 自动翻译并发数。Cloudflare Worker 会强制单线程，但仍会使用小批量打包减少请求次数。",
        "translation_timeout_seconds": "单个翻译请求超时时间，避免外部平台卡住后台操作。",
        "publication_metadata_provider": "旧版单选论文元数据平台配置，保留用于兼容。",
        "publication_metadata_providers": "论文联网查验平台，逗号分隔，如 crossref,openalex,semanticscholar。",
        "publication_display_style": "前台论文页默认显示的引用格式；复制格式仍可由访客在前台切换。",
        "patent_metadata_providers": "专利联网查验平台，逗号分隔，如 patentsview,epo_ops。",
        "patentsview_api_key": "PatentsView API Key。免费平台策略可能变化，未配置时仅尝试免密能力。",
        "epo_ops_client_id": "EPO OPS Client ID，需要账号时填写。",
        "epo_ops_client_secret": "EPO OPS Client Secret，需要账号时填写。",
        "publication_suggestion_cache_seconds": "论文编辑页历史填法提示缓存秒数，降低重复扫描。",
        "profile_suggestion_cache_seconds": "教师与团队编辑页历史填法提示缓存秒数。",
        "project_suggestion_cache_seconds": "项目编辑页历史填法提示缓存秒数。",
        "patent_suggestion_cache_seconds": "专利编辑页历史填法提示缓存秒数。",
        "student_suggestion_cache_seconds": "学生编辑页历史填法提示缓存秒数。",
        "news_suggestion_cache_seconds": "动态编辑页历史填法提示缓存秒数。",
        "course_suggestion_cache_seconds": "课程编辑页历史填法提示缓存秒数。",
        "translation_job_state": "自动翻译任务的持久状态，由系统维护；通常不需要手动修改。",
    }
    nav_items = []
    sections = []
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            extra_class = "global-wide-field" if name in {"translation_providers", "upload_allowed_extensions", "translation_job_state"} else ""
            rows = 3 if name in {"translation_providers", "upload_allowed_extensions"} else 5 if name == "translation_job_state" else None
            attrs = 'readonly aria-readonly="true"' if name == "translation_job_state" else ""
            if field.kind == "bool":
                labels.append(admin_switch_field(field, value, help_map.get(name, "")))
            else:
                labels.append(admin_field_label(field, value, help_map.get(name, ""), extra_class=extra_class, textarea_rows=rows, control_attrs=attrs))
        sections.append(f'<fieldset class="form-section global-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form global-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save"><section class="global-edit-sticky"><nav class="global-edit-nav">{"".join(nav_items)}</nav></section>{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_switch_field(field: Any, value: Any, help_text: str = "") -> str:
    checked = " checked" if truthy(value, default=False) else ""
    help_html = f'<small class="field-help" title="{esc(help_text)}">{esc(help_text)}</small>' if help_text else ""
    return f"""<label class="field-{esc(field.name)} admin-switch-field">
      <span class="field-label">{esc(field.label)}</span>
      <span class="admin-switch-control">
        <input type="hidden" name="{esc(field.name)}" value="0">
        <input type="checkbox" name="{esc(field.name)}" value="1"{checked}>
        <span></span><em></em>
      </span>
      {help_html}
    </label>"""


def admin_media_asset_form(meta: Table, row: dict[str, Any]) -> str:
    key = normalize_media_key(str(row.get("object_key") or ""))
    exists = media_file_exists(row, {})
    state = media_row_state(row, {})
    preview = media_preview(row, {})
    url = media_url(key)
    file_info = [
        ("文件状态", "文件存在" if exists else "文件缺失"),
        ("存储位置", media_storage_label(row)),
        ("访问路径", url or "无"),
        ("记录 ID", row.get("id") or ""),
        ("稳定标识", row.get("uid") or ""),
        ("大小", format_bytes(int_value(row.get("size"))) if int_value(row.get("size")) else "待检测"),
        ("校验值", row.get("checksum") or "未记录"),
    ]
    readonly = "".join(admin_readonly_field(label, value) for label, value in file_info)
    return f"""<form class="edit-form media-asset-edit-form" method="post" action="/admin/table/media_assets/save">
      <section class="form-section media-edit-preview-section">
        <div class="media-edit-preview-card">
          <a class="media-edit-preview {'is-missing' if not exists else ''}" href="{esc(url)}" target="_blank" rel="noreferrer">{preview}</a>
          <div class="media-edit-preview-actions">
            <button class="button light media-edit-current-trigger" type="button" data-media-key="{esc(key)}" data-media-url="{esc(url)}" data-media-uid="{esc(row.get("uid") or row.get("id"))}" data-media-title="{esc(row.get("title") or key)}" {'disabled' if not exists or not media_editable_image(key) else ''}>编辑当前图片</button>
            <a class="button ghost" href="{esc(url)}" target="_blank" rel="noreferrer">新标签查看</a>
          </div>
          <p class="admin-muted">图片编辑可保存为新媒体，也可替换当前源文件；非图片文件仅支持查看和修改记录信息。</p>
        </div>
        <div class="media-edit-readonly">{readonly}</div>
      </section>
      <section class="form-section media-edit-main-section">
        <h2>基础信息</h2>
        <input type="hidden" name="uid" value="{esc(row.get("uid") or "")}">
        <input type="hidden" name="id" value="{esc(row.get("id") or "")}">
        {admin_field_label(field_by_name(meta, "object_key"), key, "保存为新文件后会自动写入新的 object_key；替换源文件时该值保持不变。", "field-object_key", control_attrs='data-media-current-key')}
        {admin_field_label(field_by_name(meta, "storage_kind"), row.get("storage_kind") or media_storage_kind(row), "static 表示随 public/media 静态资源部署；local 表示 Ubuntu/本地运行时 media 目录；r2 表示 Cloudflare R2；external 表示公开外链。")}
        {admin_field_label(field_by_name(meta, "title"), row.get("title") or "", "用于后台列表和媒体选择工具显示。")}
        {admin_field_label(field_by_name(meta, "category"), row.get("category") or "", "建议按用途填写，如 profile、icons、students、news-cover。")}
        {admin_field_label(field_by_name(meta, "mime_type"), row.get("mime_type") or "", "由上传或图片编辑工具自动维护；必要时可手动修正。")}
        {admin_field_label(field_by_name(meta, "status"), row.get("status") or "active", "active 会进入媒体库；trash 会进入回收站。")}
      </section>
      <section class="form-section media-edit-system-section">
        <h2>系统字段</h2>
        {admin_readonly_input("size", row.get("size") or "", "大小")}
        {admin_readonly_input("checksum", row.get("checksum") or "", "校验值")}
      </section>
      {admin_form_actions("media_assets")}
    </form>"""


def field_by_name(meta: Table, name: str) -> Any:
    for field in meta.fields:
        if field.name == name:
            return field
    raise KeyError(name)


def admin_readonly_field(label: str, value: Any) -> str:
    return f'<div class="readonly-fact"><span>{esc(label)}</span><strong title="{esc(value)}">{esc(value)}</strong></div>'


def admin_readonly_input(name: str, value: Any, label: str) -> str:
    return f'<label class="field-{esc(name)}"><span class="field-label">{esc(label)}</span><input type="text" name="{esc(name)}" value="{esc(value)}" readonly aria-readonly="true"><small class="field-help">系统自动维护，不建议手动修改。</small></label>'


def admin_field_label(field: Any, value: Any, help_text: str = "", extra_class: str = "", textarea_rows: int | None = None, control_attrs: str = "") -> str:
    control = admin_field_control(field, value, textarea_rows, control_attrs)
    help_html = f'<small class="field-help" title="{esc(help_text)}">{esc(help_text)}</small>' if help_text else ""
    classes = " ".join(item for item in [extra_class, f"field-{field.name}"] if item)
    return f'<label class="{esc(classes)}"><span class="field-label">{esc(field.label)}</span>{control}{help_html}</label>'


def admin_field_control(field: Any, value: Any, textarea_rows: int | None = None, control_attrs: str = "") -> str:
    if textarea_rows or field.kind == "textarea":
        rows = textarea_rows or 5
        return f'<textarea name="{esc(field.name)}" rows="{rows}" {control_attrs}>{esc(value)}</textarea>'
    if field.kind == "select":
        return f'<select name="{esc(field.name)}" {control_attrs}>' + options(list(field.choices), str(value)) + "</select>"
    input_type = "number" if field.kind in {"number", "bool"} else "date" if field.kind == "date" else "url" if field.kind == "url" else "email" if field.kind == "email" else "text"
    input_html = f'<input type="{input_type}" name="{esc(field.name)}" value="{esc(value)}" {control_attrs}>'
    if field.kind == "file" or field.name in {"icon", "logo_key", "favicon_key", "og_image_key"}:
        suggestion = admin_media_suggestion(field.name)
        return f"""<div class="media-input-control">
          {admin_media_field_preview(value)}
          <div class="media-input-side">
            <div class="media-input-row">{input_html}<button class="button light media-picker-trigger" type="button" data-media-target="{esc(field.name)}" data-media-purpose="{esc(field.name)}" title="从媒体库选择、上传或裁剪生成新媒体">选择媒体</button></div>
            {admin_media_field_meta(value)}
            <div class="media-suggestion" title="{esc(suggestion)}">{esc(suggestion)}</div>
          </div>
        </div>"""
    return input_html


def admin_media_suggestion(field_name: str) -> str:
    suggestions = {
        "icon": "建议：SVG/PNG，正方形 64x64 或 128x128，比例 1:1。",
        "logo_key": "建议：透明 PNG/SVG，横向 320x96，比例约 10:3；深浅背景都要清晰。",
        "favicon_key": "建议：PNG/ICO/SVG，64x64 或 128x128，比例 1:1。",
        "og_image_key": "建议：分享图 1200x630，比例 1.91:1，文字不要贴边。",
        "avatar_key": "建议：人物照片 480x600 或 600x750，比例 4:5；主体居中。",
        "cover_key": "建议：封面图 1200x675，比例 16:9；适合动态列表和详情页裁切。",
        "pdf_key": "建议：PDF 文件，小于上传上限；文件名使用论文简称或年份。",
        "certificate_key": "建议：证书/专利扫描图 1200x900，比例 4:3，或 PDF 原件。",
        "syllabus_key": "建议：PDF/DOCX，文件名包含课程名和年份。",
        "material_key": "建议：PDF/PPT/DOCX/ZIP，文件名包含课程名、章节或版本。",
        "attachment_key": "建议：PDF/DOCX/图片，文件名清楚标识留言或材料用途。",
    }
    return suggestions.get(field_name, "建议：按当前用途选择清晰媒体；图片尽量使用 WebP/PNG/JPG，文件名简短可读。")


def admin_media_field_preview(value: Any) -> str:
    text = text_only(value, 800).strip()
    preview = admin_media_preview_markup(text)
    classes = "media-field-preview" if text else "media-field-preview is-empty"
    href = admin_media_preview_url(text) or "#"
    return f"""<div class="{classes}" data-media-preview>
      <a class="media-field-thumb" href="{esc(href)}" target="_blank" rel="noreferrer" data-media-preview-link>{preview}</a>
    </div>"""


def admin_media_field_meta(value: Any) -> str:
    text = text_only(value, 800).strip()
    label = text if text else "尚未选择媒体"
    return f"""<div class="media-field-meta">
      <strong data-media-preview-kind>{esc(admin_media_preview_kind(text))}</strong>
      <span title="{esc(label)}" data-media-preview-key>{esc(label)}</span>
    </div>"""


def admin_media_preview_markup(value: str) -> str:
    if not value:
        return '<span class="media-field-empty">无</span>'
    url = admin_media_preview_url(value)
    kind = media_preview_type(value)
    if kind == "image":
        return f'<img src="{esc(url)}" alt="媒体预览" loading="lazy" decoding="async" style="width:100%;height:100%;object-fit:contain;object-position:center center;">'
    if kind == "video":
        return f'<video src="{esc(url)}" muted playsinline preload="metadata" style="width:100%;height:100%;object-fit:contain;object-position:center center;"></video>'
    return f'<span class="media-field-file">{esc(admin_media_preview_extension(value))}</span>'


def admin_media_preview_url(value: str) -> str:
    text = text_only(value, 800).strip()
    if not text:
        return ""
    if text.startswith(("http://", "https://", "/")):
        return safe_href(text)
    return media_url(text)


def admin_media_preview_kind(value: str) -> str:
    if not value:
        return "未选择"
    kind = media_preview_type(value)
    if kind == "image":
        return "图片预览"
    if kind == "video":
        return "视频预览"
    return f"{admin_media_preview_extension(value)} 文件"


def media_preview_type(value: str) -> str:
    suffix = Path(urlparse(value).path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg", ".avif", ".bmp"}:
        return "image"
    if suffix in {".mp4", ".webm", ".ogg", ".mov", ".m4v"}:
        return "video"
    return "file"


def media_editable_image(value: str) -> bool:
    suffix = Path(urlparse(value).path).suffix.lower()
    return suffix in {".jpg", ".jpeg", ".png", ".webp"}


def admin_media_preview_extension(value: str) -> str:
    suffix = Path(urlparse(value).path).suffix.lower().lstrip(".")
    return (suffix or "file").upper()[:8]


def admin_navigation_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("基础信息", ["uid", "title", "title_en", "kind", "enabled", "sort_order"]),
        ("链接与显示位置", ["path", "url_name", "fragment", "location", "visibility"]),
        ("视觉样式", ["icon", "style"]),
    ]
    help_map = {
        "uid": "稳定标识用于数据迁移和更新，创建后尽量不要频繁修改。",
        "title": "前台中文导航文字或按钮文字。",
        "title_en": "英文模式下显示的导航文字，留空时使用中文标题。",
        "kind": "route 表示站内页面，external 表示外部链接，anchor 表示页面锚点，button 表示强调按钮。",
        "path": "站内路径如 /team、/publications；外链可填写 https://...。",
        "url_name": "可记录语义化路由名，供后续自动生成路径或迁移时识别。",
        "fragment": "页面锚点，不含 #，例如 publications；通常与 path 组合跳转到页面局部。",
        "icon": "可填写媒体 key、/media/...、https://... 或 Iconify 等公开图标路径；也可点“选择/上传”。",
        "style": "link 为普通链接，primary/secondary/ghost 用于按钮外观。",
        "location": "header 显示在顶部导航，home_hero 显示在首页主按钮区，footer 显示在页脚，admin_sidebar 显示在后台侧栏。",
        "visibility": "public 为公开可见；staff/private 可为后续权限扩展保留。",
        "enabled": "1 表示启用，0 表示停用；列表页也可以一键切换。",
        "sort_order": "数字越小越靠前；列表页按此字段从小到大显示。",
    }
    sections = []
    for title, names in groups:
        labels = []
        for name in names:
            field = field_map.get(name)
            if field:
                labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, "")))
        sections.append(f'<fieldset class="form-section nav-form-section"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form nav-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_profile_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("profile-identity", "身份与照片", ["uid", "name", "name_en", "role", "title", "organization", "lab", "avatar_key"]),
        ("profile-bio", "个人简介", ["bio", "bio_en"]),
        ("profile-contact", "联系方式", ["email", "phone", "office", "contact_visibility"]),
        ("profile-links", "学术链接", ["personal_homepage", "orcid", "google_scholar", "dblp", "github", "cnki"]),
        ("profile-career", "经历与招生", ["education", "experience", "recruiting"]),
        ("profile-display", "展示与排序", ["visibility", "is_active", "is_featured", "sort_order"]),
    ]
    help_map = {
        "uid": "稳定标识用于成员详情页路径和数据迁移，创建后尽量不要频繁修改。",
        "avatar_key": "成员照片媒体 key，可从媒体库选择、上传或裁剪；前台无图时显示姓名姓氏。",
        "bio": "前台个人详情页会优先展示个人简介，团队列表中仅显示前两行摘要。",
        "bio_en": "英文前台优先使用的个人简介。留空时英文模式会读取翻译缓存；仍没有缓存时回退显示中文简介。",
        "contact_visibility": "控制联系方式可见范围，后续接入权限时使用。",
        "visibility": "控制该成员资料是否公开显示。",
        "is_active": "1 表示在前台和后台概览中启用，0 表示停用。",
        "is_featured": "1 表示可在首页或重点区域优先展示。",
        "sort_order": "数字越小越靠前，影响团队列表和后台概览排序。",
    }
    sections = []
    nav_items = []
    suggest_fields = {"role", "title", "organization", "lab", "office"}
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if field:
                attrs = profile_suggestion_attrs(name) if name in suggest_fields else ""
                labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), control_attrs=attrs))
        sections.append(f'<fieldset class="form-section profile-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form profile-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save"><section class="profile-edit-sticky"><nav class="profile-edit-nav">{"".join(nav_items)}</nav></section>{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def profile_suggestion_attrs(name: str) -> str:
    return f'data-profile-suggest="{esc(name)}" autocomplete="off"'


def admin_research_interest_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("基础设置", ["uid", "visibility", "sort_order"]),
        ("方向名称", ["name", "name_en"]),
        ("方向描述", ["description"]),
    ]
    help_map = {
        "uid": "稳定标识用于首页研究方向标签、数据迁移和后续引用，创建后尽量不要频繁修改。",
        "name": "中文模式下显示的研究方向名称，首页会以标签形式展示。",
        "name_en": "英文模式下优先显示的研究方向名称，留空时回退到中文名称。",
        "description": "方向简介会作为首页研究方向标签的悬停说明，也可供后续研究方向详情页复用。",
        "sort_order": "数字越小越靠前，影响首页研究方向和后台概览顺序。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
    }
    sections = []
    for title, names in groups:
        labels = []
        for name in names:
            field = field_map.get(name)
            if field:
                labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, "")))
        sections.append(f'<fieldset class="form-section research-form-section"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form research-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_project_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("proj-basic", "基础信息", ["uid", "name", "project_number"]),
        ("proj-fund", "来源与经费", ["source", "fund_name", "amount"]),
        ("proj-people", "负责人和成员", ["principal", "members"]),
        ("proj-time", "时间与状态", ["start_date", "end_date", "status"]),
        ("proj-content", "项目简介", ["summary"]),
        ("proj-display", "展示与排序", ["visibility", "is_featured", "sort_order"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、数据迁移和前台引用；新增项目时会自动生成，可按需要修改。",
        "name": "项目名称，前台列表和后台搜索都会优先读取。",
        "project_number": "项目编号、合同号或任务书编号。",
        "source": "项目来源，例如国家自然科学基金、科技厅、企业合作等；支持历史填法提示。",
        "fund_name": "基金/计划名称，例如面上项目、重点研发计划、横向课题等；支持历史填法提示。",
        "amount": "项目经费金额，可填写数字，单位可在项目名称或简介中说明。",
        "principal": "项目负责人；支持从历史项目负责人中选择。",
        "members": "项目成员，多个成员可用逗号、分号或换行分隔；支持历史成员提示。",
        "start_date": "开始日期，可从日期控件选择，也可直接输入 YYYY-MM-DD。",
        "end_date": "结束日期，可从日期控件选择，也可直接输入 YYYY-MM-DD。",
        "status": "项目状态，例如在研、结题、申请中、暂停；支持历史填法提示。",
        "summary": "项目简介，前台列表会显示摘要，后续详情页可复用。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
        "is_featured": "1 表示首页展示，0 表示仅在项目页展示。",
        "sort_order": "排序一般自动使用记录 ID，通常不用调整；需要人工控制时可直接修改。",
    }
    nav_items = []
    sections = []
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            attrs = project_suggestion_attrs(name) if name in {"source", "fund_name", "status", "principal", "members"} else ""
            if name == "sort_order":
                value = row.get(name) if text_only(row.get(name), 40).strip() else row.get("id", "")
                attrs = 'placeholder="保存留空时自动使用记录 ID"'
            rows = 2 if name in {"members", "summary"} else None
            if name in {"name", "project_number"}:
                labels.append(project_duplicate_field(field, value, help_map.get(name, ""), name))
            else:
                labels.append(admin_field_label(field, value, help_map.get(name, ""), control_attrs=attrs, textarea_rows=rows))
        sections.append(f'<fieldset class="form-section project-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form project-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save"><section class="project-edit-sticky"><nav class="project-edit-nav">{"".join(nav_items)}</nav></section>{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def project_suggestion_attrs(name: str) -> str:
    return f'data-project-suggest="{esc(name)}" autocomplete="off"'


def project_duplicate_field(field: Any, value: Any, help_text: str, name: str) -> str:
    label = admin_field_label(field, value, help_text, f"project-{name}-field")
    return f"""<div class="publication-inline-lookup project-inline-lookup project-inline-lookup-{esc(name)}">
      {label}
      <div class="publication-field-actions">
        <button class="button light" type="button" data-project-check-duplicates data-project-result-target="{esc(name)}">查重</button>
      </div>
      <div class="publication-field-result project-field-result project-field-result-{esc(name)}" data-project-field-result="{esc(name)}"></div>
    </div>"""


def admin_patent_form(meta: Table, row: dict[str, Any], repo: Repository | None = None) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("pat-basic", "基础信息", ["uid", "name", "country", "patent_type"]),
        ("pat-numbers", "申请与授权", ["application_number", "application_date", "grant_number", "grant_date", "legal_status"]),
        ("pat-people", "人员与权属", ["inventors", "owner"]),
        ("pat-content", "简介与证书", ["summary", "certificate_key"]),
        ("pat-display", "展示与排序", ["visibility", "is_featured", "sort_order"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、数据迁移和前台引用；新增时会自动生成，可按需要修改。",
        "name": "专利、软著或成果名称，前台和后台搜索都会优先读取。",
        "country": "国别或地区，例如中国、美国、PCT；支持历史填法提示。",
        "patent_type": "成果类型，例如发明专利、实用新型、外观设计、软件著作权；支持历史填法提示。",
        "application_number": "申请号或登记申请编号。",
        "grant_number": "授权号、公告号或登记号。",
        "application_date": "申请日期，可从日期控件选择，也可直接输入 YYYY-MM-DD。",
        "grant_date": "授权/登记日期，可从日期控件选择，也可直接输入 YYYY-MM-DD。",
        "inventors": "发明人、作者或完成人，多个成员可用逗号、分号或换行分隔；支持历史提示。",
        "owner": "权利人、著作权人或申请人；支持历史填法提示。",
        "legal_status": "申请中、已授权、已登记、公开、驳回、转让等状态；支持历史填法提示。",
        "summary": "成果简介，前台列表可显示摘要。",
        "certificate_key": "证书、公告文件或截图的媒体 key，可从媒体库选择或上传。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
        "is_featured": "1 表示首页展示，0 表示仅在专利页展示。",
        "sort_order": "排序一般自动使用记录 ID，通常不用调整；需要人工控制时可直接修改。",
    }
    nav_items = []
    sections = []
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            attrs = patent_suggestion_attrs(name) if name in {"country", "patent_type", "legal_status", "inventors", "owner"} else ""
            if name == "sort_order":
                value = row.get(name) if text_only(row.get(name), 40).strip() else row.get("id", "")
                attrs = 'placeholder="保存留空时自动使用记录 ID"'
            rows = 2 if name in {"inventors", "summary"} else None
            if name in {"name", "application_number", "grant_number"}:
                labels.append(patent_lookup_field(field, value, help_map.get(name, ""), name))
            else:
                labels.append(admin_field_label(field, value, help_map.get(name, ""), control_attrs=attrs, textarea_rows=rows))
        sections.append(f'<fieldset class="form-section patent-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    settings = active_global(repo) if repo else {}
    return f'<form class="edit-form patent-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{patent_edit_tools("".join(nav_items), settings)}{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def patent_suggestion_attrs(name: str) -> str:
    return f'data-patent-suggest="{esc(name)}" autocomplete="off"'


def patent_lookup_field(field: Any, value: Any, help_text: str, name: str) -> str:
    label = admin_field_label(field, value, help_text, f"patent-{name}-field")
    return f"""<div class="publication-inline-lookup patent-inline-lookup patent-inline-lookup-{esc(name)}">
      {label}
      <div class="publication-field-actions">
        <button class="button light" type="button" data-patent-check-duplicates data-patent-result-target="{esc(name)}">查重</button>
        <button class="button light" type="button" data-patent-lookup-field="{esc(name)}" data-patent-result-target="{esc(name)}">修正</button>
      </div>
      <div class="publication-field-result patent-field-result patent-field-result-{esc(name)}" data-patent-field-result="{esc(name)}"></div>
    </div>"""


def patent_edit_tools(nav_html: str, settings: dict[str, Any] | None = None) -> str:
    settings = settings or {}
    providers = parse_patent_platforms(settings.get("patent_metadata_providers") or "patentsview,epo_ops")
    provider_options = "".join(f'<label><input type="checkbox" name="patent_metadata_platforms" value="{esc(provider)}" checked>{esc(patent_platform_label(provider))}</label>' for provider in providers)
    external_links = [
        ("国知局", "https://pss-system.cponline.cnipa.gov.cn", False),
        ("WIPO", "https://patentscope.wipo.int/search/en/search.jsf?queryString={query}", True),
        ("Google", "https://patents.google.com/?q={query}", True),
        ("百度", "https://www.baidu.com/s?wd={query}%20专利", True),
        ("Espacenet", "https://worldwide.espacenet.com/patent/search?q={query}", True),
        ("PatentsView", "https://search.patentsview.org/?q={query}", True),
    ]
    external_buttons = "".join(f'<button class="button light patent-external-button" type="button" data-patent-external-url="{esc(url)}" data-patent-external-query="{1 if has_query else 0}">{esc(label)}</button>' for label, url, has_query in external_links)
    return f"""<section class="publication-parse-panel patent-lookup-panel">
      <div class="publication-edit-tools patent-edit-tools">
        <div class="publication-tool-line">
          <div class="publication-platforms patent-platforms" aria-label="专利联网查验平台">{provider_options}</div>
        </div>
        <div class="patent-external-search">
          <span>外部检索</span>
          <div>{external_buttons}</div>
        </div>
        <p class="publication-tool-status admin-muted" data-patent-status>专利联网查验只在点击字段旁“修正”时运行；PatentsView/EPO OPS 的密钥可在通用设置中配置。</p>
      </div>
    </section>
    <section class="patent-edit-sticky">
      <nav class="patent-edit-nav">{nav_html}</nav>
    </section>"""


def patent_platform_label(value: str) -> str:
    return {"patentsview": "PatentsView", "epo_ops": "EPO OPS"}.get(value, value)


def admin_student_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("stu-basic", "基础身份", ["uid", "name", "name_en", "student_id", "avatar_key"]),
        ("stu-study", "学籍与分组", ["degree", "category", "grade", "status"]),
        ("stu-direction", "方向与简介", ["direction", "bio", "awards"]),
        ("stu-contact", "联系与主页", ["email", "homepage", "contact_visibility"]),
        ("stu-time", "时间与去向", ["enrollment_date", "graduation_date", "destination"]),
        ("stu-display", "展示与排序", ["visibility", "is_featured", "sort_order"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、前台引用和数据迁移；新增时会自动生成，可按需要修改。",
        "name": "学生中文姓名，前台学生页和后台检索都会优先读取。",
        "name_en": "英文模式或国际化展示时使用；留空时回退中文姓名。",
        "student_id": "学号、入组编号或内部管理编号，可用于后台检索。",
        "avatar_key": "学生照片媒体 key，可从媒体库选择、上传或裁剪；前台无图时显示姓名姓氏。",
        "degree": "培养层次，例如本科、硕士、博士、访问学生；支持历史填法提示。",
        "category": "学生分组，例如在读学生、毕业校友、助研、本科生；支持历史填法提示。",
        "grade": "年级或入学届别，例如 2024 级；支持历史填法提示。",
        "status": "当前状态，例如在读、毕业、联合培养、访学；支持历史填法提示。",
        "direction": "研究方向或课题方向，前台列表会显示；支持历史填法提示。",
        "bio": "个人简介，前台学生页会显示摘要，后续详情页可复用。",
        "awards": "获奖、荣誉、代表成果或毕业论文题目，可换行记录。",
        "email": "公开或内部可见邮箱，是否前台显示由联系方式可见性控制。",
        "homepage": "学生个人主页、GitHub、Scholar 或实验室个人页链接。",
        "contact_visibility": "控制邮箱和主页等联系方式的可见范围，便于后续接入权限。",
        "enrollment_date": "入学或加入团队日期，可从日期控件选择，也可直接输入 YYYY-MM-DD。",
        "graduation_date": "毕业或离组日期，可从日期控件选择，也可直接输入 YYYY-MM-DD。",
        "destination": "毕业去向、就业单位、升学学校或当前单位。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
        "is_featured": "1 表示首页学生区域优先展示，0 表示仅在学生页展示。",
        "sort_order": "排序一般自动使用记录 ID，通常不用调整；需要人工控制时可直接修改。",
    }
    nav_items = []
    sections = []
    suggest_fields = {"degree", "category", "grade", "status", "direction", "destination"}
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            attrs = student_suggestion_attrs(name) if name in suggest_fields else ""
            if name == "sort_order":
                value = row.get(name) if text_only(row.get(name), 40).strip() else row.get("id", "")
                attrs = 'placeholder="保存留空时自动使用记录 ID"'
            rows = 2 if name in {"bio", "awards"} else None
            labels.append(admin_field_label(field, value, help_map.get(name, ""), control_attrs=attrs, textarea_rows=rows))
        sections.append(f'<fieldset class="form-section student-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form student-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save"><section class="student-edit-sticky"><nav class="student-edit-nav">{"".join(nav_items)}</nav></section>{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def student_suggestion_attrs(name: str) -> str:
    return f'data-student-suggest="{esc(name)}" autocomplete="off"'


def admin_student_category_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("基础设置", ["uid", "key", "enabled", "display_order"]),
        ("前台显示", ["label", "label_en"]),
        ("匹配规则", ["keywords"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、导入导出和后续迁移。新增时自动生成，创建后尽量不要频繁修改。",
        "key": "分组内部 key，建议英文小写且不含空格，如 phd、master、alumni；用于程序识别和跨平台迁移。",
        "label": "中文前台显示名称，也会参与匹配学生记录中的“分组”字段；例如 在读博士、毕业生、本科生助研。",
        "label_en": "英文模式下显示的分组名称；留空时英文站点会回退显示中文标签。",
        "keywords": "匹配规则。学生记录的“分组”字段命中 key、中文标签、英文标签或这些关键词时，会归入该分组；多个词用逗号、分号或换行分隔。",
        "enabled": "是否启用该分组显示规则。1/是表示参与前台学生分组与后台统计，0/否表示临时隐藏该规则。",
        "display_order": "分组显示顺序。数字越小越靠前；新增分组默认使用当前最大排序加一，仍可手动调整。",
    }
    attrs_map = {
        "uid": 'placeholder="例如 cat-phd"',
        "key": 'placeholder="例如 phd、master、alumni"',
        "label": 'placeholder="例如 在读博士"',
        "label_en": 'placeholder="例如 PhD Students"',
        "keywords": 'placeholder="例如 博士, phd, doctoral"',
        "enabled": 'title="1 表示启用，0 表示停用"',
        "display_order": 'placeholder="数字越小越靠前"',
    }
    sections = []
    for title, names in groups:
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            rows = 3 if name == "keywords" else None
            labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), textarea_rows=rows, control_attrs=attrs_map.get(name, "")))
        sections.append(f'<fieldset class="form-section student-category-form-section"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form student-category-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_news_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("基础内容", ["uid", "title", "slug", "category", "cover_key"]),
        ("正文内容", ["content", "content_format"]),
        ("关联内容", ["related_publication_uid", "related_project_uid", "related_student_uid"]),
        ("发布与互动", ["published_at", "allow_comments"]),
        ("展示与排序", ["visibility", "is_featured", "sort_order"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、导入导出和迁移。新增时自动生成，创建后尽量不要频繁修改。",
        "title": "动态标题，前台列表、详情页标题和后台搜索都会优先读取。",
        "slug": "动态详情页 URL 标识，例如 paper-accepted；留空保存时会根据标题自动生成。",
        "category": "动态分类，例如 论文、项目、学生、课程、获奖、网站；支持历史分类提示。",
        "cover_key": "封面图媒体 key，可从媒体库选择或上传；未设置时前台列表显示轻量占位。",
        "content": "动态正文。可直接编辑源码，也可打开富文本窗口编辑；使用富文本保存时会自动切换为 html。",
        "content_format": "正文解析格式。plain 最省资源；html 支持富文本、图片和链接；markdown 预留给后续扩展。",
        "related_publication_uid": "关联论文 UID，便于未来在论文、动态之间互相跳转；可留空。",
        "related_project_uid": "关联项目 UID，便于记录项目进展和后续交叉引用；可留空。",
        "related_student_uid": "关联学生 UID，适合学生获奖、毕业去向、入组新闻等动态；可留空。",
        "allow_comments": "是否允许评论。当前前台未开放完整评论流程，可先作为后续权限/互动扩展配置。",
        "published_at": "发布时间，影响前台动态排序和首页最新动态。建议填写 YYYY-MM-DD，也可填 YYYY-MM-DD HH:MM。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
        "is_featured": "1 表示首页动态区域优先展示，0 表示仅在动态页正常显示。",
        "sort_order": "人工排序值。通常可自动使用记录 ID；需要置顶或精细排序时手动修改。",
    }
    attrs_map = {
        "title": 'placeholder="例如 团队论文被 XXX 录用"',
        "slug": 'placeholder="例如 paper-accepted-2026"',
        "category": 'placeholder="例如 论文、项目、学生、课程" data-news-suggest="category" autocomplete="off"',
        "cover_key": 'placeholder="例如 news/paper-accepted.webp"',
        "related_publication_uid": 'placeholder="例如 pub-2026-001"',
        "related_project_uid": 'placeholder="例如 project-2026-001"',
        "related_student_uid": 'placeholder="例如 student-li-si"',
        "published_at": 'placeholder="YYYY-MM-DD 或 YYYY-MM-DD HH:MM"',
        "sort_order": 'placeholder="留空保存时自动使用记录 ID"',
    }
    sections = []
    for title, names in groups:
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            attrs = attrs_map.get(name, "")
            if name == "sort_order":
                value = row.get(name) if text_only(row.get(name), 40).strip() else row.get("id", "")
            if name == "content":
                labels.append(admin_news_content_field(field, value, help_map.get(name, "")))
                continue
            rows = None
            labels.append(admin_field_label(field, value, help_map.get(name, ""), textarea_rows=rows, control_attrs=attrs))
        sections.append(f'<fieldset class="form-section news-form-section"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form news-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_news_content_field(field: Any, value: Any, help_text: str) -> str:
    control = admin_field_control(field, value, textarea_rows=6, control_attrs='data-news-rich-source')
    help_html = f'<small class="field-help" title="{esc(help_text)}">{esc(help_text)}</small>' if help_text else ""
    return f"""<label class="field-content news-rich-source-field">
      <span class="field-label">正文内容</span>
      <div class="news-rich-entry">
        <div class="news-rich-entry-actions">
          <button class="button secondary" type="button" data-news-rich-open>打开富文本编辑器</button>
          <button class="button light" type="button" data-news-rich-preview>预览 HTML</button>
          <span class="admin-muted">富文本媒体会进入媒体库统一管理。</span>
        </div>
        {control}
      </div>
      {help_html}
    </label>"""


def admin_course_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("课程基本信息", ["uid", "name", "semester", "audience"]),
        ("课程简介", ["summary", "references_text"]),
        ("资料附件", ["syllabus_key", "material_key", "material_visibility"]),
        ("展示与排序", ["visibility", "is_featured", "sort_order"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、导入导出和迁移。新增时自动生成，创建后尽量不要频繁修改。",
        "name": "课程名称，前台课程页、后台搜索和首页课程区域都会优先读取。",
        "semester": "开课学期，例如 2026 春、2025-2026 秋季；支持历史填法提示，便于保持写法一致。",
        "audience": "授课对象，例如 本科生、研究生、留学生、全校公选；支持历史填法提示。",
        "summary": "课程简介，前台课程列表会显示摘要，可写课程目标、内容主题和适合学生。",
        "references_text": "参考资料、教材、阅读材料或课程链接，可换行记录；前台或后续详情页可复用。",
        "syllabus_key": "课程大纲媒体 key，可从媒体库选择或上传 PDF、Word、图片等资料。",
        "material_key": "课件或资料包媒体 key，可从媒体库选择或上传；大文件建议后续迁移到对象存储。",
        "material_visibility": "控制课程资料可见范围。public 表示公开；hidden/staff/private 可为后续权限扩展保留。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
        "is_featured": "1 表示首页或重点区域优先展示，0 表示仅在课程页正常显示。",
        "sort_order": "人工排序值。通常可自动使用记录 ID；需要置顶或精细排序时手动修改。",
    }
    attrs_map = {
        "name": 'placeholder="例如 Web 系统设计与部署"',
        "semester": 'placeholder="例如 2026 春" data-course-suggest="semester" autocomplete="off"',
        "audience": 'placeholder="例如 研究生、本科生" data-course-suggest="audience" autocomplete="off"',
        "syllabus_key": 'placeholder="例如 courses/web-system-syllabus.pdf"',
        "material_key": 'placeholder="例如 courses/web-system-slides.pdf"',
        "sort_order": 'placeholder="留空保存时自动使用记录 ID"',
    }
    sections = []
    for title, names in groups:
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            value = row.get(name, "")
            if name == "sort_order":
                value = row.get(name) if text_only(row.get(name), 40).strip() else row.get("id", "")
            rows = 4 if name in {"summary", "references_text"} else None
            labels.append(admin_field_label(field, value, help_map.get(name, ""), textarea_rows=rows, control_attrs=attrs_map.get(name, "")))
        sections.append(f'<fieldset class="form-section course-form-section"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form course-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_message_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("来信人", ["uid", "name", "email"]),
        ("留言内容", ["message_type", "subject", "content"]),
        ("附件", ["attachment_key"]),
        ("处理与可见性", ["status", "visibility"]),
    ]
    help_map = {
        "uid": "稳定标识用于后台编辑路径、导入导出和迁移。前台提交时自动生成，人工新增时也会预填。",
        "name": "留言人的姓名或称呼，后台列表和邮件回复时会显示。",
        "email": "留言人的联系邮箱；列表页会根据该字段生成邮件回复按钮。",
        "message_type": "留言类型用于后台筛选和优先处理，例如招生、合作、论文、项目、课程或其他。",
        "subject": "留言主题，建议简洁记录咨询事项，便于后台搜索和列表快速判断。",
        "content": "留言正文。可记录前台提交内容、后续沟通摘要或管理员备注。",
        "attachment_key": "留言附件媒体 key，可从媒体库选择、上传或手动填写；留空表示没有附件。",
        "status": "处理状态。新留言、已读、已回复和已归档用于后台流转管理。",
        "visibility": "留言可见范围。建议默认 staff，后续接入权限时可限制只有后台人员查看。",
    }
    attrs_map = {
        "name": 'placeholder="例如 张三"',
        "email": 'placeholder="name@example.com"',
        "subject": 'placeholder="例如 咨询硕士招生或项目合作"',
        "attachment_key": 'placeholder="例如 messages/attachment.pdf"',
    }
    sections = []
    for title, names in groups:
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            rows = 5 if name == "content" else None
            labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), textarea_rows=rows, control_attrs=attrs_map.get(name, "")))
        sections.append(f'<fieldset class="form-section message-form-section"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    return f'<form class="edit-form message-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def admin_publication_form(meta: Table, row: dict[str, Any], repo: Repository | None = None) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("pub-basic", "核心著录", ["uid", "authors", "title", "year", "venue", "doi", "volume", "issue", "pages"]),
        ("pub-classify", "类型与展示", ["publication_type", "index_type", "author_role", "display_tags", "is_featured", "visibility", "sort_order"]),
        ("pub-link", "链接与文件", ["url", "pdf_key", "pdf_visibility"]),
        ("pub-citation", "引用文本", ["source_citation", "citation_gbt", "citation_elsevier", "citation_apa", "citation_ieee", "bibtex"]),
        ("pub-abstract", "摘要与关键词", ["abstract", "keywords"]),
    ]
    help_map = {
        "uid": "稳定标识用于数据迁移、媒体引用和后台编辑路径，创建后尽量不要频繁修改。",
        "title": "论文题名，前台自动引用生成和搜索都会读取该字段。",
        "authors": "作者列表，可用分号、逗号或换行分隔；自动生成 GB/T、APA、IEEE 和 BibTeX 时会使用。",
        "year": "发表年份，用于前台筛选、倒序显示和自动引用。",
        "venue": "期刊名、会议名或出版源，用于前台筛选和完整引文。",
        "publication_type": "如期刊论文、会议论文、预印本等；可作为前台标签显示。",
        "author_role": "标记第一作者、通讯作者或其他角色，便于前台筛选。",
        "index_type": "如 SCI、EI、SSCI、中科院一区等；可作为前台标签显示。",
        "display_tags": "前台卡片右侧展示标签，多个标签可用逗号、分号或换行分隔；留空时使用论文类型和收录字段。",
        "volume": "卷号，自动引用生成会使用。",
        "issue": "期号，自动引用生成会使用。",
        "pages": "页码范围，自动引用生成会使用。",
        "doi": "填写 DOI 本体即可，前台会自动生成 https://doi.org/ 链接。",
        "url": "论文网页、出版社页面或预印本链接。",
        "pdf_key": "PDF 媒体 key，可从媒体库选择、上传或裁剪工具中选择已有文件。",
        "pdf_visibility": "控制 PDF 附件的可见范围，便于后续接入权限。",
        "source_citation": "人工粘贴的原始引用；前台选择原始引用格式时优先显示该字段。",
        "citation_gbt": "保留人工维护的 GB/T 引用文本；当前前台默认根据结构化字段自动生成。",
        "citation_elsevier": "保留人工维护的 Elsevier Numbered 引用文本；可由结构化字段自动生成。",
        "citation_apa": "保留人工维护的 APA 引用文本；当前前台默认根据结构化字段自动生成。",
        "citation_ieee": "保留人工维护的 IEEE 引用文本；当前前台默认根据结构化字段自动生成。",
        "bibtex": "可保存从 BibTeX/DOI/ORCID 工具导入的 BibTeX 原文。",
        "abstract": "论文摘要，用于后台检索和后续详情页扩展。",
        "keywords": "关键词，多个关键词可用逗号、分号或空格分隔。",
        "visibility": "public 为前台公开显示；hidden 等非公开值会从前台公开列表中隐藏。",
        "is_featured": "1 表示代表作，可在代表论文页和首页代表论文区域优先展示。",
        "sort_order": "排序一般自动使用记录 ID，通常不用调整；需要人工控制时可直接修改。",
    }
    sections = []
    nav_items = []
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if field:
                if name in {"title", "doi"}:
                    labels.append(publication_lookup_field(field, row.get(name, ""), help_map.get(name, ""), name))
                elif name in {"authors", "display_tags"}:
                    attrs = publication_suggestion_attrs(name) if name == "display_tags" else ""
                    labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), textarea_rows=2, control_attrs=attrs))
                elif name in {"source_citation", "citation_gbt", "citation_elsevier", "citation_apa", "citation_ieee", "bibtex"}:
                    labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), textarea_rows=2))
                elif name in {"venue", "publication_type", "index_type"}:
                    labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), control_attrs=publication_suggestion_attrs(name)))
                elif name == "sort_order":
                    sort_value = row.get(name) if text_only(row.get(name), 40).strip() else row.get("id", "")
                    labels.append(admin_field_label(field, sort_value, help_map.get(name, ""), control_attrs='placeholder="保存留空时自动使用记录 ID"'))
                else:
                    labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, "")))
        legend_action = '<button class="button light publication-section-generate" type="button" data-publication-generate-citations>生成引用</button>' if section_id == "pub-citation" else ""
        sections.append(f'<fieldset class="form-section publication-form-section" id="{esc(section_id)}"><legend>{esc(title)}{legend_action}</legend>{"".join(labels)}</fieldset>')
    settings = active_global(repo) if repo else {}
    tools = publication_edit_tools(row, "".join(nav_items), settings)
    return f'<form class="edit-form publication-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save">{tools}{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def publication_suggestion_attrs(name: str) -> str:
    return f'data-publication-suggest="{esc(name)}" autocomplete="off"'


def publication_lookup_field(field: Any, value: Any, help_text: str, name: str) -> str:
    label = admin_field_label(field, value, help_text, f"publication-{name}-field", textarea_rows=2 if name == "title" else None)
    return f"""<div class="publication-inline-lookup publication-inline-lookup-{esc(name)}">
      {label}
      <div class="publication-field-actions">
        <button class="button light" type="button" data-publication-check-duplicates data-publication-result-target="{esc(name)}">查重</button>
        <button class="button light" type="button" data-publication-lookup-field="{esc(name)}" data-publication-result-target="{esc(name)}">修正</button>
      </div>
      <div class="publication-field-result publication-field-result-{esc(name)}" data-publication-field-result="{esc(name)}"></div>
    </div>"""


def publication_edit_tools(row: dict[str, Any], nav_html: str, settings: dict[str, Any] | None = None) -> str:
    settings = settings or {}
    providers = parse_platforms(settings.get("publication_metadata_providers") or settings.get("publication_metadata_provider") or "crossref,openalex,semanticscholar")
    provider_options = "".join(f'<label><input type="checkbox" name="metadata_platforms" value="{esc(provider)}" checked>{esc(publication_platform_label(provider))}</label>' for provider in providers)
    return f"""<section class="publication-parse-panel">
      <div class="publication-edit-tools">
        <label class="publication-format-select"><span>引文格式</span><select data-publication-parse-format><option value="auto">自动识别</option><option value="bibtex">BibTeX</option><option value="ieee">IEEE</option><option value="elsevier">Elsevier</option><option value="gb">GB/CNKI</option></select></label>
        <div class="publication-tool-line">
          <div class="publication-tool-buttons">
            <button type="button" class="button light" data-publication-parse>解析填充</button>
            <button type="button" class="button secondary" data-publication-generate-citations>生成引用</button>
          </div>
          <div class="publication-platforms" aria-label="联网查验平台">{provider_options}</div>
        </div>
        <label class="publication-citation-paste"><span>粘贴引文/BibTeX</span><textarea data-publication-parse-source rows="4" placeholder="粘贴 BibTeX、GB/T、IEEE、Elsevier 或普通引文，点击解析填充字段"></textarea></label>
        <p class="publication-tool-status admin-muted" data-publication-status>工具只在点击时运行；联网查验可在通用设置中配置平台。</p>
      </div>
    </section>
    <section class="publication-edit-sticky">
      <nav class="publication-edit-nav">{nav_html}</nav>
    </section>"""


def publication_platform_label(value: str) -> str:
    return {"crossref": "Crossref", "openalex": "OpenAlex", "semanticscholar": "Semantic Scholar"}.get(value, value)


def admin_translation_cache_form(meta: Table, row: dict[str, Any]) -> str:
    field_map = {field.name: field for field in meta.fields}
    groups = [
        ("tr-source", "来源定位", ["uid", "source_ref_key", "source_hash", "source_refs"]),
        ("tr-text", "原文与译文", ["source_text", "translated_text"]),
        ("tr-status", "语言与状态", ["source_lang", "target_lang", "provider", "status", "is_manual", "is_current"]),
        ("tr-error", "错误与备注", ["error_message"]),
    ]
    help_map = {
        "uid": "翻译缓存稳定标识。批量同步时根据来源 key、目标语言和原文 hash 自动生成，便于迁移和去重。",
        "source_ref_key": "来源定位由扫描数据库生成。按相同原文合并时通常为 shared:<hash>；此字段不建议人工修改。",
        "source_hash": "当前原文 hash，用于同一原文复用译文。由扫描器根据原文生成，不建议人工修改。",
        "source_refs": "来源列表由扫描器维护，记录多个后台来源和前台位置；相同原文会合并到同一条缓存。",
        "source_text": "原始文本来自源数据字段。这里修改不会回写到源数据，因此设为只读；请到来源编辑区修改原文。",
        "translated_text": "英文译文。填写后把状态设为 success，并保持当前有效，前台英文页面就会使用它。",
        "source_lang": "源语言，通常为 zh；如果原文本来就是英文，也可以写 en。",
        "target_lang": "目标语言。当前英文版统一使用 en。",
        "provider": "翻译来源，例如 manual、auto、libretranslate、deepl_free、google_translate、microsoft_translator、mymemory；用于记录译文来源。",
        "status": "pending 表示待翻译，success 表示可用于前台，failed 表示翻译失败或需要人工处理。",
        "is_manual": "1 表示人工维护译文，后续自动翻译不应轻易覆盖。",
        "is_current": "1 表示当前有效；置 0 后前台不会使用该缓存。",
        "error_message": "自动翻译或人工校验失败时记录原因，可留空。",
    }
    attrs_map = {
        "uid": 'placeholder="自动生成或手动填写稳定 ID"',
        "source_ref_key": 'placeholder="例如 profiles:profile-main-teacher:bio"',
        "source_hash": 'placeholder="自动生成的原文 hash"',
        "source_lang": 'placeholder="zh"',
        "target_lang": 'placeholder="en"',
        "provider": 'placeholder="manual"',
    }
    sections = []
    nav_items = []
    readonly_fields = {"source_text", "source_hash", "source_ref_key", "source_refs"}
    for section_id, title, names in groups:
        nav_items.append(f'<a href="#{esc(section_id)}">{esc(title)}</a>')
        labels = []
        for name in names:
            field = field_map.get(name)
            if not field:
                continue
            rows = 5 if name in {"source_text", "translated_text"} else 3 if name in {"source_refs", "error_message"} else None
            attrs = attrs_map.get(name, "")
            if name in readonly_fields:
                attrs = (attrs + " readonly").strip()
            labels.append(admin_field_label(field, row.get(name, ""), help_map.get(name, ""), textarea_rows=rows, control_attrs=attrs))
        sections.append(f'<fieldset class="form-section translation-form-section" id="{esc(section_id)}"><legend>{esc(title)}</legend>{"".join(labels)}</fieldset>')
    delete = ""
    key = row.get("uid") or row.get("id")
    if key:
        delete = f'<button class="button danger" type="submit" formaction="/admin/table/translation_cache/delete/{esc(key)}" formmethod="post" data-confirm="确定删除这条翻译缓存任务吗？原始内容不会被删除，之后可重新扫描生成。">删除缓存</button>'
    return f'<form class="edit-form translation-edit-form" method="post" action="/admin/table/{esc(meta.name)}/save"><section class="translation-edit-sticky"><nav class="translation-edit-nav">{"".join(nav_items)}</nav>{delete}</section>{"".join(sections)}{admin_form_actions(meta.name)}</form>'


def normalize_admin_data(meta: Table, data: dict[str, str]) -> dict[str, Any]:
    normalized = {field.name: text_only(data.get(field.name), 12000) for field in meta.fields}
    if not normalized.get("uid"):
        seed = normalized.get(meta.title_field) or str(time.time_ns())
        normalized["uid"] = stable_uid(meta.name.rstrip("s"), seed)
    if meta.name == "site_settings":
        raw_footer = str(data.get("footer_text") or "").strip()
        if "<" in raw_footer and ">" in raw_footer:
            normalized["footer_text"] = render_limited_html(raw_footer)
        else:
            normalized["footer_text"] = text_only(raw_footer, 12000)
    if meta.name == "news":
        normalized["slug"] = safe_slug(normalized.get("slug") or normalized.get("title", "news"))
        normalized["published_at"] = normalize_datetime_input(normalized.get("published_at"))
        content_format = text_only(data.get("content_format"), 40).strip() or "plain"
        normalized["content_format"] = content_format if content_format in {"plain", "html", "markdown"} else "plain"
        if normalized["content_format"] == "html":
            normalized["content"] = render_limited_html(data.get("content", ""))
    if meta.name == "media_assets" and not normalized.get("status"):
        normalized["status"] = "active"
    if meta.name == "media_assets":
        if not normalized.get("storage_kind"):
            normalized["storage_kind"] = media_storage_kind(normalized.get("object_key"))
        else:
            normalized["storage_kind"] = normalize_media_storage_kind(normalized.get("storage_kind"), media_storage_kind(normalized.get("object_key")))
    if meta.name == "projects":
        normalized["start_date"] = normalize_date_input(normalized.get("start_date"))
        normalized["end_date"] = normalize_date_input(normalized.get("end_date"))
    if meta.name == "patents":
        normalized["application_date"] = normalize_date_input(normalized.get("application_date"))
        normalized["grant_date"] = normalize_date_input(normalized.get("grant_date"))
    if meta.name == "students":
        normalized["enrollment_date"] = normalize_date_input(normalized.get("enrollment_date"))
        normalized["graduation_date"] = normalize_date_input(normalized.get("graduation_date"))
    if meta.name == "translation_cache":
        if not normalized.get("target_lang"):
            normalized["target_lang"] = "en"
        if not normalized.get("source_lang"):
            normalized["source_lang"] = "zh"
        if not normalized.get("provider"):
            normalized["provider"] = "manual"
        if not normalized.get("status"):
            normalized["status"] = "success" if normalized.get("translated_text") else "pending"
        if not normalized.get("source_hash") and normalized.get("source_text"):
            normalized["source_hash"] = translation_source_hash(normalized.get("source_text"))
        if not normalized.get("uid"):
            seed = f'{normalized.get("target_lang")}:{normalized.get("source_ref_key")}:{normalized.get("source_hash") or normalized.get("source_text")}'
            normalized["uid"] = stable_uid("tr", seed)
    if meta.name == "global_settings":
        normalized = global_settings_translation_defaults(normalized)
    return normalized


def normalize_date_input(value: Any) -> str:
    import re

    text = text_only(value, 40).strip()
    if not text:
        return ""
    match = re.match(r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?$", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    match = re.match(r"^(\d{4})(\d{2})(\d{2})$", text)
    if match:
        year, month, day = match.groups()
        return f"{year}-{month}-{day}"
    return text


def normalize_datetime_input(value: Any) -> str:
    import re

    text = text_only(value, 80).strip()
    if not text:
        return ""
    text = text.replace("T", " ").replace("/", "-").replace(".", "-")
    match = re.match(r"^(\d{4}[-年]\d{1,2}[-月]\d{1,2}日?)(?:\s+(\d{1,2}):(\d{1,2})(?::\d{1,2})?)?$", text)
    if match:
        date_part = normalize_date_input(match.group(1))
        if match.group(2) and match.group(3):
            return f"{date_part} {int(match.group(2)):02d}:{int(match.group(3)):02d}"
        return date_part
    match = re.match(r"^(\d{8})(?:\s*(\d{2})(\d{2}))?$", text)
    if match:
        date_part = normalize_date_input(match.group(1))
        if match.group(2) and match.group(3):
            return f"{date_part} {match.group(2)}:{match.group(3)}"
        return date_part
    return text


def active_site(repo: Repository) -> dict[str, Any]:
    rows = repo.list("site_settings", Query(limit=1))
    return rows[0] if rows else {}


def active_global(repo: Repository) -> dict[str, Any]:
    rows = repo.list("global_settings", Query(limit=1))
    return global_settings_translation_defaults(rows[0]) if rows else global_settings_translation_defaults({})


def first(repo: Repository, table: str) -> dict[str, Any]:
    rows = repo.list(table, Query(public_only=True, limit=1))
    return rows[0] if rows else {}


def nav(repo: Repository, location: str, env: dict[str, str] | None = None) -> list[dict[str, Any]]:
    if env is None:
        return repo.list("navigation_items", Query(filters={"location": location, "enabled": 1}, public_only=True, limit=20))
    return visible_list(repo, env, "navigation_items", Query(filters={"location": location, "enabled": 1}, limit=20, order_by="sort_order", descending=False))


def section(title: str, body: str) -> str:
    return f'<section class="band"><h2>{esc(title)}</h2>{body}</section>'


def section_head(title: str, href: str, env: dict[str, str] | None = None) -> str:
    env = env or {}
    return f'<div class="section-head"><h2>{esc(title)}</h2><a href="{esc(href)}">{esc(t(env, "all"))}</a></div>'


def paragraphs(value: Any) -> str:
    text = text_only(value, 1600)
    if not text:
        return ""
    return "".join(f"<p>{esc(part)}</p>" for part in text.splitlines() if part.strip())


def fact(label: str, value: Any) -> str:
    return f"<dt>{esc(label)}</dt><dd>{esc(value)}</dd>" if value else ""


def button(item: dict[str, Any], env: dict[str, str] | None = None, repo: Repository | None = None) -> str:
    env = env or {}
    label = front_value(repo, env, "navigation_items", item, "title", 120) if repo else localized_nav_title(item, current_lang(env))
    return f'<a class="button {esc(item.get("style") or "secondary")}" href="{esc(lang_url(str(item.get("path") or "/"), env))}">{esc(label)}</a>'


TRANSLATIONS = {
    "zh": {
        "admin": "后台", "login": "登录", "logout": "退出", "register": "注册", "username": "账号", "password": "密码", "password_confirm": "确认密码", "display_name": "显示名称", "login_scope_hint": "登录后可查看授权范围内的前台内容；后台编辑需要管理员或员工权限。", "academic_profile": "学术档案", "email": "邮箱", "office": "办公室", "recruiting": "招生方向",
        "username_placeholder": "请输入账号或邮箱", "password_placeholder": "请输入密码", "register_username_placeholder": "请输入注册账号", "display_name_placeholder": "请输入显示名称，可填写真实姓名", "email_placeholder": "请输入邮箱，可选", "new_password_placeholder": "请输入不少于 10 位的新密码", "password_confirm_placeholder": "请再次输入密码",
        "login_failed": "账号或密码错误。", "login_rate_limited": "登录失败次数过多，请约 {minutes} 分钟后再试。", "login_unavailable": "登录暂不可用", "admin_not_initialized": "管理员尚未初始化账号。", "no_account": "还没有账号？", "registration_closed_hint": "当前网站未开放公开注册，请联系管理员创建账号。",
        "registration_unavailable": "注册暂不可用", "registration_closed": "注册已关闭", "registration_closed_message": "管理员已关闭公开注册。你仍可使用已有账号登录，或联系管理员开通访问权限。", "back_login": "返回登录", "register_scope_hint": "公开注册账号默认是访客用户，只能查看管理员授权的前台内容，不能进入后台。", "username_required": "请填写账号。", "username_taken": "该账号已被使用。", "password_too_short": "密码至少需要 10 位。", "password_mismatch": "两次输入的密码不一致。",
        "research_interests": "研究方向", "featured_publications": "代表论文", "latest_news": "最新动态",
        "projects": "项目", "patents": "专利", "students": "学生", "team_members": "团队成员",
        "publications": "论文成果", "news": "动态", "courses": "课程", "contact": "联系",
        "role": "角色", "title": "头衔", "organization": "单位", "team": "团队", "team_search": "姓名、角色、职称、单位、简介",
        "degree": "层次", "category": "分组", "grade": "年级", "status": "状态", "direction": "方向", "student_search": "姓名、方向、年级、状态、去向",
        "student_group_mode": "显示方式", "group_by_category": "按分组显示", "group_by_degree": "按学历显示", "student_group_uncategorized": "未分组", "people_count_unit": "人",
        "patent_type": "类型", "legal_status": "状态", "country": "国家", "patent_search": "名称、申请号、授权号、发明人、权利人",
        "source": "来源", "fund_name": "基金", "project_search": "项目名称、来源、基金、项目号、负责人、成员",
        "project_number": "项目号", "project_period": "周期", "project_amount": "金额", "project_amount_unit": "万元",
        "semester": "学期", "audience": "对象", "course_search": "课程名称、学期、对象、简介",
        "year": "年份", "venue": "期刊会议", "publication_type": "类型", "author_role": "作者角色", "index_type": "收录",
        "publication_search": "题名、作者、期刊、DOI、关键词", "news_search": "标题、分类、内容、关联信息",
        "inventors": "发明人/作者", "owner": "权利人", "application_number": "申请号", "grant_number": "授权号",
        "application_info": "申请", "grant_info": "授权",
        "application_date": "申请日期", "grant_date": "授权日期",
        "bio": "个人简介", "back_team": "返回团队", "not_found": "未找到", "page_missing": "页面不存在。",
        "member_missing": "成员不存在", "news_missing": "动态不存在", "message_submitted": "留言已提交", "message_thanks": "感谢你的信息，管理员会在后台查看。",
        "back_home": "返回首页", "search": "筛选", "reset": "重置", "all": "全部", "select_all": "全选", "copy": "复制",
        "select_publication": "选择此论文", "copy_current_citation": "复制当前论文引用", "empty": "暂无内容。",
        "name": "姓名", "message_type": "类型", "subject": "主题", "content": "内容", "submit": "提交",
        "message_type_recruiting": "招生", "message_type_cooperation": "合作", "message_type_paper": "论文", "message_type_course": "课程", "message_type_other": "其他",
        "citation_source": "原始引用",
    },
    "en": {
        "admin": "Admin", "login": "Log In", "logout": "Log Out", "register": "Register", "username": "Username", "password": "Password", "password_confirm": "Confirm Password", "display_name": "Display Name", "login_scope_hint": "After logging in, you can view authorized frontend content. Admin editing requires a staff or administrator role.", "academic_profile": "Profile", "email": "Email", "office": "Office", "recruiting": "Recruiting",
        "username_placeholder": "Enter your username or email", "password_placeholder": "Enter your password", "register_username_placeholder": "Choose a username", "display_name_placeholder": "Enter a display name or real name", "email_placeholder": "Enter an email address, optional", "new_password_placeholder": "Enter a new password, at least 10 characters", "password_confirm_placeholder": "Enter the password again",
        "login_failed": "Incorrect username or password.", "login_rate_limited": "Too many failed login attempts. Please try again in about {minutes} minute(s).", "login_unavailable": "Login Unavailable", "admin_not_initialized": "The administrator account has not been initialized.", "no_account": "No account yet? ", "registration_closed_hint": "Public registration is closed. Please contact the administrator for an account.",
        "registration_unavailable": "Registration Unavailable", "registration_closed": "Registration Closed", "registration_closed_message": "Public registration has been disabled by the administrator. You can still log in with an existing account or contact the administrator for access.", "back_login": "Back to Login", "register_scope_hint": "Publicly registered accounts are visitor users by default. They can only view authorized frontend content and cannot enter the admin area.", "username_required": "Please enter a username.", "username_taken": "This username is already in use.", "password_too_short": "Password must be at least 10 characters.", "password_mismatch": "The two passwords do not match.",
        "research_interests": "Research", "featured_publications": "Selected Papers", "latest_news": "News",
        "projects": "Projects", "patents": "Patents", "students": "Students", "team_members": "Team",
        "publications": "Publications", "news": "News", "courses": "Courses", "contact": "Contact",
        "role": "Role", "title": "Title", "organization": "Organization", "team": "Group", "team_search": "Name, role, title, organization, bio",
        "degree": "Degree", "category": "Group", "grade": "Year", "status": "Status", "direction": "Direction", "student_search": "Name, direction, year, status",
        "student_group_mode": "Display", "group_by_category": "By Group", "group_by_degree": "By Degree", "student_group_uncategorized": "Uncategorized", "people_count_unit": "people",
        "patent_type": "Type", "legal_status": "Status", "country": "Country", "patent_search": "Title, application no., grant no., inventors, owner",
        "source": "Source", "fund_name": "Fund", "project_search": "Project title, source, fund, project no., PI, members",
        "project_number": "Project No.", "project_period": "Period", "project_amount": "Amount", "project_amount_unit": "CNY",
        "semester": "Semester", "audience": "Audience", "course_search": "Course name, semester, audience, summary",
        "year": "Year", "venue": "Journal / Conference", "publication_type": "Type", "author_role": "Author Role", "index_type": "Indexing",
        "publication_search": "Title, authors, venue, DOI, keywords", "news_search": "Title, category, content, related information",
        "inventors": "Inventors", "owner": "Owner", "application_number": "Application No.", "grant_number": "Grant No.",
        "application_info": "Application", "grant_info": "Grant",
        "application_date": "Application Date", "grant_date": "Grant Date",
        "bio": "Biography", "back_team": "Back to Team", "not_found": "Not Found", "page_missing": "Page not found.",
        "member_missing": "Member not found", "news_missing": "News item not found", "message_submitted": "Message Submitted", "message_thanks": "Thanks. The administrator will review your message.",
        "back_home": "Back Home", "search": "Search", "reset": "Reset", "all": "All", "select_all": "Select All", "copy": "Copy",
        "select_publication": "Select this publication", "copy_current_citation": "Copy citation", "empty": "No content yet.",
        "name": "Name", "message_type": "Type", "subject": "Subject", "content": "Message", "submit": "Submit",
        "message_type_recruiting": "Recruiting", "message_type_cooperation": "Collaboration", "message_type_paper": "Paper", "message_type_course": "Course", "message_type_other": "Other",
        "citation_source": "Source Citation",
    },
}


def current_lang(env: dict[str, str]) -> str:
    return "en" if env.get("_LANG") == "en" else "zh"


def i18n_dictionary_paths() -> list[Path]:
    paths: list[Path] = [Path(I18N_DICTIONARY_FILENAME)]
    try:
        current = Path(__file__).resolve()
        for parent in current.parents:
            paths.append(parent / I18N_DICTIONARY_FILENAME)
    except Exception:
        pass
    unique: list[Path] = []
    seen: set[str] = set()
    for path in paths:
        key = str(path)
        if key not in seen:
            unique.append(path)
            seen.add(key)
    return unique


def normalize_i18n_dictionary(data: Any) -> dict[str, Any]:
    entries = data.get("entries", data) if isinstance(data, dict) else data
    normalized: dict[str, dict[str, str]] = {}
    if isinstance(entries, dict):
        iterable = entries.items()
    elif isinstance(entries, list):
        iterable = []
        for index, item in enumerate(entries):
            if isinstance(item, dict):
                key = text_only(item.get("key") or item.get("uid") or f"entry-{index + 1}", 160).strip()
                iterable.append((key, item))
    else:
        iterable = []
    for raw_key, raw_value in iterable:
        key = text_only(raw_key, 180).strip()
        if not key:
            continue
        if isinstance(raw_value, dict):
            zh = text_only(raw_value.get("zh") or raw_value.get("cn") or raw_value.get("source") or raw_value.get("source_text"), 12000).strip()
            en = text_only(raw_value.get("en") or raw_value.get("english") or raw_value.get("target") or raw_value.get("translated_text"), 12000).strip()
            context = text_only(raw_value.get("context") or raw_value.get("source_ref_key") or raw_value.get("note"), 300).strip()
        else:
            value = text_only(raw_value, 12000).strip()
            if has_cjk(key):
                zh, en = key, value
            else:
                zh, en = value, ""
            context = ""
        if not zh and has_cjk(key):
            zh = key
        if not zh and not en:
            continue
        normalized[key] = {"zh": zh, "en": en, "context": context}
    return {"version": int_value((data or {}).get("version") if isinstance(data, dict) else 1, 1), "entries": normalized}


def load_i18n_dictionary(env: dict[str, str] | None = None) -> dict[str, Any]:
    env = env or {}
    raw = text_only(env.get("_I18N_DICTIONARY_JSON"), 2_000_000).strip()
    if raw:
        try:
            return normalize_i18n_dictionary(json.loads(raw))
        except Exception:
            return {"version": 1, "entries": {}}
    for path in i18n_dictionary_paths():
        try:
            if not path.is_file():
                continue
            stat = path.stat()
            cache_path = str(path.resolve())
            if I18N_DICTIONARY_CACHE.get("path") == cache_path and I18N_DICTIONARY_CACHE.get("mtime") == stat.st_mtime and isinstance(I18N_DICTIONARY_CACHE.get("data"), dict):
                return I18N_DICTIONARY_CACHE["data"]
            data = normalize_i18n_dictionary(json.loads(path.read_text(encoding="utf-8")))
            I18N_DICTIONARY_CACHE.update({"path": cache_path, "mtime": stat.st_mtime, "data": data})
            return data
        except Exception:
            continue
    return {"version": 1, "entries": {}}


def i18n_dictionary_entries(env: dict[str, str] | None = None) -> dict[str, dict[str, str]]:
    data = load_i18n_dictionary(env)
    entries = data.get("entries") if isinstance(data, dict) else {}
    return entries if isinstance(entries, dict) else {}


def i18n_dictionary_lookup_key(env: dict[str, str], key: str, lang: str) -> str:
    entry = i18n_dictionary_entries(env).get(key)
    if not isinstance(entry, dict):
        return ""
    value = text_only(entry.get(lang), 12000).strip()
    if value:
        return value
    if lang == "zh":
        return text_only(entry.get("source") or entry.get("zh"), 12000).strip()
    return ""


def i18n_dictionary_lookup_source(env: dict[str, str], source: str, lang: str = "en") -> str:
    source = text_only(source, 12000).strip()
    if not source:
        return ""
    source_casefold = source.casefold()
    for entry in i18n_dictionary_entries(env).values():
        if not isinstance(entry, dict):
            continue
        zh = text_only(entry.get("zh"), 12000).strip()
        en = text_only(entry.get("en"), 12000).strip()
        if zh and zh.casefold() == source_casefold:
            return text_only(entry.get(lang), 12000).strip()
        if en and en.casefold() == source_casefold and lang == "zh":
            return zh
    return ""


def t(env: dict[str, str], key: str) -> str:
    lang = current_lang(env)
    dictionary_value = i18n_dictionary_lookup_key(env, key, lang)
    if dictionary_value:
        return dictionary_value
    return TRANSLATIONS.get(lang, TRANSLATIONS["zh"]).get(key, TRANSLATIONS["zh"].get(key, key))


def translation_ref_key(table: str, row: dict[str, Any], field: str) -> str:
    key = text_only(row.get("uid") or row.get("slug") or row.get("id") or "row", 200).strip() or "row"
    return f"{table}:{key}:{field}"


def translation_source_hash(value: Any) -> str:
    text = text_only(value, 12000).strip()
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16] if text else ""


def translation_cache_uid(ref_key: str, source_hash: str, target_lang: str = "en") -> str:
    return stable_uid("tr", f"{target_lang}:{ref_key}:{source_hash}")


def translation_cache_index(repo: Repository, env: dict[str, str]) -> dict[tuple[str, str], dict[str, Any]]:
    cache_key = "_TRANSLATION_CACHE_INDEX"
    existing = env.get(cache_key)
    if isinstance(existing, dict):
        return existing
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for row in repo.list("translation_cache", Query(limit=1000)):
        if text_only(row.get("target_lang"), 20).strip() not in {"en", "EN", "english", "English"}:
            continue
        if not truthy(row.get("is_current"), default=True):
            continue
        translated = text_only(row.get("translated_text"), 12000).strip()
        if not translated:
            continue
        status = text_only(row.get("status"), 40).strip() or "success"
        if status in {"failed", "stale", "deleted"}:
            continue
        ref_key = text_only(row.get("source_ref_key"), 300).strip()
        source_hash = text_only(row.get("source_hash"), 80).strip()
        source_text = text_only(row.get("source_text"), 12000).strip()
        if ref_key and source_hash:
            translation_index_put(index, (ref_key, source_hash), row)
        if source_hash:
            translation_index_put(index, ("", source_hash), row)
        if ref_key and source_text:
            translation_index_put(index, (ref_key, translation_source_hash(source_text)), row)
    env[cache_key] = index  # type: ignore[assignment]
    return index


def translation_index_put(index: dict[tuple[str, str], dict[str, Any]], key: tuple[str, str], row: dict[str, Any]) -> None:
    current = index.get(key)
    if current is None or translation_cache_rank(row) >= translation_cache_rank(current):
        index[key] = row


def translation_cache_rank(row: dict[str, Any]) -> int:
    status = text_only(row.get("status"), 40).strip()
    if status in {"success", "reviewed"} and text_only(row.get("translated_text"), 12000).strip():
        return 4
    if text_only(row.get("translated_text"), 12000).strip():
        return 3
    if status == "pending":
        return 2
    if status == "failed":
        return 1
    return 0


def truthy(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    return str(value).strip().casefold() in {"1", "true", "yes", "on", "y"}


def front_value(repo: Repository, env: dict[str, str], table: str, row: dict[str, Any], field: str, limit: int = 1600) -> str:
    source = text_only(row.get(field), limit).strip()
    if current_lang(env) != "en":
        return source
    english_field = ENGLISH_FIELD_OVERRIDES.get((table, field))
    if english_field:
        direct = text_only(row.get(english_field), limit).strip()
        if direct:
            return direct
    if not source:
        return ""
    if should_preserve_english_source(table, field, source):
        return source
    dictionary_value = i18n_dictionary_lookup_source(env, source, "en")
    if dictionary_value:
        return text_only(dictionary_value, limit).strip()
    ref_key = translation_ref_key(table, row, field)
    source_hash = translation_source_hash(source)
    cached = translation_cache_index(repo, env).get((ref_key, source_hash)) or translation_cache_index(repo, env).get(("", source_hash))
    if cached:
        translated = text_only(cached.get("translated_text"), limit).strip()
        if translated:
            return translated
    return source


def should_preserve_english_source(table: str, field: str, source: str) -> bool:
    if has_cjk(source):
        return False
    if table == "publications":
        return field in {"title", "source_citation", "authors", "venue", "publication_type", "author_role", "index_type", "display_tags"}
    return False


def front_row(repo: Repository, env: dict[str, str], table: str, row: dict[str, Any], fields: tuple[str, ...] | list[str] | None = None) -> dict[str, Any]:
    if current_lang(env) != "en":
        return row
    localized = dict(row)
    for field in fields or FRONTEND_TRANSLATION_FIELDS.get(table, ()):
        localized[field] = front_value(repo, env, table, row, field, 12000)
    return localized


def front_paragraphs(repo: Repository, env: dict[str, str], table: str, row: dict[str, Any], field: str) -> str:
    return paragraphs(front_value(repo, env, table, row, field, 5000))


def localized_value(row: dict[str, Any], field: str, lang: str) -> str:
    if lang == "en":
        value = text_only(row.get(f"{field}_en"), 500).strip()
        if value:
            return value
    return text_only(row.get(field), 500).strip()


def localized_name(row: dict[str, Any], lang: str) -> str:
    return localized_value(row, "name", lang) or text_only(row.get("name"), 500).strip()


def localized_site_name(site: dict[str, Any], lang: str) -> str:
    if lang == "en":
        return text_only(site.get("site_name_en"), 200).strip() or text_only(site.get("site_name"), 200).strip() or "Academic Website"
    return text_only(site.get("site_name"), 200).strip() or "教师个人网站"


def localized_nav_title(item: dict[str, Any], lang: str) -> str:
    if lang == "en":
        return text_only(item.get("title_en"), 120).strip() or text_only(item.get("title"), 120).strip()
    return text_only(item.get("title"), 120).strip()


def lang_url(path: str, env: dict[str, str]) -> str:
    if current_lang(env) != "en" or path.startswith(("http://", "https://", "mailto:")):
        return safe_href(path)
    separator = "&" if "?" in path else "?"
    return safe_href(f"{path}{separator}lang=en")


def site_absolute_url(site_url: str, path: str) -> str:
    base = (site_url or "").rstrip("/")
    clean = path or "/"
    if clean.startswith(("http://", "https://")):
        return clean
    if not clean.startswith("/"):
        clean = "/" + clean
    return f"{base}{clean}" if base else clean


def english_path(path: str) -> str:
    separator = "&" if "?" in path else "?"
    return f"{path}{separator}lang=en"


def seo_language_links(env: dict[str, str]) -> str:
    site_url = str(env.get("SITE_URL") or "").rstrip("/")
    path = str(env.get("_PATH") or "/")
    zh_url = site_absolute_url(site_url, path)
    en_url = site_absolute_url(site_url, english_path(path))
    canonical = en_url if current_lang(env) == "en" else zh_url
    return "\n".join(
        [
            f'<link rel="canonical" href="{esc(canonical)}">',
            f'<link rel="alternate" hreflang="zh-CN" href="{esc(zh_url)}">',
            f'<link rel="alternate" hreflang="en" href="{esc(en_url)}">',
            f'<link rel="alternate" hreflang="x-default" href="{esc(zh_url)}">',
        ]
    )


def language_switch(env: dict[str, str]) -> str:
    path = env.get("_PATH") or "/"
    if current_lang(env) == "en":
        return f'<a class="lang-switch" href="{esc(safe_href(path))}" lang="zh-CN">中文</a>'
    return f'<a class="lang-switch" href="{esc(safe_href(path + "?lang=en"))}" lang="en">EN</a>'


def external_links(profile: dict[str, Any]) -> str:
    links = []
    for label, key in [("主页", "personal_homepage"), ("Scholar", "google_scholar"), ("DBLP", "dblp"), ("GitHub", "github"), ("CNKI", "cnki")]:
        if profile.get(key):
            links.append(f'<a href="{esc(safe_href(profile.get(key)))}" target="_blank" rel="noreferrer">{esc(label)}</a>')
    return '<div class="external-links">' + "".join(links) + "</div>" if links else ""


def profile_links(row: dict[str, Any]) -> str:
    links = []
    if row.get("email"):
        links.append(contact_chip("mail", "邮箱", row.get("email"), safe_href("mailto:" + str(row.get("email")))))
    if row.get("phone"):
        links.append(contact_chip("tel", "电话", row.get("phone")))
    homepage = row.get("personal_homepage") or row.get("homepage")
    if homepage:
        links.append(contact_chip("home", "主页", homepage_label(homepage), safe_href(homepage)))
    if row.get("google_scholar"):
        links.append(contact_chip("scholar", "学术", "Google Scholar", safe_href(row.get("google_scholar"))))
    if row.get("orcid"):
        links.append(contact_chip("orcid", "ORCID", row.get("orcid"), orcid_href(row.get("orcid"))))
    if row.get("dblp"):
        links.append(contact_chip("dblp", "DBLP", "DBLP", safe_href(row.get("dblp"))))
    if row.get("github"):
        links.append(contact_chip("github", "代码", github_label(row.get("github")), safe_href(row.get("github"))))
    if row.get("cnki"):
        links.append(contact_chip("cnki", "知网", "CNKI", safe_href(row.get("cnki"))))
    return "".join(links)


def contact_chip(kind: str, label: str, value: Any, href: str = "") -> str:
    content = esc(text_only(value, 120).strip())
    icon = f'<span class="contact-icon contact-icon-{esc(kind)}" aria-hidden="true"><img class="contact-icon-img" src="{esc(icon_url(kind))}" alt="" loading="lazy"><span class="contact-icon-fallback">{esc(label[:2])}</span></span>'
    inner = f'{icon}<span class="contact-separator">:</span><span class="contact-value">{content}</span>'
    if href:
        return f'<a class="contact-chip" href="{esc(safe_href(href))}" target="_blank" rel="noreferrer">{inner}</a>'
    return f'<span class="contact-chip">{inner}</span>'


def icon_url(kind: str) -> str:
    icons = {
        "mail": "lucide/mail",
        "tel": "lucide/phone",
        "home": "lucide/globe",
        "scholar": "simple-icons/googlescholar",
        "orcid": "simple-icons/orcid",
        "dblp": "simple-icons/dblp",
        "github": "simple-icons/github",
        "cnki": "simple-icons/cnki",
    }
    icon = icons.get(kind, "lucide/link")
    return f"https://api.iconify.design/{icon}.svg?height=12&color=%2312645d"


def homepage_label(value: Any) -> str:
    text = text_only(value, 160).strip()
    if not text:
        return "个人主页"
    parsed = urlparse(text if "://" in text else "https://" + text)
    host = parsed.netloc or parsed.path
    path = parsed.path.strip("/")
    label = host + (f"/{path}" if path else "")
    return label[:80] if label else text[:80]


def team_identity(row: dict[str, Any]) -> str:
    badges = []
    seen = set()
    for key in ["role", "title", "organization", "lab"]:
        value = text_only(row.get(key), 80).strip()
        if value and value not in seen:
            badges.append(f'<span>{esc(value)}</span>')
            seen.add(value)
    return '<div class="team-badges">' + "".join(badges) + "</div>" if badges else ""


def profile_detail_sections(row: dict[str, Any]) -> str:
    groups = [
        ("基本信息", ["name_en", "role", "title", "organization", "lab", "office"]),
        ("联系方式", ["email", "phone"]),
        ("学术链接", ["personal_homepage", "orcid", "google_scholar", "dblp", "github", "cnki"]),
        ("经历与招生", ["education", "experience", "recruiting"]),
    ]
    sections = []
    for title, names in groups:
        items = profile_detail_items(row, names)
        if items:
            sections.append(f'<section class="team-detail-section"><h2>{esc(title)}</h2><dl class="profile-detail-grid team-detail-grid">{items}</dl></section>')
    return "".join(sections) or empty()


def profile_detail_items(row: dict[str, Any], names: list[str] | None = None) -> str:
    skip = {"uid", "name", "bio", "avatar_key", "contact_visibility", "visibility", "is_active", "is_featured", "sort_order"}
    items = []
    fields = TABLE_MAP["profiles"].fields
    if names is not None:
        by_name = {field.name: field for field in fields}
        fields = tuple(by_name[name] for name in names if name in by_name)
    for field in fields:
        if field.name in skip:
            continue
        value = row.get(field.name)
        if value in (None, ""):
            continue
        items.append(f'<div class="profile-detail-item"><dt>{esc(field.label)}</dt><dd>{profile_detail_value(field.name, value, field.kind)}</dd></div>')
    return "".join(items)


def profile_detail_value(name: str, value: Any, kind: str) -> str:
    text = text_only(value, 3000)
    if name == "email":
        return f'<a href="{esc(safe_href("mailto:" + text))}">{esc(text)}</a>'
    if name == "orcid":
        return f'<a href="{esc(orcid_href(text))}" target="_blank" rel="noreferrer">{esc(text)}</a>'
    if kind == "url":
        return f'<a href="{esc(safe_href(text))}" target="_blank" rel="noreferrer">{esc(text)}</a>'
    if kind == "textarea":
        return esc(text).replace("\n", "<br>")
    return esc(text)


def orcid_href(value: Any) -> str:
    text = text_only(value, 80).strip()
    if text.startswith(("http://", "https://")):
        return safe_href(text)
    return safe_href("https://orcid.org/" + text)


def github_label(value: Any) -> str:
    text = text_only(value, 120).strip().rstrip("/")
    if "/" in text:
        return text.rsplit("/", 1)[-1] or "GitHub"
    return text or "GitHub"


def latest_publications(rows: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
    def key(row: dict[str, Any]) -> tuple[int, int, int, str]:
        return (
            int_value(row.get("year"), 0),
            int_value(row.get("id"), 0),
            int_value(row.get("sort_order"), 0),
            text_only(row.get("updated_at"), 80),
        )

    return sorted(rows, key=key, reverse=True)[: max(1, limit)]


def publication_list(rows: list[dict[str, Any]], selectable: bool = False, compact: bool = False, display_style: str = "gbt", repo: Repository | None = None, env: dict[str, str] | None = None) -> str:
    items = []
    total = len(rows)
    for index, row in enumerate(rows, 1):
        display_row = front_row(repo, env or {}, "publications", row) if repo and env else row
        number = total - index + 1 if compact else index
        citations = publication_citations(display_row)
        pdf = f'<a href="{esc(media_url(row.get("pdf_key")))}">PDF</a>' if row.get("pdf_key") else ""
        doi = f'<a href="https://doi.org/{esc(row.get("doi"))}" target="_blank" rel="noreferrer">DOI</a>' if row.get("doi") else ""
        citation_attrs = citation_data_attrs(citations)
        select_label = t(env or {}, "select_publication")
        checkbox = f'<label class="citation-check-wrap" title="{esc(select_label)}"><input type="checkbox" class="copy-check" {citation_attrs}><span class="sr-only">{esc(select_label)}</span></label>' if selectable else ""
        display_citation = publication_display_citation(display_row, citations, display_style)
        if compact:
            tags = publication_tags(display_row)
            copy_title = t(env or {}, "copy_current_citation")
            copy_button = f'<button type="button" class="citation-copy-one" {citation_attrs} aria-label="{esc(copy_title)}" title="{esc(copy_title)}">⧉</button>'
            items.append(f"""<article class="citation-item publication-card">
              <div class="citation-index"><span class="item-number">{number}</span>{checkbox}</div>
              <div class="citation-body">
                <div class="publication-copy-zone">
                  <p class="pub-citation" {citation_attrs}>{esc(display_citation)}</p>
                </div>
                <div class="citation-tools publication-tools no-copy"><div class="publication-links">{doi}{pdf}</div><div class="publication-tool-right">{tags}{copy_button}</div></div>
              </div>
            </article>""")
        else:
            items.append(f'<article class="publication-item">{checkbox}<span class="index">{number}</span><div><p>{esc(display_citation)}</p><p class="meta">{esc(row.get("year"))} / {esc(display_row.get("venue"))} / {esc(display_row.get("publication_type"))} {doi} {pdf}</p></div></article>')
    class_name = "citation-list classified-list" if compact else "publication-list"
    return f'<section class="{class_name}">' + ("".join(items) or empty(env or {})) + "</section>"


def publication_citations(row: dict[str, Any]) -> dict[str, str]:
    generated = generated_publication_citations(row)
    source = text_only(row.get("source_citation"), 3000).strip() or generated["gbt"]
    return {
        "gbt": generated["gbt"],
        "elsevier": generated["elsevier"],
        "apa": generated["apa"],
        "ieee": generated["ieee"],
        "bibtex": generated["bibtex"],
        "source": source,
    }


def generated_publication_citations(row: dict[str, Any]) -> dict[str, str]:
    authors = split_authors(row.get("authors"))
    title = text_only(row.get("title"), 500).strip()
    venue = text_only(row.get("venue"), 300).strip()
    year = text_only(row.get("year"), 40).strip()
    volume = text_only(row.get("volume"), 80).strip()
    issue = text_only(row.get("issue"), 80).strip()
    pages = text_only(row.get("pages"), 120).strip()
    doi = display_doi(row.get("doi"))
    pub_type = text_only(row.get("publication_type"), 80).strip()
    marker = publication_marker(pub_type)
    citation_title = sentence_case_title(title)
    is_article = is_article_number(pages)
    pages_display = normalize_page_range(pages)
    pages_en_dash = pages_display.replace("-", "–")
    year_part = year or "n.d."

    gbt_authors = format_gbt_authors(authors)
    gbt_source = ", ".join(part for part in [venue, year] if part)
    gbt_vol_issue = volume + (f"({issue})" if issue else "") if volume else (f"({issue})" if issue else "")
    if gbt_vol_issue or pages_display:
        gbt_source = f"{gbt_source}, {gbt_vol_issue}: {pages_display}" if gbt_source and gbt_vol_issue and pages_display else f"{gbt_source}, {gbt_vol_issue}" if gbt_source and gbt_vol_issue else f"{gbt_source}: {pages_display}" if gbt_source and pages_display else gbt_vol_issue or pages_display
    gbt = sentence_join(f"{gbt_authors}. {citation_title}{marker}. {gbt_source}.", f"DOI:{doi}." if doi else "")

    ieee_authors = format_ieee_authors(authors)
    ieee_parts = [venue, f"vol. {volume}" if volume else "", f"no. {issue}" if issue else ""]
    ieee_parts.append(f"Art. no. {pages_display}" if is_article and pages_display else f"pp. {pages_display}" if pages_display else "")
    ieee_parts.append(year_part)
    ieee_middle = ", ".join(part for part in ieee_parts if part)
    ieee = sentence_join(f'{ieee_authors}, "{citation_title}," {ieee_middle},' if ieee_middle else f'{ieee_authors}, "{citation_title},"', f"doi: {doi}." if doi else "")

    elsevier_authors = format_elsevier_authors(authors)
    elsevier_volume = volume + (f" ({issue})" if issue else "") if volume else (f"({issue})" if issue else "")
    elsevier_source_parts = [venue, elsevier_volume, f"({year_part})", pages_en_dash]
    elsevier_source = " ".join(part for part in elsevier_source_parts if part)
    elsevier = sentence_join(f"{elsevier_authors}, {citation_title}, {elsevier_source}.", doi_url(doi) + "." if doi else "")

    apa_authors = format_apa_authors(authors)
    apa_volume = f"*{volume}*" + (f"({issue})" if issue else "") if volume else (f"({issue})" if issue else "")
    apa_tail = f"Article {pages_display}." if is_article and pages_display else f"{pages_en_dash}." if pages_display else ""
    apa_source = ", ".join(part for part in [f"*{venue}*" if venue else "", apa_volume, apa_tail] if part)
    apa = sentence_join(f"{apa_authors} ({year_part}). {citation_title}. {apa_source}", doi_url(doi) if doi else "", final_period=False)

    bibtex = generated_bibtex(row, normalize_authors(row.get("authors")), title, venue, year, volume, issue, pages, doi, pub_type)
    return {"gbt": cleanup_citation(gbt), "elsevier": cleanup_citation(elsevier), "apa": cleanup_citation(apa, strip_final_doi_period=True), "ieee": cleanup_citation(ieee), "bibtex": bibtex}


def split_authors(value: Any) -> list[str]:
    text = text_only(value, 1200).strip()
    if not text:
        return ["Unknown"]
    text = text.replace("\r", "\n").replace("；", ";")
    separators = [";", "\n"]
    if ";" not in text and "\n" not in text:
        if " and " in text:
            return [part.strip(" ,") for part in text.split(" and ") if part.strip(" ,")]
        if "，" in text:
            return [part.strip(" ，,") for part in text.split("，") if part.strip(" ，,")]
    for sep in separators:
        text = text.replace(sep, ";")
    return [part.strip(" ，,") for part in text.split(";") if part.strip(" ，,")] or ["Unknown"]


def normalize_authors(value: Any) -> str:
    return "; ".join(split_authors(value))


def has_cjk(text: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in text)


def name_parts(author: str) -> tuple[str, list[str]]:
    author = " ".join(author.replace(".", ". ").split())
    if "," in author:
        surname, given = author.split(",", 1)
        return surname.strip(), [part.strip(".") for part in given.split() if part.strip(" .")]
    parts = [part.strip(".") for part in author.split() if part.strip(" .")]
    if len(parts) <= 1:
        return author.strip(), []
    return parts[-1], parts[:-1]


def initials(parts: list[str], dotted: bool = True) -> str:
    letters = [part[0].upper() for part in parts if part]
    if not letters:
        return ""
    return " ".join(f"{letter}." for letter in letters) if dotted else "".join(letters)


def format_ieee_author(author: str) -> str:
    if has_cjk(author) or author == "Unknown":
        return author
    surname, given = name_parts(author)
    prefix = initials(given)
    return f"{prefix} {surname}".strip()


def format_ieee_authors(authors: list[str]) -> str:
    formatted = [format_ieee_author(author) for author in authors]
    if len(formatted) <= 1:
        return formatted[0] if formatted else "Unknown"
    return ", ".join(formatted[:-1]) + f" and {formatted[-1]}"


def format_elsevier_authors(authors: list[str]) -> str:
    return ", ".join(format_ieee_author(author) for author in authors)


def format_apa_author(author: str) -> str:
    if has_cjk(author) or author == "Unknown":
        return author
    surname, given = name_parts(author)
    prefix = initials(given)
    return f"{surname}, {prefix}".strip(" ,")


def format_apa_authors(authors: list[str]) -> str:
    formatted = [format_apa_author(author) for author in authors]
    if len(formatted) <= 1:
        return formatted[0] if formatted else "Unknown"
    if len(formatted) == 2:
        return f"{formatted[0]}, & {formatted[1]}"
    return ", ".join(formatted[:-1]) + f", & {formatted[-1]}"


def format_gbt_author(author: str) -> str:
    if has_cjk(author) or author == "Unknown":
        return author
    surname, given = name_parts(author)
    return f"{surname} {initials(given, dotted=False)}".strip()


def format_gbt_authors(authors: list[str]) -> str:
    shown = authors[:3]
    suffix = "等" if len(authors) > 3 and any(has_cjk(author) for author in authors) else "et al." if len(authors) > 3 else ""
    text = ", ".join(format_gbt_author(author) for author in shown)
    return f"{text}, {suffix}" if suffix else text


def sentence_case_title(title: str) -> str:
    text = title.strip().strip(".")
    if not text:
        return ""
    return text[0].upper() + text[1:]


def normalize_page_range(pages: str) -> str:
    return text_only(pages, 120).strip().replace("–", "-").replace("—", "-")


def is_article_number(pages: str) -> bool:
    value = normalize_page_range(pages)
    return bool(value and "-" not in value)


def doi_url(doi: str) -> str:
    value = display_doi(doi)
    return f"https://doi.org/{value}" if value else ""


def display_doi(value: Any) -> str:
    text = text_only(value, 200).strip()
    text = text.removeprefix("https://doi.org/").removeprefix("http://doi.org/")
    if text.lower().startswith("doi:"):
        text = text[4:]
    return text.strip().rstrip(".,;")


def sentence_join(*parts: str, final_period: bool = True) -> str:
    text = " ".join(part.strip() for part in parts if part and part.strip())
    text = text.replace(" ,", ",").replace(" .", ".")
    if final_period and text and not text.endswith("."):
        text += "."
    return text


def cleanup_citation(text: str, strip_final_doi_period: bool = False) -> str:
    while "  " in text:
        text = text.replace("  ", " ")
    text = text.replace("..", ".").replace("., .", ".").replace(", .", ".").replace(",.", ".").replace(" ,", ",")
    text = text.strip(" ,")
    if strip_final_doi_period and "https://doi.org/" in text and text.endswith("."):
        text = text[:-1]
    return text.strip()


def publication_marker(pub_type: str) -> str:
    if "会议" in pub_type or "conference" in pub_type.lower():
        return "[C]"
    if "书" in pub_type or "book" in pub_type.lower():
        return "[M]"
    if "专利" in pub_type:
        return "[P]"
    return "[J]" if pub_type else ""


def generated_bibtex(row: dict[str, Any], authors: str, title: str, venue: str, year: str, volume: str, issue: str, pages: str, doi: str, pub_type: str) -> str:
    entry_type = "inproceedings" if ("会议" in pub_type or "conference" in pub_type.lower()) else "article"
    key_seed = text_only(row.get("uid") or title or "publication", 120).replace(" ", "_")
    author_value = " and ".join(part.strip() for part in authors.split(";") if part.strip())
    fields = [
        ("title", title),
        ("author", author_value),
        ("journal" if entry_type == "article" else "booktitle", venue),
        ("year", year),
        ("volume", volume),
        ("number", issue),
        ("pages", pages),
        ("doi", doi),
    ]
    body = ",\n".join(f"  {name} = {{{value}}}" for name, value in fields if value)
    return f"@{entry_type}{{{key_seed},\n{body}\n}}"


def citation_data_attrs(citations: dict[str, str]) -> str:
    return " ".join(f'data-{esc(style)}="{esc(value)}"' for style, value in citations.items())


def citation_style_options(selected: str, env: dict[str, str] | None = None) -> str:
    env = env or {}
    labels = [("gbt", "GB/T"), ("elsevier", "Elsevier"), ("apa", "APA"), ("ieee", "IEEE"), ("bibtex", "BibTeX"), ("source", t(env, "citation_source"))]
    selected = selected if selected in {key for key, _label in labels} else "gbt"
    return "".join(f'<option value="{esc(key)}" {"selected" if key == selected else ""}>{esc(label)}</option>' for key, label in labels)


def publication_display_citation(row: dict[str, Any], citations: dict[str, str], display_style: str) -> str:
    style = display_style if display_style in {"gbt", "elsevier", "apa", "ieee", "bibtex", "source"} else "gbt"
    if style == "source":
        source = text_only(row.get("source_citation"), 3000).strip()
        if source:
            return source
    return citations.get(style) or citations.get("gbt") or ""


def publication_source(row: dict[str, Any]) -> str:
    parts = []
    if row.get("venue"):
        parts.append(f'<span class="pub-venue">{esc(row.get("venue"))}</span>')
    details = []
    for label, key in [("年", "year"), ("卷", "volume"), ("期", "issue"), ("页", "pages")]:
        value = row.get(key)
        if value not in (None, ""):
            details.append(f'{label} {value}')
    if row.get("doi"):
        details.append(f'DOI {row.get("doi")}')
    if details:
        parts.append(f'<span class="pub-source-detail">{esc(" / ".join(details))}</span>')
    return " ".join(parts)


def publication_tags(row: dict[str, Any]) -> str:
    tags = []
    custom = text_only(row.get("display_tags"), 500).strip()
    if custom:
        raw_tags = [part.strip() for part in custom.replace("；", ";").replace("，", ",").replace("\n", ",").replace(";", ",").split(",")]
        for value in raw_tags:
            if value:
                tags.append(f'<span class="publication-tag">{esc(value[:40])}</span>')
    else:
        for key in ["publication_type", "index_type"]:
            value = text_only(row.get(key), 80).strip()
            if value:
                tags.append(f'<span class="publication-tag">{esc(value)}</span>')
    return '<div class="publication-tags">' + "".join(tags) + "</div>" if tags else ""


def home_publication_list(rows: list[dict[str, Any]], display_style: str, repo: Repository, env: dict[str, str]) -> str:
    items = []
    total = len(rows)
    for index, row in enumerate(rows, 1):
        display_row = front_row(repo, env, "publications", row)
        citations = publication_citations(display_row)
        number = total - index + 1
        citation = publication_display_citation(display_row, citations, display_style)
        tags = publication_tags(display_row)
        items.append(f"""<article class="home-pub-item">
          <span class="home-item-number">{number}</span>
          <div class="home-pub-body">
            <p>{esc(citation)}</p>
            {tags}
          </div>
        </article>""")
    return f'<section class="home-pub-list">{"".join(items) or empty(env)}</section>'


def home_project_items(rows: list[dict[str, Any]], repo: Repository, env: dict[str, str]) -> str:
    parts = []
    for row in rows:
        display_row = front_row(repo, env, "projects", row)
        source = text_only(display_row.get("source"), 120).strip()
        fund = text_only(display_row.get("fund_name"), 140).strip()
        funding = " - ".join(item for item in (source, fund) if item)
        details = home_inline_facts([
            (t(env, "project_period"), front_project_period(row.get("start_date"), row.get("end_date"), env)),
            (t(env, "project_number"), row.get("project_number")),
            (t(env, "project_amount"), front_project_amount(row.get("amount"), env)),
            (t(env, "status"), display_row.get("status")),
        ])
        parts.append(f"""<article class="home-mini-card home-project-card">
          {f'<div class="home-mini-kicker">{esc(funding)}</div>' if funding else ''}
          <strong>{esc(display_row.get("name"))}</strong>
          {details}
        </article>""")
    return f'<div class="home-mini-list">{"".join(parts) or empty(env)}</div>'


def home_patent_items(rows: list[dict[str, Any]], repo: Repository, env: dict[str, str]) -> str:
    parts = []
    for row in rows:
        display_row = front_row(repo, env, "patents", row)
        tags = patent_tags(display_row)
        details = home_inline_facts([
            (t(env, "inventors"), display_row.get("inventors")),
            (t(env, "owner"), display_row.get("owner")),
            (t(env, "application_number"), row.get("application_number")),
            (t(env, "grant_number"), row.get("grant_number")),
        ])
        parts.append(f"""<article class="home-mini-card home-patent-card">
          <div class="home-mini-title-row"><strong>{esc(display_row.get("name"))}</strong>{tags}</div>
          {details}
        </article>""")
    return f'<div class="home-mini-list">{"".join(parts) or empty(env)}</div>'


def home_student_items(rows: list[dict[str, Any]], repo: Repository, env: dict[str, str]) -> str:
    parts = []
    lang = current_lang(env)
    for row in rows:
        display_row = front_row(repo, env, "students", row)
        name = front_value(repo, env, "students", row, "name", 500) or localized_name(row, lang)
        details = home_inline_facts([
            (t(env, "degree"), display_row.get("degree")),
            (t(env, "grade"), display_row.get("grade")),
            (t(env, "status"), display_row.get("status")),
            (t(env, "direction"), display_row.get("direction")),
        ])
        summary = text_only(display_row.get("bio") or display_row.get("destination"), 180).strip()
        parts.append(f"""<article class="home-mini-card home-student-card">
          <strong>{esc(name)}</strong>
          {details}
          {f'<p class="home-mini-summary">{esc(summary)}</p>' if summary else ''}
        </article>""")
    return f'<div class="home-mini-list">{"".join(parts) or empty(env)}</div>'


def home_inline_facts(items: list[tuple[str, Any]]) -> str:
    chips = []
    for label, value in items:
        text = text_only(value, 220).strip()
        if text:
            chips.append(f'<span><em>{esc(label)}</em>{esc(text)}</span>')
    return f'<div class="home-mini-facts">{"".join(chips)}</div>' if chips else ""


def simple_items(rows: list[dict[str, Any]], title_field: str, meta_field: str, repo: Repository | None = None, env: dict[str, str] | None = None, table: str = "") -> str:
    parts = []
    for row in rows:
        display_row = front_row(repo, env or {}, table, row) if repo and env and table else row
        parts.append(f'<p><strong>{esc(display_row.get(title_field))}</strong><span>{esc(display_row.get(meta_field))}</span></p>')
    return '<div class="mini-list">' + "".join(parts) + "</div>"


def news_list(rows: list[dict[str, Any]], detail: bool = False, repo: Repository | None = None, env: dict[str, str] | None = None) -> str:
    items = []
    for index, row in enumerate(rows, 1):
        display_row = front_row(repo, env or {}, "news", row) if repo and env else row
        href = f'/news/{safe_slug(str(row.get("slug") or row.get("title") or ""))}'
        if env:
            href = lang_url(href, env)
        if detail:
            summary = f'<p class="news-excerpt">{esc(text_only(display_row.get("content"), 180))}</p>'
            cover = '<div class="news-cover placeholder">N</div>'
            items.append(f'<article class="news-row"><span class="item-number">{index}</span>{cover}<div class="news-content"><div class="news-meta-line"><time>{esc(row.get("published_at"))}</time><span>{esc(display_row.get("category"))}</span></div><h2><a href="{esc(href)}">{esc(display_row.get("title"))}</a></h2>{summary}</div></article>')
        else:
            items.append(f'<article class="news-item"><time>{esc(row.get("published_at"))}</time><a href="{esc(href)}">{esc(display_row.get("title"))}</a></article>')
    class_name = "news-list compact-news-list" if detail else "news-list"
    return f'<section class="{class_name}">' + ("".join(items) or empty(env or {})) + "</section>"


def compact_filter_form(query: dict[str, str], placeholder: str, filter_groups: list[tuple[str, str, list[Any]]], env: dict[str, str] | None = None) -> str:
    env = env or {}
    reset_href = "?lang=en" if current_lang(env) == "en" else "?"
    return f"""<form class="filters compact-filterbar" method="get">
      {lang_hidden(env)}
      <input class="filter-search" name="q" value="{esc(query.get("q", ""))}" placeholder="{esc(placeholder)}">
      {filter_selects(query, filter_groups)}
      <button>{esc(t(env, "search"))}</button>
      <a class="button ghost filter-reset" href="{esc(reset_href)}">{esc(t(env, "reset"))}</a>
    </form>"""


def lang_hidden(env: dict[str, str]) -> str:
    return '<input type="hidden" name="lang" value="en">' if current_lang(env) == "en" else ""


def filter_selects(query: dict[str, str], filter_groups: list[tuple[str, str, list[Any]]]) -> str:
    controls = []
    for name, label, values in filter_groups:
        selected = text_only(query.get(name, ""), 200).strip()
        if not values and not selected:
            continue
        option_values = list(values)
        if selected:
            raw_values = [str(item[0]) if isinstance(item, (tuple, list)) and item else str(item) for item in option_values]
            if selected not in raw_values:
                option_values.insert(0, (selected, selected))
        style = filter_select_style(label, option_values, selected)
        title = filter_select_title(label, option_values)
        controls.append(f'<select class="filter-select filter-select-{esc(name)}" name="{esc(name)}" style="{style}" title="{esc(title)}"><option value="">{esc(label)}</option>{select_options(option_values, selected)}</select>')
    return "".join(controls)


def filter_select_style(label: str, values: list[Any], selected: str = "") -> str:
    pairs = [option_pair(item) for item in values]
    selected_label = next((item_label for item_value, item_label in pairs if str(item_value) == str(selected)), "")
    closed_text = selected_label or label
    max_label = max([label, closed_text] + [item_label for _item_value, item_label in pairs], key=visual_text_units, default=label)
    closed_em = max(4.8, min(12.8, visual_text_units(closed_text) * 0.54 + 2.2))
    option_em = max(closed_em + 1.0, min(26.0, visual_text_units(max_label) * 0.54 + 3.6))
    return f"--filter-select-width:{closed_em:.1f}em;--filter-option-width:{option_em:.1f}em;"


def filter_select_title(label: str, values: list[Any]) -> str:
    labels = [option_pair(item)[1] for item in values]
    longest = max(labels, key=visual_text_units, default="")
    return f"{label}: {longest}" if longest else label


def option_pair(item: Any) -> tuple[str, str]:
    if isinstance(item, (tuple, list)) and item:
        value = str(item[0])
        label = str(item[1] if len(item) > 1 else item[0])
        return value, label
    value = str(item)
    return value, value


def visual_text_units(value: Any) -> float:
    units = 0.0
    for char in text_only(value, 300):
        code = ord(char)
        if char.isspace():
            units += 0.45
        elif code <= 0x007F:
            units += 0.62
        elif 0xFF00 <= code <= 0xFFEF:
            units += 1.0
        elif 0x4E00 <= code <= 0x9FFF or 0x3040 <= code <= 0x30FF or 0xAC00 <= code <= 0xD7AF:
            units += 1.0
        else:
            units += 0.86
    return max(units, 1.0)


def filter_options(rows: list[dict[str, Any]], specs: list[tuple[str, str]], repo: Repository | None = None, env: dict[str, str] | None = None, table: str = "") -> list[tuple[str, str, list[Any]]]:
    groups = []
    for name, label in specs:
        values = sorted(
            {text_only(row.get(name), 120).strip() for row in rows if text_only(row.get(name), 120).strip()},
            key=lambda value: filter_option_sort_key(name, value),
        )
        if repo and env and table and current_lang(env) == "en":
            options_with_labels = []
            for value in values[:80]:
                sample = next((row for row in rows if text_only(row.get(name), 120).strip() == value), {})
                display = front_value(repo, env, table, sample, name, 300) if sample else ""
                fallback = localized_option_label(env, value)
                options_with_labels.append((value, fallback if display == value and fallback != value else display or fallback))
            groups.append((name, label, options_with_labels))
        else:
            groups.append((name, label, values[:80]))
    return groups


def filter_option_sort_key(field: str, value: str) -> tuple[Any, ...]:
    text = text_only(value, 200).strip()
    folded = text.casefold()
    if field in {"year", "grade"} or "year" in field or "date" in field:
        number = first_int(text)
        return (0, -number, folded) if number is not None else (1, folded)
    if field in {"degree", "education", "education_level", "academic_degree", "category"}:
        return (degree_rank(text), folded)
    if field in {"title", "role"}:
        return (academic_title_rank(text), folded)
    if field in {"author_role"}:
        order = {"first": 0, "corresponding": 1, "other": 2}
        return (order.get(folded, 50), folded)
    if field in {"status", "legal_status"}:
        return (status_rank(text), folded)
    return (0, folded)


def first_int(value: str) -> int | None:
    digits = ""
    for char in value:
        if char.isdigit():
            digits += char
        elif digits:
            break
    return int(digits) if digits else None


def degree_rank(value: str) -> int:
    text = text_only(value, 200).strip().casefold()
    rules = [
        (0, ("postdoc", "post-doc", "博士后")),
        (1, ("phd", "ph.d", "doctor", "doctoral", "博士")),
        (2, ("master", "硕士", "研究生")),
        (3, ("bachelor", "undergraduate", "本科", "学士")),
        (4, ("associate", "专科")),
        (5, ("alumni", "graduate", "毕业")),
    ]
    for rank, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return rank
    return 50


def academic_title_rank(value: str) -> int:
    text = text_only(value, 200).strip().casefold()
    rules = [
        (0, ("academician", "院士")),
        (2, ("associate professor", "副教授")),
        (3, ("assistant professor", "讲师", "助理教授")),
        (1, ("professor", "教授")),
        (4, ("postdoc", "博士后")),
        (5, ("phd", "doctor", "博士")),
        (6, ("master", "硕士")),
        (7, ("undergraduate", "本科")),
    ]
    for rank, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return rank
    return 50


def status_rank(value: str) -> int:
    text = text_only(value, 200).strip().casefold()
    rules = [
        (0, ("active", "ongoing", "在研", "进行", "有效", "授权")),
        (1, ("pending", "submitted", "申请", "受理", "待")),
        (2, ("completed", "closed", "finished", "结题", "完成", "毕业")),
        (3, ("inactive", "hidden", "archived", "停用", "隐藏", "归档")),
    ]
    for rank, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return rank
    return 50


def query_filters(query: dict[str, str], names: list[str]) -> dict[str, str]:
    return {name: query[name] for name in names if query.get(name)}


def options(values: list[str], selected: str) -> str:
    return "".join(f'<option value="{esc(value)}" {"selected" if value == selected else ""}>{esc(value or "未设置")}</option>' for value in values)


def select_options(values: list[Any], selected: str) -> str:
    parts = []
    for item in values:
        if isinstance(item, (tuple, list)) and item:
            value = str(item[0])
            label = str(item[1] if len(item) > 1 else item[0])
        else:
            value = str(item)
            label = value
        parts.append(f'<option value="{esc(value)}" {"selected" if value == selected else ""}>{esc(label)}</option>')
    return "".join(parts)


def localized_option_label(env: dict[str, str], value: Any) -> str:
    text = text_only(value, 200).strip()
    if current_lang(env) != "en":
        return text
    labels = {
        "first": "First Author",
        "corresponding": "Corresponding Author",
        "other": "Other",
        "public": "Public",
        "authenticated": "Authenticated",
        "staff": "Staff",
        "owner": "Owner",
        "hidden": "Hidden",
        "new": "New",
        "read": "Read",
        "replied": "Replied",
        "archived": "Archived",
        "active": "Active",
        "trash": "Trash",
    }
    return labels.get(text, text)


def empty(env: dict[str, str] | None = None) -> str:
    return f'<p class="empty">{esc(t(env or {}, "empty"))}</p>'


def robots_txt(site_url: str) -> str:
    sitemap = site_absolute_url(site_url, "/sitemap.xml")
    sitemap_index = site_absolute_url(site_url, "/sitemap-index.xml")
    lines = [
        "User-agent: *",
        "Allow: /",
        "Allow: /assets/",
        "Allow: /media/",
        "Disallow: /admin",
        "Disallow: /api",
        "Disallow: /api/",
        "Disallow: /login",
        "Disallow: /register",
        "Disallow: /logout",
        "Disallow: /exports/",
        "Disallow: /data/",
        "Disallow: /.cache/",
        "Disallow: /migrations/",
        "Disallow: /seed/",
        "Disallow: /src/",
        "Disallow: /tools/",
        "Disallow: /*?next=",
        "Disallow: /*?token=",
        "Disallow: /*?preview=",
        "",
        f"Sitemap: {sitemap}",
        f"Sitemap: {sitemap_index}",
        "",
    ]
    return "\n".join(lines)


def security_txt(site_url: str) -> str:
    contact = site_absolute_url(site_url, "/contact")
    canonical = site_absolute_url(site_url, "/.well-known/security.txt")
    expires = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime(time.time() + 365 * 24 * 60 * 60))
    lines = [
        f"Contact: {contact}",
        f"Expires: {expires}",
        "Preferred-Languages: zh, en",
        f"Canonical: {canonical}",
        "",
        "# This file is for responsible security reporting.",
        "# Do not include passwords, tokens, database files, exports, or private data in reports unless explicitly requested.",
        "",
    ]
    return "\n".join(lines)


def llms_txt(repo: Repository, site_url: str) -> str:
    site = active_site(repo)
    site_name = text_only(site.get("site_name_en") or site.get("site_name"), 200).strip() or "Academic Website"
    description = text_only(site.get("seo_description") or site.get("hero_subtitle"), 500).strip()
    public_links = "\n".join(f"- {site_absolute_url(site_url, path)}" for path in public_sitemap_paths(repo))
    return "\n".join(
        [
            f"# {site_name}",
            "",
            description or "Teacher academic website with public profile, research, publications, projects, patents, students, courses, news, and contact information.",
            "",
            "## Public Content",
            public_links,
            "",
            "## Machine-Readable Indexes",
            f"- Sitemap XML: {site_absolute_url(site_url, '/sitemap.xml')}",
            f"- Sitemap Text: {site_absolute_url(site_url, '/sitemap.txt')}",
            "",
            "## Access Boundaries",
            "- Use only public front-end pages and public media referenced by those pages.",
            "- Do not crawl admin, API, login, registration, export, cache, database, migration, source, seed, or tooling paths.",
            "",
        ]
    )


def sitemap_index_xml(site_url: str) -> str:
    sitemap = site_absolute_url(site_url, "/sitemap.xml")
    lastmod = time.strftime("%Y-%m-%d", time.gmtime())
    return f'<?xml version="1.0" encoding="UTF-8"?><sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9"><sitemap><loc>{esc(sitemap)}</loc><lastmod>{lastmod}</lastmod></sitemap></sitemapindex>'


def sitemap_txt(repo: Repository, site_url: str) -> str:
    lines: list[str] = []
    for path in public_sitemap_paths(repo):
        lines.append(site_absolute_url(site_url, path))
        lines.append(site_absolute_url(site_url, english_path(path)))
    return "\n".join(lines) + "\n"


def public_sitemap_paths(repo: Repository) -> list[str]:
    urls = ["/", "/team", "/publications", "/featured-publications", "/projects", "/patents", "/students", "/news", "/courses", "/contact"]
    urls += [f'/team/{safe_slug(str(row.get("uid") or row.get("id") or ""))}' for row in repo.list("profiles", Query(public_only=True, limit=500)) if row.get("uid") or row.get("id")]
    urls += [f'/news/{safe_slug(str(row.get("slug") or row.get("title") or ""))}' for row in repo.list("news", Query(public_only=True, limit=500)) if row.get("slug") or row.get("title")]
    unique_urls: list[str] = []
    for url in urls:
        if url and url not in unique_urls and not private_index_path(url):
            unique_urls.append(url)
    return unique_urls


def private_index_path(path: str) -> bool:
    clean = "/" + str(path or "").lstrip("/")
    blocked_prefixes = ("/admin", "/api/", "/login", "/register", "/logout", "/exports/", "/data/", "/.cache/", "/migrations/", "/seed/", "/src/", "/tools/")
    return clean.startswith(blocked_prefixes)


def sitemap_xml(repo: Repository, site_url: str) -> str:
    unique_urls = public_sitemap_paths(repo)
    entries = []
    for path in unique_urls:
        zh_url = site_absolute_url(site_url, path)
        en_url = site_absolute_url(site_url, english_path(path))
        alternates = (
            f'<xhtml:link rel="alternate" hreflang="zh-CN" href="{esc(zh_url)}"/>'
            f'<xhtml:link rel="alternate" hreflang="en" href="{esc(en_url)}"/>'
            f'<xhtml:link rel="alternate" hreflang="x-default" href="{esc(zh_url)}"/>'
        )
        entries.append(f"<url><loc>{esc(zh_url)}</loc>{alternates}</url>")
        entries.append(f"<url><loc>{esc(en_url)}</loc>{alternates}</url>")
    body = "".join(entries)
    return f'<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" xmlns:xhtml="http://www.w3.org/1999/xhtml">{body}</urlset>'


def html_response(content: str, status: int = 200) -> ResponseTuple:
    return status, security_headers() + [("content-type", "text/html; charset=utf-8")], content.encode("utf-8")


def text_response(content: str, status: int = 200) -> ResponseTuple:
    return status, security_headers() + [("content-type", "text/plain; charset=utf-8")], content.encode("utf-8")


def xml_response(content: str, status: int = 200) -> ResponseTuple:
    return status, security_headers() + [("content-type", "application/xml; charset=utf-8")], content.encode("utf-8")


def json_response(data: Any, status: int = 200) -> ResponseTuple:
    return status, security_headers() + [("content-type", "application/json; charset=utf-8")], json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def media_json_response(data: Any, status: int = 200, refresh: bool = False) -> ResponseTuple:
    cache_control = "no-store" if refresh else f"private, max-age={MEDIA_STATS_CACHE_TTL_SECONDS}, stale-while-revalidate={MEDIA_STATS_CACHE_TTL_SECONDS}"
    return status, security_headers() + [
        ("content-type", "application/json; charset=utf-8"),
        ("cache-control", cache_control),
    ], json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def redirect(location: str) -> ResponseTuple:
    return 302, security_headers() + [("location", location)], b""


def security_headers() -> list[tuple[str, str]]:
    return [
        ("x-content-type-options", "nosniff"),
        ("x-frame-options", "DENY"),
        ("referrer-policy", "strict-origin-when-cross-origin"),
        ("permissions-policy", "camera=(), microphone=(), geolocation=(), payment=()"),
        (
            "content-security-policy",
            "default-src 'self'; img-src 'self' data: https://api.iconify.design; media-src 'self'; style-src 'self' 'unsafe-inline'; script-src 'self'; connect-src 'self'; object-src 'none'; frame-src 'none'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'",
        ),
    ]


def _query(query_string: str) -> dict[str, str]:
    parsed = parse_qs(query_string, keep_blank_values=True)
    return {key: values[-1] if values else "" for key, values in parsed.items()}


def _form(body: bytes) -> dict[str, str]:
    return _query(body.decode("utf-8", "ignore"))


def _form_multi(body: bytes) -> dict[str, list[str]]:
    return parse_qs(body.decode("utf-8", "ignore"), keep_blank_values=True)
